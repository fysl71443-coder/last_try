import json
import os
from datetime import datetime, date, timedelta
from datetime import date as _date
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_from_directory, send_file
from flask_login import login_user, logout_user, login_required, current_user
from sqlalchemy import func, inspect, text, or_
from sqlalchemy.exc import IntegrityError

from app import db, csrf
from app import __init__ as app_init  # for template globals, including can()
ext_db = None
from app.models import AppKV, TableLayout
from models import User
from models import OrderInvoice
from models import MenuCategory, MenuItem, SalesInvoice, SalesInvoiceItem, Customer, PurchaseInvoice, PurchaseInvoiceItem, ExpenseInvoice, ExpenseInvoiceItem, Settings, Meal, MealIngredient, RawMaterial, Supplier, Employee, Salary, Payment, EmployeeSalaryDefault, DepartmentRate, EmployeeHours, Account, LedgerEntry, JournalEntry, JournalLine, JournalAudit
from models import get_saudi_now, KSA_TZ
from forms import SalesInvoiceForm, EmployeeForm, ExpenseInvoiceForm, PurchaseInvoiceForm, MealForm, RawMaterialForm

main = Blueprint('main', __name__)
@main.before_app_request
def _block_employee_screens():
    try:
        p = (request.path or '').lower()
    except Exception:
        p = ''
    legacy_prefixes = ['/employees/payroll', '/employees/payroll/detailed', '/salaries', '/salary', '/print/salary']
    if any(p.startswith(pref) for pref in legacy_prefixes):
        return ('', 404)
# --- Employees: Detailed Payroll Report (across months with totals) ---
@main.route('/employees/payroll/detailed', methods=['GET'], endpoint='payroll_detailed')
@login_required
def payroll_detailed():
    return ('', 404)

# --- Unified Visual Dashboard for Employees ---
@main.route('/employee-uvd', methods=['GET'], endpoint='employee_uvd')
@login_required
def employee_uvd():
    try:
        today = get_saudi_now().date()
        year = today.year
        month = today.month
        emp_count = int(db.session.query(func.coalesce(func.count(Employee.id), 0)).scalar() or 0)
        sal_sum = float(db.session.query(func.coalesce(func.sum(Salary.total_salary), 0)).
                        filter(Salary.year == year, Salary.month == month).scalar() or 0)
        last_pay = db.session.query(Payment.payment_date).filter(Payment.invoice_type == 'salary').order_by(Payment.payment_date.desc()).first()
        last_pay_dt = (last_pay[0].date() if last_pay and last_pay[0] else None)
        je_cnt = int(db.session.query(func.coalesce(func.count(JournalEntry.id), 0)).
                     filter(func.extract('year', JournalEntry.date) == year,
                            func.extract('month', JournalEntry.date) == month).scalar() or 0)
        adv_acc = Account.query.filter_by(code='1030').first()
        adv_out = 0.0
        if adv_acc:
            rows = db.session.query(func.coalesce(func.sum(LedgerEntry.debit), 0), func.coalesce(func.sum(LedgerEntry.credit), 0)).filter(LedgerEntry.account_id == adv_acc.id).first()
            d_sum = float(rows[0] or 0)
            c_sum = float(rows[1] or 0)
            adv_out = max(d_sum - c_sum, 0.0)
    except Exception:
        emp_count = 0
        sal_sum = 0.0
        adv_out = 0.0
        last_pay_dt = None
        je_cnt = 0

    try:
        employees = Employee.query.order_by(Employee.full_name.asc()).limit(200).all()
    except Exception:
        employees = []
    try:
        prev_salaries = Salary.query.order_by(Salary.year.desc(), Salary.month.desc()).limit(20).all()
    except Exception:
        prev_salaries = []
    try:
        advances = db.session.query(JournalLine).join(Account).\
            filter(Account.code == '1030').order_by(JournalLine.line_date.desc()).limit(20).all()
    except Exception:
        advances = []
    try:
        ledgers = JournalEntry.query.order_by(JournalEntry.date.desc()).limit(20).all()
    except Exception:
        ledgers = []

    try:
        dept_counts = {}
        for e in employees:
            k = (getattr(e, 'department', '') or '').strip().lower()
            if not k:
                k = 'other'
            dept_counts[k] = dept_counts.get(k, 0) + 1
    except Exception:
        dept_counts = {}

    emp_metrics = {}
    try:
        from models import EmployeeSalaryDefault, Salary, JournalLine, Account
        from sqlalchemy import and_
        sdefs = EmployeeSalaryDefault.query.filter(EmployeeSalaryDefault.employee_id.in_([int(getattr(e,'id',0) or 0) for e in employees])).all()
        base_map = {int(d.employee_id): float(d.base_salary or 0) for d in sdefs}
        for e in employees:
            eid = int(getattr(e,'id',0) or 0)
            last_total = 0.0
            try:
                s = Salary.query.filter_by(employee_id=eid).order_by(Salary.year.desc(), Salary.month.desc()).first()
                if s:
                    last_total = float(getattr(s,'total_salary',0) or 0)
            except Exception:
                last_total = 0.0
            adv_total = 0.0
            try:
                adv_acc = Account.query.filter_by(code='1030').first()
                if adv_acc:
                    rows = JournalLine.query.filter_by(employee_id=eid, account_id=int(adv_acc.id)).all()
                    dsum = sum([float(getattr(r,'debit',0) or 0) for r in rows])
                    csum = sum([float(getattr(r,'credit',0) or 0) for r in rows])
                    adv_total = max(dsum - csum, 0.0)
            except Exception:
                adv_total = 0.0
            emp_metrics[eid] = {
                'basic': float(base_map.get(eid, 0.0)),
                'last_salary': float(last_total or 0.0),
                'advance': float(adv_total or 0.0),
            }
    except Exception as ex:
        # Log the error for debugging
        import traceback
        print(f"Error building emp_metrics: {ex}")
        traceback.print_exc()
        emp_metrics = {}

    auto_open_create = (str(request.args.get('mode') or '').lower() == 'create')
    return render_template('employee_uvd.html',
                           emp_count=emp_count,
                           sal_sum=sal_sum,
                           adv_out=adv_out,
                           last_pay_dt=last_pay_dt,
                           je_cnt=je_cnt,
                           employees=employees,
                           prev_salaries=prev_salaries,
                           advances=advances,
                           ledgers=ledgers,
                           dept_counts=dept_counts,
                           emp_metrics=emp_metrics,
                           auto_open_create=auto_open_create)



@main.route('/suppliers/edit/<int:sid>', methods=['GET', 'POST'], endpoint='suppliers_edit')
@login_required
def suppliers_edit(sid):
    supplier = Supplier.query.get_or_404(sid)
    if request.method == 'POST':
        supplier.name = request.form.get('name', supplier.name)
        supplier.contact_person = request.form.get('contact_person', supplier.contact_person)
        supplier.phone = request.form.get('phone', supplier.phone)
        supplier.email = request.form.get('email', supplier.email)
        supplier.tax_number = request.form.get('tax_number', supplier.tax_number)
        supplier.address = request.form.get('address', supplier.address)
        supplier.notes = request.form.get('notes', supplier.notes)
        supplier.active = True if str(request.form.get('active', supplier.active)).lower() in ['1','true','yes','on'] else False
        try:
            db.session.commit()
            flash('✅ Supplier updated successfully', 'success')
            return redirect(url_for('main.suppliers'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating supplier: {e}', 'danger')
    return render_template('supplier_edit.html', supplier=supplier)

# --- Simple helpers / constants ---
BRANCH_LABELS = {
    'place_india': 'Place India',
    'china_town': 'China Town',
}

# --- Helper functions ---
def safe_table_number(table_number) -> int:
    """Safely convert table_number to int, default to 0 if None/invalid"""
    try:
        return int(table_number or 0)
    except (ValueError, TypeError):
        return 0

# --- Permission helpers (reuse AppKV like template can()) ---
def _normalize_scope(s: str) -> str:
    s = (s or '').strip().lower()
    if s in ('place','palace','india','palace_india'): return 'place_india'
    if s in ('china','china town','chinatown'): return 'china_town'
    return s or 'all'


def _read_user_perms(uid: int, scope: str):
    try:
        k = f"user_perms:{scope}:{int(uid)}"
        row = AppKV.query.filter_by(k=k).first()
        if not row:
            return {}
        data = json.loads(row.v)
        items = data.get('items') or []
        out = {}
        for it in items:
            key = (it.get('screen_key') or '').strip()
            if not key:
                continue
            out[key] = {
                'view': bool(it.get('view')),
                'add': bool(it.get('add')),
                'edit': bool(it.get('edit')),
                'delete': bool(it.get('delete')),
                'print': bool(it.get('print')),
            }
        return out
    except Exception:
        return {}


def user_can(screen: str, action: str = 'view', branch_scope: str = None) -> bool:
    try:
        if not getattr(current_user, 'is_authenticated', False):
            return False
        # Admin bypass
        if getattr(current_user, 'username', '') == 'admin' or getattr(current_user, 'id', None) == 1:
            return True
        if getattr(current_user, 'role', '') == 'admin':
            return True
        scopes = []
        if branch_scope:
            scopes = [_normalize_scope(branch_scope), 'all']
        else:
            scopes = ['all']
        for sc in scopes:
            perms = _read_user_perms(current_user.id, sc)
            scr = perms.get(screen)
            if scr and scr.get(action):
                return True
        return False
    except Exception:
        return False



# Safe helper: current time in Saudi Arabia timezone
try:
    import pytz as _pytz
except Exception:  # pragma: no cover
    _pytz = None
from datetime import datetime as _dt

# use get_saudi_now from models

# ---------- Table status helpers (transactional & safe for concurrency) ----------
def _set_table_status_concurrent(branch_code: str, table_number: str, status: str) -> None:
    """Set a table's status with a short transaction and row-level lock.
    Creates the row if missing. Safe for concurrent updates under Postgres.
    """
    try:
        from models import Table
        tbl_no = str(table_number)
        # Short, isolated transaction
        with db.session.begin():
            # Lock the target row if exists; otherwise create
            t = (db.session.query(Table)
                 .with_for_update(of=Table, nowait=False)
                 .filter_by(branch_code=branch_code, table_number=tbl_no)
                 .first())
            if not t:
                t = Table(branch_code=branch_code, table_number=tbl_no, status=status)
                db.session.add(t)
            else:
                t.status = status
                t.updated_at = get_saudi_now()
    except Exception:
        # Do not propagate table-status errors to main flow
        try: db.session.rollback()
        except Exception: pass

def kv_get(key, default=None):
    rec = AppKV.query.filter_by(k=key).first()
    if not rec:
        return default
    try:
        return json.loads(rec.v)
    except Exception:
        return default

def kv_set(key, value):
    data = json.dumps(value or {})
    rec = AppKV.query.filter_by(k=key).first()
    if rec:
        rec.v = data
    else:
        rec = AppKV(k=key, v=data)
        db.session.add(rec)
    db.session.commit()

@main.route('/api/purchase-categories', methods=['GET'])
def api_purchase_categories():
    try:
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        data_path = os.path.join(base_dir, 'data', 'purchase_categories.json')
        with open(data_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        return jsonify(payload)
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'error': str(e), 'purchase_categories': []}), 500


# Ensure tables exist (safe for first runs) and seed a demo menu if empty
_DEF_MENU = {
    'Starters': [
        ('Spring Rolls', 8.0), ('Samosa', 6.5)
    ],
    'Main Courses': [
        ('Butter Chicken', 18.0), ('Chicken Chow Mein', 16.0)
    ],
    'Biryani': [
        ('Chicken Biryani', 15.0), ('Veg Biryani', 13.0)
    ],
    'Noodles': [
        ('Hakka Noodles', 12.0), ('Singapore Noodles', 13.5)
    ],
    'Drinks': [
        ('Lassi', 5.0), ('Iced Tea', 4.0)
    ],
    'Desserts': [
        ('Gulab Jamun', 6.0), ('Ice Cream', 4.5)
    ],
}

def ensure_menu_sort_order_column():
    """Ensure menu_categories.sort_order exists; add it if missing (safe no-op otherwise)."""
    try:
        insp = inspect(db.engine)
        cols = [c['name'] for c in insp.get_columns('menu_categories')]
        if 'sort_order' not in cols:
            try:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE menu_categories ADD COLUMN sort_order INTEGER DEFAULT 0'))
                    conn.execute(text('UPDATE menu_categories SET sort_order = COALESCE(sort_order, 0)'))
            except Exception:
                # If ALTER fails (e.g., permissions), just continue; fallbacks in queries handle it
                pass
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
def ensure_menuitem_compat_columns():
    """Ensure menu_items has backward-compatible columns (name, price, display_order)."""
    try:
        insp = inspect(db.engine)
        cols = {c['name'] for c in insp.get_columns('menu_items')}
        with db.engine.begin() as conn:
            if 'name' not in cols:
                try:
                    conn.execute(text('ALTER TABLE menu_items ADD COLUMN name VARCHAR(150)'))
                except Exception:
                    pass
            if 'price' not in cols:
                try:
                    conn.execute(text('ALTER TABLE menu_items ADD COLUMN price FLOAT DEFAULT 0.0'))
                except Exception:
                    pass
            if 'display_order' not in cols:
                try:
                    conn.execute(text('ALTER TABLE menu_items ADD COLUMN display_order INTEGER'))
                except Exception:
                    pass
    except Exception:
        # If introspection fails, ignore; query fallbacks may still work
        pass


def ensure_tables():
    try:
        db.create_all()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
    # Ensure new columns exist for compatibility with older DBs
    ensure_menu_sort_order_column()
    ensure_menuitem_compat_columns()


def seed_menu_if_empty():
    try:
        if MenuCategory.query.count() > 0:
            return
        # create categories in a deterministic order
        order = 0
        for cat_name, items in _DEF_MENU.items():
            order += 10
            cat = MenuCategory(name=cat_name, sort_order=order)
            db.session.add(cat)
            db.session.flush()
            for nm, pr in items:
                db.session.add(MenuItem(name=nm, price=float(pr), category_id=cat.id))
        db.session.commit()
        try:
            total_amt = float(inv.total_after_tax_discount or 0.0)
            db.session.add(Payment(
                invoice_id=inv.id,
                invoice_type='sales',
                amount_paid=total_amt,
                payment_method=(payment_method or '').lower() or 'cash',
                payment_date=get_saudi_now()
            ))
            inv.status = 'paid'
            db.session.commit()
        except Exception:
            db.session.rollback()
        try:
            total_amt = float(inv.total_after_tax_discount or 0.0)
            db.session.add(Payment(
                invoice_id=inv.id,
                invoice_type='sales',
                amount_paid=total_amt,
                payment_method=(payment_method or '').lower() or 'cash',
                payment_date=get_saudi_now()
            ))
            inv.status = 'paid'
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception:
        db.session.rollback()
        # ignore seed errors in production path
        pass


# Warmup DB only once per process to avoid heavy create_all/seed on hot paths
_DB_WARMED_UP = False

def warmup_db_once():
    global _DB_WARMED_UP
    if _DB_WARMED_UP:
        return
    try:
        ensure_tables()
        ensure_purchaseinvoice_compat_columns()
        seed_menu_if_empty()
        seed_chart_of_accounts()
    except Exception:
        # ignore errors to avoid blocking requests
        pass
    finally:
        _DB_WARMED_UP = True

@main.route('/')
@login_required
def home():
    # Redirect authenticated users to dashboard for main control screen
    return redirect(url_for('main.dashboard'))

@csrf.exempt
@main.route('/login', methods=['GET', 'POST'])
def login():
    # Ensure DB tables exist on first local run to avoid 'no such table: user'
    try:
        ensure_tables()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()
        user = None
        if username:
            try:
                user = User.query.filter(func.lower(User.username) == username.lower()).first()
            except Exception:
                user = User.query.filter_by(username=username).first()

        # Safe bootstrap: if DB has no users at all, allow creating default admin/admin123 on first login
        if not user:
            try:
                total_users = User.query.count()
            except Exception:
                total_users = 0
            if total_users == 0 and username == 'admin' and password == 'admin123':
                try:
                    new_admin = User(username='admin')
                    new_admin.set_password('admin123')
                    db.session.add(new_admin)
                    db.session.commit()
                    login_user(new_admin)
                    flash('تم إنشاء مستخدم المدير الافتراضي بنجاح', 'success')
                    return redirect(url_for('main.dashboard'))
                except Exception:
                    db.session.rollback()
                    flash('خطأ في تهيئة المستخدم الافتراضي', 'danger')

        if user and password and user.check_password(password):
            login_user(user)
            return redirect(url_for('main.dashboard'))
        else:
            flash('خطأ في اسم المستخدم أو كلمة المرور', 'danger')
    return render_template('login.html')

@main.route('/debug/db', methods=['GET'], endpoint='debug_db')
def debug_db():
    try:
        uri = current_app.config.get('SQLALCHEMY_DATABASE_URI')
    except Exception:
        uri = None
    try:
        cnt = int(User.query.count())
    except Exception:
        cnt = -1
    return jsonify({'uri': uri, 'users_count': cnt})

@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))


# ---------- Main application pages (simple render-only) ----------
@main.route('/dashboard', endpoint='dashboard')
@login_required
def dashboard():
    try:
        from sqlalchemy import func
        from models import SalesInvoice
        from models import get_saudi_now
        today = get_saudi_now().date()

        q = SalesInvoice.query.with_entities(
            func.coalesce(func.count(SalesInvoice.id), 0).label('cnt'),
            func.coalesce(func.sum(SalesInvoice.total_after_tax_discount), 0).label('sum'),
            func.coalesce(func.avg(SalesInvoice.total_after_tax_discount), 0).label('avg'),
        ).filter(SalesInvoice.date == today)
        row = q.first()

        orders_count = int(row.cnt or 0)
        total_earnings = float(row.sum or 0)
        avg_order_value = float(row.avg or 0)

        if orders_count == 0:
            from datetime import timedelta
            start_dt = today - timedelta(days=30)
            row2 = SalesInvoice.query.with_entities(
                func.coalesce(func.count(SalesInvoice.id), 0).label('cnt'),
                func.coalesce(func.sum(SalesInvoice.total_after_tax_discount), 0).label('sum'),
                func.coalesce(func.avg(SalesInvoice.total_after_tax_discount), 0).label('avg'),
            ).filter(SalesInvoice.date.between(start_dt, today)).first()
            orders_count = int((row2.cnt or 0))
            total_earnings = float(row2.sum or 0)
            avg_order_value = float(row2.avg or 0)

        stats = {
            'orders_count': orders_count,
            'total_earnings': total_earnings,
            'avg_order_value': avg_order_value,
        }
    except Exception as e:
        # Log the error for debugging
        import traceback
        print(f"Error in dashboard function: {e}")
        traceback.print_exc()
        stats = {'orders_count': 0, 'total_earnings': 0.0, 'avg_order_value': 0.0}
    return render_template('dashboard.html', stats=stats)

# Orders Invoices screen
@main.route('/orders', endpoint='order_invoices_list')
@login_required
def order_invoices_list():
    # Enforce permission on Orders screen
    # Note: Permission checks for this screen are enforced in the UI via Jinja 'can()'.
    # Avoid calling can() directly here to prevent NameError in Python context.
    try:
        # Optional branch filter
        branch_f = (request.args.get('branch') or '').strip()
        # Fetch orders (limit for safety)
        q_orders = OrderInvoice.query
        if branch_f:
            q_orders = q_orders.filter(OrderInvoice.branch == branch_f)
        orders = q_orders.order_by(OrderInvoice.invoice_date.desc()).limit(500).all()

        # Totals using SQL functions
        from sqlalchemy import func
        q_totals = OrderInvoice.query.with_entities(
            func.coalesce(func.sum(OrderInvoice.subtotal), 0).label('subtotal_sum'),
            func.coalesce(func.sum(OrderInvoice.discount), 0).label('discount_sum'),
            func.coalesce(func.sum(OrderInvoice.vat), 0).label('vat_sum'),
            func.coalesce(func.sum(OrderInvoice.total), 0).label('total_sum'),
        )
        if branch_f:
            q_totals = q_totals.filter(OrderInvoice.branch == branch_f)
        totals_row = q_totals.first()
        totals = {
            'subtotal_sum': float(getattr(totals_row, 'subtotal_sum', 0) or 0),
            'discount_sum': float(getattr(totals_row, 'discount_sum', 0) or 0),
            'vat_sum': float(getattr(totals_row, 'vat_sum', 0) or 0),
            'total_sum': float(getattr(totals_row, 'total_sum', 0) or 0),
        }

        # Items summary in Python (DB-agnostic)
        items_summary_map = {}
        for o in orders:
            try:
                for it in (o.items or []):
                    name = (it.get('name') if isinstance(it, dict) else None) or '-'
                    qty_val = it.get('qty') if isinstance(it, dict) else 0
                    try:
                        qty = float(qty_val or 0)
                    except Exception:
                        qty = 0.0
                    items_summary_map[name] = items_summary_map.get(name, 0.0) + qty
            except Exception:
                continue
        items_summary = [
            {'item_name': k, 'total_qty': v}
            for k, v in sorted(items_summary_map.items(), key=lambda x: (-x[1], x[0]))
        ]

        return render_template('order_invoices.html', orders=orders, totals=totals, items_summary=items_summary, current_branch=branch_f)
    except Exception as e:
        current_app.logger.error('order_invoices_list error: %s', e)
        flash('Failed to load order invoices', 'danger')
        return render_template('order_invoices.html', orders=[], totals={'subtotal_sum':0,'discount_sum':0,'vat_sum':0,'total_sum':0}, items_summary=[])

@main.route('/orders/clear', methods=['POST'], endpoint='order_invoices_clear_all')
@login_required
def order_invoices_clear_all():
    try:
        from models import OrderInvoice
        num = db.session.query(OrderInvoice).delete(synchronize_session=False)
        db.session.commit()
        try:
            flash(f'تم حذف {num} من فواتير الطلبات نهائياً', 'success')
        except Exception:
            flash('تم حذف جميع فواتير الطلبات نهائياً', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('order_invoices_clear_all error: %s', e)
        flash('فشل حذف فواتير الطلبات', 'danger')
    return redirect(url_for('main.order_invoices_list'))

@main.route('/sales', endpoint='sales')
@login_required
def sales():
    # Modern flow: show branches, then tables, then table invoice
    branches = [
        {'code': 'china_town', 'label': 'CHINA TOWN', 'url': url_for('main.sales_tables', branch_code='china_town')},
        {'code': 'place_india', 'label': 'PALACE INDIA', 'url': url_for('main.sales_tables', branch_code='place_india')},
    ]
    # Filter branches by user permissions (allow via specific branch or global 'all')
    branches = [b for b in branches if user_can('sales', 'view', b['code'])]
    return render_template('sales_branches.html', branches=branches)

@main.route('/purchases', methods=['GET','POST'], endpoint='purchases')
@login_required
def purchases():
    warmup_db_once()
    # Prepare form and lists
    form = PurchaseInvoiceForm()
    try:
        if not form.date.data:
            form.date.data = get_saudi_now().date()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
    suppliers = []
    try:
        if 'Supplier' in globals():
            suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    except Exception:
        suppliers = []
    suppliers_json = []
    for s in suppliers:
        cr = getattr(s, 'cr_number', None)
        iban = getattr(s, 'iban', None)
        if (not cr or not iban):
            try:
                notes = (getattr(s, 'notes', '') or '')
                if notes:
                    for part in notes.split('|'):
                        p = (part or '').strip()
                        if (not cr) and p[:3].upper() == 'CR:':
                            cr = p[3:].strip()
                        if (not iban) and p[:5].upper() == 'IBAN:':
                            iban = p[5:].strip()
            except Exception:
                pass
        suppliers_json.append({
            'id': s.id,
            'name': s.name,
            'phone': getattr(s, 'phone', None),
            'tax_number': getattr(s, 'tax_number', None),
            'address': getattr(s, 'address', None),
            'cr_number': cr,
            'iban': iban,
            'active': getattr(s, 'active', True),
        })
    # Ensure RawMaterial table contains bilingual items from purchase_categories
    try:
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        data_path = os.path.join(base_dir, 'data', 'purchase_categories.json')
        if os.path.exists(data_path):
            with open(data_path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            pcs = payload.get('purchase_categories') or []
            existing = RawMaterial.query.all() if 'RawMaterial' in globals() else []
            # Build lookup by name and by name_ar (lowercased)
            name_lut = { (getattr(m,'name','') or '').strip().lower(): m for m in existing }
            ar_lut = { (getattr(m,'name_ar','') or '').strip().lower(): m for m in existing }
            def bi(sen, sar):
                sen = (sen or '').strip()
                sar = (sar or '').strip()
                if sen and sar:
                    return f"{sen} / {sar}"
                return sen or sar
            def _map_unit(nm):
                s = (nm or '').strip().lower()
                if any(k in s for k in ['meat','beef','lamb','chicken','fish','seafood','لحم','بقر','غنم','دجاج','سمك','بحرية']):
                    return 'kg'
                if any(k in s for k in ['vegetable','veg','greens','herb','خضار','أعشاب']):
                    return 'kg'
                if any(k in s for k in ['spice','spices','powder','tea','coffee','بهار','بهارات','بودرة','شاي','قهوة']):
                    return 'gram'
                if any(k in s for k in ['oil','sauce','liquid','milk','drink','juice','زيت','صلصة','سائل','حليب','مشروب','عصير']):
                    return 'liter'
                if any(k in s for k in ['rice','grain','pulses','lentil','flour','sugar','salt','أرز','حبوب','بقول','عدس','طحين','سكر','ملح']):
                    return 'kg'
                return 'piece'
            for cat in pcs:
                subs = cat.get('subcategories') or []
                for sub in subs:
                    nm = sub.get('name') or {}
                    items = sub.get('items') or []
                    for it in items:
                        if isinstance(it, dict):
                            en = (it.get('en') or '').strip()
                            ar = (it.get('ar') or '').strip()
                        else:
                            # if string, use as both
                            en = str(it).strip()
                            ar = str(it).strip()
                        disp = bi(en, ar)
                        key_en = disp.strip().lower()
                        key_ar = ar.strip().lower()
                        rm = name_lut.get(key_en) or ar_lut.get(key_ar)
                        if not rm:
                            try:
                                cat_name = bi((nm.get('en') if isinstance(nm, dict) else str(nm)), (nm.get('ar') if isinstance(nm, dict) else str(nm)))
                                unit = _map_unit(cat_name)
                                rm = RawMaterial(
                                    name=disp,
                                    name_ar=ar or None,
                                    unit=unit,
                                    cost_per_unit=0,
                                    category=cat_name,
                                )
                                db.session.add(rm)
                                db.session.flush()
                                name_lut[key_en] = rm
                                if key_ar:
                                    ar_lut[key_ar] = rm
                            except Exception:
                                try:
                                    db.session.rollback()
                                except Exception:
                                    pass
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
    except Exception:
        pass

    try:
        materials = RawMaterial.query.filter_by(active=True).order_by(RawMaterial.name.asc()).all() if 'RawMaterial' in globals() else []
    except Exception:
        materials = []
    materials_json = []
    for m in materials:
        name = m.name
        materials_json.append({
            'id': m.id,
            'name': name,
            'unit': m.unit,
            'cost_per_unit': float(m.cost_per_unit or 0),
            'stock_quantity': float(m.stock_quantity or 0),
            'category': getattr(m, 'category', None) or '',
        })

    if request.method == 'POST':
        pm = (request.form.get('payment_method') or 'cash').strip().lower()
        date_str = request.form.get('date') or get_saudi_now().date().isoformat()
        inv_type = (request.form.get('invoice_type') or 'VAT').strip().upper()
        supplier_name = (request.form.get('supplier_name') or '').strip() or None
        supplier_free = (request.form.get('supplier_name_free') or '').strip()
        if inv_type == 'NO_VAT' and supplier_free:
            supplier_name = supplier_free
        supplier_id = request.form.get('supplier_id', type=int)
        supplier_invoice_number = (request.form.get('supplier_invoice_number') or '').strip()
        notes_val = (request.form.get('notes') or '').strip()
        try:
            inv = PurchaseInvoice(
                invoice_number=f"INV-PUR-{get_saudi_now().year}-{(PurchaseInvoice.query.count()+1):04d}",
                date=datetime.strptime(date_str, '%Y-%m-%d').date(),
                supplier_name=supplier_name,
                supplier_id=supplier_id,
                payment_method=pm,
                user_id=getattr(current_user, 'id', 1)
            )
            try:
                inv.supplier_invoice_number = supplier_invoice_number
                inv.notes = notes_val
            except Exception:
                pass
            inv.status = (request.form.get('status') or (getattr(PurchaseInvoice, 'status', None) and 'unpaid') or 'unpaid').strip().lower()
            if inv.status not in ('paid','partial','unpaid'):
                inv.status = 'unpaid'
            # Initialize totals; will be computed from items below
            inv.total_before_tax = 0.0
            inv.tax_amount = 0.0
            inv.discount_amount = 0.0
            inv.total_after_tax_discount = 0.0
            if ext_db is not None:
                ext_db.session.add(inv); ext_db.session.commit()
            else:
                db.session.add(inv); db.session.commit()
        except Exception as e:
            if ext_db is not None:
                ext_db.session.rollback()
            else:
                db.session.rollback()
            flash(f'Could not save purchase invoice: {e}', 'danger')
            return redirect(url_for('main.purchases'))

        # Save posted purchase items and update inventory
        try:
            from decimal import Decimal
            def _to_ascii_digits(s):
                try:
                    if s is None:
                        return ''
                    s = str(s)
                    # Arabic-Indic and Persian digits to ASCII
                    trans = str.maketrans('٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹', '01234567890123456789')
                    s = s.translate(trans)
                    # Arabic decimal separator
                    s = s.replace('٫', '.')
                    s = s.replace('،', '.')
                    return s
                except Exception:
                    return str(s or '')
            idxs = set()
            for k in request.form.keys():
                if k.startswith('items-') and k.endswith('-raw_material_id'):
                    try:
                        idxs.add(int(k.split('-')[1]))
                    except Exception:
                        pass
            total_before_tax_dec = Decimal('0')
            total_tax_dec = Decimal('0')
            total_discount_dec = Decimal('0')
            items_saved = 0
            VAT_RATE = Decimal('0.15') if inv_type == 'VAT' else Decimal('0')
            for i in sorted(idxs):
                try:
                    rm_id = int(request.form.get(f'items-{i}-raw_material_id') or 0)
                except Exception:
                    rm_id = 0
                qty_raw = _to_ascii_digits(request.form.get(f'items-{i}-quantity'))
                price_raw = _to_ascii_digits(request.form.get(f'items-{i}-price_before_tax'))
                disc_raw = _to_ascii_digits(request.form.get(f'items-{i}-discount'))
                try:
                    qty = Decimal(str(qty_raw)) if (qty_raw not in (None, '')) else Decimal('0')
                except Exception:
                    qty = Decimal('0')
                try:
                    unit_price = Decimal(str(price_raw)) if (price_raw not in (None, '')) else Decimal('0')
                except Exception:
                    unit_price = Decimal('0')
                try:
                    disc = Decimal(str(disc_raw)) if (disc_raw not in (None, '')) else Decimal('0')
                except Exception:
                    disc = Decimal('0')
                if rm_id and qty > 0:
                    raw_material = RawMaterial.query.get(rm_id)
                    if raw_material:
                        line_base = (qty * unit_price) - disc
                        if line_base < 0:
                            line_base = Decimal('0')
                        line_tax = (line_base * VAT_RATE) if VAT_RATE > 0 else Decimal('0')
                        line_total = line_base + line_tax
                        prev_qty = Decimal(str(raw_material.stock_quantity or 0))
                        prev_cost = Decimal(str(raw_material.cost_per_unit or 0))
                        new_total_qty = prev_qty + qty
                        if new_total_qty > 0:
                            new_total_cost = (prev_cost * prev_qty) + (unit_price * qty)
                            raw_material.cost_per_unit = (new_total_cost / new_total_qty).quantize(Decimal('0.0001'))
                            raw_material.stock_quantity = new_total_qty
                        else:
                            raw_material.stock_quantity = prev_qty + qty
                        inv_item = PurchaseInvoiceItem(
                            invoice_id=inv.id,
                            raw_material_id=raw_material.id,
                            raw_material_name=getattr(raw_material, 'display_name', raw_material.name),
                            quantity=float(qty),
                            price_before_tax=float(unit_price),
                            tax=float(line_tax),
                            discount=float(disc),
                            total_price=float(line_total)
                        )
                        db.session.add(inv_item)
                        total_before_tax_dec += line_base
                        total_tax_dec += line_tax
                        total_discount_dec += disc
                        items_saved += 1
            inv.total_before_tax = float(total_before_tax_dec)
            inv.tax_amount = float(total_tax_dec)
            inv.discount_amount = float(total_discount_dec)
            inv.total_after_tax_discount = float(total_before_tax_dec + total_tax_dec - total_discount_dec)
            if items_saved > 0:
                try:
                    inv.notes = ((inv.notes or '') + f" | ITEMS:{items_saved}").strip(' |')
                except Exception:
                    pass
            db.session.commit()
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(f'Failed to save purchase items: {e}', 'warning')

        # Auto-create payment and ledger outside main try
        st = (getattr(inv, 'status', '') or '').lower()
        total_amt = float(inv.total_after_tax_discount or 0.0)
        base_amt = float(getattr(inv, 'total_before_tax', 0) or 0.0) - float(getattr(inv, 'discount_amount', 0) or 0.0)
        tax_amt = float(getattr(inv, 'tax_amount', 0) or 0.0)
        extra = []
        if supplier_invoice_number:
            extra.append(f"SUPREF:{supplier_invoice_number}")
        if notes_val:
            extra.append(f"NOTE:{notes_val[:50]}")
        ref = f"PUR {inv.invoice_number}" + ((' ' + ' '.join(extra)) if extra else '')
        production_flag = ((request.form.get('production_flag') or '').strip().lower() in ('1','true','yes','on'))
        inv_kind = (request.form.get('inventory_kind') or '').strip().upper()
        inv_acc = '1020' if inv_kind == 'FOOD' else ('1025' if inv_kind == 'OTHER' else None)
        debit_code = None
        if inv_type == 'VAT':
            debit_code = inv_acc if production_flag and inv_acc else '5000'
        else:
            debit_code = inv_acc if production_flag and inv_acc else 'COGS'
        if debit_code:
            # Use resolved COA mapping for debit
            nm = CHART_OF_ACCOUNTS.get(debit_code, {}).get('name', 'Purchases/COGS')
            tp = CHART_OF_ACCOUNTS.get(debit_code, {}).get('type', 'expense')
            _post_ledger(inv.date, debit_code, nm, tp, base_amt, 0.0, ref)
        if tax_amt > 0:
            _post_ledger(inv.date, 'VAT_IN', 'VAT Input', 'tax', tax_amt, 0.0, ref)
        _post_ledger(inv.date, 'AP', 'Accounts Payable', 'liability', 0.0, base_amt + tax_amt, ref)
        if st == 'paid' and total_amt > 0:
            db.session.add(Payment(invoice_id=inv.id, invoice_type='purchase', amount_paid=total_amt, payment_method=(pm or 'CASH').upper()))
            db.session.commit()
            cash_acc = _pm_account(pm)
            _post_ledger(inv.date, 'AP', 'Accounts Payable', 'liability', total_amt, 0.0, f'PAY PUR {inv.invoice_number}')
            if cash_acc:
                _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', 0.0, total_amt, f'PAY PUR {inv.invoice_number}')
                try:
                    if (pm or '').strip().lower() == 'bank':
                        fee = round(total_amt * 0.02, 2)
                        if fee > 0:
                            _post_ledger(inv.date, '5025', 'عمولات بنكية', 'expense', fee, 0.0, f'BANK FEE {inv.invoice_number}')
                            _post_ledger(inv.date, '1010', 'البنك', 'asset', 0.0, fee, f'BANK FEE {inv.invoice_number}')
                except Exception:
                    pass
        elif st == 'partial':
            amt_raw = request.form.get('partial_paid_amount')
            amt = float(amt_raw or 0.0)
            if total_amt > 0 and amt > 0:
                if amt > total_amt:
                    amt = total_amt
                db.session.add(Payment(invoice_id=inv.id, invoice_type='purchase', amount_paid=amt, payment_method=(pm or 'CASH').upper()))
                db.session.commit()
                cash_acc = _pm_account(pm)
                _post_ledger(inv.date, 'AP', 'Accounts Payable', 'liability', amt, 0.0, f'PAY PUR {inv.invoice_number}')
                if cash_acc:
                    _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', 0.0, amt, f'PAY PUR {inv.invoice_number}')
                    try:
                        if (pm or '').strip().lower() == 'bank':
                            fee = round(amt * 0.02, 2)
                            if fee > 0:
                                _post_ledger(inv.date, '5025', 'عمولات بنكية', 'expense', fee, 0.0, f'BANK FEE {inv.invoice_number}')
                                _post_ledger(inv.date, '1010', 'البنك', 'asset', 0.0, fee, f'BANK FEE {inv.invoice_number}')
                    except Exception:
                        pass
        try:
            _create_purchase_journal(inv)
        except Exception:
            pass
        flash('Purchase invoice saved', 'success')
        return redirect(url_for('main.purchases'))

    return render_template('purchases.html', form=form, suppliers_list=suppliers, suppliers_json=suppliers_json, materials_json=materials_json)

@main.route('/raw-materials', methods=['GET', 'POST'], endpoint='raw_materials')
@login_required
def raw_materials():
    form = RawMaterialForm()
    mode = (request.args.get('mode') or '').strip().lower()
    if request.method == 'POST' and (request.form.get('action') or '') == 'bulk_update_quantities':
        try:
            mats = RawMaterial.query.filter_by(active=True).all()
            for m in mats:
                qv = request.form.get(f'qty_{int(m.id)}')
                cv = request.form.get(f'cost_{int(m.id)}')
                try:
                    q = float(qv or 0)
                except Exception:
                    q = 0.0
                try:
                    c = float(cv or 0)
                except Exception:
                    c = 0.0
                if q > 0:
                    m.stock_quantity = float(m.stock_quantity or 0.0) + q
                if c > 0:
                    m.cost_per_unit = c
            db.session.commit()
            flash('تم تحديث الكميات بنجاح', 'success')
        except Exception:
            db.session.rollback()
            flash('فشل تحديث الكميات', 'danger')
        return redirect(url_for('main.raw_materials', mode='quantities'))
    if request.method == 'POST' and form.validate_on_submit():
        try:
            rm = RawMaterial(
                name=form.name.data,
                name_ar=form.name_ar.data,
                unit=form.unit.data,
                cost_per_unit=form.cost_per_unit.data,
                category=form.category.data,
            )
            db.session.add(rm)
            db.session.commit()
            flash('تم حفظ المادة الخام بنجاح', 'success')
            return redirect(url_for('main.raw_materials'))
        except Exception as e:
            db.session.rollback()
            flash('فشل حفظ المادة الخام', 'danger')
    materials = RawMaterial.query.filter_by(active=True).all()
    return render_template('raw_materials.html', form=form, materials=materials, mode=mode)

@main.route('/api/raw_materials', methods=['GET'], endpoint='api_raw_materials')
@login_required
def api_raw_materials():
    try:
        cat = (request.args.get('category') or '').strip()
        mats_q = RawMaterial.query.filter_by(active=True)
        if cat:
            mats_q = mats_q.filter(RawMaterial.category == cat)
        mats = mats_q.order_by(RawMaterial.name.asc()).all()
        data = [
            {
                'id': int(m.id),
                'name': (m.display_name if hasattr(m, 'display_name') else m.name),
                'name_ar': (getattr(m, 'name_ar', '') or ''),
                'unit': (m.unit or ''),
                'stock_quantity': float(getattr(m, 'stock_quantity', 0.0) or 0.0),
                'cost_per_unit': float(getattr(m, 'cost_per_unit', 0.0) or 0.0),
                'category': (m.category or '')
            }
            for m in mats
        ]
        return jsonify({'ok': True, 'items': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/raw_materials/categories', methods=['GET'], endpoint='api_raw_materials_categories')
@login_required
def api_raw_materials_categories():
    try:
        cats = db.session.query(RawMaterial.category).filter(RawMaterial.active == True).distinct().all()
        items = [
            {
                'name': (c[0] or ''),
                'label': (c[0] or '')
            }
            for c in cats if c and c[0]
        ]
        return jsonify({'ok': True, 'categories': items})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/meals', methods=['GET', 'POST'], endpoint='meals')
@login_required
def meals():
    # Build form and context similar to monolith, but within blueprint
    raw_materials = RawMaterial.query.filter_by(active=True).all()
    material_choices = [(m.id, m.display_name) for m in raw_materials]

    form = MealForm()
    # Ensure ingredient subforms have choices
    for ingredient_form in form.ingredients:
        ingredient_form.raw_material_id.choices = material_choices

    materials_json = json.dumps([
        {
            'id': m.id,
            'name': m.display_name,
            'cost_per_unit': float(m.cost_per_unit),
            'unit': m.unit,
        }
        for m in raw_materials
    ])

    if form.validate_on_submit():
        try:
            from decimal import Decimal
            meal = Meal(
                name=form.name.data,
                name_ar=form.name_ar.data,
                description=form.description.data,
                category=form.category.data,
                profit_margin_percent=form.profit_margin_percent.data,
                user_id=current_user.id,
            )
            db.session.add(meal)
            db.session.flush()  # get meal.id

            total_cost = 0
            # Parse dynamic ingredient rows from POST keys
            idxs = set()
            for k in request.form.keys():
                if k.startswith('ingredients-') and k.endswith('-raw_material_id'):
                    try:
                        idxs.add(int(k.split('-')[1]))
                    except Exception:
                        pass
            for i in sorted(idxs):
                try:
                    rm_id = int(request.form.get(f'ingredients-{i}-raw_material_id') or 0)
                    qty_raw = request.form.get(f'ingredients-{i}-quantity')
                    qty = Decimal(qty_raw) if qty_raw not in (None, '') else Decimal('0')
                except Exception:
                    rm_id, qty = 0, Decimal('0')
                if rm_id and qty > 0:
                    raw_material = RawMaterial.query.get(rm_id)
                    if raw_material:
                        ing = MealIngredient(meal_id=meal.id, raw_material_id=raw_material.id, quantity=qty)
                        try:
                            ing_cost = qty * raw_material.cost_per_unit
                            ing.total_cost = ing_cost
                        except Exception:
                            ing.total_cost = 0
                        db.session.add(ing)
                        total_cost += float(ing.total_cost)

            meal.total_cost = total_cost
            meal.calculate_selling_price()
            db.session.commit()
            flash('تم إنشاء الوجبة بنجاح', 'success')
            return redirect(url_for('main.meals'))
        except Exception as e:
            db.session.rollback()
            flash('فشل حفظ الوجبة', 'danger')

    all_meals = Meal.query.filter_by(active=True).all()
    return render_template('meals.html', form=form, meals=all_meals, materials_json=materials_json)

# -------- Meals import (Excel/CSV): Name, Name (Arabic), Selling Price --------
@main.route('/meals/import', methods=['POST'], endpoint='meals_import')
@login_required
def meals_import():
    import os, csv
    from io import TextIOWrapper
    file = request.files.get('file')
    if not file or not file.filename:
        flash('لم يتم اختيار ملف', 'warning')
        return redirect(url_for('main.menu'))

    # إن وُجد قسم حالي مرسل من النموذج نعيد التوجيه إليه بعد الاستيراد
    cat_id = request.form.get('cat_id', type=int)

    ext = os.path.splitext(file.filename)[1].lower()

    def upsert_meal(name_en, name_ar, price_val):
        name_en = (name_en or '').strip()
        name_ar = (name_ar or '').strip() or None
        try:
            price = float(str(price_val).replace(',', '').strip()) if price_val is not None else 0.0
        except Exception:
            price = 0.0
        # محاولة إيجاد وجبة موجودة بنفس الاسم (و/أو الاسم العربي)
        existing = None
        if name_en and name_ar:
            existing = Meal.query.filter_by(name=name_en, name_ar=name_ar).first()
        if not existing and name_en:
            existing = Meal.query.filter_by(name=name_en).first()
        if existing:
            # تأكد من وجود user_id، ثم حدّث سعر البيع فقط
            try:
                if getattr(existing, 'user_id', None) is None:
                    existing.user_id = current_user.id
            except Exception:
                pass
            try:
                existing.selling_price = price
            except Exception:
                pass
            return existing
        # إنشاء وجبة جديدة مع تعيين user_id
        m = Meal(name=name_en or 'Unnamed', name_ar=name_ar, description=None, category=None, user_id=current_user.id)
        try:
            m.selling_price = price
        except Exception:
            pass
        db.session.add(m)
        return m

    imported, errors = 0, 0

    if ext == '.csv':
        try:
            stream = TextIOWrapper(file.stream, encoding='utf-8')
            reader = csv.DictReader(stream)
            def norm(s):
                return (s or '').strip().lower()
            for row in reader:
                cols = {norm(k): v for k, v in row.items()}
                name = cols.get('name') or cols.get('اسم') or cols.get('product')
                name_ar = cols.get('name (arabic)') or cols.get('name_ar') or cols.get('arabic') or cols.get('الاسم العربي')
                price = cols.get('selling price') or cols.get('price') or cols.get('السعر')
                if not name and not name_ar:
                    continue
                try:
                    upsert_meal(name, name_ar, price)
                    imported += 1
                except Exception:
                    errors += 1
            db.session.commit()
            flash(f'تم استيراد {imported} وجبة بنجاح' + (f'، أخطاء: {errors}' if errors else ''), 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'فشل استيراد CSV: {e}', 'danger')
        return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))

    elif ext in ('.xlsx', '.xls'):
        try:
            try:
                import openpyxl
            except Exception:
                flash('لا يمكن قراءة ملفات Excel بدون تثبيت openpyxl. يمكنك رفع CSV بدلاً من ذلك أو اسمح لي بتثبيت openpyxl.', 'warning')
                return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
                headers.append((cell or '').strip().lower())
            index = {h: i for i, h in enumerate(headers)}
            def get_val(row, keys):
                for k in keys:
                    if k in index:
                        return row[index[k]]
                return None
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = get_val(row, ['name', 'اسم'])
                name_ar = get_val(row, ['name (arabic)', 'name_ar', 'arabic', 'الاسم العربي'])
                price = get_val(row, ['selling price', 'price', 'السعر'])
                if not name and not name_ar:
                    continue
                try:
                    upsert_meal(name, name_ar, price)
                    imported += 1
                except Exception:
                    errors += 1
            db.session.commit()
            flash(f'تم استيراد {imported} وجبة من Excel' + (f'، أخطاء: {errors}' if errors else ''), 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'فشل استيراد Excel: {e}', 'danger')
        return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))

    else:
        flash('صيغة الملف غير مدعومة. الرجاء رفع ملف CSV أو Excel (.xlsx/.xls).', 'warning')
        return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))

@main.route('/inventory', endpoint='inventory')
@login_required
def inventory():
    from sqlalchemy import func
    raw_materials = RawMaterial.query.filter_by(active=True).all()
    meals = Meal.query.filter_by(active=True).all()

    # Build inventory cost ledger from purchases
    ledger_rows = []
    try:
        q = db.session.query(
            PurchaseInvoiceItem.raw_material_id.label('rm_id'),
            func.max(PurchaseInvoice.date).label('last_date'),
            func.sum(PurchaseInvoiceItem.quantity).label('qty'),
            func.sum(PurchaseInvoiceItem.total_price).label('total_cost')
        ).join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
        q = q.group_by(PurchaseInvoiceItem.raw_material_id)
        rm_map = {m.id: m for m in raw_materials}
        for r in q.all():
            rm = rm_map.get(int(r.rm_id)) if r.rm_id is not None else None
            name = (rm.display_name if rm else '-')
            unit = (rm.unit if rm else '-')
            # Quantities and costs
            qty = float(r.qty or 0)
            total_cost = float(r.total_cost or 0)
            avg_cost = (total_cost / qty) if qty else 0.0
            # Current stock equals cumulative purchased quantity (no consumption tracking here)
            current_stock = qty
            stock_value = current_stock * avg_cost
            ledger_rows.append({
                'material': name,
                'unit': unit,
                'purchased_qty': qty,
                'avg_cost': avg_cost,
                'total_cost': total_cost,
                'current_stock': current_stock,
                'stock_value': stock_value,
                'last_date': r.last_date.strftime('%Y-%m-%d') if r.last_date else ''
            })
        ledger_rows.sort(key=lambda x: (x['material'] or '').lower())
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        ledger_rows = []

    purchases = []
    try:
        invs = PurchaseInvoice.query.order_by(PurchaseInvoice.id.desc()).limit(1000).all()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        invs = []
    items_map = {}
    try:
        ids = [inv.id for inv in invs]
        if ids:
            rows = PurchaseInvoiceItem.query.filter(PurchaseInvoiceItem.invoice_id.in_(ids)).order_by(PurchaseInvoiceItem.id.asc()).all()
            for it in rows:
                items_map.setdefault(it.invoice_id, []).append(it)
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        items_map = {}
    for inv in invs:
        items_ctx = []
        for it in items_map.get(inv.id, []):
            items_ctx.append({
                'name': it.raw_material_name,
                'quantity': float(it.quantity or 0),
                'price_before_tax': float(it.price_before_tax or 0),
                'discount': float(it.discount or 0),
                'tax': float(it.tax or 0),
                'total_price': float(it.total_price or 0),
            })
        purchases.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'date': inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '',
            'supplier_name': getattr(inv, 'supplier_name', '') or '',
            'payment_method': (getattr(inv, 'payment_method', '') or '').upper(),
            'status': getattr(inv, 'status', '') or '',
            'total_before_tax': float(getattr(inv, 'total_before_tax', 0.0) or 0.0),
            'tax_amount': float(getattr(inv, 'tax_amount', 0.0) or 0.0),
            'discount_amount': float(getattr(inv, 'discount_amount', 0.0) or 0.0),
            'total_after_tax_discount': float(getattr(inv, 'total_after_tax_discount', 0.0) or 0.0),
            'items': items_ctx,
        })
    try:
        today = get_saudi_now().date()
        start_month = today.replace(day=1)
        def latest_cost_per_unit(rm_id: int):
            try:
                row = (
                    db.session.query(PurchaseInvoiceItem.price_before_tax, PurchaseInvoice.date)
                    .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                    .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                    .order_by(PurchaseInvoice.date.desc(), PurchaseInvoiceItem.id.desc())
                    .first()
                )
                if row and row[0] is not None:
                    return float(row[0])
            except Exception:
                pass
            try:
                avg = (
                    db.session.query(func.coalesce(func.avg(PurchaseInvoiceItem.price_before_tax), 0))
                    .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                    .scalar()
                ) or 0
                if float(avg) > 0:
                    return float(avg)
            except Exception:
                pass
            try:
                rm = RawMaterial.query.get(int(rm_id))
                return float(getattr(rm, 'cost_per_unit', 0) or 0)
            except Exception:
                return 0.0

        total_inventory_cost = 0.0
        for rm in (raw_materials or []):
            unit_cost = latest_cost_per_unit(int(getattr(rm, 'id', 0) or 0))
            qty = float(getattr(rm, 'stock_quantity', 0) or 0)
            total_inventory_cost += unit_cost * qty

        try:
            total_purchases_month = float(
                db.session.query(func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0))
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .scalar() or 0
            )
        except Exception:
            total_purchases_month = 0.0

        try:
            sales_invoices_q = (
                db.session.query(SalesInvoice.id, SalesInvoice.date, SalesInvoice.total_after_tax_discount)
                .filter(SalesInvoice.date >= start_month, SalesInvoice.date <= today)
            )
            sales_invoice_ids = [int(r[0]) for r in sales_invoices_q.all()]
            total_sales_revenue_month = float(
                sum([float(r[2] or 0) for r in sales_invoices_q.all()])
            )
            sales_qty_rows = (
                db.session.query(func.coalesce(func.sum(SalesInvoiceItem.quantity), 0))
                .filter(SalesInvoiceItem.invoice_id.in_(sales_invoice_ids))
                .scalar() or 0
            )
            total_meals_sold_month = float(sales_qty_rows or 0)
        except Exception:
            total_sales_revenue_month = 0.0
            total_meals_sold_month = 0.0

        meals_list = Meal.query.filter_by(active=True).all()
        meal_by_name = { (getattr(m, 'name', '') or '').strip(): m for m in (meals_list or []) }
        sales_qty_by_meal = {}
        try:
            rows = (
                db.session.query(SalesInvoiceItem.product_name, func.coalesce(func.sum(SalesInvoiceItem.quantity), 0))
                .filter(SalesInvoiceItem.invoice_id.in_(sales_invoice_ids))
                .group_by(SalesInvoiceItem.product_name)
                .all()
            )
            for name, qty in (rows or []):
                k = (name or '').strip()
                sales_qty_by_meal[k] = float(qty or 0)
        except Exception:
            sales_qty_by_meal = {}

        estimated_total_meal_cost = 0.0
        meals_analysis = []
        outdated_meals = []
        def latest_purchase_date(rm_id: int):
            try:
                d = (
                    db.session.query(PurchaseInvoice.date)
                    .join(PurchaseInvoiceItem, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                    .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                    .order_by(PurchaseInvoice.date.desc(), PurchaseInvoiceItem.id.desc())
                    .first()
                )
                return d[0] if d and d[0] else None
            except Exception:
                return None
        OUTDATED_DAYS = 60
        for m_name, sold_qty in sales_qty_by_meal.items():
            meal = meal_by_name.get(m_name)
            if not meal:
                continue
            total_cost = 0.0
            outdated_flag = False
            try:
                ingrs = MealIngredient.query.filter_by(meal_id=int(meal.id)).all()
            except Exception:
                ingrs = []
            for ing in (ingrs or []):
                unit_cost = latest_cost_per_unit(int(getattr(ing, 'raw_material_id', 0) or 0))
                last_dt = latest_purchase_date(int(getattr(ing, 'raw_material_id', 0) or 0))
                if not last_dt or (today - last_dt).days > OUTDATED_DAYS:
                    outdated_flag = True
                qty_needed = float(getattr(ing, 'quantity', 0) or 0)
                total_cost += unit_cost * qty_needed * float(sold_qty or 0)
            estimated_total_meal_cost += total_cost
            revenue = float(getattr(meal, 'selling_price', 0) or 0) * float(sold_qty or 0)
            profit = revenue - total_cost
            meals_analysis.append({
                'meal_name': m_name,
                'sold_qty': float(sold_qty or 0),
                'consumption_cost': float(total_cost or 0),
                'revenue': float(revenue or 0),
                'profit': float(profit or 0),
            })
            if outdated_flag:
                outdated_meals.append(m_name)

        total_profit_generated = float(total_sales_revenue_month or 0) - float(estimated_total_meal_cost or 0)

        purchases_rows = []
        try:
            invs_m = (
                PurchaseInvoice.query
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .order_by(PurchaseInvoice.date.desc(), PurchaseInvoice.id.desc())
                .limit(500)
                .all()
            )
            for inv in (invs_m or []):
                items_count = int(PurchaseInvoiceItem.query.filter_by(invoice_id=inv.id).count() or 0)
                purchases_rows.append({
                    'date': inv.date.isoformat() if getattr(inv, 'date', None) else '',
                    'invoice_number': getattr(inv, 'invoice_number', ''),
                    'supplier': getattr(inv, 'supplier_name', '') or '-',
                    'payment_method': getattr(inv, 'payment_method', '') or '-',
                    'status': getattr(inv, 'status', '') or '-',
                    'subtotal': float(getattr(inv, 'total_before_tax', 0) or 0),
                    'vat': float(getattr(inv, 'tax_amount', 0) or 0),
                    'discount': float(getattr(inv, 'discount_amount', 0) or 0),
                    'final_total': float(getattr(inv, 'total_after_tax_discount', 0) or 0),
                    'items_count': items_count,
                })
        except Exception:
            purchases_rows = []

        top_ingredients = []
        purchase_trend = []
        try:
            rows_t = (
                db.session.query(PurchaseInvoiceItem.raw_material_name, func.coalesce(func.sum(PurchaseInvoiceItem.total_price), 0))
                .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .group_by(PurchaseInvoiceItem.raw_material_name)
                .order_by(func.coalesce(func.sum(PurchaseInvoiceItem.total_price), 0).desc())
                .limit(10)
                .all()
            )
            top_ingredients = [{'name': n, 'total': float(t or 0)} for n, t in (rows_t or [])]
            trows = (
                db.session.query(PurchaseInvoice.date, func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0))
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .group_by(PurchaseInvoice.date)
                .order_by(PurchaseInvoice.date.asc())
                .all()
            )
            purchase_trend = [{'date': (d.isoformat() if d else ''), 'total': float(t or 0)} for d, t in (trows or [])]
        except Exception:
            top_ingredients = []
            purchase_trend = []

        stock_rows = []
        try:
            p_qty_rows = (
                db.session.query(PurchaseInvoiceItem.raw_material_id, func.coalesce(func.sum(PurchaseInvoiceItem.quantity), 0))
                .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .group_by(PurchaseInvoiceItem.raw_material_id)
                .all()
            )
            p_qty_map = {int(rm_id): float(qty or 0) for rm_id, qty in (p_qty_rows or []) if rm_id is not None}
            usage_map = {}
            for m in (meals_list or []):
                sold = float(sales_qty_by_meal.get((getattr(m, 'name', '') or '').strip(), 0) or 0)
                if sold <= 0:
                    continue
                ingrs = MealIngredient.query.filter_by(meal_id=int(m.id)).all()
                for ing in (ingrs or []):
                    usage_map[int(getattr(ing, 'raw_material_id', 0) or 0)] = usage_map.get(int(getattr(ing, 'raw_material_id', 0) or 0), 0.0) + (float(getattr(ing, 'quantity', 0) or 0) * sold)
            for rm in (raw_materials or []):
                rid = int(getattr(rm, 'id', 0) or 0)
                opening_qty = float(getattr(rm, 'stock_quantity', 0) or 0)
                purchases_qty = float(p_qty_map.get(rid, 0) or 0)
                estimated_usage = float(usage_map.get(rid, 0) or 0)
                expected_stock = max(opening_qty + purchases_qty - estimated_usage, 0.0)
                risk = 'low' if expected_stock < max(1.0, opening_qty * 0.1) else ('excess' if purchases_qty > estimated_usage * 1.5 else 'ok')
                stock_rows.append({
                    'ingredient': rm.display_name,
                    'opening_qty': opening_qty,
                    'purchases_qty': purchases_qty,
                    'estimated_usage': estimated_usage,
                    'expected_stock': expected_stock,
                    'risk': risk,
                })
        except Exception:
            stock_rows = []

        return render_template(
            'inventory.html',
            raw_materials=raw_materials,
            meals=meals,
            ledger_rows=ledger_rows,
            purchases=purchases,
            _=(lambda s, **kw: s),
            kpi={
                'total_inventory_cost': float(total_inventory_cost or 0),
                'total_purchases_month': float(total_purchases_month or 0),
                'total_meals_sold_month': float(total_meals_sold_month or 0),
                'estimated_total_meal_cost': float(estimated_total_meal_cost or 0),
                'total_profit_generated': float(total_profit_generated or 0),
            },
            purchases_summary=purchases_rows,
            top_ingredients=top_ingredients,
            purchase_trend=purchase_trend,
            meals_analysis=meals_analysis,
            outdated_meals=outdated_meals,
            stock_rows=stock_rows,
        )
    except Exception:
        return render_template(
            'inventory.html',
            raw_materials=raw_materials,
            meals=meals,
            ledger_rows=ledger_rows,
            purchases=purchases,
            _=(lambda s, **kw: s),
            kpi={
                'total_inventory_cost': 0.0,
                'total_purchases_month': 0.0,
                'total_meals_sold_month': 0.0,
                'estimated_total_meal_cost': 0.0,
                'total_profit_generated': 0.0,
            },
            purchases_summary=[],
            top_ingredients=[],
            purchase_trend=[],
            meals_analysis=[],
            outdated_meals=[],
            stock_rows=[],
        )

@main.route('/inventory-intelligence', methods=['GET'], endpoint='inventory_intelligence')
@login_required
def inventory_intelligence():
    try:
        today = get_saudi_now().date()
        start_month = today.replace(day=1)
        def latest_cost_per_unit(rm_id: int):
            try:
                row = (
                    db.session.query(PurchaseInvoiceItem.price_before_tax, PurchaseInvoice.date)
                    .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                    .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                    .order_by(PurchaseInvoice.date.desc(), PurchaseInvoiceItem.id.desc())
                    .first()
                )
                if row and row[0] is not None:
                    return float(row[0])
            except Exception:
                pass
            try:
                avg = (
                    db.session.query(func.coalesce(func.avg(PurchaseInvoiceItem.price_before_tax), 0))
                    .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                    .scalar()
                ) or 0
                if float(avg) > 0:
                    return float(avg)
            except Exception:
                pass
            try:
                rm = RawMaterial.query.get(int(rm_id))
                return float(getattr(rm, 'cost_per_unit', 0) or 0)
            except Exception:
                return 0.0

        raw_materials = RawMaterial.query.filter_by(active=True).all()
        total_inventory_cost = 0.0
        for rm in (raw_materials or []):
            unit_cost = latest_cost_per_unit(int(getattr(rm, 'id', 0) or 0))
            qty = float(getattr(rm, 'stock_quantity', 0) or 0)
            total_inventory_cost += unit_cost * qty

        try:
            total_purchases_month = float(
                db.session.query(func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0))
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .scalar() or 0
            )
        except Exception:
            total_purchases_month = 0.0

        try:
            sales_invoices_q = (
                db.session.query(SalesInvoice.id, SalesInvoice.date, SalesInvoice.total_after_tax_discount)
                .filter(SalesInvoice.date >= start_month, SalesInvoice.date <= today)
            )
            sales_invoice_ids = [int(r[0]) for r in sales_invoices_q.all()]
            total_sales_revenue_month = float(
                sum([float(r[2] or 0) for r in sales_invoices_q.all()])
            )
            sales_qty_rows = (
                db.session.query(func.coalesce(func.sum(SalesInvoiceItem.quantity), 0))
                .filter(SalesInvoiceItem.invoice_id.in_(sales_invoice_ids))
                .scalar() or 0
            )
            total_meals_sold_month = float(sales_qty_rows or 0)
        except Exception:
            total_sales_revenue_month = 0.0
            total_meals_sold_month = 0.0

        meals = Meal.query.filter_by(active=True).all()
        meal_by_name = { (getattr(m, 'name', '') or '').strip(): m for m in (meals or []) }
        sales_qty_by_meal = {}
        try:
            rows = (
                db.session.query(SalesInvoiceItem.product_name, func.coalesce(func.sum(SalesInvoiceItem.quantity), 0))
                .filter(SalesInvoiceItem.invoice_id.in_(sales_invoice_ids))
                .group_by(SalesInvoiceItem.product_name)
                .all()
            )
            for name, qty in (rows or []):
                k = (name or '').strip()
                sales_qty_by_meal[k] = float(qty or 0)
        except Exception:
            sales_qty_by_meal = {}

        estimated_total_meal_cost = 0.0
        meals_analysis = []
        outdated_meals = []
        def latest_purchase_date(rm_id: int):
            try:
                d = (
                    db.session.query(PurchaseInvoice.date)
                    .join(PurchaseInvoiceItem, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                    .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                    .order_by(PurchaseInvoice.date.desc(), PurchaseInvoiceItem.id.desc())
                    .first()
                )
                return d[0] if d and d[0] else None
            except Exception:
                return None
        OUTDATED_DAYS = 60
        for m_name, sold_qty in sales_qty_by_meal.items():
            meal = meal_by_name.get(m_name)
            if not meal:
                continue
            total_cost = 0.0
            outdated_flag = False
            try:
                ingrs = MealIngredient.query.filter_by(meal_id=int(meal.id)).all()
            except Exception:
                ingrs = []
            for ing in (ingrs or []):
                unit_cost = latest_cost_per_unit(int(getattr(ing, 'raw_material_id', 0) or 0))
                last_dt = latest_purchase_date(int(getattr(ing, 'raw_material_id', 0) or 0))
                if not last_dt or (today - last_dt).days > OUTDATED_DAYS:
                    outdated_flag = True
                qty_needed = float(getattr(ing, 'quantity', 0) or 0)
                total_cost += unit_cost * qty_needed * float(sold_qty or 0)
            estimated_total_meal_cost += total_cost
            revenue = float(getattr(meal, 'selling_price', 0) or 0) * float(sold_qty or 0)
            profit = revenue - total_cost
            meals_analysis.append({
                'meal_name': m_name,
                'sold_qty': float(sold_qty or 0),
                'consumption_cost': float(total_cost or 0),
                'revenue': float(revenue or 0),
                'profit': float(profit or 0),
            })
            if outdated_flag:
                outdated_meals.append(m_name)

        total_profit_generated = float(total_sales_revenue_month or 0) - float(estimated_total_meal_cost or 0)

        purchases_rows = []
        try:
            invs = (
                PurchaseInvoice.query
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .order_by(PurchaseInvoice.date.desc(), PurchaseInvoice.id.desc())
                .limit(500)
                .all()
            )
            for inv in (invs or []):
                items_count = int(PurchaseInvoiceItem.query.filter_by(invoice_id=inv.id).count() or 0)
                purchases_rows.append({
                    'date': inv.date.isoformat() if getattr(inv, 'date', None) else '',
                    'invoice_number': getattr(inv, 'invoice_number', ''),
                    'supplier': getattr(inv, 'supplier_name', '') or '-',
                    'payment_method': getattr(inv, 'payment_method', '') or '-',
                    'status': getattr(inv, 'status', '') or '-',
                    'subtotal': float(getattr(inv, 'total_before_tax', 0) or 0),
                    'vat': float(getattr(inv, 'tax_amount', 0) or 0),
                    'discount': float(getattr(inv, 'discount_amount', 0) or 0),
                    'final_total': float(getattr(inv, 'total_after_tax_discount', 0) or 0),
                    'items_count': items_count,
                })
        except Exception:
            purchases_rows = []

        # Top purchased ingredients and trend (this month)
        top_ingredients = []
        purchase_trend = []
        try:
            rows = (
                db.session.query(PurchaseInvoiceItem.raw_material_name, func.coalesce(func.sum(PurchaseInvoiceItem.total_price), 0))
                .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .group_by(PurchaseInvoiceItem.raw_material_name)
                .order_by(func.coalesce(func.sum(PurchaseInvoiceItem.total_price), 0).desc())
                .limit(10)
                .all()
            )
            top_ingredients = [{'name': n, 'total': float(t or 0)} for n, t in (rows or [])]
            trows = (
                db.session.query(PurchaseInvoice.date, func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0))
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .group_by(PurchaseInvoice.date)
                .order_by(PurchaseInvoice.date.asc())
                .all()
            )
            purchase_trend = [{'date': (d.isoformat() if d else ''), 'total': float(t or 0)} for d, t in (trows or [])]
        except Exception:
            top_ingredients = []
            purchase_trend = []

        stock_rows = []
        try:
            p_qty_rows = (
                db.session.query(PurchaseInvoiceItem.raw_material_id, func.coalesce(func.sum(PurchaseInvoiceItem.quantity), 0))
                .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                .filter(PurchaseInvoice.date >= start_month, PurchaseInvoice.date <= today)
                .group_by(PurchaseInvoiceItem.raw_material_id)
                .all()
            )
            p_qty_map = {int(rm_id): float(qty or 0) for rm_id, qty in (p_qty_rows or []) if rm_id is not None}
            usage_map = {}
            for m in (meals or []):
                sold = float(sales_qty_by_meal.get((getattr(m, 'name', '') or '').strip(), 0) or 0)
                if sold <= 0:
                    continue
                ingrs = MealIngredient.query.filter_by(meal_id=int(m.id)).all()
                for ing in (ingrs or []):
                    usage_map[int(getattr(ing, 'raw_material_id', 0) or 0)] = usage_map.get(int(getattr(ing, 'raw_material_id', 0) or 0), 0.0) + (float(getattr(ing, 'quantity', 0) or 0) * sold)
            for rm in (raw_materials or []):
                rid = int(getattr(rm, 'id', 0) or 0)
                opening_qty = float(getattr(rm, 'stock_quantity', 0) or 0)
                purchases_qty = float(p_qty_map.get(rid, 0) or 0)
                estimated_usage = float(usage_map.get(rid, 0) or 0)
                expected_stock = max(opening_qty + purchases_qty - estimated_usage, 0.0)
                risk = 'low' if expected_stock < max(1.0, opening_qty * 0.1) else ('excess' if purchases_qty > estimated_usage * 1.5 else 'ok')
                stock_rows.append({
                    'ingredient': rm.display_name,
                    'opening_qty': opening_qty,
                    'purchases_qty': purchases_qty,
                    'estimated_usage': estimated_usage,
                    'expected_stock': expected_stock,
                    'risk': risk,
                })
        except Exception:
            stock_rows = []

        return render_template(
            'inventory_intelligence.html',
            kpi={
                'total_inventory_cost': float(total_inventory_cost or 0),
                'total_purchases_month': float(total_purchases_month or 0),
                'total_meals_sold_month': float(total_meals_sold_month or 0),
                'estimated_total_meal_cost': float(estimated_total_meal_cost or 0),
                'total_profit_generated': float(total_profit_generated or 0),
            },
            purchases=purchases_rows,
            top_ingredients=top_ingredients,
            purchase_trend=purchase_trend,
            meals_analysis=meals_analysis,
            outdated_meals=outdated_meals,
            stock_rows=stock_rows,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'فشل تحميل لوحة الذكاء: {e}', 'danger')
        return redirect(url_for('main.inventory'))

@main.route('/api/inventory/intelligence', methods=['GET'], endpoint='api_inventory_intelligence')
@login_required
def api_inventory_intelligence():
    try:
        import math
        # Feature flag and access control
        if not current_app.config.get('INVENTORY_INTEL_ENABLED', False):
            return jsonify({"message": "⚠ Inventory Intelligence is disabled"}), 403
        role = (getattr(current_user, 'role', '') or '').strip().lower()
        username = (getattr(current_user, 'username', '') or '').strip().lower()
        if role != 'admin' and username != 'admin' and getattr(current_user, 'id', None) != 1:
            return jsonify({"error": "Access denied"}), 403

        cost_method = (request.args.get('method') or 'avg').strip().lower()
        locale = (request.args.get('locale') or 'ar').strip().lower()
        sd = (request.args.get('start_date') or '').strip()
        ed = (request.args.get('end_date') or '').strip()
        today = get_saudi_now().date()
        start_date = today.replace(day=1)
        end_date = today
        try:
            if sd:
                start_date = datetime.strptime(sd, '%Y-%m-%d').date()
        except Exception:
            start_date = today.replace(day=1)
        try:
            if ed:
                end_date = datetime.strptime(ed, '%Y-%m-%d').date()
        except Exception:
            end_date = today

        # If not preview mode, return lightweight KPI-only response
        if (request.args.get('preview') or '').strip() != '1':
            try:
                month_purchases_total = float(
                    db.session.query(func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0))
                    .filter(PurchaseInvoice.date >= start_date, PurchaseInvoice.date <= end_date)
                    .scalar() or 0
                )
            except Exception:
                month_purchases_total = 0.0
            try:
                inv_ids = [int(r[0]) for r in db.session.query(SalesInvoice.id).filter(SalesInvoice.date >= start_date, SalesInvoice.date <= end_date).all()] or []
                meals_sold = float(
                    db.session.query(func.coalesce(func.sum(SalesInvoiceItem.quantity), 0))
                    .filter(SalesInvoiceItem.invoice_id.in_(inv_ids) if inv_ids else text('1=0'))
                    .scalar() or 0
                )
            except Exception:
                meals_sold = 0.0
            kpi = {
                'total_inventory_value': 0.0,
                'month_purchases_total': month_purchases_total,
                'meals_sold': meals_sold,
                'estimated_production_cost': 0.0,
                'total_profit': 0.0,
                'label_en': 'KPI Overview',
                'label_ar': 'المؤشرات الرئيسية',
                'note': 'Preview mode required for full analysis (use preview=1)'
            }
            return jsonify({
                'kpi': kpi,
                'purchases_summary': [],
                'meal_analysis': [],
                'stock_analysis': [],
                'alerts': []
            })

        def latest_unit_cost_map():
            rows = (
                db.session.query(PurchaseInvoiceItem.raw_material_id, PurchaseInvoiceItem.price_before_tax, PurchaseInvoice.date)
                .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                .order_by(PurchaseInvoiceItem.raw_material_id.asc(), PurchaseInvoice.date.desc(), PurchaseInvoiceItem.id.desc())
                .all()
            )
            m = {}
            seen = set()
            for rid, price, dt in rows:
                if rid is None:
                    continue
                if rid in seen:
                    continue
                m[int(rid)] = float(price or 0)
                seen.add(int(rid))
            return m

        def avg_unit_cost_map():
            rows = (
                db.session.query(PurchaseInvoiceItem.raw_material_id, func.coalesce(func.sum(PurchaseInvoiceItem.quantity * PurchaseInvoiceItem.price_before_tax), 0), func.coalesce(func.sum(PurchaseInvoiceItem.quantity), 0))
                .group_by(PurchaseInvoiceItem.raw_material_id)
                .all()
            )
            m = {}
            for rid, total_cost, total_qty in rows:
                if rid is None:
                    continue
                denom = float(total_qty or 0)
                m[int(rid)] = (float(total_cost or 0) / denom) if denom > 0 else 0.0
            return m

        def get_purchase_lots(rm_id: int):
            lots = (
                db.session.query(PurchaseInvoiceItem.quantity, PurchaseInvoiceItem.price_before_tax)
                .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
                .filter(PurchaseInvoiceItem.raw_material_id == rm_id)
                .order_by(PurchaseInvoice.date.asc(), PurchaseInvoiceItem.id.asc())
                .all()
            )
            return [{'qty': float(q or 0), 'price': float(p or 0)} for q, p in (lots or [])]

        def fifo_cost(usage_qty: float, lots: list):
            remaining = float(usage_qty or 0)
            cost = 0.0
            for lot in lots:
                if remaining <= 0:
                    break
                take = min(remaining, float(lot.get('qty') or 0))
                cost += take * float(lot.get('price') or 0)
                remaining -= take
            return cost

        latest_cost = latest_unit_cost_map()
        avg_cost = avg_unit_cost_map()
        cost_map = avg_cost if cost_method == 'fifo' and not avg_cost else (avg_cost if cost_method == 'avg' else latest_cost)

        invs = (
            db.session.query(PurchaseInvoice)
            .filter(PurchaseInvoice.date >= start_date, PurchaseInvoice.date <= end_date)
            .order_by(PurchaseInvoice.date.desc(), PurchaseInvoice.id.desc())
            .limit(500)
            .all()
        )
        purchases_summary = []
        for inv in (invs or []):
            items_count = int(PurchaseInvoiceItem.query.filter_by(invoice_id=inv.id).count() or 0)
            purchases_summary.append({
                'date': inv.date.isoformat() if getattr(inv, 'date', None) else '',
                'invoice_no': getattr(inv, 'invoice_number', ''),
                'supplier': getattr(inv, 'supplier_name', '') or '-',
                'subtotal': float(getattr(inv, 'total_before_tax', 0) or 0),
                'vat': float(getattr(inv, 'tax_amount', 0) or 0),
                'discount': float(getattr(inv, 'discount_amount', 0) or 0),
                'final_total': float(getattr(inv, 'total_after_tax_discount', 0) or 0),
                'items_count': items_count,
                'label_en': 'Purchases Summary',
                'label_ar': 'ملخص المشتريات',
            })

        sales_ids = [int(r[0]) for r in (
            db.session.query(SalesInvoice.id)
            .filter(SalesInvoice.date >= start_date, SalesInvoice.date <= end_date)
            .all()
        )]

        name_qty_rev = {}
        if sales_ids:
            rows = (
                db.session.query(SalesInvoiceItem.product_name, func.coalesce(func.sum(SalesInvoiceItem.quantity), 0), func.coalesce(func.sum(SalesInvoiceItem.quantity * SalesInvoiceItem.price_before_tax), 0))
                .filter(SalesInvoiceItem.invoice_id.in_(sales_ids))
                .group_by(SalesInvoiceItem.product_name)
                .all()
            )
            for name, qty, rev in (rows or []):
                k = (name or '').strip()
                name_qty_rev[k] = {'qty_sold': float(qty or 0), 'revenue': float(rev or 0)}

        meals = Meal.query.filter_by(active=True).all()
        def _norm_txt(x: str) -> str:
            import unicodedata, re
            t = (x or '').strip().lower()
            t = t.translate(str.maketrans('٠١٢٣٤٥٦٧٨٩٬٫', '0123456789,.'))
            repl = {'أ':'ا','إ':'ا','آ':'ا','ى':'ي','ؤ':'و','ئ':'ي','ة':'ه'}
            t = ''.join(repl.get(ch, ch) for ch in t)
            t = unicodedata.normalize('NFD', t)
            t = ''.join(ch for ch in t if unicodedata.category(ch) != 'Mn')
            t = re.sub(r"[^0-9a-z\u0621-\u064A ]+", " ", t)
            t = re.sub(r"\s+", " ", t).strip()
            return t
        def _meal_index(meals_list):
            idx = {}
            for m in (meals_list or []):
                n1 = (getattr(m, 'name', '') or '').strip()
                n2 = (getattr(m, 'name_ar', '') or '').strip()
                if n1:
                    idx[_norm_txt(n1)] = m
                if n2:
                    idx[_norm_txt(n2)] = m
                dn = (getattr(m, 'display_name', '') or '').strip()
                if dn:
                    for part in [p.strip() for p in dn.split('/') if p.strip()]:
                        idx[_norm_txt(part)] = m
            return idx
        _idx = _meal_index(meals)
        from difflib import SequenceMatcher
        def _resolve_meal(name):
            if not (name or '').strip():
                return None
            q = _norm_txt(name)
            m = _idx.get(q)
            if m:
                return m
            best = None
            best_score = 0.0
            for k, mv in _idx.items():
                r = SequenceMatcher(None, q, k).ratio()
                if r > best_score:
                    best = mv
                    best_score = r
            if best_score >= 0.86:
                return best
            tokens_q = set(q.split())
            for k, mv in _idx.items():
                toks_k = set(k.split())
                inter = len(tokens_q & toks_k)
                base = max(len(tokens_q), 1)
                if base > 0 and (inter / base) >= 0.7:
                    return mv
            return None
        meal_analysis = []
        total_profit = 0.0
        for m_name, s in name_qty_rev.items():
            meal = _resolve_meal(m_name)
            if not meal:
                continue
            ingrs = MealIngredient.query.filter_by(meal_id=int(meal.id)).all()
            cost_val = 0.0
            for ing in (ingrs or []):
                rid = int(getattr(ing, 'raw_material_id', 0) or 0)
                usage = float(s['qty_sold'] or 0) * float(getattr(ing, 'quantity', 0) or 0)
                if cost_method == 'fifo':
                    cost_val += fifo_cost(usage, get_purchase_lots(rid))
                else:
                    unit_cost = float(cost_map.get(rid, latest_cost.get(rid, 0)) or 0)
                    cost_val += usage * unit_cost
            revenue = float(s['revenue'] or 0)
            profit = revenue - cost_val
            total_profit += profit
            meal_analysis.append({
                'meal_id': int(getattr(meal, 'id', 0) or 0),
                'meal_name': m_name,
                'qty_sold': float(s['qty_sold'] or 0),
                'consumption_cost': float(cost_val or 0),
                'revenue': float(revenue or 0),
                'profit': float(profit or 0),
            })
        for m in meal_analysis:
            rev = float(m.get('revenue') or 0)
            prof = float(m.get('profit') or 0)
            m['margin_pct'] = (prof / rev) if rev > 0 else 0.0
            m['contribution_pct'] = (prof / total_profit) if total_profit > 0 else 0.0

        usage_map = {}
        for m_name, s in name_qty_rev.items():
            meal = _resolve_meal(m_name)
            if not meal:
                continue
            ingrs = MealIngredient.query.filter_by(meal_id=int(meal.id)).all()
            for ing in (ingrs or []):
                rid = int(getattr(ing, 'raw_material_id', 0) or 0)
                usage_map[rid] = usage_map.get(rid, 0.0) + (float(s['qty_sold'] or 0) * float(getattr(ing, 'quantity', 0) or 0))

        opening_map = {int(getattr(rm, 'id', 0) or 0): float(getattr(rm, 'stock_quantity', 0) or 0) for rm in RawMaterial.query.filter_by(active=True).all()}
        p_qty_rows = (
            db.session.query(PurchaseInvoiceItem.raw_material_id, func.coalesce(func.sum(PurchaseInvoiceItem.quantity), 0))
            .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
            .filter(PurchaseInvoice.date >= start_date, PurchaseInvoice.date <= end_date)
            .group_by(PurchaseInvoiceItem.raw_material_id)
            .all()
        )
        purchases_qty = {int(rm_id): float(qty or 0) for rm_id, qty in (p_qty_rows or []) if rm_id is not None}
        last_batch_rows = (
            db.session.query(PurchaseInvoiceItem.raw_material_id, PurchaseInvoiceItem.quantity)
            .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
            .order_by(PurchaseInvoiceItem.raw_material_id.asc(), PurchaseInvoice.date.desc(), PurchaseInvoiceItem.id.desc())
            .all()
        )
        last_batch_qty = {}
        seen = set()
        for rid, q in (last_batch_rows or []):
            if rid is None:
                continue
            if int(rid) in seen:
                continue
            last_batch_qty[int(rid)] = float(q or 0)
            seen.add(int(rid))

        stock_analysis = []
        for rid in set(list(opening_map.keys()) + list(purchases_qty.keys()) + list(usage_map.keys())):
            opening_qty = float(opening_map.get(rid, 0) or 0)
            purchases_q = float(purchases_qty.get(rid, 0) or 0)
            usage_q = float(usage_map.get(rid, 0) or 0)
            expected = opening_qty + purchases_q - usage_q
            last_q = float(last_batch_qty.get(rid, 0) or 0)
            thr = last_q * 0.10
            risk = 'ok'
            if expected < 0:
                risk = 'negative'
            elif expected <= thr:
                risk = 'low'
            stock_analysis.append({
                'ingredient_id': rid,
                'opening_qty': opening_qty,
                'purchases_qty': purchases_q,
                'estimated_usage': usage_q,
                'expected_stock': expected,
                'risk': risk,
                'label_en': 'Stock Analysis',
                'label_ar': 'تحليل المخزون',
            })

        total_inventory_value = 0.0
        for s in stock_analysis:
            rid = int(s.get('ingredient_id') or 0)
            unit_cost = float(latest_cost.get(rid, avg_cost.get(rid, 0)) or 0)
            total_inventory_value += max(float(s.get('expected_stock') or 0), 0.0) * unit_cost

        kpi = {
            'total_inventory_value': float(total_inventory_value or 0),
            'month_purchases_total': float(sum([ps['final_total'] for ps in purchases_summary]) or 0),
            'meals_sold': float(sum([s['qty_sold'] for s in name_qty_rev.values()]) or 0),
            'estimated_production_cost': float(sum([m['consumption_cost'] for m in meal_analysis]) or 0),
            'total_profit': float(sum([m['profit'] for m in meal_analysis]) or 0),
            'label_en': 'KPI Overview',
            'label_ar': 'المؤشرات الرئيسية',
        }

        outdated_ids = []
        rows_last = (
            db.session.query(PurchaseInvoiceItem.raw_material_id, PurchaseInvoice.date)
            .join(PurchaseInvoice, PurchaseInvoice.id == PurchaseInvoiceItem.invoice_id)
            .order_by(PurchaseInvoiceItem.raw_material_id.asc(), PurchaseInvoice.date.desc())
            .all()
        )
        seen = set()
        for rid, dt in (rows_last or []):
            if rid is None:
                continue
            if int(rid) in seen:
                continue
            seen.add(int(rid))
            if not dt or (today - dt).days > 60:
                outdated_ids.append(int(rid))
        alerts = []
        for m in meal_analysis:
            if float(m.get('margin_pct', 0) or 0) < 0.15:
                alerts.append({'type': 'low_margin', 'meal_id': int(m.get('meal_id') or 0), 'label_en': 'Low Margin', 'label_ar': 'هامش منخفض'})
        for rid in outdated_ids:
            alerts.append({'type': 'outdated_cost', 'ingredient_id': rid, 'label_en': 'Outdated Cost', 'label_ar': 'سعر قديم'})
        for s in stock_analysis:
            if (s.get('risk') or '') in ('negative','low'):
                alerts.append({'type': 'stock_risk', 'ingredient_id': int(s.get('ingredient_id') or 0), 'risk': s.get('risk'), 'label_en': 'Stock Risk', 'label_ar': 'مخاطر المخزون'})

        payload = {
            'kpi': kpi,
            'purchases_summary': purchases_summary,
            'meal_analysis': meal_analysis,
            'stock_analysis': stock_analysis,
            'alerts': alerts,
        }
        return jsonify(payload)
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/expenses', methods=['GET', 'POST'], endpoint='expenses')
@login_required
def expenses():
    form = ExpenseInvoiceForm()
    try:
        form.date.data = get_saudi_now().date()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
    if request.method == 'POST':
        try:
            date_str = request.form.get('date') or get_saudi_now().date().isoformat()
            pm = (request.form.get('payment_method') or 'CASH').strip().upper()
            if pm not in ('CASH','BANK'):
                pm = 'CASH'
            status_val = (request.form.get('status') or 'paid').strip().lower()
            if status_val not in ('paid','partial','unpaid'):
                status_val = 'paid'
            idx = 0
            from decimal import Decimal
            def _parse_decimal(s):
                try:
                    t = (s or '').strip()
                    if not t:
                        return Decimal('0')
                    trans = str.maketrans('٠١٢٣٤٥٦٧٨٩٬٫', '0123456789,.')
                    t = t.translate(trans).replace(',', '')
                    return Decimal(t)
                except Exception:
                    return Decimal('0')
            total_before = Decimal('0.00')
            total_tax = Decimal('0.00')
            total_disc = Decimal('0.00')
            items_buffer = []
            while True:
                prefix = f"items-{idx}-"
                desc = request.form.get(prefix + 'description')
                qty = request.form.get(prefix + 'quantity')
                price = request.form.get(prefix + 'price_before_tax')
                tax = request.form.get(prefix + 'tax')
                disc = request.form.get(prefix + 'discount')
                acc_code = request.form.get(prefix + 'account_code')
                if desc is None and qty is None and price is None and tax is None and disc is None:
                    break
                if not desc and not qty and not price:
                    idx += 1
                    continue
                try:
                    qf = _parse_decimal(qty)
                    pf = _parse_decimal(price)
                    tf = _parse_decimal(tax)
                    df = _parse_decimal(disc)
                except Exception:
                    qf = Decimal('0'); pf = Decimal('0'); tf = Decimal('0'); df = Decimal('0')
                line_total = (qf * pf) - df + tf
                total_before += (qf * pf)
                total_tax += tf
                total_disc += df
                try:
                    current_app.logger.info(f"Expense item: desc={(desc or '').strip()} qty={qf} price={pf} tax={tf} disc={df} acc={(acc_code or '').strip().upper()}")
                except Exception:
                    pass
                items_buffer.append({
                    'description': (desc or '').strip(),
                    'quantity': qf,
                    'price_before_tax': pf,
                    'tax': tf,
                    'discount': df,
                    'total_price': line_total,
                    'account_code': (acc_code or '').strip().upper(),
                })
                idx += 1
            try:
                last = ExpenseInvoice.query.order_by(ExpenseInvoice.id.desc()).first()
                seq = (int(getattr(last, 'id', 0)) + 1) if last and getattr(last, 'id', None) else 1
                inv_no = f"INV-EXP-{get_saudi_now().year}-{seq:04d}"
            except Exception:
                inv_no = f"INV-EXP-{get_saudi_now().strftime('%Y%m%d%H%M%S')}"
            inv = ExpenseInvoice(
                invoice_number=inv_no,
                date=datetime.strptime(date_str, '%Y-%m-%d').date(),
                payment_method=pm,
                total_before_tax=total_before,
                tax_amount=total_tax,
                discount_amount=total_disc,
                total_after_tax_discount=(total_before - total_disc + total_tax),
                status=status_val,
                user_id=getattr(current_user, 'id', 1)
            )
            db.session.add(inv)
            db.session.flush()
            for row in items_buffer:
                item = ExpenseInvoiceItem(
                    invoice_id=inv.id,
                    description=row['description'],
                    quantity=row['quantity'] or Decimal('0'),
                    price_before_tax=row['price_before_tax'] or Decimal('0'),
                    tax=row['tax'] or Decimal('0'),
                    discount=row['discount'] or Decimal('0'),
                    total_price=row['total_price'] or Decimal('0')
                )
                db.session.add(item)
            db.session.commit()
            try:
                tax_amt = float(inv.tax_amount or 0.0)
                for row in items_buffer:
                    try:
                        base = float((row.get('quantity',Decimal('0')) or Decimal('0')) * (row.get('price_before_tax',Decimal('0')) or Decimal('0')) - (row.get('discount',Decimal('0')) or Decimal('0')))
                        sel = _expense_account_by_code(row.get('account_code'))
                        if not sel:
                            sel = _expense_account_for(row.get('description',''))
                        acc_code, acc_name, acc_type = sel
                        if base > 0:
                            _post_ledger(inv.date, acc_code, acc_name, acc_type, base, 0.0, f'EXP {inv.invoice_number}')
                    except Exception:
                        continue
                if tax_amt > 0:
                    _post_ledger(inv.date, 'VAT_IN', 'VAT Input', 'tax', tax_amt, 0.0, f'EXP {inv.invoice_number}')
                total_base = float(inv.total_before_tax or 0.0) - float(inv.discount_amount or 0.0)
                _post_ledger(inv.date, 'AP', 'Accounts Payable', 'liability', 0.0, total_base + tax_amt, f'EXP {inv.invoice_number}')
            except Exception:
                pass
            try:
                pay_amt = 0.0
                if status_val == 'paid':
                    pay_amt = float(inv.total_after_tax_discount or 0.0)
                elif status_val == 'partial':
                    pv = request.form.get('partial_paid_amount')
                    try:
                        pv_dec = _parse_decimal(pv)
                        pay_amt = float(pv_dec or Decimal('0'))
                    except Exception:
                        pay_amt = 0.0
                    total_inc_tax = float(inv.total_after_tax_discount or 0.0)
                    if pay_amt > total_inc_tax:
                        pay_amt = total_inc_tax
                if pay_amt > 0.0:
                    db.session.add(Payment(invoice_id=inv.id, invoice_type='expense', amount_paid=pay_amt, payment_method=pm))
                    db.session.commit()
                    _post_ledger(inv.date, 'AP', 'Accounts Payable', 'liability', pay_amt, 0.0, f'PAY EXP {inv.invoice_number}')
                    ca = _pm_account(pm)
                if ca:
                    _post_ledger(inv.date, ca.code, ca.name, 'asset', 0.0, pay_amt, f'PAY EXP {inv.invoice_number}')
            except Exception:
                try:
                    db.session.rollback()
                except Exception:
                    pass
            try:
                _create_expense_journal(inv)
            except Exception:
                pass
            flash('Expense saved', 'success')
        except Exception as e:
            db.session.rollback()
            try:
                current_app.logger.exception('Failed to save expense: %s', e)
                try:
                    current_app.logger.info(f"Expense form meta: date={date_str} pm={pm} status={status_val}")
                except Exception:
                    pass
            except Exception:
                pass
            flash('Failed to save expense', 'danger')
        return redirect(url_for('main.expenses'))
    invs = ExpenseInvoice.query.order_by(ExpenseInvoice.date.desc()).limit(50).all()
    invs_json = []
    try:
        for inv in invs:
            items = []
            for it in (getattr(inv, 'items', []) or []):
                items.append({
                    'description': getattr(it, 'description', ''),
                    'quantity': float(getattr(it, 'quantity', 0) or 0),
                    'price_before_tax': float(getattr(it, 'price_before_tax', 0) or 0),
                    'tax': float(getattr(it, 'tax', 0) or 0),
                    'discount': float(getattr(it, 'discount', 0) or 0),
                    'total_price': float(getattr(it, 'total_price', 0) or 0),
                })
            invs_json.append({
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'date': inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '',
                'payment_method': getattr(inv, 'payment_method', ''),
                'total': float(getattr(inv, 'total_after_tax_discount', 0) or 0),
                'status': getattr(inv, 'status', ''),
                'items': items,
            })
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
    return render_template('expenses.html', form=form, invoices=invs, invoices_json=invs_json)

@main.route('/expenses/delete/<int:eid>', methods=['POST'], endpoint='expense_delete')
@login_required
def expense_delete(eid):
	try:
		inv = ExpenseInvoice.query.get(int(eid))
		if not inv:
			flash('Expense invoice not found', 'warning')
			return redirect(url_for('main.expenses'))
		# Delete children items first
		for it in (inv.items or []):
			try:
				db.session.delete(it)
			except Exception:
				pass
		try:
			Payment.query.filter(Payment.invoice_id == inv.id, Payment.invoice_type == 'expense').delete(synchronize_session=False)
		except Exception:
			pass
		try:
			from models import JournalEntry, JournalLine
			rows = JournalEntry.query.filter(JournalEntry.description.ilike('%{}%'.format(inv.invoice_number))).all()
			for je in (rows or []):
				JournalLine.query.filter(JournalLine.journal_id == je.id).delete(synchronize_session=False)
				db.session.delete(je)
		except Exception:
			pass
		try:
			db.session.query(LedgerEntry).filter(LedgerEntry.description.ilike('%{}%'.format(inv.invoice_number))).delete(synchronize_session=False)
		except Exception:
			pass
		db.session.delete(inv)
		db.session.commit()
		flash('Expense invoice deleted', 'success')
	except Exception:
		db.session.rollback()
		flash('Failed to delete expense invoice', 'danger')
	return redirect(url_for('main.expenses'))

@main.route('/invoices', endpoint='invoices')
@login_required
def invoices():
    t = (request.args.get('type') or 'sales').strip().lower()
    if t not in ('sales','purchases','expenses','all'):
        t = 'sales'

    # Build invoices list for template to render checkboxes and actions
    invoices = []
    try:
        # Helper to compute paid sum
        def paid_sum(kind: str, ids: list):
            if not ids:
                return {}
            rows = db.session.query(
                Payment.invoice_id,
                func.coalesce(func.sum(Payment.amount_paid), 0)
            ).filter(
                Payment.invoice_type == kind,
                Payment.invoice_id.in_(ids)
            ).group_by(Payment.invoice_id).all()
            return {int(i): float(p or 0.0) for i, p in rows}

        if t in ('all', 'sales'):
            try:
                sales = SalesInvoice.query.order_by(SalesInvoice.date.desc()).limit(500).all()
            except Exception:
                sales = []
            s_ids = [int(getattr(s, 'id', 0) or 0) for s in (sales or [])]
            s_paid = paid_sum('sales', s_ids)
            for s in (sales or []):
                cust = (getattr(s, 'customer_name', '') or '').strip().lower()
                def _norm_group(n: str):
                    x = (n or '').lower()
                    if ('hunger' in x) or ('هنقر' in x) or ('هونقر' in x):
                        return 'hunger'
                    if ('keeta' in x) or ('كيتا' in x) or ('كيت' in x):
                        return 'keeta'
                    return ''
                grp = _norm_group(cust)
                total = float(getattr(s, 'total_after_tax_discount', 0) or 0)
                paid = float(s_paid.get(int(getattr(s, 'id', 0) or 0), 0.0))
                remaining = max(total - paid, 0.0)
                # في تبويب المبيعات: اعرض فقط فواتير المنصات غير المدفوعة
                if t == 'sales':
                    if grp not in ('keeta','hunger'):
                        continue
                    if remaining <= 0.0:
                        continue
                    status_calc = 'partial' if paid > 0 else 'unpaid'
                else:
                    # في تبويب all اعرض جميع فواتير المبيعات بالحالة الفعلية
                    if remaining <= 0.0 and (getattr(s, 'status', '') or '').lower() == '':
                        status_calc = 'paid'
                    else:
                        # احترم حالة السجل إن وجدت وإلا احسب من المدفوع
                        st = (getattr(s, 'status', '') or '').lower()
                        if st in ('paid','partial','unpaid'):
                            status_calc = st
                        else:
                            status_calc = ('paid' if remaining <= 0.0 and total > 0 else ('partial' if paid > 0 else 'unpaid'))
                invoices.append({
                    'id': int(getattr(s, 'id', 0) or 0),
                    'invoice_number': getattr(s, 'invoice_number', None) or f"S-{getattr(s, 'id', '')}",
                    'invoice_type': 'sales',
                    'customer_supplier': getattr(s, 'customer_name', None) or '-',
                    'total_amount': total,
                    'paid_amount': paid,
                    'remaining_amount': remaining,
                    'status': status_calc,
                    'due_date': None,
                })

        if t in ('all', 'purchases'):
            try:
                purchases = PurchaseInvoice.query.order_by(PurchaseInvoice.date.desc()).limit(500).all()
            except Exception:
                purchases = []
            p_ids = [int(getattr(p, 'id', 0) or 0) for p in (purchases or [])]
            p_paid = paid_sum('purchase', p_ids)
            for p in (purchases or []):
                total = float(getattr(p, 'total_after_tax_discount', 0) or 0)
                paid = float(p_paid.get(int(getattr(p, 'id', 0) or 0), 0.0))
                invoices.append({
                    'id': int(getattr(p, 'id', 0) or 0),
                    'invoice_number': getattr(p, 'invoice_number', None) or f"P-{getattr(p, 'id', '')}",
                    'invoice_type': 'purchases',
                    'customer_supplier': getattr(p, 'supplier_name', None) or '-',
                    'total_amount': total,
                    'paid_amount': paid,
                    'remaining_amount': max(total - paid, 0.0),
                    'status': (getattr(p, 'status', '') or 'unpaid').lower(),
                    'due_date': None,
                })

        if t in ('all', 'expenses'):
            try:
                expenses = ExpenseInvoice.query.order_by(ExpenseInvoice.date.desc()).limit(500).all()
            except Exception:
                expenses = []
            e_ids = [int(getattr(e, 'id', 0) or 0) for e in (expenses or [])]
            e_paid = paid_sum('expense', e_ids)
            for e in (expenses or []):
                total = float(getattr(e, 'total_after_tax_discount', 0) or 0)
                paid = float(e_paid.get(int(getattr(e, 'id', 0) or 0), 0.0))
                invoices.append({
                    'id': int(getattr(e, 'id', 0) or 0),
                    'invoice_number': getattr(e, 'invoice_number', None) or f"E-{getattr(e, 'id', '')}",
                    'invoice_type': 'expenses',
                    'customer_supplier': 'Expense',
                    'total_amount': total,
                    'paid_amount': paid,
                    'remaining_amount': max(total - paid, 0.0),
                    'status': (getattr(e, 'status', '') or 'paid').lower(),
                    'due_date': None,
                })
    except Exception:
        invoices = []

    return render_template('invoices.html', current_type=t, invoices=invoices)
@main.route('/reports/print/payments', methods=['GET'], endpoint='reports_print_payments')
@login_required
def reports_print_payments():
    try:
        from datetime import datetime as _dt, date as _date
        from sqlalchemy import func
        from models import PurchaseInvoice, ExpenseInvoice, Payment, Settings

        inv_type = (request.args.get('type') or 'purchase').strip().lower()
        if inv_type not in ('purchase','expense'):
            inv_type = 'purchase'

        start_s = request.args.get('start_date') or ''
        end_s = request.args.get('end_date') or ''
        try:
            start_d = _dt.strptime(start_s, '%Y-%m-%d').date() if start_s else _date.min
        except Exception:
            start_d = _date.min
        try:
            end_d = _dt.strptime(end_s, '%Y-%m-%d').date() if end_s else _date.max
        except Exception:
            end_d = _date.max

        rows = []
        totals_by_party = {}

        if inv_type == 'purchase':
            q = PurchaseInvoice.query.filter(PurchaseInvoice.date.between(start_d, end_d))
            for inv in q.order_by(PurchaseInvoice.date.desc(), PurchaseInvoice.id.desc()).all():
                total = float(inv.total_after_tax_discount or 0.0)
                paid = float(
                    db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                      .filter(Payment.invoice_type=='purchase', Payment.invoice_id==inv.id).scalar() or 0.0
                )
                remaining = max(total - paid, 0.0)
                party = getattr(inv, 'supplier_name', '-') or '-'
                rows.append({
                    'BILL NO': inv.invoice_number,
                    'DATE': inv.date.strftime('%Y-%m-%d') if inv.date else '',
                    'SUPPLIER': party,
                    'ITEMS/QTY/TOTAL': f"{len(getattr(inv,'items',[]) or [])} items",
                    'PAID': paid,
                    'REMAINING': remaining,
                    'METHOD': (inv.payment_method or '').upper(),
                    'STATUS': (inv.status or '').upper()
                })
                t = totals_by_party.setdefault(party, {'TOTAL':0.0,'PAID':0.0,'REMAINING':0.0})
                t['TOTAL'] += total; t['PAID'] += paid; t['REMAINING'] += remaining
        else:
            q = ExpenseInvoice.query.filter(ExpenseInvoice.date.between(start_d, end_d))
            for inv in q.order_by(ExpenseInvoice.date.desc(), ExpenseInvoice.id.desc()).all():
                total = float(inv.total_after_tax_discount or 0.0)
                paid = float(
                    db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                      .filter(Payment.invoice_type=='expense', Payment.invoice_id==inv.id).scalar() or 0.0
                )
                remaining = max(total - paid, 0.0)
                party = 'Expense'
                rows.append({
                    'BILL NO': inv.invoice_number,
                    'DATE': inv.date.strftime('%Y-%m-%d') if inv.date else '',
                    'SUPPLIER': party,
                    'ITEMS/QTY/TOTAL': f"{len(getattr(inv,'items',[]) or [])} items",
                    'PAID': paid,
                    'REMAINING': remaining,
                    'METHOD': (inv.payment_method or '').upper(),
                    'STATUS': (inv.status or '').upper()
                })
                t = totals_by_party.setdefault(party, {'TOTAL':0.0,'PAID':0.0,'REMAINING':0.0})
                t['TOTAL'] += total; t['PAID'] += paid; t['REMAINING'] += remaining

        columns = ['BILL NO','DATE','SUPPLIER','ITEMS/QTY/TOTAL','PAID','REMAINING','METHOD','STATUS']
        data = rows
        totals = {
            'PAID': sum(r.get('PAID',0.0) for r in rows),
            'REMAINING': sum(r.get('REMAINING',0.0) for r in rows),
        }
        settings = Settings.query.first()
        # Render with payment totals by method per the template optional block
        payment_totals = {}
        for r in rows:
            pm = r.get('METHOD') or 'UNKNOWN'
            payment_totals[pm] = payment_totals.get(pm, 0.0) + float(r.get('PAID') or 0.0)

        # Supplier totals: show remaining by supplier; adapt labels
        supplier_totals = { k: v['REMAINING'] for k, v in totals_by_party.items() }

        return render_template('print_report.html',
                               report_title=("Purchases" if inv_type=='purchase' else "Expenses"),
                               columns=columns,
                               data=data,
                               totals=totals,
                               totals_columns=['PAID','REMAINING'],
                               totals_colspan=4,
                               settings=settings,
                               start_date=start_s, end_date=end_s,
                               payment_method=request.args.get('payment_method') or 'all',
                               generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
                               item_totals=supplier_totals,
                               item_totals_title=('Supplier Totals' if inv_type=='purchase' else 'Expense Totals'),
                               item_totals_name_label=('Supplier' if inv_type=='purchase' else 'Expense'),
                               item_totals_value_label='Remaining',
                               payment_totals=payment_totals)
    except Exception:
        return redirect(url_for('payments'))


@main.route('/invoices/delete', methods=['POST'], endpoint='invoices_delete')
@login_required
def invoices_delete():
    scope = (request.form.get('scope') or 'selected').strip().lower()
    inv_type = (request.form.get('invoice_type') or request.form.get('current_type') or 'sales').strip().lower()
    ids = [int(x) for x in request.form.getlist('invoice_ids') if str(x).isdigit()]
    deleted = 0
    # If user chose to delete selected invoices but none were selected, do not proceed with bulk delete
    if scope == 'selected' and not ids:
        try:
            flash('لم يتم تحديد أي فاتورة للحذف', 'warning')
        except Exception:
            flash('No selected invoices to delete', 'warning')
        return redirect(url_for('main.invoices', type=inv_type))
    try:
        if inv_type == 'sales':
            q = SalesInvoice.query
            if scope == 'selected' and ids:
                q = q.filter(SalesInvoice.id.in_(ids))
            rows = q.all()
            for inv in rows:
                try:
                    SalesInvoiceItem.query.filter_by(invoice_id=inv.id).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    Payment.query.filter(Payment.invoice_id == inv.id, Payment.invoice_type == 'sales').delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    db.session.query(LedgerEntry).filter(LedgerEntry.description.ilike('%{}%'.format(inv.invoice_number))).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    from models import JournalEntry, JournalLine
                    rows = JournalEntry.query.filter(JournalEntry.description.ilike('%{}%'.format(inv.invoice_number))).all()
                    for je in (rows or []):
                        JournalLine.query.filter(JournalLine.journal_id == je.id).delete(synchronize_session=False)
                        db.session.delete(je)
                except Exception:
                    pass
                db.session.delete(inv)
                deleted += 1
        elif inv_type == 'purchases':
            q = PurchaseInvoice.query
            if scope == 'selected' and ids:
                q = q.filter(PurchaseInvoice.id.in_(ids))
            rows = q.all()
            for inv in rows:
                try:
                    PurchaseInvoiceItem.query.filter_by(invoice_id=inv.id).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    Payment.query.filter(Payment.invoice_id == inv.id, Payment.invoice_type == 'purchase').delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    db.session.query(LedgerEntry).filter(LedgerEntry.description.ilike('%{}%'.format(inv.invoice_number))).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    from models import JournalEntry, JournalLine
                    rows = JournalEntry.query.filter(JournalEntry.description.ilike('%{}%'.format(inv.invoice_number))).all()
                    for je in (rows or []):
                        JournalLine.query.filter(JournalLine.journal_id == je.id).delete(synchronize_session=False)
                        db.session.delete(je)
                except Exception:
                    pass
                db.session.delete(inv)
                deleted += 1
        elif inv_type == 'expenses':
            q = ExpenseInvoice.query
            if scope == 'selected' and ids:
                q = q.filter(ExpenseInvoice.id.in_(ids))
            rows = q.all()
            for inv in rows:
                try:
                    ExpenseInvoiceItem.query.filter_by(invoice_id=inv.id).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    Payment.query.filter(Payment.invoice_id == inv.id, Payment.invoice_type == 'expense').delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    db.session.query(LedgerEntry).filter(LedgerEntry.description.ilike('%{}%'.format(inv.invoice_number))).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    from models import JournalEntry, JournalLine
                    rows = JournalEntry.query.filter(JournalEntry.description.ilike('%{}%'.format(inv.invoice_number))).all()
                    for je in (rows or []):
                        JournalLine.query.filter(JournalLine.journal_id == je.id).delete(synchronize_session=False)
                        db.session.delete(je)
                except Exception:
                    pass
                db.session.delete(inv)
                deleted += 1
        db.session.commit()
        flash(f"Deleted {deleted} invoice(s)", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Delete failed: {e}", 'danger')
    return redirect(url_for('main.invoices', type=inv_type))






@main.route('/employees/<int:eid>/delete', methods=['POST'], endpoint='employee_delete')
@main.route('/employees/delete', methods=['POST'], endpoint='employee_delete_by_query')
@login_required
def employee_delete(eid=None):
    if not user_can('employees','delete'):
        flash('You do not have permission / لا تملك صلاحية الوصول', 'danger')
        return redirect(url_for('main.dashboard'))
    if eid is None:
        eid = request.args.get('id', type=int)
    eid = int(eid)
    try:
        emp = Employee.query.get_or_404(eid)
        # Delete related salary payments, salaries, and defaults, then employee
        sal_rows = Salary.query.filter_by(employee_id=eid).all()
        sal_ids = [s.id for s in sal_rows]
        if sal_ids:
            try:
                Payment.query.filter(Payment.invoice_type == 'salary', Payment.invoice_id.in_(sal_ids)).delete(synchronize_session=False)
            except Exception:
                pass
            Salary.query.filter(Salary.id.in_(sal_ids)).delete(synchronize_session=False)
        try:
            EmployeeSalaryDefault.query.filter_by(employee_id=eid).delete(synchronize_session=False)
        except Exception:
            pass
        db.session.delete(emp)
        db.session.commit()
        flash('Employee and all related records deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting employee: {e}', 'danger')
    return redirect(url_for('main.dashboard'))

@main.route('/employees/create-salary', methods=['POST'], endpoint='employees_create_salary')
@login_required
def employees_create_salary():
    if not user_can('employees','add') and not user_can('salaries','add'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    try:
        emp_id = int(request.form.get('employee_id') or 0)
        year = int(request.form.get('year') or get_saudi_now().year)
        month = int(request.form.get('month') or get_saudi_now().month)
        base = float(request.form.get('basic_salary') or 0)
        allow = float(request.form.get('allowances') or 0)
        ded = float(request.form.get('deductions') or 0)
        prev_due = float(request.form.get('previous_salary_due') or 0)
        total = base + allow - ded + prev_due
        if total < 0:
            total = 0
        emp = None
        try:
            emp = Employee.query.get(emp_id)
        except Exception:
            emp = None
        if not emp:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 400
        sal = Salary(employee_id=emp_id, year=year, month=month,
                     basic_salary=base, allowances=allow, deductions=ded,
                     previous_salary_due=prev_due, total_salary=total,
                     status='due')
        try:
            session = Employee.query.session
        except Exception:
            try:
                session = Salary.query.session
            except Exception:
                session = (ext_db.session if ext_db is not None else db.session)
        session.add(sal)
        session.commit()
        return jsonify({'ok': True, 'salary_id': sal.id}), 200
    except Exception as e:
        try:
            (ext_db.session if ext_db is not None else db.session).rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/salaries/pay', methods=['GET','POST'], endpoint='salaries_pay')
@main.route('/salaries/pay/', methods=['GET','POST'])
@login_required
def salaries_pay():
    # Create/update salary for a month and optionally record a payment
    if request.method == 'POST':
        try:
            emp_id_raw = request.form.get('employee_id')
            month_raw = (request.form.get('month') or request.form.get('pay_month') or get_saudi_now().strftime('%Y-%m'))
            month_str = (month_raw or '').strip()
            year = None
            month = None
            try:
                if '-' in month_str and month_str.count('-') == 1:
                    y, m = month_str.split('-')
                    year = int(y)
                    month = int(m)
                else:
                    from calendar import month_name
                    parts = month_str.replace('/', ' ').split()
                    if len(parts) >= 2:
                        name = parts[0].strip().lower()
                        yval = int(parts[-1])
                        idx = None
                        for i in range(1, 13):
                            if month_name[i].lower() == name:
                                idx = i
                                break
                        if idx:
                            year = yval
                            month = idx
                if not year or not month or month < 1 or month > 12:
                    raise ValueError('bad month')
            except Exception:
                if request.is_json or ('application/json' in (request.headers.get('Accept') or '').lower()):
                    return jsonify({'error': True, 'message': 'صيغة الشهر غير صحيحة. استخدم YYYY-MM مثل 2025-11'}), 400
                year = get_saudi_now().year
                month = get_saudi_now().month
            amount = float(request.form.get('paid_amount') or 0)
            method_raw = (request.form.get('payment_method') or 'cash').strip().lower()
            if method_raw in ('cash','نقدي'):
                method = 'cash'
            elif method_raw in ('bank','بنك','card','visa','بطاقة'):
                method = 'bank'
            else:
                method = 'cash'
            if not (amount > 0):
                error_msg = 'قيمة السداد مطلوبة'
                if request.is_json or ('application/json' in (request.headers.get('Accept') or '').lower()):
                    return jsonify({'error': True, 'message': error_msg}), 400
                flash(error_msg, 'danger')
                return redirect(url_for('main.payroll', year=year, month=month))

            # Validate employee id for single-payment mode
            if (emp_id_raw or '').strip().lower() != 'all':
                if not str(emp_id_raw or '').strip().isdigit():
                    error_msg = 'الموظف غير محدد أو رقم غير صالح'
                    if request.is_json or ('application/json' in (request.headers.get('Accept') or '').lower()):
                        return jsonify({'error': True, 'message': error_msg}), 400
                    flash(error_msg, 'danger')
                    return redirect(url_for('main.payroll', year=year, month=month))
                try:
                    emp_check = Employee.query.get(int(emp_id_raw))
                except Exception:
                    emp_check = None
                if not emp_check:
                    error_msg = 'الموظف غير موجود'
                    if request.is_json or ('application/json' in (request.headers.get('Accept') or '').lower()):
                        return jsonify({'error': True, 'message': error_msg}), 404
                    flash(error_msg, 'danger')
                    return redirect(url_for('main.payroll', year=year, month=month))
            basic_override = None
            allow_override = None
            ded_override   = None
            prev_override  = None

            

            # Determine employees to pay: single or all
            if (emp_id_raw or '').strip().lower() == 'all':
                dept_filter = (request.form.get('department') or '').strip().lower()
                if dept_filter:
                    from sqlalchemy import func
                    employee_ids = [int(e.id) for e in Employee.query.filter(func.lower(Employee.department) == dept_filter).all()]
                else:
                    employee_ids = [int(e.id) for e in Employee.query.all()]
            else:
                employee_ids = [int(emp_id_raw)]

            created_payment_ids = []

            # FIFO distribute payment across arrears -> current -> advance
            if amount > 0:
                # Build ordered list of months from hire date up to current month
                def _start_from(emp_id: int):
                    try:
                        emp_obj = Employee.query.get(emp_id)
                    except Exception:
                        emp_obj = None
                    sy, sm = year, month
                    try:
                        if getattr(emp_obj, 'hire_date', None):
                            sy = int(emp_obj.hire_date.year)
                            sm = int(emp_obj.hire_date.month)
                    except Exception:
                        pass
                    return sy, sm
                # helper iterate months
                def month_iter(y0, m0, y1, m1):
                    y, m = y0, m0
                    cnt = 0
                    while (y < y1) or (y == y1 and m <= m1):
                        yield y, m
                        m += 1
                        if m > 12:
                            m = 1; y += 1
                        cnt += 1
                        if cnt > 240:
                            break
                for _emp_id in employee_ids:
                    remaining_payment = float(amount or 0)
                    start_y, start_m = _start_from(_emp_id)
                    for yy, mm in month_iter(start_y, start_m, year, month):
                        row = Salary.query.filter_by(employee_id=_emp_id, year=yy, month=mm).first()
                        if not row:
                            base = allow = ded = 0.0
                            try:
                                from models import EmployeeSalaryDefault, DepartmentRate, EmployeeHours
                                d = EmployeeSalaryDefault.query.filter_by(employee_id=_emp_id).first()
                                allow = float(getattr(d, 'allowances', 0) or 0)
                                ded = float(getattr(d, 'deductions', 0) or 0)
                                hrs_row = EmployeeHours.query.filter_by(employee_id=_emp_id, year=yy, month=mm).first()
                                emp_o = Employee.query.get(_emp_id)
                                dept_name = (getattr(emp_o, 'department', '') or '').lower()
                                rate_row = DepartmentRate.query.filter(DepartmentRate.name == dept_name).first()
                                hourly_rate = float(getattr(rate_row, 'hourly_rate', 0) or 0)
                                try:
                                    from app.models import AppKV
                                    kv = AppKV.get(f"emp_settings:{int(_emp_id)}") or {}
                                    kv_type = str(kv.get('salary_type','') or '').lower()
                                    kv_rate = float(kv.get('hourly_rate') or 0.0)
                                    if kv_rate > 0:
                                        hourly_rate = kv_rate
                                    if kv_type == 'hourly':
                                        hour_based_base = float(getattr(hrs_row, 'hours', 0) or 0) * hourly_rate if hrs_row else 0.0
                                        base = float(hour_based_base or 0.0)
                                    else:
                                        base = float(getattr(d, 'base_salary', 0) or 0)
                                except Exception:
                                    hour_based_base = float(getattr(hrs_row, 'hours', 0) or 0) * hourly_rate if hrs_row else 0.0
                                    default_base = float(getattr(d, 'base_salary', 0) or 0)
                                    base = hour_based_base if hour_based_base > 0 else default_base
                            except Exception:
                                pass
                            prev_component = 0.0
                            total_amount = max(0.0, base + allow - ded + prev_component)
                            row = Salary(employee_id=_emp_id, year=yy, month=mm,
                                         basic_salary=base, allowances=allow, deductions=ded,
                                         previous_salary_due=prev_component, total_salary=total_amount,
                                         status='due')
                            db.session.add(row)
                            db.session.flush()
                        already_paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0)).
                            filter(Payment.invoice_type=='salary', Payment.invoice_id==row.id).scalar() or 0.0)
                        month_due = max(0.0, float(row.total_salary or 0) - already_paid)
                        if month_due <= 0:
                            row.status = 'paid'
                            continue
                        if remaining_payment <= 0:
                            break
                        pay_amount = remaining_payment if remaining_payment < month_due else month_due
                        if pay_amount > 0:
                            p = Payment(invoice_id=row.id, invoice_type='salary', amount_paid=pay_amount, payment_method=method)
                            db.session.add(p)
                            db.session.flush()
                            try:
                                from models import JournalEntry, JournalLine
                                cash_acc = _pm_account(method)
                                pay_liab = _account(SHORT_TO_NUMERIC['PAYROLL_LIAB'][0], CHART_OF_ACCOUNTS['2130']['name'], CHART_OF_ACCOUNTS['2130']['type'])
                                je = JournalEntry(entry_number=f"JE-SALPAY-{row.id}", date=get_saudi_now().date(), branch_code=None, description=f"Salary payment {row.year}-{row.month} EMP {row.employee_id}", status='posted', total_debit=pay_amount, total_credit=pay_amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None), salary_id=row.id)
                                db.session.add(je); db.session.flush()
                                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=pay_liab.id, debit=pay_amount, credit=0, description='Payroll liability', line_date=get_saudi_now().date(), employee_id=row.employee_id))
                                if cash_acc:
                                    db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=cash_acc.id, debit=0, credit=pay_amount, description='Cash/Bank', line_date=get_saudi_now().date(), employee_id=row.employee_id))
                                db.session.commit()
                            except Exception:
                                db.session.rollback()
                            try:
                                cash_acc = _pm_account(method)
                                # Clear liability on payment: DR Payroll Liabilities, CR Cash/Bank
                                _post_ledger(get_saudi_now().date(), 'PAYROLL_LIAB', 'مستحقات رواتب الموظفين', 'liability', pay_amount, 0.0, f'PAY SAL {row.year}-{row.month} EMP {row.employee_id}')
                                if cash_acc:
                                    _post_ledger(get_saudi_now().date(), cash_acc.code, cash_acc.name, 'asset', 0.0, pay_amount, f'PAY SAL {row.year}-{row.month} EMP {row.employee_id}')
                            except Exception:
                                pass
                            try:
                                created_payment_ids.append(int(p.id))
                            except Exception:
                                pass
                            remaining_payment -= pay_amount
                            already_paid += pay_amount
                            if already_paid >= float(row.total_salary or 0) and float(row.total_salary or 0) > 0:
                                row.status = 'paid'
                            else:
                                row.status = 'partial'
                    if remaining_payment > 0:
                        adv_y = year + (1 if month == 12 else 0)
                        adv_m = 1 if month == 12 else month + 1
                        adv_row = Salary.query.filter_by(employee_id=_emp_id, year=adv_y, month=adv_m).first()
                        if not adv_row:
                            base = allow = ded = 0.0
                            try:
                                from models import EmployeeSalaryDefault
                                d = EmployeeSalaryDefault.query.filter_by(employee_id=_emp_id).first()
                                if d:
                                    base = float(d.base_salary or 0)
                                    allow = float(d.allowances or 0)
                                    ded = float(d.deductions or 0)
                            except Exception:
                                pass
                            adv_row = Salary(employee_id=_emp_id, year=adv_y, month=adv_m,
                                             basic_salary=base, allowances=allow, deductions=ded,
                                             previous_salary_due=0.0,
                                             total_salary=max(0.0, base + allow - ded), status='partial')
                            db.session.add(adv_row)
                            db.session.flush()
                        p2 = Payment(invoice_id=adv_row.id, invoice_type='salary', amount_paid=remaining_payment, payment_method=method)
                        db.session.add(p2)
                        db.session.flush()
                        try:
                            cash_acc = _pm_account(method)
                            # Record employee advance: DR Employee Advances (asset), CR Cash/Bank
                            _post_ledger(get_saudi_now().date(), 'EMP_ADV', 'سلف للموظفين', 'asset', remaining_payment, 0.0, f'ADV EMP {adv_row.employee_id} {adv_row.year}-{adv_row.month}')
                            if cash_acc:
                                _post_ledger(get_saudi_now().date(), cash_acc.code, cash_acc.name, 'asset', 0.0, remaining_payment, f'ADV EMP {adv_row.employee_id} {adv_row.year}-{adv_row.month}')
                            try:
                                from models import JournalEntry, JournalLine
                                emp_adv_acc = _account(SHORT_TO_NUMERIC['EMP_ADV'][0], CHART_OF_ACCOUNTS['1030']['name'], CHART_OF_ACCOUNTS['1030']['type'])
                                je2 = JournalEntry(entry_number=f"JE-ADV-{adv_row.employee_id}-{adv_row.year}{adv_row.month:02d}", date=get_saudi_now().date(), branch_code=None, description=f"Employee advance {adv_row.employee_id} {adv_row.year}-{adv_row.month}", status='posted', total_debit=remaining_payment, total_credit=remaining_payment, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None), salary_id=adv_row.id)
                                db.session.add(je2); db.session.flush()
                                db.session.add(JournalLine(journal_id=je2.id, line_no=1, account_id=emp_adv_acc.id, debit=remaining_payment, credit=0, description='Employee advance', line_date=get_saudi_now().date(), employee_id=adv_row.employee_id))
                                if cash_acc:
                                    db.session.add(JournalLine(journal_id=je2.id, line_no=2, account_id=cash_acc.id, debit=0, credit=remaining_payment, description='Cash/Bank', line_date=get_saudi_now().date(), employee_id=adv_row.employee_id))
                                db.session.commit()
                            except Exception:
                                db.session.rollback()
                        except Exception:
                            pass
                        try:
                            created_payment_ids.append(int(p2.id))
                        except Exception:
                            pass

            # Recompute current month sal status for immediate UI feedback
            # Recompute status for the current salary of the first employee (for UI feedback)
            try:
                first_emp = employee_ids[0]
                sal = Salary.query.filter_by(employee_id=first_emp, year=year, month=month).first() or _ensure_salary_for(first_emp)
                paid_sum = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0)).
                    filter(Payment.invoice_type == 'salary', Payment.invoice_id == sal.id).scalar() or 0)
                total_due = float(sal.total_salary or 0)
                if paid_sum >= total_due and total_due > 0:
                    sal.status = 'paid'
                elif paid_sum > 0:
                    sal.status = 'partial'
                else:
                    sal.status = 'due'
            except Exception:
                pass

            db.session.commit()
            success_msg = 'تم تسجيل السداد'
            if request.is_json or ('application/json' in (request.headers.get('Accept') or '').lower()):
                return jsonify({'success': True, 'message': success_msg, 'payment_ids': created_payment_ids})
            flash(success_msg, 'success')
        except Exception as e:
            db.session.rollback()
            error_msg = f'خطأ في حفظ الراتب/الدفع: {str(e)}'
            if request.is_json or ('application/json' in (request.headers.get('Accept') or '').lower()):
                return jsonify({'error': True, 'message': error_msg}), 400
            flash(error_msg, 'danger')
        # Redirect back to logical screen
        try:
            if (emp_id_raw or '').strip().lower() != 'all' and str(emp_id_raw).isdigit():
                return redirect(url_for('main.pay_salary', emp_id=int(emp_id_raw), year=year, month=month))
        except Exception:
            pass
        return redirect(url_for('main.payroll', year=year, month=month))

    # GET
    try:
        employees = Employee.query.order_by(Employee.full_name.asc()).all()
    except Exception:
        employees = []
    selected_month = request.args.get('month') or get_saudi_now().strftime('%Y-%m')
    selected_employee = request.args.get('employee', type=int)
    # Build defaults map and previous due per employee for the selected month
    defaults_map = {}
    prev_due_map = {}
    try:
        from models import EmployeeSalaryDefault
        for d in EmployeeSalaryDefault.query.all():
            defaults_map[int(d.employee_id)] = {
                'basic_salary': float(d.base_salary or 0.0),
                'allowances': float(d.allowances or 0.0),
                'deductions': float(d.deductions or 0.0)
            }
    except Exception:
        defaults_map = {}
    try:
        y, m = selected_month.split('-'); end_y, end_m = int(y), int(m)
        # previous period is the month before selected
        if end_m == 1:
            prev_y, prev_m = end_y - 1, 12
        else:
            prev_y, prev_m = end_y, end_m - 1

        # Preload defaults for fallback months without Salary rows
        try:
            from models import EmployeeSalaryDefault
            defaults = {int(d.employee_id): d for d in EmployeeSalaryDefault.query.all()}
        except Exception:
            defaults = {}

        def month_iter(y0, m0, y1, m1):
            y, m = y0, m0
            count = 0
            while (y < y1) or (y == y1 and m <= m1):
                yield y, m
                m += 1
                if m > 12:
                    m = 1; y += 1
                count += 1
                if count > 240:  # guard: max 20 years
                    break

        for emp in (employees or []):
            try:
                # No hire date -> no arrears
                if not getattr(emp, 'hire_date', None):
                    prev_due_map[int(emp.id)] = 0.0
                    continue
                start_y, start_m = int(emp.hire_date.year), int(emp.hire_date.month)
                # If hire date after previous period -> nothing due
                if (start_y > prev_y) or (start_y == prev_y and start_m > prev_m):
                    prev_due_map[int(emp.id)] = 0.0
                    continue

                due_sum = 0.0
                paid_sum = 0.0
                for yy, mm in month_iter(start_y, start_m, prev_y, prev_m):
                    s = Salary.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                    if s:
                        basic = float(s.basic_salary or 0.0)
                        allow = float(s.allowances or 0.0)
                        ded = float(s.deductions or 0.0)
                        # Monthly gross for the period only (exclude carry-over to avoid double counting)
                        total = max(0.0, basic + allow - ded)
                        due_sum += total
                        paid = db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0)). \
                            filter(Payment.invoice_type == 'salary', Payment.invoice_id == s.id).scalar() or 0.0
                        paid_sum += float(paid or 0.0)
                    else:
                        d = defaults.get(int(emp.id))
                        if d:
                            base = float(getattr(d, 'base_salary', 0.0) or 0.0)
                            allow = float(getattr(d, 'allowances', 0.0) or 0.0)
                            ded = float(getattr(d, 'deductions', 0.0) or 0.0)
                            due_sum += max(0.0, base + allow - ded)
                        # if no defaults, count as 0 for that month
                prev_due_map[int(emp.id)] = max(0.0, due_sum - paid_sum)
            except Exception:
                prev_due_map[int(emp.id)] = 0.0
    except Exception:
        prev_due_map = {}

    # Optional: list current month salaries
    try:
        y, m = selected_month.split('-'); year = int(y); month = int(m)
        current_salaries = Salary.query.filter_by(year=year, month=month).all()
    except Exception:
        current_salaries = []

    # Selected employee's salary for this month (to prefill form with existing values)
    selected_salary = None
    try:
        if selected_employee and current_salaries:
            for s in current_salaries:
                if int(getattr(s, 'employee_id', 0) or 0) == int(selected_employee):
                    selected_salary = s
                    break
    except Exception:
        selected_salary = None

    # Compute paid sum per salary for coloring Paid/Remaining columns
    paid_map = {}
    try:
        if current_salaries:
            ids = [int(s.id) for s in current_salaries if getattr(s, 'id', None)]
            if ids:
                rows = db.session.query(
                    Payment.invoice_id,
                    func.coalesce(func.sum(Payment.amount_paid), 0)
                ).filter(
                    Payment.invoice_type == 'salary',
                    Payment.invoice_id.in_(ids)
                ).group_by(Payment.invoice_id).all()
                for sid, paid in rows:
                    paid_map[int(sid)] = float(paid or 0.0)
    except Exception:
        paid_map = {}

    return render_template(
        'salaries_pay.html',
        employees=employees,
        month=selected_month,
        salaries=current_salaries,
        selected_employee=selected_employee,
        selected_salary=selected_salary,
        defaults_map=defaults_map,
        prev_due_map=prev_due_map,
        paid_map=paid_map
    )


@main.route('/print/salary-receipt')
@login_required
def salary_receipt():
    # Accept multiple payment ids: ?pids=1,2,3
    pids_param = (request.args.get('pids') or '').strip()
    try:
        ids = [int(x) for x in pids_param.split(',') if x.strip().isdigit()]
    except Exception:
        ids = []
    payments = []
    if ids:
        try:
            payments = Payment.query.filter(Payment.id.in_(ids)).all()
        except Exception:
            payments = []
    # Group by employee and month for display
    items = []
    try:
        for p in (payments or []):
            s = Salary.query.get(int(p.invoice_id)) if getattr(p, 'invoice_type', '') == 'salary' else None
            if not s:
                continue
            emp = Employee.query.get(int(s.employee_id)) if getattr(s, 'employee_id', None) else None
            items.append({
                'employee': getattr(emp, 'full_name', 'Employee'),
                'year': int(getattr(s, 'year', 0) or 0),
                'month': int(getattr(s, 'month', 0) or 0),
                'amount': float(getattr(p, 'amount_paid', 0) or 0),
                'method': getattr(p, 'payment_method', '-') or '-',
                'payment_date': getattr(p, 'payment_date', None),
            })
    except Exception:
        items = []
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    return render_template('print/payroll.html', items=items, settings=settings)


# Alias route to handle legacy or shortened path usage
@main.route('/pay', methods=['POST'], endpoint='pay_alias')
@login_required
def pay_alias():
    return ('', 404)


@main.route('/salaries/statements', methods=['GET'], endpoint='salaries_statements')
@login_required
def salaries_statements():
    # Accept either ?month=YYYY-MM or ?year=&month=
    month_param = (request.args.get('month') or '').strip()
    status_f = (request.args.get('status') or '').strip().lower()
    if '-' in month_param:
        try:
            y, m = month_param.split('-'); year = int(y); month = int(m)
        except Exception:
            year = get_saudi_now().year; month = get_saudi_now().month
    else:
        year = request.args.get('year', type=int) or get_saudi_now().year
        month = request.args.get('month', type=int) or get_saudi_now().month

    from sqlalchemy import func
    rows = []
    totals = {'basic': 0.0, 'allow': 0.0, 'ded': 0.0, 'prev': 0.0, 'total': 0.0, 'paid': 0.0, 'remaining': 0.0}

    try:
        # Load all employees to always show everyone
        employees = Employee.query.order_by(Employee.full_name.asc()).all()

        # Defaults map
        try:
            from models import EmployeeSalaryDefault
            defaults_map = {d.employee_id: d for d in EmployeeSalaryDefault.query.all()}
        except Exception:
            defaults_map = {}

        # Helper month iterator
        def month_iter(y0, m0, y1, m1):
            y, m = y0, m0
            cnt = 0
            while (y < y1) or (y == y1 and m <= m1):
                yield y, m
                m += 1
                if m > 12:
                    m = 1; y += 1
                cnt += 1
                if cnt > 240:
                    break

        for emp in (employees or []):
            # Determine start month
            if getattr(emp, 'hire_date', None):
                y0, m0 = int(emp.hire_date.year), int(emp.hire_date.month)
            else:
                y0, m0 = year, month

            # Build dues per month from hire date to current
            dues_map = {}
            month_keys = []
            sal_ids = []
            for yy, mm in month_iter(y0, m0, year, month):
                month_keys.append((yy, mm))
                srow = Salary.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                if srow:
                    base = float(srow.basic_salary or 0.0)
                    allow = float(srow.allowances or 0.0)
                    ded = float(srow.deductions or 0.0)
                    due_amt = max(0.0, base + allow - ded)
                    sal_ids.append(int(srow.id))
                else:
                    d = defaults_map.get(int(emp.id))
                    base = float(getattr(d, 'base_salary', 0.0) or 0.0)
                    allow = float(getattr(d, 'allowances', 0.0) or 0.0)
                    ded = float(getattr(d, 'deductions', 0.0) or 0.0)
                    due_amt = max(0.0, base + allow - ded)
                dues_map[(yy, mm)] = {'base': base, 'allow': allow, 'ded': ded, 'due': due_amt}

            # Total payments recorded across all months for this employee
            total_paid_all = 0.0
            if sal_ids:
                total_paid_all = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                                       .filter(Payment.invoice_type == 'salary', Payment.invoice_id.in_(sal_ids))
                                       .scalar() or 0.0)

            # FIFO allocation: previous months first, then current
            remaining_payment = total_paid_all
            # previous months
            prev_due_sum = 0.0
            prev_paid_alloc = 0.0
            for yy, mm in month_keys[:-1]:
                due_amt = float(dues_map[(yy, mm)]['due'] or 0.0)
                prev_due_sum += due_amt
                if remaining_payment <= 0:
                    continue
                pay = due_amt if remaining_payment >= due_amt else remaining_payment
                prev_paid_alloc += pay
                remaining_payment -= pay

            # current month
            c_base = float(dues_map[(year, month)]['base'] or 0.0)
            c_allow = float(dues_map[(year, month)]['allow'] or 0.0)
            c_ded = float(dues_map[(year, month)]['ded'] or 0.0)
            c_due = float(dues_map[(year, month)]['due'] or 0.0)
            c_paid_alloc = 0.0
            if remaining_payment > 0 and c_due > 0:
                c_paid_alloc = c_due if remaining_payment >= c_due else remaining_payment
                remaining_payment -= c_paid_alloc

            prev_remaining = max(0.0, prev_due_sum - prev_paid_alloc)
            total = max(0.0, c_due + prev_remaining)
            paid_display = c_paid_alloc
            remaining_total = max(0.0, total - paid_display)

            status = 'paid' if (total > 0 and remaining_total <= 0.01) else ('partial' if paid_display > 0 else ('paid' if total == 0 else 'due'))

            rows.append({
                'employee_id': int(emp.id),
                'employee_name': emp.full_name,
                'basic': c_base,
                'allow': c_allow,
                'ded': c_ded,
                'prev': prev_remaining,
                'total': total,
                'paid': paid_display,
                'remaining': remaining_total,
                'status': status,
                'prev_details': []
            })
            totals['basic'] += c_base
            totals['allow'] += c_allow
            totals['ded'] += c_ded
            totals['prev'] += prev_remaining
            totals['total'] += total
            totals['paid'] += paid_display
            totals['remaining'] += remaining_total
    except Exception:
        pass

    # Apply optional status filter after computing per-employee rows
    if status_f in ('paid','due','partial'):
        try:
            rows = [r for r in rows if (str(r.get('status') or '').lower() == status_f)]
            totals = {
                'basic': sum(float(r.get('basic') or 0) for r in rows),
                'allow': sum(float(r.get('allow') or 0) for r in rows),
                'ded': sum(float(r.get('ded') or 0) for r in rows),
                'prev': sum(float(r.get('prev') or 0) for r in rows),
                'total': sum(float(r.get('total') or 0) for r in rows),
                'paid': sum(float(r.get('paid') or 0) for r in rows),
                'remaining': sum(float(r.get('remaining') or 0) for r in rows),
            }
        except Exception:
            pass

    return render_template('salaries_statements.html', year=year, month=month, rows=rows, totals=totals, status_f=status_f)



@main.route('/api/salaries/statements', methods=['GET'], endpoint='api_salaries_statements')
@login_required
def api_salaries_statements():
    month_param = (request.args.get('month') or '').strip()
    status_f = (request.args.get('status') or '').strip().lower()
    emp_id_f = request.args.get('emp_id', type=int)
    dept_f = (request.args.get('dept') or '').strip()
    if '-' in month_param:
        try:
            y, m = month_param.split('-'); year = int(y); month = int(m)
        except Exception:
            year = get_saudi_now().year; month = get_saudi_now().month
    else:
        year = request.args.get('year', type=int) or get_saudi_now().year
        month = request.args.get('month', type=int) or get_saudi_now().month
    from sqlalchemy import func
    rows = []
    totals = {'basic': 0.0, 'allow': 0.0, 'ded': 0.0, 'prev': 0.0, 'total': 0.0, 'paid': 0.0, 'remaining': 0.0}
    try:
        employees_q = Employee.query.order_by(Employee.full_name.asc())
        if dept_f:
            employees_q = employees_q.filter(func.lower(Employee.department) == dept_f.lower())
        employees = employees_q.all()
        try:
            from models import EmployeeSalaryDefault
            defaults_map = {d.employee_id: d for d in EmployeeSalaryDefault.query.all()}
        except Exception:
            defaults_map = {}
        def month_iter(y0, m0, y1, m1):
            y, m = y0, m0
            cnt = 0
            while (y < y1) or (y == y1 and m <= m1):
                yield y, m
                m += 1
                if m > 12:
                    m = 1; y += 1
                cnt += 1
                if cnt > 240:
                    break
        for emp in (employees or []):
            if emp_id_f and int(emp.id) != int(emp_id_f):
                continue
            if getattr(emp, 'hire_date', None):
                y0, m0 = int(emp.hire_date.year), int(emp.hire_date.month)
            else:
                y0, m0 = year, month
            dues_map = {}
            month_keys = []
            sal_ids = []
            for yy, mm in month_iter(y0, m0, year, month):
                month_keys.append((yy, mm))
                srow = Salary.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                if srow:
                    base = float(srow.basic_salary or 0.0)
                    allow = float(srow.allowances or 0.0)
                    ded = float(srow.deductions or 0.0)
                    due_amt = max(0.0, base + allow - ded)
                    sal_ids.append(int(srow.id))
                else:
                    d = defaults_map.get(int(emp.id))
                    base = float(getattr(d, 'base_salary', 0.0) or 0.0)
                    allow = float(getattr(d, 'allowances', 0.0) or 0.0)
                    ded = float(getattr(d, 'deductions', 0.0) or 0.0)
                    due_amt = max(0.0, base + allow - ded)
                dues_map[(yy, mm)] = {'base': base, 'allow': allow, 'ded': ded, 'due': due_amt}
            total_paid_all = 0.0
            if sal_ids:
                total_paid_all = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                                       .filter(Payment.invoice_type == 'salary', Payment.invoice_id.in_(sal_ids))
                                       .scalar() or 0.0)
            remaining_payment = total_paid_all
            prev_due_sum = 0.0
            prev_paid_alloc = 0.0
            for yy, mm in month_keys[:-1]:
                due_amt = float(dues_map[(yy, mm)]['due'] or 0.0)
                prev_due_sum += due_amt
                if remaining_payment <= 0:
                    continue
                pay = due_amt if remaining_payment >= due_amt else remaining_payment
                prev_paid_alloc += pay
                remaining_payment -= pay
            c_base = float(dues_map[(year, month)]['base'] or 0.0)
            c_allow = float(dues_map[(year, month)]['allow'] or 0.0)
            c_ded = float(dues_map[(year, month)]['ded'] or 0.0)
            c_due = float(dues_map[(year, month)]['due'] or 0.0)
            c_paid_alloc = 0.0
            if remaining_payment > 0 and c_due > 0:
                c_paid_alloc = c_due if remaining_payment >= c_due else remaining_payment
                remaining_payment -= c_paid_alloc
            prev_remaining = max(0.0, prev_due_sum - prev_paid_alloc)
            total = max(0.0, c_due + prev_remaining)
            paid_display = c_paid_alloc
            remaining_total = max(0.0, total - paid_display)
            status = 'paid' if (total > 0 and remaining_total <= 0.01) else ('partial' if paid_display > 0 else ('paid' if total == 0 else 'due'))
            row = {
                'employee_id': int(emp.id),
                'employee_name': emp.full_name,
                'month_label': f"{year:04d}-{month:02d}",
                'year': year,
                'month': month,
                'basic': c_base,
                'allow': c_allow,
                'ded': c_ded,
                'prev': prev_remaining,
                'total': total,
                'paid': paid_display,
                'remaining': remaining_total,
                'status': status
            }
            rows.append(row)
            totals['basic'] += c_base
            totals['allow'] += c_allow
            totals['ded'] += c_ded
            totals['prev'] += prev_remaining
            totals['total'] += total
            totals['paid'] += paid_display
            totals['remaining'] += remaining_total
    except Exception:
        pass
    if status_f in ('paid','due','partial'):
        try:
            rows = [r for r in rows if (str(r.get('status') or '').lower() == status_f)]
            totals = {
                'basic': sum(float(r.get('basic') or 0) for r in rows),
                'allow': sum(float(r.get('allow') or 0) for r in rows),
                'ded': sum(float(r.get('ded') or 0) for r in rows),
                'prev': sum(float(r.get('prev') or 0) for r in rows),
                'total': sum(float(r.get('total') or 0) for r in rows),
                'paid': sum(float(r.get('paid') or 0) for r in rows),
                'remaining': sum(float(r.get('remaining') or 0) for r in rows),
            }
        except Exception:
            pass
    return jsonify({'ok': True, 'year': year, 'month': month, 'rows': rows, 'totals': totals})

@main.route('/api/salaries/upsert', methods=['POST'], endpoint='api_salaries_upsert')
@login_required
def api_salaries_upsert():
    try:
        emp_id = request.form.get('employee_id', type=int)
        month_raw = (request.form.get('month') or '').strip() or get_saudi_now().strftime('%Y-%m')
        absence_hours = float(request.form.get('absence_hours') or 0)
        overtime_hours = float(request.form.get('overtime_hours') or 0)
        bonus_amount = float(request.form.get('bonus_amount') or 0)
        deduction_amount = float(request.form.get('deduction_amount') or 0)
        if not emp_id:
            return jsonify({'ok': False, 'error': 'employee_required'}), 400
        try:
            y, m = month_raw.split('-'); year = int(y); month = int(m)
        except Exception:
            year = get_saudi_now().year; month = get_saudi_now().month
        emp = Employee.query.get(emp_id)
        if not emp:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 404
        from models import EmployeeSalaryDefault, DepartmentRate, EmployeeHours
        d = EmployeeSalaryDefault.query.filter_by(employee_id=int(emp_id)).first()
        base = float(getattr(d, 'base_salary', 0.0) or 0.0)
        dept_name = (getattr(emp, 'department', '') or '').lower()
        rate_row = DepartmentRate.query.filter(DepartmentRate.name == dept_name).first()
        hourly_rate = float(getattr(rate_row, 'hourly_rate', 0.0) or 0.0)
        try:
            from app.models import AppKV
            kv = AppKV.get(f"emp_settings:{int(emp_id)}") or {}
            kv_type = str(kv.get('salary_type','') or '').lower()
            kv_rate = float(kv.get('hourly_rate') or 0.0)
            if kv_rate > 0:
                hourly_rate = kv_rate
            if kv_type == 'hourly':
                hrs = EmployeeHours.query.filter_by(employee_id=int(emp_id), year=year, month=month).first()
                base = float(getattr(hrs, 'hours', 0.0) or 0.0) * float(hourly_rate or 0.0)
        except Exception:
            pass
        if hourly_rate <= 0 and base > 0:
            hourly_rate = base / 240.0
        allow = max(0.0, overtime_hours * hourly_rate) + max(0.0, bonus_amount)
        ded = max(0.0, absence_hours * hourly_rate) + max(0.0, deduction_amount)
        s = Salary.query.filter_by(employee_id=emp_id, year=year, month=month).first()
        prev_due = float(getattr(s, 'previous_salary_due', 0.0) or 0.0) if s else 0.0
        total = max(0.0, base + allow - ded + prev_due)
        if not s:
            s = Salary(employee_id=emp_id, year=year, month=month,
                       basic_salary=base, allowances=allow, deductions=ded,
                       previous_salary_due=prev_due, total_salary=total,
                       status='due')
            db.session.add(s)
        else:
            s.basic_salary = base
            s.allowances = allow
            s.deductions = ded
            s.total_salary = total
            s.status = 'due'
        db.session.commit()
        return jsonify({'ok': True, 'salary_id': int(s.id), 'year': year, 'month': month, 'basic': base, 'allowances': allow, 'deductions': ded, 'total': total})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/salaries/<int:sid>/update', methods=['POST'], endpoint='api_salaries_update')
@login_required
def api_salaries_update(sid: int):
    try:
        if not user_can('salaries', 'edit'):
            return jsonify({'ok': False, 'error': 'forbidden'}), 403
        s = Salary.query.get_or_404(int(sid))
        basic = request.form.get('basic_salary', type=float)
        allow = request.form.get('allowances', type=float)
        ded = request.form.get('deductions', type=float)
        prev = request.form.get('previous_salary_due', type=float)
        status = (request.form.get('status') or '').strip().lower()
        if basic is not None:
            s.basic_salary = float(basic or 0.0)
        if allow is not None:
            s.allowances = float(allow or 0.0)
        if ded is not None:
            s.deductions = float(ded or 0.0)
        if prev is not None:
            s.previous_salary_due = float(prev or 0.0)
        total = max(0.0, float(s.basic_salary or 0.0) + float(s.allowances or 0.0) - float(s.deductions or 0.0) + float(s.previous_salary_due or 0.0))
        s.total_salary = total
        if status in ('paid','due','partial'):
            s.status = status
        db.session.commit()
        return jsonify({'ok': True, 'salary_id': int(s.id), 'total': float(s.total_salary or 0.0), 'status': str(s.status or 'due')})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/salaries/<int:sid>/delete', methods=['POST'], endpoint='api_salaries_delete')
@login_required
def api_salaries_delete(sid: int):
    try:
        if not user_can('salaries', 'delete'):
            return jsonify({'ok': False, 'error': 'forbidden'}), 403
        s = Salary.query.get_or_404(int(sid))
        try:
            Payment.query.filter(Payment.invoice_type == 'salary', Payment.invoice_id == int(sid)).delete(synchronize_session=False)
        except Exception:
            pass
        try:
            from models import JournalEntry
            JournalEntry.query.filter(JournalEntry.salary_id == int(sid)).delete(synchronize_session=False)
        except Exception:
            pass
        db.session.delete(s)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/payroll/seed-test', methods=['POST'], endpoint='api_payroll_seed_test')
@csrf.exempt
def api_payroll_seed_test():
    try:
        from datetime import date as _date, timedelta as _td
        y = get_saudi_now().year
        m = get_saudi_now().month
        created = []
        for idx, cfg in enumerate([
            {'name': 'موظف تجريبي 1', 'nid': f'TST{y}A', 'dept': 'شيف', 'base': 1800.0},
            {'name': 'موظف تجريبي 2', 'nid': f'TST{y}B', 'dept': 'صالة', 'base': 1600.0},
        ]):
            try:
                emp = Employee.query.filter_by(full_name=cfg['name']).first()
            except Exception:
                emp = None
            if not emp:
                hd = (get_saudi_now().date() - _td(days=365))
                emp = Employee(full_name=cfg['name'], national_id=cfg['nid'], department=cfg['dept'], position='موظف', phone='', email='', hire_date=hd, status='active', active=True)
                db.session.add(emp); db.session.flush()
                try:
                    from models import EmployeeSalaryDefault
                    db.session.add(EmployeeSalaryDefault(employee_id=int(emp.id), base_salary=cfg['base'], allowances=0.0, deductions=0.0))
                except Exception:
                    pass
            srow = Salary.query.filter_by(employee_id=int(emp.id), year=y, month=m).first()
            if not srow:
                srow = Salary(employee_id=int(emp.id), year=y, month=m, basic_salary=cfg['base'], allowances=0.0, deductions=0.0, previous_salary_due=0.0, total_salary=cfg['base'], status='due')
                db.session.add(srow); db.session.flush()
            created.append(int(emp.id))
        db.session.commit()
        for i, emp_id in enumerate(created):
            amt_adv = 300.0 if i == 0 else 0.0
            amt_pay = 900.0 if i == 0 else 0.0
            amt_repay = 200.0 if i == 1 else 0.0
            try:
                if amt_adv > 0:
                    _post_ledger(get_saudi_now().date(), 'EMP_ADV', 'سلف للموظفين', 'asset', amt_adv, 0.0, f'ADV EMP {emp_id}')
                if amt_repay > 0:
                    _post_ledger(get_saudi_now().date(), 'EMP_ADV', 'سلف للموظفين', 'asset', 0.0, amt_repay, f'ADV REPAY {emp_id}')
            except Exception:
                pass
            try:
                srow = Salary.query.filter_by(employee_id=int(emp_id), year=y, month=m).first()
                if srow and amt_pay > 0:
                    p = Payment(invoice_id=int(srow.id), invoice_type='salary', amount_paid=amt_pay, payment_method='cash')
                    db.session.add(p); db.session.flush()
                    try:
                        from models import JournalEntry, JournalLine
                        cash_acc = _pm_account('cash')
                        pay_liab = _account(SHORT_TO_NUMERIC['PAYROLL_LIAB'][0], CHART_OF_ACCOUNTS['2130']['name'], CHART_OF_ACCOUNTS['2130']['type'])
                        je = JournalEntry(entry_number=f"JE-SALPAY-{srow.id}", date=get_saudi_now().date(), branch_code=None, description=f"Salary payment {y}-{m} EMP {emp_id}", status='posted', total_debit=amt_pay, total_credit=amt_pay, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None), salary_id=srow.id)
                        db.session.add(je); db.session.flush()
                        db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=pay_liab.id, debit=amt_pay, credit=0.0, description='Payroll liability', line_date=get_saudi_now().date(), employee_id=emp_id))
                        if cash_acc:
                            db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=cash_acc.id, debit=0.0, credit=amt_pay, description='Cash/Bank', line_date=get_saudi_now().date(), employee_id=emp_id))
                        db.session.commit()
                    except Exception:
                        db.session.rollback()
                    try:
                        cash_acc = _pm_account('cash')
                        _post_ledger(get_saudi_now().date(), 'PAYROLL_LIAB', 'مستحقات رواتب الموظفين', 'liability', amt_pay, 0.0, f'PAY SAL {y}-{m} EMP {emp_id}')
                        if cash_acc:
                            _post_ledger(get_saudi_now().date(), cash_acc.code, cash_acc.name, 'asset', 0.0, amt_pay, f'PAY SAL {y}-{m} EMP {emp_id}')
                    except Exception:
                        pass
                    paid_sum = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0)).filter(Payment.invoice_type=='salary', Payment.invoice_id==srow.id).scalar() or 0.0)
                    total_due = float(srow.total_salary or 0.0)
                    if paid_sum >= total_due and total_due > 0:
                        srow.status = 'paid'
                    elif paid_sum > 0:
                        srow.status = 'partial'
                    else:
                        srow.status = 'due'
                    db.session.commit()
            except Exception:
                try:
                    db.session.rollback()
                except Exception:
                    pass
        return jsonify({'ok': True, 'employees': created})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/reports/print/salaries', methods=['GET'], endpoint='reports_print_salaries')
@login_required
def reports_print_salaries():
    # Redirect to detailed payroll statement to ensure new screen is used everywhere
    try:
        q = request.query_string.decode('utf-8') if request.query_string else ''
    except Exception:
        q = ''
    target = url_for('main.reports_print_salaries_detailed')
    if q:
        target = f"{target}?{q}"
    return redirect(target)


@main.route('/reports/print/salaries_detailed', methods=['GET'], endpoint='reports_print_salaries_detailed')
@login_required
def reports_print_salaries_detailed():
    # period input
    month_param = (request.args.get('month') or '').strip()
    if '-' in month_param:
        try:
            y, m = month_param.split('-'); year = int(y); month = int(m)
        except Exception:
            year = get_saudi_now().year; month = get_saudi_now().month
    else:
        year = request.args.get('year', type=int) or get_saudi_now().year
        month = request.args.get('month', type=int) or get_saudi_now().month

    # header
    try:
        settings = Settings.query.first()
        company_name = settings.company_name or 'Company'
    except Exception:
        company_name = 'Company'
    period_label = f"{year:04d}-{month:02d}"

    # Build detailed month-by-month allocations (FIFO: pay oldest dues first)
    employees_data = []
    try:
        # defaults
        try:
            from models import EmployeeSalaryDefault
            defaults_map = {d.employee_id: d for d in EmployeeSalaryDefault.query.all()}
        except Exception:
            defaults_map = {}

        # helper iterator
        def month_iter(y0, m0, y1, m1):
            y, m = y0, m0
            cnt = 0
            while (y < y1) or (y == y1 and m <= m1):
                yield y, m
                m += 1
                if m > 12:
                    m = 1; y += 1
                cnt += 1
                if cnt > 240:
                    break

        for emp in Employee.query.order_by(Employee.full_name.asc()).all():
            # range from hire to current+1 (to place credit if needed)
            if getattr(emp, 'hire_date', None):
                y0, m0 = int(emp.hire_date.year), int(emp.hire_date.month)
            else:
                y0, m0 = year, month

            # gather dues per month and salary ids
            dues = []
            sal_ids = []
            for yy, mm in month_iter(y0, m0, year, month):
                srow = Salary.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                if srow:
                    base = float(srow.basic_salary or 0.0)
                    allow = float(srow.allowances or 0.0)
                    ded = float(srow.deductions or 0.0)
                    due_amt = max(0.0, base + allow - ded)
                    sal_ids.append(int(srow.id))
                else:
                    d = defaults_map.get(int(emp.id))
                    base = float(getattr(d, 'base_salary', 0.0) or 0.0)
                    allow = float(getattr(d, 'allowances', 0.0) or 0.0)
                    ded = float(getattr(d, 'deductions', 0.0) or 0.0)
                    due_amt = max(0.0, base + allow - ded)
                dues.append({'y': yy, 'm': mm, 'base': base, 'allow': allow, 'ded': ded, 'due': due_amt})

            total_paid_all = 0.0
            if sal_ids:
                total_paid_all = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                                       .filter(Payment.invoice_type == 'salary', Payment.invoice_id.in_(sal_ids))
                                       .scalar() or 0.0)

            # FIFO allocate against previous dues first
            remaining_payment = total_paid_all
            detailed_rows = []
            prev_running_due = 0.0
            for d in dues:
                month_due = float(d['due'] or 0.0)
                prev_for_row = float(prev_running_due or 0.0)
                total_for_row = prev_for_row + month_due

                pay = 0.0
                if remaining_payment > 0 and total_for_row > 0:
                    pay = total_for_row if remaining_payment >= total_for_row else remaining_payment
                    remaining_payment -= pay

                remaining = total_for_row - pay

                # Status with tolerance for tiny residuals
                tol = 0.01
                if total_for_row <= tol and pay <= tol:
                    status = 'paid'
                elif remaining <= tol:
                    status = 'paid'
                elif pay > tol:
                    status = 'partial'
                else:
                    status = 'due'

                detailed_rows.append(type('Row', (), {
                    'month': date(d['y'], d['m'], 1),
                    'basic': d['base'],
                    'allowances': d['allow'],
                    'deductions': d['ded'],
                    'prev_due': prev_for_row,
                    'total': total_for_row,
                    'paid_amount': pay,
                    'remaining_amount': remaining,
                    'status': status,
                }))

                # Carry forward any unpaid balance as previous due for next month
                prev_running_due = remaining

            # leftover payment -> credit month (next month)
            if remaining_payment > 0:
                from calendar import monthrange
                nm_y = year + (1 if month == 12 else 0)
                nm_m = 1 if month == 12 else month + 1
                detailed_rows.append(type('Row', (), {
                    'month': date(nm_y, nm_m, 1),
                    'basic': 0.0,
                    'allowances': 0.0,
                    'deductions': 0.0,
                    'prev_due': 0.0,
                    'total': 0.0,
                    'paid_amount': remaining_payment,
                    'remaining_amount': -remaining_payment,
                    'status': 'credit',
                }))

            employees_data.append(type('Emp', (), {
                'name': emp.full_name,
                'salaries': detailed_rows
            }))
    except Exception:
        employees_data = []

    return render_template(
        'print/payroll_statement.html',
        company_name=company_name,
        period_label=period_label,
        generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
        employees=employees_data
    )



@main.route('/employees/payroll', methods=['GET'], endpoint='payroll')
@login_required
def payroll():
    # Exact context and data shape per provided template
    from calendar import month_name
    years = list(range(get_saudi_now().year - 2, get_saudi_now().year + 3))
    year = request.args.get('year', type=int) or get_saudi_now().year
    month = request.args.get('month', type=int) or get_saudi_now().month
    status = (request.args.get('status') or 'all').strip().lower()

    # months as value/label pairs
    months = [{ 'value': i, 'label': month_name[i] } for i in range(1, 13)]
    # department filter
    dept_f = (request.args.get('dept') or '').strip().lower()
    # build department options
    try:
        from sqlalchemy import func as _func
        rows_dept = db.session.query(_func.lower(Employee.department)).filter(Employee.department.isnot(None)).distinct().all()
        dept_options = sorted([str(r[0]) for r in rows_dept if r and r[0]])
    except Exception:
        dept_options = []

    # Build payrolls rows
    payrolls = []
    try:
        from sqlalchemy import func
        q_emps = Employee.query.order_by(Employee.full_name.asc())
        if dept_f:
            q_emps = q_emps.filter(func.lower(Employee.department) == dept_f)
        emps = q_emps.all()
        from models import EmployeeSalaryDefault
        defaults_map = {int(d.employee_id): d for d in EmployeeSalaryDefault.query.all()}
        for emp in (emps or []):
            # current month components
            s = Salary.query.filter_by(employee_id=emp.id, year=year, month=month).first()
            if s:
                basic = float(s.basic_salary or 0)
                allow = float(s.allowances or 0)
                ded = float(s.deductions or 0)
            else:
                d = defaults_map.get(int(emp.id))
                basic = float(getattr(d, 'base_salary', 0) or 0)
                allow = float(getattr(d, 'allowances', 0) or 0)
                ded = float(getattr(d, 'deductions', 0) or 0)
            current_due = max(0.0, basic + allow - ded)

            # previous dues sum (exclude current)
            prev_due = 0.0
            if getattr(emp, 'hire_date', None):
                sy, sm = int(emp.hire_date.year), int(emp.hire_date.month)
            else:
                sy, sm = year, month
            yy, mm = sy, sm
            guard = 0
            while (yy < year) or (yy == year and mm < month):
                s2 = Salary.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                if s2:
                    bb = float(s2.basic_salary or 0); aa = float(s2.allowances or 0); dd = float(s2.deductions or 0)
                else:
                    d2 = defaults_map.get(int(emp.id))
                    bb = float(getattr(d2, 'base_salary', 0) or 0)
                    aa = float(getattr(d2, 'allowances', 0) or 0)
                    dd = float(getattr(d2, 'deductions', 0) or 0)
                prev_due += max(0.0, bb + aa - dd)
                mm += 1
                if mm > 12:
                    mm = 1; yy += 1
                guard += 1
                if guard > 240:
                    break

            # total paid overall across existing salary rows
            sal_ids = [int(x.id) for x in Salary.query.filter_by(employee_id=emp.id).all() if getattr(x, 'id', None)]
            paid_all = 0.0
            if sal_ids:
                paid_all = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                                 .filter(Payment.invoice_type == 'salary', Payment.invoice_id.in_(sal_ids)).scalar() or 0.0)

            # FIFO: allocate to previous first
            remaining_payment = paid_all
            prev_alloc = min(prev_due, remaining_payment)
            remaining_payment -= prev_alloc
            current_alloc = min(current_due, remaining_payment)

            total = current_due + max(0.0, prev_due - prev_alloc)
            remaining = max(0.0, total - current_alloc)
            status_v = 'paid' if (total > 0 and remaining <= 0.01) else ('partial' if current_alloc > 0 else ('paid' if total == 0 else 'due'))

            payrolls.append(type('Row', (), {
                'employee': type('E', (), {'id': int(emp.id), 'full_name': emp.full_name}),
                'basic': basic,
                'allowances': allow,
                'deductions': ded,
                'prev_due': max(0.0, prev_due - prev_alloc),
                'total': total,
                'paid': current_alloc,
                'remaining': remaining,
                'status': status_v,
            }))
    except Exception:
        payrolls = []

    if status in ('paid','due','partial'):
        payrolls = [r for r in payrolls if (str(getattr(r, 'status', '')).lower() == status)]

    # Summary metrics for header cards
    journal_count = 0
    emp_adv_total = 0.0
    try:
        from models import JournalEntry, LedgerEntry, Account
        from sqlalchemy import func as _func
        journal_count = int(JournalEntry.query.filter(JournalEntry.date.between(date(year, month, 1), get_saudi_now().date())).count() or 0)
        adv_code = SHORT_TO_NUMERIC.get('EMP_ADV', ('1030',))[0]
        acc = Account.query.filter(Account.code == adv_code).first()
        if acc:
            row = db.session.query(
                _func.coalesce(_func.sum(LedgerEntry.debit), 0),
                _func.coalesce(_func.sum(LedgerEntry.credit), 0)
            ).filter(LedgerEntry.account_id == acc.id).first()
            dsum, csum = row or (0,0)
            emp_adv_total = max(0.0, float(dsum or 0) - float(csum or 0))
    except Exception:
        journal_count = 0
        emp_adv_total = 0.0

    return render_template('payroll.html', years=years, months=months, year=year, month=month, status=status, payrolls=payrolls, journal_count=journal_count, emp_adv_total=emp_adv_total, dept_options=dept_options, dept_selected=dept_f)


@main.route('/employees/<int:emp_id>/pay', methods=['GET', 'POST'], endpoint='pay_salary')
@login_required
def pay_salary(emp_id):
    emp = Employee.query.get_or_404(emp_id)

    # Support provided template shape and query params (employee_id, year, month)
    selected_employee_id = request.args.get('employee_id', type=int) or int(emp.id)
    year = request.args.get('year', type=int) or get_saudi_now().year
    month = request.args.get('month', type=int) or get_saudi_now().month

    if request.method == 'POST':
        amount = request.form.get('paid_amount', type=float) or 0.0
        method = (request.form.get('payment_method') or 'cash')
        month_str = f"{year:04d}-{month:02d}"
        with current_app.test_request_context('/salaries/pay', method='POST', data={
            'employee_id': str(int(selected_employee_id)),
            'month': month_str,
            'paid_amount': str(amount),
            'payment_method': method,
        }):
            try:
                return salaries_pay()
            except Exception as e:
                flash(f'Error saving salary/payment: {e}', 'danger')
                return redirect(url_for('main.pay_salary', emp_id=emp.id, employee_id=selected_employee_id, year=year, month=month))

    # Build employees list for dropdown
    try:
        employees = Employee.query.order_by(Employee.full_name.asc()).all()
    except Exception:
        employees = []

    # Build salary summary for selected employee
    # Prefer hour-based calculation for the selected period if available
    from models import EmployeeSalaryDefault, DepartmentRate, EmployeeHours
    d = EmployeeSalaryDefault.query.filter_by(employee_id=selected_employee_id).first()
    # Defaults
    default_allow = float(getattr(d, 'allowances', 0) or 0)
    default_ded = float(getattr(d, 'deductions', 0) or 0)
    default_base = float(getattr(d, 'base_salary', 0) or 0)
    # Hour-based override
    try:
        hrs_row = EmployeeHours.query.filter_by(employee_id=selected_employee_id, year=year, month=month).first()
        dept_name = (emp.department or '').lower()
        rate_row = DepartmentRate.query.filter(DepartmentRate.name == dept_name).first()
        hourly_rate = float(getattr(rate_row, 'hourly_rate', 0) or 0)
        hour_based_base = float(getattr(hrs_row, 'hours', 0) or 0) * hourly_rate if hrs_row else 0.0
    except Exception:
        hour_based_base = 0.0
    base = hour_based_base if hour_based_base > 0 else default_base
    allow = default_allow
    ded = default_ded
    # Previous due: sum of previous months' totals minus sum of payments (outstanding arrears)
    from sqlalchemy import func as _func
    try:
        # Sum totals and paid amounts for months strictly before the selected month
        rem_q = db.session.query(
            _func.coalesce(_func.sum(Salary.total_salary), 0),
            _func.coalesce(_func.sum(Payment.amount_paid), 0)
        ).outerjoin(
            Payment, (Payment.invoice_type == 'salary') & (Payment.invoice_id == Salary.id)
        ).filter(
            Salary.employee_id == selected_employee_id
        ).filter(
            (Salary.year < year) | ((Salary.year == year) & (Salary.month < month))
        )
        # Respect hire date if available
        try:
            emp_obj = Employee.query.get(selected_employee_id)
            if getattr(emp_obj, 'hire_date', None):
                hd_y = int(emp_obj.hire_date.year)
                hd_m = int(emp_obj.hire_date.month)
                rem_q = rem_q.filter((Salary.year > hd_y) | ((Salary.year == hd_y) & (Salary.month >= hd_m)))
        except Exception:
            pass
        tot_sum, paid_sum = rem_q.first()
        prev_due = max(0.0, float(tot_sum or 0) - float(paid_sum or 0))
    except Exception:
        prev_due = 0.0

    salary_vm = {
        'basic': base,
        'allowances': allow,
        'deductions': ded,
        'prev_due': prev_due,
        'total': max(0.0, base + allow - ded + prev_due),
    }

    # Determine oldest unpaid month for this employee (FIFO across months)
    next_due_year = year
    next_due_month = month
    try:
        # Determine start (hire date or current selection)
        if getattr(emp, 'hire_date', None):
            start_y, start_m = int(emp.hire_date.year), int(emp.hire_date.month)
        else:
            start_y, start_m = year, month

        # Sum of all payments against this employee's salary rows
        sal_ids_q = Salary.query.with_entities(Salary.id).filter_by(employee_id=selected_employee_id).all()
        sal_id_list = [int(sid) for (sid,) in sal_ids_q] if sal_ids_q else []
        total_paid_all = 0.0
        if sal_id_list:
            total_paid_all = float(db.session.query(_func.coalesce(_func.sum(Payment.amount_paid), 0))
                                   .filter(Payment.invoice_type == 'salary', Payment.invoice_id.in_(sal_id_list))
                                   .scalar() or 0.0)

        # Iterate months from start to selected period; carry previous remaining
        def _month_iter(y0, m0, y1, m1):
            y, m = y0, m0
            cnt = 0
            while (y < y1) or (y == y1 and m <= m1):
                yield y, m
                m += 1
                if m > 12:
                    m = 1; y += 1
                cnt += 1
                if cnt > 240:
                    break

        remaining_payment = total_paid_all
        prev_running = 0.0
        found = False
        for yy, mm in _month_iter(start_y, start_m, year, month):
            srow = Salary.query.filter_by(employee_id=selected_employee_id, year=yy, month=mm).first()
            if srow:
                _b = float(srow.basic_salary or 0.0)
                _a = float(srow.allowances or 0.0)
                _d = float(srow.deductions or 0.0)
                month_due = max(0.0, _b + _a - _d)
            else:
                # Fallback to defaults for months without a stored Salary row
                month_due = max(0.0, base + allow - ded)

            total_for_row = prev_running + month_due
            pay_here = min(remaining_payment, total_for_row)
            remaining_payment -= pay_here
            remaining_row = max(0.0, total_for_row - pay_here)
            if (remaining_row > 0) and (not found):
                next_due_year, next_due_month = yy, mm
                found = True
                break
            prev_running = remaining_row
    except Exception:
        next_due_year, next_due_month = year, month

    from calendar import month_name as _mn
    next_due_label = f"{_mn[next_due_month]}, {next_due_year}"

    # Current month salaries table
    current_month_salaries = []
    try:
        rows = Salary.query.filter_by(year=year, month=month).all()
    except Exception:
        rows = []
    # Build paid map for current month rows
    paid_map = {}
    try:
        ids = [int(r.id) for r in rows if getattr(r, 'id', None)]
        if ids:
            for sid, paid in db.session.query(Payment.invoice_id, _func.coalesce(_func.sum(Payment.amount_paid), 0)).\
                    filter(Payment.invoice_type=='salary', Payment.invoice_id.in_(ids)).group_by(Payment.invoice_id):
                paid_map[int(sid)] = float(paid or 0)
    except Exception:
        paid_map = {}
    for r in (rows or []):
        bb = float(r.basic_salary or 0); aa = float(r.allowances or 0); dd = float(r.deductions or 0)
        prev = float(r.previous_salary_due or 0)
        net = max(0.0, bb + aa - dd + prev)
        paid = float(paid_map.get(int(r.id), 0) or 0)
        remaining = max(0.0, net - paid)
        st = (r.status or '-')
        emp_obj = Employee.query.get(int(r.employee_id)) if getattr(r, 'employee_id', None) else None
        current_month_salaries.append(type('Row', (), {
            'employee': type('E', (), {'full_name': getattr(emp_obj, 'full_name', f'#{r.employee_id}') }),
            'basic': bb,
            'allowances': aa,
            'deductions': dd,
            'prev_due': prev,
            'total': net,
            'paid': paid,
            'remaining': remaining,
            'status': st,
        }))

    from calendar import month_name
    month_name_str = month_name[month]
    return render_template('pay_salary.html',
                           employees=employees,
                           selected_employee_id=selected_employee_id,
                           salary=salary_vm,
                           year=year,
                           month=month,
                           month_name=month_name_str,
                           current_month_salaries=current_month_salaries,
                           next_due_year=next_due_year,
                           next_due_month=next_due_month,
                           next_due_label=next_due_label)


# --- Printing: Payroll summary for a month ---
@main.route('/payroll/print/<int:year>/<int:month>', methods=['GET'], endpoint='payroll_print')
@login_required
def payroll_print(year: int, month: int):
    try:
        rows = Salary.query.filter_by(year=year, month=month).all()
    except Exception:
        rows = []
    # Map Salary rows to view rows expected by template (basic, allowances, deductions, prev_due, total, paid, remaining, status)
    from sqlalchemy import func as _func
    view_rows = []
    try:
        ids = [int(r.id) for r in rows if getattr(r, 'id', None)]
        paid_map = {}
        if ids:
            for sid, paid in db.session.query(Payment.invoice_id, _func.coalesce(_func.sum(Payment.amount_paid), 0)).\
                    filter(Payment.invoice_type=='salary', Payment.invoice_id.in_(ids)).group_by(Payment.invoice_id):
                paid_map[int(sid)] = float(paid or 0)
        for r in (rows or []):
            emp_obj = Employee.query.get(int(r.employee_id)) if getattr(r, 'employee_id', None) else None
            base = float(r.basic_salary or 0); allow = float(r.allowances or 0); ded = float(r.deductions or 0)
            prev = float(r.previous_salary_due or 0)
            total = max(0.0, float(r.total_salary or (base + allow - ded + prev)))
            paid = float(paid_map.get(int(r.id), 0) or 0)
            remaining = max(0.0, total - paid)
            status_v = 'paid' if (total > 0 and remaining <= 0.01) else ('partial' if paid > 0 else ('paid' if total == 0 else 'due'))
            view_rows.append(type('Row', (), {
                'employee': type('E', (), {'full_name': getattr(emp_obj, 'full_name', f'#{r.employee_id}') }),
                'basic': base,
                'allowances': allow,
                'deductions': ded,
                'prev_due': prev,
                'total': total,
                'paid': paid,
                'remaining': remaining,
                'status': status_v,
            }))
    except Exception:
        view_rows = []
    return render_template('payroll-reports.html',
                           mode='monthly',
                           payrolls=view_rows,
                           year=year,
                           month=month,
                           report_title='🧾 Monthly Payroll Statement / كشف رواتب الشهر')


# --- Employees Settings: Hour-based payroll ---
@main.route('/employees/settings', methods=['GET', 'POST'], endpoint='employees_settings')
@login_required
def employees_settings():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    try:
        from models import EmployeeSalaryDefault
    except Exception:
        EmployeeSalaryDefault = None

    # Department management: add/update/delete
    mode = (request.form.get('mode') or '').strip().lower() if request.method == 'POST' else ''
    if request.method == 'POST' and mode in ('add_dept','delete_dept',''):
        # add/update via form (empty mode or add_dept), delete via delete_dept
        name = (request.form.get('dept_name') or request.form.get('new_dept_name') or '').strip().lower()
        rate = request.form.get('hourly_rate', type=float)
        if rate is None:
            rate = request.form.get('new_hourly_rate', type=float) or 0.0
        y_recalc = request.form.get('year', type=int) or year
        m_recalc = request.form.get('month', type=int) or month
        if name:
            if mode == 'delete_dept':
                row = DepartmentRate.query.filter_by(name=name).first()
                if row:
                    db.session.delete(row)
                    db.session.commit()
            else:
                row = DepartmentRate.query.filter_by(name=name).first()
                if row:
                    row.hourly_rate = rate
                else:
                    row = DepartmentRate(name=name, hourly_rate=rate)
                    db.session.add(row)
                db.session.commit()
        # Recalculate all employees for selected period using updated rates
        dept_rates = {dr.name: float(dr.hourly_rate or 0.0) for dr in DepartmentRate.query.all()}
        emps = Employee.query.order_by(Employee.full_name.asc()).all()
        for emp in emps:
            hrs = EmployeeHours.query.filter_by(employee_id=emp.id, year=y_recalc, month=m_recalc).first()
            rate_emp = dept_rates.get((emp.department or '').lower(), 0.0)
            total_basic = float(hrs.hours) * float(rate_emp) if hrs else 0.0
            d = EmployeeSalaryDefault.query.filter_by(employee_id=emp.id).first() if EmployeeSalaryDefault else None
            allow = float(getattr(d, 'allowances', 0.0) or 0.0)
            ded = float(getattr(d, 'deductions', 0.0) or 0.0)
            prev_due = 0.0
            total = max(0.0, total_basic + allow - ded + prev_due)
            s = Salary.query.filter_by(employee_id=emp.id, year=y_recalc, month=m_recalc).first()
            if not s:
                s = Salary(employee_id=emp.id, year=y_recalc, month=m_recalc,
                           basic_salary=total_basic, allowances=allow,
                           deductions=ded, previous_salary_due=prev_due,
                           total_salary=total, status='due')
                db.session.add(s)
            else:
                s.basic_salary = total_basic
                s.allowances = allow
                s.deductions = ded
                s.previous_salary_due = prev_due
                s.total_salary = total
        db.session.commit()
        return redirect(url_for('main.employees_settings', year=y_recalc, month=m_recalc))

    # Save monthly hours and recalc salaries
    if request.method == 'POST' and mode == 'hours':
        year = request.form.get('year', type=int) or year
        month = request.form.get('month', type=int) or month
        dept_rates = {dr.name: float(dr.hourly_rate or 0.0) for dr in DepartmentRate.query.all()}
        emps = Employee.query.order_by(Employee.full_name.asc()).all()
        for emp in emps:
            hours_val = request.form.get(f'hours_{emp.id}', type=float)
            allow_def = request.form.get(f'allow_default_{emp.id}', type=float)
            ded_def = request.form.get(f'ded_default_{emp.id}', type=float)
            if hours_val is None and allow_def is None and ded_def is None:
                continue
            if hours_val is not None:
                rec = EmployeeHours.query.filter_by(employee_id=emp.id, year=year, month=month).first()
                if not rec:
                    rec = EmployeeHours(employee_id=emp.id, year=year, month=month, hours=hours_val)
                    db.session.add(rec)
                else:
                    rec.hours = hours_val
            if (allow_def is not None) or (ded_def is not None):
                d = EmployeeSalaryDefault.query.filter_by(employee_id=emp.id).first() if EmployeeSalaryDefault else None
                if not d and EmployeeSalaryDefault:
                    d = EmployeeSalaryDefault(employee_id=emp.id, base_salary=0.0, allowances=0.0, deductions=0.0)
                    db.session.add(d); db.session.flush()
                if allow_def is not None:
                    d.allowances = float(allow_def or 0.0)
                if ded_def is not None:
                    d.deductions = float(ded_def or 0.0)
        db.session.commit()

        for emp in emps:
            hrs = EmployeeHours.query.filter_by(employee_id=emp.id, year=year, month=month).first()
            rate = dept_rates.get((emp.department or '').lower(), 0.0)
            total_basic = float(getattr(hrs, 'hours', 0.0) or 0.0) * float(rate)
            d = EmployeeSalaryDefault.query.filter_by(employee_id=emp.id).first() if EmployeeSalaryDefault else None
            abs_hours = request.form.get(f'absent_hours_{emp.id}', type=float) or 0.0
            ot_hours = request.form.get(f'overtime_hours_{emp.id}', type=float) or 0.0
            abs_ded = float(abs_hours) * float(rate)
            ot_allow = float(ot_hours) * float(rate)
            allow = float(getattr(d, 'allowances', 0.0) or 0.0) + ot_allow
            ded = float(getattr(d, 'deductions', 0.0) or 0.0) + abs_ded
            prev_due = 0.0
            total = max(0.0, total_basic + allow - ded + prev_due)

            s = Salary.query.filter_by(employee_id=emp.id, year=year, month=month).first()
            if not s:
                s = Salary(employee_id=emp.id, year=year, month=month,
                           basic_salary=total_basic, allowances=allow,
                           deductions=ded, previous_salary_due=prev_due,
                           total_salary=total, status='due')
                db.session.add(s)
            else:
                s.basic_salary = total_basic
                s.allowances = allow
                s.deductions = ded
                s.previous_salary_due = prev_due
                s.total_salary = total
            # Create accrual journal entry automatically if not exists and amount > 0
            try:
                if total > 0:
                    from models import JournalEntry, JournalLine
                    existing = JournalEntry.query.filter(JournalEntry.salary_id == s.id, JournalEntry.description.ilike('%Payroll accrual%')).first()
                    if not existing:
                        exp_acc = _account(SHORT_TO_NUMERIC['SAL_EXP'][0], CHART_OF_ACCOUNTS['5310']['name'], CHART_OF_ACCOUNTS['5310']['type'])
                        liab_acc = _account(SHORT_TO_NUMERIC['PAYROLL_LIAB'][0], CHART_OF_ACCOUNTS['2130']['name'], CHART_OF_ACCOUNTS['2130']['type'])
                        je = JournalEntry(entry_number=f"JE-SALACC-{emp.id}-{year}{month:02d}", date=get_saudi_now().date(), branch_code=None, description=f"Payroll accrual {emp.id} {year}-{month}", status='posted', total_debit=total, total_credit=total, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None), salary_id=s.id)
                        db.session.add(je); db.session.flush()
                        if exp_acc:
                            db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=exp_acc.id, debit=total, credit=0, description='Payroll expense', line_date=get_saudi_now().date(), employee_id=emp.id))
                        if liab_acc:
                            db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=liab_acc.id, debit=0, credit=total, description='Payroll liability', line_date=get_saudi_now().date(), employee_id=emp.id))
                        try:
                            _post_ledger(get_saudi_now().date(), 'SAL_EXP', 'مصروف رواتب', 'expense', total, 0.0, f'ACCRUAL {emp.id} {year}-{month}')
                            _post_ledger(get_saudi_now().date(), 'PAYROLL_LIAB', 'رواتب مستحقة', 'liability', 0.0, total, f'ACCRUAL {emp.id} {year}-{month}')
                        except Exception:
                            pass
            except Exception:
                pass
        db.session.commit()

        flash('تم حفظ الساعات وتحديث الرواتب', 'success')
        return redirect(url_for('main.employees_settings', year=year, month=month))

    # GET: render settings
    dept_rates = DepartmentRate.query.order_by(DepartmentRate.name.asc()).all()
    employees = Employee.query.order_by(Employee.full_name.asc()).all()
    hours = EmployeeHours.query.filter_by(year=year, month=month).all()
    hours_map = {h.employee_id: float(h.hours or 0.0) for h in hours}
    dept_rate_map = {dr.name: float(dr.hourly_rate or 0.0) for dr in dept_rates}
    defaults_map = {int(d.employee_id): d for d in (EmployeeSalaryDefault.query.all() if EmployeeSalaryDefault else [])}
    # Build status map for current month salaries
    status_map = {}
    try:
        rows = Salary.query.filter_by(year=year, month=month).all()
        ids = [int(r.id) for r in rows if getattr(r, 'id', None)]
        paid_map = {}
        if ids:
            from sqlalchemy import func as _func
            for sid, paid in db.session.query(Payment.invoice_id, _func.coalesce(_func.sum(Payment.amount_paid), 0)).\
                    filter(Payment.invoice_type=='salary', Payment.invoice_id.in_(ids)).group_by(Payment.invoice_id):
                paid_map[int(sid)] = float(paid or 0.0)
        for r in (rows or []):
            total = max(0.0, float(r.basic_salary or 0.0) + float(r.allowances or 0.0) - float(r.deductions or 0.0) + float(r.previous_salary_due or 0.0))
            paid = float(paid_map.get(int(r.id), 0.0) or 0.0)
            remaining = max(0.0, total - paid)
            st = 'paid' if (total > 0 and remaining <= 0.01) else ('partial' if paid > 0 else ('paid' if total == 0 else 'due'))
            try:
                status_map[int(r.employee_id)] = st
            except Exception:
                pass
    except Exception:
        status_map = {}
    return render_template('employees_settings.html',
                           dept_rates=dept_rates,
                           employees=employees,
                           hours_map=hours_map,
                           dept_rate_map=dept_rate_map,
                           defaults_map=defaults_map,
                           status_map=status_map,
                           year=year, month=month,
                           hide_back_button=True,
                           hide_logout_button=True)

@main.route('/api/employee/settings', methods=['GET'], endpoint='api_employee_settings_get')
@login_required
def api_employee_settings_get():
    try:
        emp_id = request.args.get('emp_id', type=int)
        if not emp_id:
            return jsonify({'ok': False, 'error': 'employee_required'}), 400
        emp = Employee.query.get(emp_id)
        if not emp:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 404
        s = kv_get(f"emp_settings:{int(emp_id)}", {}) or {}
        work_type = (s.get('work_type') or 'hourly')
        pay_cycle = (s.get('pay_cycle') or 'monthly')
        payment_method = (s.get('payment_method') or 'cash')
        ot_rate = float(s.get('ot_rate') or 0.0)
        allow_allowances = bool(s.get('allow_allowances', True))
        allow_bonuses = bool(s.get('allow_bonuses', True))
        show_in_reports = bool(s.get('show_in_reports', True))
        try:
            dept = (emp.department or '').strip().lower()
            dr = DepartmentRate.query.filter_by(name=dept).first()
            dept_rate = float(getattr(dr, 'hourly_rate', 0.0) or 0.0)
        except Exception:
            dept_rate = 0.0
        hourly_rate_employee = float(s.get('hourly_rate_employee') or 0.0) or dept_rate
        monthly_hours = float(s.get('monthly_hours') or (getattr(emp, 'work_hours', 0) or 0))
        status = (s.get('status') or (emp.status or 'active'))
        return jsonify({'ok': True, 'settings': {
            'work_type': work_type,
            'monthly_hours': monthly_hours,
            'hourly_rate_employee': hourly_rate_employee,
            'status': status,
            'pay_cycle': pay_cycle,
            'payment_method': payment_method,
            'ot_rate': ot_rate,
            'allow_allowances': allow_allowances,
            'allow_bonuses': allow_bonuses,
            'show_in_reports': show_in_reports,
            'dept_rate_fallback': dept_rate
        }})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/employee/settings', methods=['POST'], endpoint='api_employee_settings_save')
@login_required
@csrf.exempt
def api_employee_settings_save():
    try:
        emp_id = request.form.get('emp_id', type=int)
        if not emp_id:
            return jsonify({'ok': False, 'error': 'employee_required'}), 400
        emp = Employee.query.get(emp_id)
        if not emp:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 404
        cur = kv_get(f"emp_settings:{int(emp_id)}", {}) or {}
        fields = ['work_type','monthly_hours','hourly_rate_employee','status','pay_cycle','payment_method','ot_rate','allow_allowances','allow_bonuses','show_in_reports']
        for f in fields:
            if f in request.form:
                val = request.form.get(f)
                if f in ('monthly_hours','hourly_rate_employee','ot_rate'):
                    try:
                        val = float(val)
                    except Exception:
                        val = 0.0
                elif f in ('allow_allowances','allow_bonuses','show_in_reports'):
                    val = True if str(val).lower() in ['1','true','yes','on'] else False
                else:
                    val = (val or '').strip()
                cur[f] = val
        if 'status' in cur:
            emp.status = cur['status'] or (emp.status or 'active')
        if 'monthly_hours' in cur:
            try:
                emp.work_hours = float(cur['monthly_hours'] or 0.0)
            except Exception:
                pass
        db.session.commit()
        kv_set(f"emp_settings:{int(emp_id)}", cur)
        return jsonify({'ok': True, 'settings': cur})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

# --- Department Rates API ---
@main.route('/api/dept-rates', methods=['GET'], endpoint='api_dept_rates_get')
@login_required
def api_dept_rates_get():
    try:
        rows = DepartmentRate.query.order_by(DepartmentRate.name.asc()).all()
        data = []
        for r in rows:
            name = (r.name or '').strip().lower()
            mh_raw = kv_get(f"dept_hours:{name}", None)
            mh = 0.0
            try:
                if isinstance(mh_raw, dict):
                    mh = float(mh_raw.get('monthly_hours') or 0.0)
                elif mh_raw is None:
                    mh = 0.0
                else:
                    mh = float(mh_raw or 0.0)
            except Exception:
                mh = 0.0
            data.append({'name': r.name, 'hourly_rate': float(r.hourly_rate or 0.0), 'monthly_hours': mh})
        return jsonify({'ok': True, 'rows': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/dept-rates', methods=['POST'], endpoint='api_dept_rates_set')
@login_required
@csrf.exempt
def api_dept_rates_set():
    try:
        name = (request.form.get('name') or '').strip().lower()
        rate = request.form.get('hourly_rate', type=float) or 0.0
        monthly_hours = request.form.get('monthly_hours', type=float)
        mode = (request.form.get('mode') or '').strip().lower()
        month_param = (request.form.get('month') or '').strip()
        apply_future = True if str(request.form.get('apply_future', 'false')).lower() in ['1','true','yes','on'] else False
        if '-' in month_param:
            try:
                y, m = month_param.split('-'); year = int(y); month = int(m)
            except Exception:
                year = get_saudi_now().year; month = get_saudi_now().month
        else:
            year = request.form.get('year', type=int) or get_saudi_now().year
            month = request.form.get('month', type=int) or get_saudi_now().month

        if not name:
            return jsonify({'ok': False, 'error': 'dept_name_required'}), 400

        if mode == 'delete':
            row = DepartmentRate.query.filter_by(name=name).first()
            if row:
                db.session.delete(row)
                db.session.commit()
            return jsonify({'ok': True, 'deleted': True})
        else:
            row = DepartmentRate.query.filter_by(name=name).first()
            if row:
                row.hourly_rate = rate
            else:
                row = DepartmentRate(name=name, hourly_rate=rate)
                db.session.add(row)
            db.session.commit()
            if monthly_hours is not None:
                try:
                    kv_set(f"dept_hours:{name}", float(monthly_hours or 0.0))
                except Exception:
                    pass

        # Recalculate salaries for the selected period, only for unpaid ones
        emps = Employee.query.filter(func.lower(Employee.department) == name).all()
        for emp in emps:
            try:
                hrs = EmployeeHours.query.filter_by(employee_id=emp.id, year=year, month=month).first()
                total_basic = float(getattr(hrs, 'hours', 0.0) or 0.0) * float(rate)
                d = EmployeeSalaryDefault.query.filter_by(employee_id=emp.id).first()
                allow = float(getattr(d, 'allowances', 0.0) or 0.0)
                ded = float(getattr(d, 'deductions', 0.0) or 0.0)
                prev_due = 0.0
                total = max(0.0, total_basic + allow - ded + prev_due)
                s = Salary.query.filter_by(employee_id=emp.id, year=year, month=month).first()
                if not s:
                    s = Salary(employee_id=emp.id, year=year, month=month,
                               basic_salary=total_basic, allowances=allow,
                               deductions=ded, previous_salary_due=prev_due,
                               total_salary=total, status='due')
                    db.session.add(s)
                else:
                    if str(getattr(s, 'status', 'due')).lower() != 'paid':
                        s.basic_salary = total_basic
                        s.allowances = allow
                        s.deductions = ded
                        s.previous_salary_due = prev_due
                        s.total_salary = total
            except Exception:
                pass
        db.session.commit()

        # Optionally apply to next period (future)
        if apply_future:
            yy, mm = year, month + 1
            if mm > 12:
                mm = 1; yy += 1
            for emp in emps:
                try:
                    hrs = EmployeeHours.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                    total_basic = float(getattr(hrs, 'hours', 0.0) or 0.0) * float(rate)
                    d = EmployeeSalaryDefault.query.filter_by(employee_id=emp.id).first()
                    allow = float(getattr(d, 'allowances', 0.0) or 0.0)
                    ded = float(getattr(d, 'deductions', 0.0) or 0.0)
                    prev_due = 0.0
                    total = max(0.0, total_basic + allow - ded + prev_due)
                    s = Salary.query.filter_by(employee_id=emp.id, year=yy, month=mm).first()
                    if not s:
                        s = Salary(employee_id=emp.id, year=yy, month=mm,
                                   basic_salary=total_basic, allowances=allow,
                                   deductions=ded, previous_salary_due=prev_due,
                                   total_salary=total, status='due')
                        db.session.add(s)
                    else:
                        if str(getattr(s, 'status', 'due')).lower() != 'paid':
                            s.basic_salary = total_basic
                            s.allowances = allow
                            s.deductions = ded
                            s.previous_salary_due = prev_due
                            s.total_salary = total
                except Exception:
                    pass
            db.session.commit()

        return jsonify({'ok': True, 'name': name, 'hourly_rate': rate, 'year': year, 'month': month, 'apply_future': apply_future})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/dept-reassign', methods=['POST'], endpoint='api_dept_reassign')
@login_required
@csrf.exempt
def api_dept_reassign():
    try:
        from_dept = (request.form.get('from_dept') or '').strip().lower()
        to_dept = (request.form.get('to_dept') or '').strip().lower()
        if not from_dept or not to_dept:
            return jsonify({'ok': False, 'error': 'invalid_dept'}), 400
        emps = Employee.query.filter(func.lower(Employee.department) == from_dept).all()
        for emp in emps:
            emp.department = to_dept
        db.session.commit()
        return jsonify({'ok': True, 'reassigned': len(emps)})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/employee/note', methods=['POST'], endpoint='api_employee_note')
@login_required
@csrf.exempt
def api_employee_note():
    try:
        emp_id = request.form.get('emp_id', type=int)
        note = (request.form.get('note') or '').strip()
        if not emp_id:
            return jsonify({'ok': False, 'error': 'employee_required'}), 400
        kv_set(f"emp_note:{int(emp_id)}", {'note': note})
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# --- Employee monthly hours API ---
@main.route('/api/employee-hours', methods=['GET'], endpoint='api_employee_hours_get')
@login_required
def api_employee_hours_get():
    try:
        month_param = (request.args.get('month') or '').strip()
        if '-' in month_param:
            try:
                y, m = month_param.split('-'); year = int(y); month = int(m)
            except Exception:
                year = get_saudi_now().year; month = get_saudi_now().month
        else:
            year = request.args.get('year', type=int) or get_saudi_now().year
            month = request.args.get('month', type=int) or get_saudi_now().month
        rows = EmployeeHours.query.filter_by(year=year, month=month).all()
        data = [{'employee_id': r.employee_id, 'hours': float(r.hours or 0.0)} for r in rows]
        return jsonify({'ok': True, 'year': year, 'month': month, 'rows': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# --- Printing: Bulk salary receipt for selected employees (current period) ---
@main.route('/salary/receipt', methods=['POST'], endpoint='salary_receipt_bulk')
@login_required
@csrf.exempt
def salary_receipt_bulk():
    from calendar import month_name as _month_name
    # Determine target period
    year = request.form.get('year', type=int) or get_saudi_now().year
    month = request.form.get('month', type=int) or get_saudi_now().month
    selected_ids = request.form.getlist('employee_ids') or []
    try:
        employee_ids = [int(i) for i in selected_ids if str(i).isdigit()]
    except Exception:
        employee_ids = []

    # Build payments-like rows for printing only
    payments = []
    try:
        # Defaults map
        try:
            from models import EmployeeSalaryDefault
            defaults_map = {int(d.employee_id): d for d in EmployeeSalaryDefault.query.all()}
        except Exception:
            defaults_map = {}

        for emp_id in (employee_ids or []):
            emp = Employee.query.get(emp_id)
            if not emp:
                continue
            # Current month salary components
            s = Salary.query.filter_by(employee_id=emp_id, year=year, month=month).first()
            if s:
                base = float(s.basic_salary or 0)
                allow = float(s.allowances or 0)
                ded = float(s.deductions or 0)
                prev = float(s.previous_salary_due or 0)
                total = max(0.0, base + allow - ded + prev)
                # Sum paid against this salary record
                from sqlalchemy import func as _func
                paid = float(db.session.query(_func.coalesce(_func.sum(Payment.amount_paid), 0))
                             .filter(Payment.invoice_type=='salary', Payment.invoice_id==s.id)
                             .scalar() or 0.0)
            else:
                d = defaults_map.get(emp_id)
                base = float(getattr(d, 'base_salary', 0) or 0)
                allow = float(getattr(d, 'allowances', 0) or 0)
                ded = float(getattr(d, 'deductions', 0) or 0)
                prev = 0.0
                total = max(0.0, base + allow - ded)
                paid = 0.0
            net_remaining = max(0.0, total - paid)

            payments.append(type('P', (), {
                'employee': emp,
                'basic': base,
                'allowances': allow,
                'deductions': ded,
                'prev_due': prev,
                'paid_amount': paid,
                # In receipt, "Net Salary" should reflect the month's net total (not remaining)
                'net_salary': total,
                # Optionally expose remaining for templates that may show it later
                'remaining': net_remaining,
                'method': 'cash',
            }))
    except Exception:
        payments = []

    # Company name
    try:
        settings = Settings.query.first()
        company_name = settings.company_name or 'Company'
    except Exception:
        company_name = 'Company'

    return render_template(
        'payroll-reports.html',
        mode='receipt',
        payments=payments,
        company_name=company_name,
        payment_date=get_saudi_now().date(),
        now=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
        report_title='🧾 Salary Payment Receipt / إيصال سداد رواتب'
    )
@main.route('/reports/print/salaries', methods=['GET'], endpoint='reports_print_salaries_legacy')
@login_required
def reports_print_salaries_legacy():
    """Back-compat: redirect to new detailed payroll statement with same query params."""
    try:
        q = request.query_string.decode('utf-8') if request.query_string else ''
    except Exception:
        q = ''
    target = url_for('main.reports_print_salaries_detailed')
    if q:
        target = f"{target}?{q}"
    return redirect(target)

@main.route('/payments', endpoint='payments')
@login_required
def payments():
    # Build unified list of purchase and expense invoices with paid totals
    status_f = (request.args.get('status') or '').strip().lower()
    type_f = (request.args.get('type') or '').strip().lower()
    supplier_f = (request.args.get('supplier') or '').strip().lower()
    group_f = (request.args.get('cust_group') or 'all').strip().lower()
    invoices = []
    PAYMENT_METHODS = ['CASH','CARD','BANK','ONLINE','MADA','VISA']
    sd_str = (request.args.get('start_date') or '2025-10-01').strip()
    ed_str = (request.args.get('end_date') or get_saudi_now().date().isoformat()).strip()
    def _to_ascii_digits(s: str):
        try:
            arabic = '٠١٢٣٤٥٦٧٨٩'
            for i, d in enumerate('0123456789'):
                s = s.replace(arabic[i], d)
            return s
        except Exception:
            return s
    def _parse_date(s: str):
        s = _to_ascii_digits((s or '').strip())
        try:
            if '-' in s:
                return datetime.fromisoformat(s)
            if '/' in s:
                try:
                    return datetime.strptime(s, '%d/%m/%Y')
                except Exception:
                    return datetime.strptime(s, '%m/%d/%Y')
        except Exception:
            pass
        return None
    sd_dt = _parse_date(sd_str) or datetime(get_saudi_now().year, 10, 1)
    ed_dt = _parse_date(ed_str) or get_saudi_now()

    try:
        built_counts = {'purchase':0,'expense':0,'sales':0}
        # Use decimal rounding to avoid float precision issues causing wrong statuses
        from decimal import Decimal, ROUND_HALF_UP
        def to_cents(value):
            try:
                return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            except Exception:
                return Decimal('0.00')
        def compute_status(total_amt, paid_amt):
            total_c = to_cents(total_amt)
            paid_c = to_cents(paid_amt)
            # Treat small residuals (<= 0.01) as fully paid
            if (total_c - paid_c) <= Decimal('0.01'):
                return 'paid'
            if paid_c > Decimal('0.00'):
                return 'partial'
            return 'unpaid'
        q = PurchaseInvoice.query
        try:
            q = q.filter(or_(PurchaseInvoice.created_at.between(sd_dt, ed_dt), PurchaseInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        p_rows = q.order_by(PurchaseInvoice.created_at.desc()).limit(1000).all()
        p_ids = [int(inv.id) for inv in p_rows]
        p_paid = {}
        if p_ids:
            _rows = (db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0))
                     .filter(Payment.invoice_type == 'purchase', Payment.invoice_id.in_(p_ids))
                     .group_by(Payment.invoice_id)
                     .all())
            for i, s in _rows:
                p_paid[int(i)] = float(s or 0.0)
        for inv in p_rows:
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(p_paid.get(int(inv.id), 0.0))
            status_calc = compute_status(total, paid)
            invoices.append({'id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'type': 'purchase', 'party': inv.supplier_name or 'Supplier', 'total': total, 'paid': paid, 'status': status_calc, 'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')})
            built_counts['purchase'] += 1
    except Exception:
        pass

    try:
        q = ExpenseInvoice.query
        try:
            q = q.filter(or_(ExpenseInvoice.created_at.between(sd_dt, ed_dt), ExpenseInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        e_rows = q.order_by(ExpenseInvoice.created_at.desc()).limit(1000).all()
        e_ids = [int(inv.id) for inv in e_rows]
        e_paid = {}
        if e_ids:
            _rows = (db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0))
                     .filter(Payment.invoice_type == 'expense', Payment.invoice_id.in_(e_ids))
                     .group_by(Payment.invoice_id)
                     .all())
            for i, s in _rows:
                e_paid[int(i)] = float(s or 0.0)
        for inv in e_rows:
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(e_paid.get(int(inv.id), 0.0))
            status_calc = compute_status(total, paid)
            invoices.append({'id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'type': 'expense', 'party': 'Expense', 'total': total, 'paid': paid, 'status': status_calc, 'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')})
            built_counts['expense'] += 1
    except Exception:
        pass

    try:
        import re
        def _norm_group(n: str):
            raw = (n or '').lower()
            if ('هنقر' in raw) or ('هونقر' in raw) or ('هَنقَر' in raw):
                return 'hunger'
            if ('كيتا' in raw) or ('كيت' in raw):
                return 'keeta'
            s = re.sub(r'[^a-z]', '', raw)
            if s.startswith('hunger'):
                return 'hunger'
            if s.startswith('keeta') or s.startswith('keet'):
                return 'keeta'
            return ''
        q = SalesInvoice.query
        try:
            q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        s_rows = q.order_by(SalesInvoice.created_at.desc()).limit(1000).all()
        s_ids = [int(inv.id) for inv in s_rows]
        s_paid = {}
        if s_ids:
            _rows = (db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0))
                     .filter(Payment.invoice_type == 'sales', Payment.invoice_id.in_(s_ids))
                     .group_by(Payment.invoice_id)
                     .all())
            for i, s in _rows:
                s_paid[int(i)] = float(s or 0.0)
        for inv in s_rows:
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(s_paid.get(int(inv.id), 0.0))
            status_calc = compute_status(total, paid)
            name = (inv.customer_name or '').strip() or 'Customer'
            grp = _norm_group(name)
            if grp not in ('keeta','hunger'):
                continue
            if group_f in ('keeta','hunger') and grp != group_f:
                continue
            invoices.append({'id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'type': 'sales', 'party': name, 'total': total, 'paid': paid, 'status': status_calc, 'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')})
            built_counts['sales'] += 1
    except Exception:
        pass

    suppliers_groups = []
    suppliers_totals = {'total': 0.0, 'paid': 0.0, 'remaining': 0.0, 'count': 0}
    try:
        by_supplier = {}
        q = PurchaseInvoice.query
        try:
            q = q.filter(or_(PurchaseInvoice.created_at.between(sd_dt, ed_dt), PurchaseInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        q = q.order_by(PurchaseInvoice.created_at.desc()).limit(2000)
        invs = q.all()
        ids = [int(inv.id) for inv in invs]
        paid_map = {}
        if ids:
            _rows = (db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0))
                     .filter(Payment.invoice_type == 'purchase', Payment.invoice_id.in_(ids))
                     .group_by(Payment.invoice_id)
                     .all())
            for i, s in _rows:
                paid_map[int(i)] = float(s or 0.0)
        for inv in invs:
            name = (inv.supplier_name or (getattr(inv, 'supplier', None).name if getattr(inv, 'supplier', None) else '') or 'Supplier').strip()
            if supplier_f and (name.lower().find(supplier_f) == -1):
                continue
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(paid_map.get(int(inv.id), 0.0))
            remaining = max(0.0, total - paid)
            status_calc = compute_status(total, paid)
            if status_f in ('due','unpaid') and status_calc == 'paid':
                continue
            if status_f == 'paid' and status_calc != 'paid':
                continue
            by_supplier.setdefault(name, {'supplier': name, 'invoices': [], 'total': 0.0, 'paid': 0.0, 'remaining': 0.0})
            by_supplier[name]['invoices'].append({
                'id': inv.id,
                'invoice_number': getattr(inv, 'invoice_number', None) or inv.id,
                'type': 'purchase',
                'party': name,
                'total': total,
                'paid': paid,
                'remaining': remaining,
                'status': status_calc,
                'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')
            })
            by_supplier[name]['total'] += total
            by_supplier[name]['paid'] += paid
            by_supplier[name]['remaining'] += remaining
        suppliers_groups = sorted(by_supplier.values(), key=lambda g: g['remaining'], reverse=True)
        suppliers_totals['total'] = sum(g['total'] for g in suppliers_groups)
        suppliers_totals['paid'] = sum(g['paid'] for g in suppliers_groups)
        suppliers_totals['remaining'] = sum(g['remaining'] for g in suppliers_groups)
        suppliers_totals['count'] = len(suppliers_groups)
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        suppliers_groups = []
        suppliers_totals = {'total': 0.0, 'paid': 0.0, 'remaining': 0.0, 'count': 0}

    debtors_groups = []
    debtors_totals = {'total': 0.0, 'paid': 0.0, 'remaining': 0.0, 'count': 0}
    try:
        import re
        def normalize_group(n: str):
            raw = (n or '').lower()
            if ('هنقر' in raw) or ('هونقر' in raw) or ('هَنقَر' in raw):
                return 'hunger'
            if ('كيتا' in raw) or ('كيت' in raw):
                return 'keeta'
            s = re.sub(r'[^a-z]', '', raw)
            if s.startswith('hunger'):
                return 'hunger'
            if s.startswith('keeta') or s.startswith('keet'):
                return 'keeta'
            return ''
        by_group = { 'keeta': {'group': 'keeta', 'invoices': [], 'total': 0.0, 'paid': 0.0, 'remaining': 0.0},
                     'hunger': {'group': 'hunger', 'invoices': [], 'total': 0.0, 'paid': 0.0, 'remaining': 0.0} }
        q = SalesInvoice.query
        try:
            q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        q = q.order_by(SalesInvoice.created_at.desc()).limit(2000)
        invs = q.all()
        ids = [int(inv.id) for inv in invs]
        paid_map = {}
        if ids:
            _rows = (db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0))
                     .filter(Payment.invoice_type == 'sales', Payment.invoice_id.in_(ids))
                     .group_by(Payment.invoice_id)
                     .all())
            for i, s in _rows:
                paid_map[int(i)] = float(s or 0.0)
        for inv in invs:
            grp = normalize_group(inv.customer_name or '')
            if grp not in by_group:
                continue
            if group_f in ('keeta','hunger') and grp != group_f:
                continue
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(paid_map.get(int(inv.id), 0.0))
            remaining = max(0.0, total - paid)
            status_calc = compute_status(total, paid)
            by_group[grp]['invoices'].append({
                'id': inv.id,
                'invoice_number': getattr(inv, 'invoice_number', None) or inv.id,
                'type': 'sales',
                'party': grp,
                'total': total,
                'paid': paid,
                'remaining': remaining,
                'status': status_calc,
                'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')
            })
            by_group[grp]['total'] += total
            by_group[grp]['paid'] += paid
            by_group[grp]['remaining'] += remaining
        debtors_groups = [by_group['keeta'], by_group['hunger']]
        debtors_totals['total'] = sum(g['total'] for g in debtors_groups)
        debtors_totals['paid'] = sum(g['paid'] for g in debtors_groups)
        debtors_totals['remaining'] = sum(g['remaining'] for g in debtors_groups)
        debtors_totals['count'] = sum(1 for g in debtors_groups if g['invoices'])
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        debtors_groups = []
        debtors_totals = {'total': 0.0, 'paid': 0.0, 'remaining': 0.0, 'count': 0}

    # Apply filters
    if status_f:
        invoices = [i for i in invoices if (i.get('status') or '').lower() == status_f]
    if type_f:
        # normalize: purchases->purchase, expenses->expense, sales ignored
        if type_f == 'purchases': type_f = 'purchase'
        if type_f == 'expenses': type_f = 'expense'
        invoices = [i for i in invoices if (i.get('type') or '') == type_f]

    # Fallback: if no rows (or Sales selected), load sales invoices unconditionally within date range
    if ((type_f or '') == 'sales' or not type_f) and not invoices:
        try:
            q = SalesInvoice.query
            try:
                q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt.date(), ed_dt.date())))
            except Exception:
                pass
            invs = q.order_by(SalesInvoice.created_at.desc()).limit(1000).all()
            ids = [int(inv.id) for inv in invs]
            paid_map = {}
            if ids:
                _rows = (db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0))
                         .filter(Payment.invoice_type == 'sales', Payment.invoice_id.in_(ids))
                         .group_by(Payment.invoice_id)
                         .all())
                for i, s in _rows:
                    paid_map[int(i)] = float(s or 0.0)
            for inv in invs:
                total = float(inv.total_after_tax_discount or 0.0)
                paid = float(paid_map.get(int(inv.id), 0.0))
                status_calc = compute_status(total, paid)
                name = (inv.customer_name or '').strip() or 'Customer'
                import re
                raw = name.lower()
                def _norm_group_fallback(n: str):
                    if ('هنقر' in n) or ('هونقر' in n) or ('هَنقَر' in n):
                        return 'hunger'
                    if ('كيتا' in n) or ('كيت' in n):
                        return 'keeta'
                    s = re.sub(r'[^a-z]', '', n)
                    if s.startswith('hunger'):
                        return 'hunger'
                    if s.startswith('keeta') or s.startswith('keet'):
                        return 'keeta'
                    return ''
                g = _norm_group_fallback(raw)
                if g not in ('keeta','hunger'):
                    continue
                if group_f in ('keeta','hunger') and g != group_f:
                    continue
                invoices.append({
                    'id': inv.id,
                    'invoice_number': getattr(inv, 'invoice_number', None) or inv.id,
                    'type': 'sales',
                    'party': name,
                    'total': total,
                    'paid': paid,
                    'status': status_calc,
                    'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')
                })
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
            pass

        # Re-apply status/type filters after fallback to honor user selection
        if status_f:
            invoices = [i for i in invoices if (i.get('status') or '').lower() == status_f]
        if type_f:
            invoices = [i for i in invoices if (i.get('type') or '') == type_f]

    try:
        current_app.logger.info(f"payments range sd={sd_dt} ed={ed_dt} built_counts={built_counts} final_count={len(invoices)} type_f={type_f} status_f={status_f} group_f={group_f}")
    except Exception as e:
        # Log the error for debugging
        import traceback
        print(f"Error in payments logging: {e}")
        traceback.print_exc()
        pass

    return render_template('payments.html', invoices=invoices, PAYMENT_METHODS=PAYMENT_METHODS, status_f=status_f, type_f=type_f,
                           suppliers_groups=suppliers_groups, suppliers_totals=suppliers_totals,
                           debtors_groups=debtors_groups, debtors_totals=debtors_totals, start_date=sd_dt.date().isoformat(), end_date=ed_dt.date().isoformat(), group_f=group_f, supplier_f=supplier_f)

@main.route('/payments.json', methods=['GET'], endpoint='payments_json')
@login_required
def payments_json():
    status_f = (request.args.get('status') or '').strip().lower()
    type_f = (request.args.get('type') or '').strip().lower()
    supplier_f = (request.args.get('supplier') or '').strip().lower()
    group_f = (request.args.get('cust_group') or 'all').strip().lower()
    invoices = []
    sd_str = (request.args.get('start_date') or '2025-10-01').strip()
    ed_str = (request.args.get('end_date') or get_saudi_now().date().isoformat()).strip()
    def _to_ascii_digits(s: str):
        try:
            arabic = '٠١٢٣٤٥٦٧٨٩'
            for i, d in enumerate('0123456789'):
                s = s.replace(arabic[i], d)
            return s
        except Exception:
            return s
    def _parse_date(s: str):
        s = _to_ascii_digits((s or '').strip())
        try:
            if '-' in s:
                return datetime.fromisoformat(s)
            if '/' in s:
                try:
                    return datetime.strptime(s, '%d/%m/%Y')
                except Exception:
                    return datetime.strptime(s, '%m/%d/%Y')
        except Exception:
            pass
        return None
    sd_dt = _parse_date(sd_str) or datetime(get_saudi_now().year, 10, 1)
    ed_dt = _parse_date(ed_str) or get_saudi_now()
    try:
        from decimal import Decimal, ROUND_HALF_UP
        def to_cents(value):
            try:
                return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            except Exception:
                return Decimal('0.00')
        def compute_status(total_amt, paid_amt):
            total_c = to_cents(total_amt)
            paid_c = to_cents(paid_amt)
            if (total_c - paid_c) <= Decimal('0.01'):
                return 'paid'
            if paid_c > Decimal('0.00'):
                return 'partial'
            return 'unpaid'
        q = PurchaseInvoice.query
        try:
            q = q.filter(or_(PurchaseInvoice.created_at.between(sd_dt, ed_dt), PurchaseInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        p_rows = q.order_by(PurchaseInvoice.created_at.desc()).limit(1000).all()
        p_ids = [int(inv.id) for inv in p_rows]
        p_paid = {}
        if p_ids:
            for i, s in db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0)).filter(Payment.invoice_type == 'purchase', Payment.invoice_id.in_(p_ids)).group_by(Payment.invoice_id).all():
                p_paid[int(i)] = float(s or 0.0)
        for inv in p_rows:
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(p_paid.get(int(inv.id), 0.0))
            invoices.append({'id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'type': 'purchase', 'party': inv.supplier_name or 'Supplier', 'total': total, 'paid': paid, 'status': compute_status(total, paid), 'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')})
    except Exception:
        pass
    try:
        q = ExpenseInvoice.query
        try:
            q = q.filter(or_(ExpenseInvoice.created_at.between(sd_dt, ed_dt), ExpenseInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        e_rows = q.order_by(ExpenseInvoice.created_at.desc()).limit(1000).all()
        e_ids = [int(inv.id) for inv in e_rows]
        e_paid = {}
        if e_ids:
            for i, s in db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0)).filter(Payment.invoice_type == 'expense', Payment.invoice_id.in_(e_ids)).group_by(Payment.invoice_id).all():
                e_paid[int(i)] = float(s or 0.0)
        for inv in e_rows:
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(e_paid.get(int(inv.id), 0.0))
            invoices.append({'id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'type': 'expense', 'party': 'Expense', 'total': total, 'paid': paid, 'status': compute_status(total, paid), 'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')})
    except Exception:
        pass
    try:
        import re
        def _norm_group(n: str):
            raw = (n or '').lower()
            if ('هنقر' in raw) or ('هونقر' in raw) or ('هَنقَر' in raw):
                return 'hunger'
            if ('كيتا' in raw) or ('كيت' in raw):
                return 'keeta'
            s = re.sub(r'[^a-z]', '', raw)
            if s.startswith('hunger'):
                return 'hunger'
            if s.startswith('keeta') or s.startswith('keet'):
                return 'keeta'
            return ''
        q = SalesInvoice.query
        try:
            q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        s_rows = q.order_by(SalesInvoice.created_at.desc()).limit(1000).all()
        s_ids = [int(inv.id) for inv in s_rows]
        s_paid = {}
        if s_ids:
            for i, s in db.session.query(Payment.invoice_id, func.coalesce(func.sum(Payment.amount_paid), 0)).filter(Payment.invoice_type == 'sales', Payment.invoice_id.in_(s_ids)).group_by(Payment.invoice_id).all():
                s_paid[int(i)] = float(s or 0.0)
        for inv in s_rows:
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(s_paid.get(int(inv.id), 0.0))
            name = (inv.customer_name or '').strip() or 'Customer'
            grp = _norm_group(name)
            if grp not in ('keeta','hunger'):
                continue
            if group_f in ('keeta','hunger') and grp != group_f:
                continue
            invoices.append({'id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'type': 'sales', 'party': name, 'total': total, 'paid': paid, 'status': compute_status(total, paid), 'date': (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')})
    except Exception:
        pass
    if status_f:
        invoices = [i for i in invoices if (i.get('status') or '').lower() == status_f]
    if type_f:
        if type_f == 'purchases': type_f = 'purchase'
        if type_f == 'expenses': type_f = 'expense'
        invoices = [i for i in invoices if (i.get('type') or '') == type_f]
    return jsonify({'ok': True, 'count': len(invoices), 'invoices': invoices})

@main.route('/payments/export', methods=['GET'], endpoint='payments_export')
@login_required
def payments_export():
    fmt = (request.args.get('format') or 'pdf').strip().lower()
    group_f = (request.args.get('cust_group') or 'all').strip().lower()
    sd_str = (request.args.get('start_date') or '2025-10-01').strip()
    ed_str = (request.args.get('end_date') or get_saudi_now().date().isoformat()).strip()
    def _to_ascii_digits(s: str):
        try:
            arabic = '٠١٢٣٤٥٦٧٨٩'
            for i, d in enumerate('0123456789'):
                s = s.replace(arabic[i], d)
            return s
        except Exception:
            return s
    def _parse_date(s: str):
        s = _to_ascii_digits((s or '').strip())
        try:
            if '-' in s:
                return datetime.fromisoformat(s)
            if '/' in s:
                try:
                    return datetime.strptime(s, '%d/%m/%Y')
                except Exception:
                    return datetime.strptime(s, '%m/%d/%Y')
        except Exception:
            pass
        return None
    sd_dt = _parse_date(sd_str) or datetime(get_saudi_now().year, 10, 1)
    ed_dt = _parse_date(ed_str) or get_saudi_now()

    groups = {'keeta': {'rows': [], 'sums': {'Amount': 0.0, 'Discount': 0.0, 'VAT': 0.0, 'Total': 0.0}},
              'hunger': {'rows': [], 'sums': {'Amount': 0.0, 'Discount': 0.0, 'VAT': 0.0, 'Total': 0.0}}}
    try:
        from sqlalchemy import or_
        import re
        items_sub = db.session.query(
            SalesInvoiceItem.invoice_id.label('inv_id'),
            func.count(SalesInvoiceItem.id).label('items_count'),
            func.sum(SalesInvoiceItem.price_before_tax * SalesInvoiceItem.quantity).label('amount_sum'),
            func.sum(SalesInvoiceItem.tax).label('tax_sum')
        ).group_by(SalesInvoiceItem.invoice_id).subquery()

        def norm_group(n: str):
            raw = (n or '').lower()
            if ('هنقر' in raw) or ('هونقر' in raw) or ('هَنقَر' in raw):
                return 'hunger'
            if ('كيتا' in raw) or ('كيت' in raw):
                return 'keeta'
            s = re.sub(r'[^a-z]', '', raw)
            if s.startswith('hunger'):
                return 'hunger'
            if s.startswith('keeta') or s.startswith('keet'):
                return 'keeta'
            return ''

        q = db.session.query(
            SalesInvoice,
            items_sub.c.items_count,
            items_sub.c.amount_sum,
            items_sub.c.tax_sum
        ).outerjoin(items_sub, items_sub.c.inv_id == SalesInvoice.id)
        try:
            q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        q = q.order_by(SalesInvoice.created_at.desc()).limit(5000)
        for inv, items_count, amount_sum, tax_sum in q.all():
            grp = norm_group(inv.customer_name or '')
            if grp not in ('keeta','hunger'):
                continue
            if group_f in ('keeta','hunger') and grp != group_f:
                continue
            amount = float(amount_sum or 0.0)
            discount = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            vat = float(tax_sum or 0.0)
            total = float(amount - discount + vat)
            dt = getattr(inv, 'date', None)
            date_str = dt.strftime('%Y-%m-%d') if dt else ''
            branch = getattr(inv, 'branch', None) or ''
            row = {
                'Invoice': getattr(inv, 'invoice_number', None) or f"S-{inv.id}",
                'Items': int(items_count or 0),
                'Amount': amount,
                'Discount': discount,
                'Total': total,
                'VAT': vat,
                'Date': date_str,
                'Branch': branch,
            }
            groups[grp]['rows'].append(row)
            groups[grp]['sums']['Amount'] += amount
            groups[grp]['sums']['Discount'] += discount
            groups[grp]['sums']['VAT'] += vat
            groups[grp]['sums']['Total'] += total
    except Exception:
        groups = {'keeta': {'rows': [], 'sums': {'Amount': 0.0, 'Discount': 0.0, 'VAT': 0.0, 'Total': 0.0}},
                  'hunger': {'rows': [], 'sums': {'Amount': 0.0, 'Discount': 0.0, 'VAT': 0.0, 'Total': 0.0}}}

    if fmt == 'excel':
        try:
            import pandas as pd
            from io import BytesIO
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                for name in ['keeta','hunger']:
                    if group_f in ('keeta','hunger') and name != group_f:
                        continue
                    df = pd.DataFrame(groups[name]['rows'], columns=['Invoice','Items','Amount','Discount','Total','VAT','Date','Branch'])
                    if not df.empty:
                        totals_row = {'Invoice': 'Totals', 'Items': '', 'Amount': groups[name]['sums']['Amount'], 'Discount': groups[name]['sums']['Discount'], 'Total': groups[name]['sums']['Total'], 'VAT': groups[name]['sums']['VAT'], 'Date': '', 'Branch': ''}
                        df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
                    df.to_excel(writer, index=False, sheet_name=name.upper())
                sum_df = pd.DataFrame([
                    {'Client': 'KEETA', 'Amount': groups['keeta']['sums']['Amount'], 'Discount': groups['keeta']['sums']['Discount'], 'VAT': groups['keeta']['sums']['VAT'], 'Total': groups['keeta']['sums']['Total']},
                    {'Client': 'HUNGER', 'Amount': groups['hunger']['sums']['Amount'], 'Discount': groups['hunger']['sums']['Discount'], 'VAT': groups['hunger']['sums']['VAT'], 'Total': groups['hunger']['sums']['Total']},
                    {'Client': 'GRAND', 'Amount': groups['keeta']['sums']['Amount'] + groups['hunger']['sums']['Amount'], 'Discount': groups['keeta']['sums']['Discount'] + groups['hunger']['sums']['Discount'], 'VAT': groups['keeta']['sums']['VAT'] + groups['hunger']['sums']['VAT'], 'Total': groups['keeta']['sums']['Total'] + groups['hunger']['sums']['Total']}
                ])
                sum_df.to_excel(writer, index=False, sheet_name='SUMMARY')
            buf.seek(0)
            filename = f"payments_keeta_hunger_{sd_dt.date().isoformat()}_{ed_dt.date().isoformat()}.xlsx"
            return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception:
            return redirect(url_for('main.payments', cust_group=group_f, start_date=sd_dt.date().isoformat(), end_date=ed_dt.date().isoformat(), type='sales'))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from io import BytesIO
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        title = Paragraph('Payments — keeta/hunger', styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        for name in ['keeta','hunger']:
            if group_f in ('keeta','hunger') and name != group_f:
                continue
            elements.append(Paragraph(name.upper(), styles['Heading2']))
            data = [['Invoice','Items','Amount','Discount','Total','VAT','Date','Branch']]
            for r in groups[name]['rows']:
                data.append([
                    r.get('Invoice',''),
                    str(r.get('Items',0)),
                    f"{float(r.get('Amount',0) or 0):.2f}",
                    f"{float(r.get('Discount',0) or 0):.2f}",
                    f"{float(r.get('Total',0) or 0):.2f}",
                    f"{float(r.get('VAT',0) or 0):.2f}",
                    r.get('Date',''),
                    r.get('Branch',''),
                ])
            totals_row = ['Totals','', f"{groups[name]['sums']['Amount']:.2f}", f"{groups[name]['sums']['Discount']:.2f}", f"{groups[name]['sums']['Total']:.2f}", f"{groups[name]['sums']['VAT']:.2f}", '', '']
            data.append(totals_row)
            table = LongTable(data, colWidths=[90,50,70,70,70,70,80,80], repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))
        grand = {
            'Amount': groups['keeta']['sums']['Amount'] + groups['hunger']['sums']['Amount'],
            'Discount': groups['keeta']['sums']['Discount'] + groups['hunger']['sums']['Discount'],
            'VAT': groups['keeta']['sums']['VAT'] + groups['hunger']['sums']['VAT'],
            'Total': groups['keeta']['sums']['Total'] + groups['hunger']['sums']['Total']
        }
        elements.append(Paragraph('GRAND TOTALS', styles['Heading2']))
        gdata = [['Amount','Discount','Total','VAT'], [f"{grand['Amount']:.2f}", f"{grand['Discount']:.2f}", f"{grand['Total']:.2f}", f"{grand['VAT']:.2f}"]]
        gtable = LongTable(gdata, colWidths=[100,100,100,100], repeatRows=1)
        gtable.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(gtable)
        doc.build(elements)
        buf.seek(0)
        filename = f"payments_keeta_hunger_{sd_dt.date().isoformat()}_{ed_dt.date().isoformat()}.pdf"
        return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')
    except Exception:
        return redirect(url_for('main.payments', cust_group=group_f, start_date=sd_dt.date().isoformat(), end_date=ed_dt.date().isoformat(), type='sales'))

@main.route('/reports', endpoint='reports')
@login_required
def reports():
    return render_template('reports.html')

@main.route('/reports/monthly', methods=['GET'], endpoint='reports_monthly')
@login_required
def reports_monthly():
    try:
        from models import Employee
        return render_template('reports_monthly.html', employees=Employee.query.all())
    except Exception:
        return render_template('reports_monthly.html', employees=[])

@main.route('/api/reports/monthly', methods=['GET'], endpoint='api_reports_monthly')
@login_required
def api_reports_monthly():
    try:
        from datetime import date
        from models import Salary, Employee, JournalLine, JournalEntry, Payment
        month = (request.args.get('month') or '').strip()
        year = request.args.get('year', type=int)
        emp_id = request.args.get('emp_id', type=int) or request.args.get('emp', type=int)
        emp_q = (request.args.get('emp') or '').strip().lower() if not emp_id else ''
        dept_q = (request.args.get('dept') or '').strip()
        start_dt = None; end_dt = None
        if month:
            y, m = month.split('-'); y = int(y); m = int(m)
            start_dt = date(y, m, 1)
            end_dt = date(y + (1 if m==12 else 0), 1 if m==12 else m+1, 1)
        elif year:
            start_dt = date(year, 1, 1)
            end_dt = date(year+1, 1, 1)
        emps = {}
        try:
            for e in Employee.query.all():
                emps[int(e.id)] = {'name': e.full_name, 'dept': (e.department or '').strip()}
        except Exception:
            pass
        sq = db.session.query(Salary)
        if month:
            y, m = month.split('-'); y = int(y); m = int(m)
            sq = sq.filter(Salary.year == y, Salary.month == m)
        elif year:
            sq = sq.filter(Salary.year == year)
        elif start_dt and end_dt:
            sq = sq.filter(Salary.month_date >= start_dt, Salary.month_date < end_dt)
        srows = sq.all()
        # Resolve Employee Advances account (EMP_ADV -> '1030')
        try:
            adv_code = SHORT_TO_NUMERIC.get('EMP_ADV', ('1030',))[0]
            adv_acc = _account(adv_code, CHART_OF_ACCOUNTS[adv_code]['name'], CHART_OF_ACCOUNTS[adv_code]['type'])
        except Exception:
            adv_acc = None
        # Deductions within selected period (based on description keywords)
        jlq = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id)
        if start_dt and end_dt:
            jlq = jlq.filter(JournalEntry.date >= start_dt, JournalEntry.date < end_dt)
        jrows = jlq.all()
        ded_hist_by_emp = {}
        try:
            dq = db.session.query(Salary.employee_id, func.coalesce(func.sum(Salary.deductions), 0.0))
            if end_dt:
                dq = dq.filter(Salary.month_date < end_dt)
            dq = dq.group_by(Salary.employee_id)
            for emp_id_val, ded_sum in dq.all():
                ded_hist_by_emp[int(emp_id_val or 0)] = float(ded_sum or 0.0)
        except Exception:
            ded_hist_by_emp = {}
        # Advances as historical outstanding up to end of period: sum EMP_ADV debits minus credits up to end_dt
        adv_debit_by_emp = {}
        adv_credit_by_emp = {}
        if adv_acc:
            adv_q = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id)
            adv_q = adv_q.filter(JournalLine.account_id == adv_acc.id)
            if end_dt:
                adv_q = adv_q.filter(JournalEntry.date < end_dt)
            for jl, je in adv_q.all():
                eid = int(getattr(jl, 'employee_id', 0) or 0)
                adv_debit_by_emp[eid] = adv_debit_by_emp.get(eid, 0.0) + float(getattr(jl, 'debit', 0) or 0)
                adv_credit_by_emp[eid] = adv_credit_by_emp.get(eid, 0.0) + float(getattr(jl, 'credit', 0) or 0)
        rows = []
        payroll_total = 0.0
        for s in srows:
            eid = int(getattr(s, 'employee_id', 0) or 0)
            nm = emps.get(eid, {}).get('name', '')
            dp = (emps.get(eid, {}).get('dept', '') or '').strip()
            if emp_id and eid != emp_id:
                continue
            if emp_q and nm.lower().find(emp_q) < 0:
                continue
            if dept_q and dp.lower() != dept_q.lower():
                continue
            basic = float(getattr(s, 'basic_salary', 0) or 0)
            bonus = float(getattr(s, 'allowances', 0) or 0)
            total = float(getattr(s, 'total_salary', 0) or 0)
            ded_month = float(getattr(s, 'deductions', 0) or 0)
            # Net monthly advances (granted - repaid) within the selected period
            adv_deb = float(adv_debit_by_emp.get(eid, 0.0) or 0.0)
            adv_cre = float(adv_credit_by_emp.get(eid, 0.0) or 0.0)
            advances = max(0.0, adv_deb - adv_cre)
            # Actual payments against this salary
            try:
                paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                             .filter(Payment.invoice_type=='salary', Payment.invoice_id==int(getattr(s,'id',0) or 0)).scalar() or 0.0)
            except Exception:
                paid = 0.0
            remaining = max(0.0, total - paid)
            # Normalize status to avoid marking zero totals as paid
            st = (getattr(s, 'status', '') or '').strip().lower() or 'due'
            if total <= 0:
                st = 'due'
            else:
                if paid >= total:
                    st = 'paid'
                elif paid > 0:
                    st = 'partial'
                else:
                    st = 'due'
            # Skip completely empty rows (no totals, no payments, no advances, no deductions)
            if total <= 0 and paid <= 0 and advances <= 0 and ded_month <= 0:
                continue
            net = max(0.0, total - advances - ded_month)
            rows.append({
                'employee_id': eid,
                'name': nm,
                'dept': dp,
                'basic': basic,
                'ot': 0.0,
                'bonus': bonus,
                'total': total,
                'advances': advances,
                'deductions': float(ded_hist_by_emp.get(eid, 0.0) or 0.0),
                'net': net,
                'status': st,
                'month': f"{int(getattr(s,'year',0) or 0)}-{str(int(getattr(s,'month',0) or 0)).zfill(2)}"
            })
            payroll_total += total
        # KPI totals should reflect only employees included in rows
        adv_total = float(sum([r.get('advances', 0.0) for r in rows]) or 0.0)
        ded_total = float(sum([r.get('deductions', 0.0) for r in rows]) or 0.0)
        entries_count = len(jrows)
        series = {
            'payroll': [float(getattr(sr, 'total_salary', 0) or 0) for sr in srows][0:24],
            'advances': [float(r.get('advances', 0.0)) for r in rows][0:24],
            'deductions': [float(r.get('deductions', 0.0)) for r in rows][0:24],
        }
        return jsonify({'ok': True, 'kpis': { 'payroll_total': payroll_total, 'advances_total': adv_total, 'deductions_total': ded_total, 'entries_count': entries_count }, 'rows': rows, 'series': series})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/reports/print/customer-sales', methods=['GET'], endpoint='reports_print_customer_sales')
@login_required
def reports_print_customer_sales():
    customer = (request.args.get('customer') or '').strip()
    customers_param = (request.args.get('customers') or '').strip()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    branch = (request.args.get('branch') or 'all').strip()
    rows = []
    totals = {'Amount': 0.0, 'Discount': 0.0, 'VAT': 0.0, 'Total': 0.0}
    payment_totals = {}
    try:
        from sqlalchemy import or_, and_
        import re
        def normalize_group(n: str):
            s = re.sub(r'[^a-z]', '', (n or '').lower())
            if s.startswith('hunger'): return 'hunger'
            if s.startswith('keeta') or s.startswith('keet'): return 'keeta'
            if s.startswith('noon'): return 'noon'
            return s
        # Aggregate items per invoice for amount and VAT
        items_sub = db.session.query(
            SalesInvoiceItem.invoice_id.label('inv_id'),
            func.sum(SalesInvoiceItem.price_before_tax * SalesInvoiceItem.quantity).label('amount_sum'),
            func.sum(SalesInvoiceItem.tax).label('vat_sum')
        ).group_by(SalesInvoiceItem.invoice_id).subquery()

        q = db.session.query(
            SalesInvoice,
            items_sub.c.amount_sum,
            items_sub.c.vat_sum
        ).outerjoin(items_sub, items_sub.c.inv_id == SalesInvoice.id)
        q = q.filter(SalesInvoice.status == 'paid')
        customers_list = []
        if customers_param:
            customers_list = [normalize_group(s) for s in customers_param.split(',') if s.strip()]
        base_filters = []
        if customers_list:
            for base in customers_list:
                if base:
                    base_filters.append(SalesInvoice.customer_name.ilike(base + '%'))
        elif customer:
            base = normalize_group(customer)
            if base:
                base_filters.append(SalesInvoice.customer_name.ilike(base + '%'))
        if base_filters:
            q = q.filter(or_(*base_filters))
        if branch and branch != 'all':
            q = q.filter(SalesInvoice.branch == branch)
        # Date range
        sd_dt = None; ed_dt = None
        try:
            sd_dt = datetime.fromisoformat(start_date) if start_date else None
        except Exception:
            sd_dt = None
        try:
            ed_dt = datetime.fromisoformat(end_date) if end_date else None
        except Exception:
            ed_dt = None
        if sd_dt and ed_dt:
            q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt, ed_dt)))
        elif sd_dt and (not ed_dt):
            q = q.filter(or_(SalesInvoice.created_at >= sd_dt, SalesInvoice.date >= sd_dt))
        elif ed_dt and (not sd_dt):
            q = q.filter(or_(SalesInvoice.created_at <= ed_dt, SalesInvoice.date <= ed_dt))

        q = q.order_by(SalesInvoice.created_at.desc(), SalesInvoice.date.desc()).limit(2000)
        for inv, amount_sum, vat_sum in q.all():
            dt = getattr(inv, 'created_at', None) or getattr(inv, 'date', None) or get_saudi_now()
            day_name = dt.strftime('%A')
            amount = float(amount_sum or 0.0)
            discount = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            vat = float(vat_sum or 0.0)
            total = float(amount - discount + vat)
            pm = (inv.payment_method or '').upper()
            rows.append({
                'Date': dt.strftime('%Y-%m-%d'),
                'Day': day_name,
                'Invoice': inv.invoice_number,
                'Payment': pm,
                'Amount': amount,
                'Discount': discount,
                'VAT': vat,
                'Total': total,
            })
            totals['Amount'] += amount
            totals['Discount'] += discount
            totals['VAT'] += vat
            totals['Total'] += total
            payment_totals[pm] = payment_totals.get(pm, 0.0) + total
    except Exception:
        pass

    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    title_customer = (normalize_group(customer) if customer else '')
    if customers_list:
        try:
            title_customer = ', '.join(customers_list)
        except Exception:
            title_customer = title_customer or 'multiple'
    meta = {
        'title': f"Customer Sales — {title_customer}",
        'customer': title_customer,
        'branch': branch,
        'start_date': start_date,
        'end_date': end_date,
        'generated_at': datetime.utcnow().strftime('%Y-%m-%d %H:%M')
    }
    columns = ['Date','Day','Invoice','Payment','Amount','Discount','VAT','Total']
    return render_template('print_report.html', report_title=meta['title'], settings=settings,
                           generated_at=meta['generated_at'], start_date=meta['start_date'], end_date=meta['end_date'],
                           payment_method='all', branch=meta['branch'],
                           columns=columns, data=rows, totals=totals, totals_columns=['Amount','Discount','VAT','Total'],
                           totals_colspan=4, payment_totals=payment_totals)

@main.route('/admin/reclassify-keeta-hunger', methods=['POST'], endpoint='admin_reclassify_keeta_hunger')
@login_required
def admin_reclassify_keeta_hunger():
    start_date = (request.form.get('start_date') or '2025-10-01').strip()
    end_date = (request.form.get('end_date') or get_saudi_now().date().isoformat()).strip()
    changed = 0
    removed = 0
    try:
        import re
        from sqlalchemy import or_
        sd_dt = datetime.fromisoformat(start_date)
        ed_dt = datetime.fromisoformat(end_date)
        def norm(n: str):
            s = re.sub(r'[^a-z]', '', (n or '').lower())
            if s.startswith('hunger'):
                return 'hunger'
            if s.startswith('keeta') or s.startswith('keet'):
                return 'keeta'
            return ''
        q = SalesInvoice.query
        try:
            q = q.filter(or_(SalesInvoice.created_at.between(sd_dt, ed_dt), SalesInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        for inv in q.order_by(SalesInvoice.created_at.desc()).limit(5000).all():
            g = norm(inv.customer_name or '')
            if g not in ('keeta','hunger'):
                continue
            for p in db.session.query(Payment).filter(Payment.invoice_type=='sales', Payment.invoice_id==inv.id).all():
                try:
                    db.session.delete(p)
                    removed += 1
                except Exception:
                    pass
            inv.status = 'unpaid'
            changed += 1
        db.session.commit()
        return jsonify({'ok': True, 'changed': changed, 'removed': removed})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500

# ---------- POS/Tables: basic navigation ----------
@main.route('/sales/<branch_code>/tables', endpoint='sales_tables')
@login_required
def sales_tables(branch_code):
    if not user_can('sales','view', branch_code):
        flash('لا تملك صلاحية الوصول لفرع المبيعات هذا', 'warning')
        return redirect(url_for('main.sales'))
    branch_label = BRANCH_LABELS.get(branch_code, branch_code)

    grouped_tables = []
    tables = []

    try:
        from models import TableSection, TableSectionAssignment

        def decode_name(name: str):
            name = name or ''
            visible = name
            layout = ''
            if '[rows:' in name and name.endswith(']'):
                try:
                    before, bracket = name.rsplit('[rows:', 1)
                    visible = before.rstrip()
                    layout = bracket[:-1].strip()
                except Exception:
                    pass
            return visible, layout

        sections = TableSection.query.filter_by(branch_code=branch_code).order_by(TableSection.sort_order, TableSection.id).all()
        assignments = TableSectionAssignment.query.filter_by(branch_code=branch_code).all()

        # Build status map based on draft orders (occupied / available)
        status_map = {}
        for assignment in assignments:
            number = safe_table_number(assignment.table_number)
            if number <= 0:
                continue
            draft = kv_get(f'draft:{branch_code}:{number}', {}) or {}
            status_map[number] = 'occupied' if (draft.get('items') or []) else 'available'

        assignments_by_section = {}
        for assignment in assignments:
            assignments_by_section.setdefault(assignment.section_id, []).append(assignment)

        # Build grouped tables honoring saved layout and showing empty sections as headers
        for section in sections:
            visible, layout = decode_name(section.name)

            # Collect tables for this section (sorted by table number)
            section_tables = []
            section_assignments = assignments_by_section.get(section.id, [])
            section_assignments.sort(key=lambda a: safe_table_number(a.table_number))
            for assignment in section_assignments:
                number = safe_table_number(assignment.table_number)
                if number <= 0:
                    continue
                section_tables.append({
                    'number': number,
                    'status': status_map.get(number, 'available')
                })

            # Split into rows according to saved layout "1,3,4" etc.
            rows = []
            if layout:
                try:
                    counts = [int(x.strip()) for x in layout.split(',') if x.strip()]
                except Exception:
                    counts = []
                i = 0
                for cnt in counts:
                    if cnt <= 0:
                        continue
                    rows.append(section_tables[i:i+cnt])
                    i += cnt
                if i < len(section_tables):
                    rows.append(section_tables[i:])
            else:
                # If no layout saved and tables exist, place all tables in a single row
                if section_tables:
                    rows = [section_tables]
                else:
                    rows = []

            section_entry = {
                'section': visible or branch_label,
                'rows': rows,
                # Keep flat list for template fallback when rows is empty
                'tables': section_tables
            }

            # Always append the section so saved sections appear even without assignments
            grouped_tables.append(section_entry)

    except Exception as e:
        print(f"[sales_tables] Error loading grouped tables for {branch_code}: {e}")
        grouped_tables = []
        tables = []

    if not grouped_tables:
        settings = kv_get('table_settings', {}) or {}
        default_count = 20
        if branch_code == 'china_town':
            count = int((settings.get('china') or {}).get('count', default_count))
        elif branch_code == 'place_india':
            count = int((settings.get('india') or {}).get('count', default_count))
        else:
            count = default_count
        for i in range(1, count + 1):
            draft = kv_get(f'draft:{branch_code}:{i}', {}) or {}
            status = 'occupied' if (draft.get('items') or []) else 'available'
            tables.append({'number': i, 'status': status})

    return render_template('sales_tables.html', branch_code=branch_code, branch_label=branch_label, tables=tables, grouped_tables=grouped_tables or None)

@main.route('/sales/china_town', endpoint='sales_china')
@login_required
def sales_china():
    return redirect(url_for('main.sales_tables', branch_code='china_town'))


@main.route('/sales/place_india', endpoint='sales_india')
@login_required
def sales_india():
    return redirect(url_for('main.sales_tables', branch_code='place_india'))


@main.route('/pos/<branch_code>', endpoint='pos_home')
@login_required
def pos_home(branch_code):
    if not user_can('sales','view', branch_code):
        flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u0627\u0644\u0648\u0635\u0648\u0644 \u0644\u0641\u0631\u0639 \u0627\u0644\u0645\u0628\u064a\u0639\u0627\u062a \u0647\u0630\u0627', 'warning')
        return redirect(url_for('main.sales'))

    return redirect(url_for('main.sales_tables', branch_code=branch_code))


@main.route('/pos/<branch_code>/table/<int:table_number>', endpoint='pos_table')
@login_required
def pos_table(branch_code, table_number):
    if not user_can('sales','view', branch_code):
        # If XHR requested HTML (POS page) but user is not allowed, redirect to login preserving next
        return redirect(url_for('main.login', next=request.path))
    branch_label = BRANCH_LABELS.get(branch_code, branch_code)
    vat_rate = 15
    # Load any existing draft for this table
    draft = kv_get(f'draft:{branch_code}:{table_number}', {}) or {}
    draft_items = json.dumps(draft.get('items') or [])
    current_draft = type('Obj', (), {'id': draft.get('draft_id')}) if draft.get('draft_id') else None
    # Warmup DB (avoid heavy create_all/seed on every request)
    warmup_db_once()
    # Load categories from DB for UI and provide a name->id map
    try:


        cats = MenuCategory.query.order_by(MenuCategory.sort_order, MenuCategory.name).all()
    except Exception:
        try:


            cats = MenuCategory.query.order_by(MenuCategory.name).all()
        except Exception:
            cats = MenuCategory.query.all()
    categories = [c.name for c in cats]
    cat_map = {}
    def _slug(s):
        try:
            import re
            s = (s or '').strip().lower()
            s = re.sub(r"[^a-z0-9]+", "-", s)
            s = re.sub(r"-+", "-", s).strip('-')
            return s or 'category'
        except Exception:
            return 'category'
    def _static_image_url(*candidates):
        try:
            import os
            for rel in candidates:
                if not rel:
                    continue
                fp = os.path.join(current_app.static_folder, rel.replace('/', os.sep))
                if os.path.exists(fp):
                    return '/static/' + rel
        except Exception:
            pass
        return None
    cat_image_map = {}
    try:
        _def_cat_img = kv_get('menu:default_category_image', None) or kv_get('menu:default_image', None) or '/static/logo.svg'
    except Exception:
        _def_cat_img = '/static/logo.svg'
    for c in cats:
        cat_map[c.name] = c.id
        cat_map[c.name.upper()] = c.id
        try:
            u = kv_get(f"menu:category_image:{c.id}", None) or kv_get(f"menu:category_image_by_name:{_slug(c.name)}", None)
        except Exception:
            u = None
        if not u:
            sslug = _slug(c.name)
            u = _static_image_url(f"images/categories/{sslug}.webp", f"images/categories/{sslug}.jpg", f"images/categories/{sslug}.png") or _def_cat_img
        cat_image_map[c.name] = u or _def_cat_img
    cat_map_json = json.dumps(cat_map)
    cat_image_map_json = json.dumps(cat_image_map)
    today = get_saudi_now().date().isoformat()
    # Load settings for currency icon, etc.
    try:
        s = Settings.query.first()
    except Exception:
        s = None
    return render_template('sales_table_invoice.html',
                           branch_code=branch_code,
                           branch_label=branch_label,
                           table_number=table_number,
                           vat_rate=vat_rate,
                           draft_items=draft_items,
                           current_draft=current_draft,
                           categories=categories,
                           cat_map_json=cat_map_json,
                           cat_image_map_json=cat_image_map_json,
                           today=today,
                           settings=s)


# ---------- Lightweight APIs used by front-end JS ----------
@main.route('/api/table-settings', methods=['GET', 'POST'], endpoint='api_table_settings')
@login_required
def api_table_settings():
    if request.method == 'GET':
        settings = kv_get('table_settings', {}) or {}
        if not settings:
            settings = {'china': {'count': 20, 'numbering': 'numeric'}, 'india': {'count': 20, 'numbering': 'numeric'}}
        return jsonify({'success': True, 'settings': settings})
    # POST
    try:
        data = request.get_json(force=True) or {}

        kv_set('table_settings', data)
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@main.route('/api/table-sections/<branch_code>', methods=['GET', 'POST'], endpoint='api_table_sections')
@login_required
def api_table_sections(branch_code):
    if not user_can('sales','view', branch_code):
        return jsonify({'success': False, 'error': 'forbidden'}), 403
    
    if request.method == 'GET':
        # Read from the same system as table-layout
        try:
            from models import TableSection, TableSectionAssignment
            
            def decode_name(name: str):
                name = name or ''
                visible = name
                layout = ''
                if '[rows:' in name and name.endswith(']'):
                    try:
                        before, bracket = name.rsplit('[rows:', 1)
                        visible = before.rstrip()
                        layout = bracket[:-1].strip()
                    except Exception:
                        pass
                return visible, layout

            sections = TableSection.query.filter_by(branch_code=branch_code).order_by(TableSection.sort_order, TableSection.id).all()
            assignments = TableSectionAssignment.query.filter_by(branch_code=branch_code).all()
            
            sections_data = []
            for s in sections:
                visible, layout = decode_name(s.name)
                sections_data.append({
                    'id': s.id,
                    'name': visible,
                    'sort_order': s.sort_order,
                    'layout': layout
                })
            
            assignments_data = []
            for a in assignments:
                assignments_data.append({
                    'table_number': a.table_number,
                    'section_id': a.section_id
                })
            
            return jsonify({'success': True, 'sections': sections_data, 'assignments': assignments_data})
        except Exception as e:
            return jsonify({'success': True, 'sections': [], 'assignments': []})
    
    try:
        # Save to the same system as table-layout
        from models import TableSection, TableSectionAssignment
        
        payload = request.get_json(force=True) or {}
        sections_data = payload.get('sections') or []
        assignments_data = payload.get('assignments') or []
        
        # Clear existing sections and assignments for this branch
        TableSectionAssignment.query.filter_by(branch_code=branch_code).delete()
        TableSection.query.filter_by(branch_code=branch_code).delete()
        
        # Create new sections
        for section_idx, section in enumerate(sections_data):
            section_name = section.get('name', '')
            layout = section.get('layout', '')
            encoded_name = f"{section_name} [rows:{layout}]" if layout else section_name
            
            new_section = TableSection(
                branch_code=branch_code,
                name=encoded_name,
                sort_order=section.get('sort_order', section_idx)
            )
            db.session.add(new_section)
            db.session.flush()
            
            # Create assignments for this section
            for assignment in assignments_data:
                if assignment.get('section_id') == section.get('id'):
                    assignment_obj = TableSectionAssignment(
                        branch_code=branch_code,
                        table_number=str(assignment.get('table_number')),
                        section_id=new_section.id
                    )
                    db.session.add(assignment_obj)
        
        db.session.commit()
        return jsonify({'success': True, 'sections': sections_data})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@main.route('/api/tables/<branch_code>', methods=['GET'], endpoint='api_tables_status')
@login_required

def api_tables_status(branch_code):
    # Read drafts to mark occupied tables
    if not user_can('sales','view', branch_code):
        return jsonify({'success': False, 'error': 'forbidden'}), 403
    settings = kv_get('table_settings', {}) or {}
    if branch_code == 'china_town':
        count = int((settings.get('china') or {}).get('count', 20))
    elif branch_code == 'place_india':
        count = int((settings.get('india') or {}).get('count', 20))
    else:
        count = 20
    items = []
    # Build a quick map of DB table statuses for this branch
    try:
        from models import Table
        rows = Table.query.filter_by(branch_code=branch_code).all()
        db_status = { (str(r.table_number).strip()): (r.status or 'available') for r in (rows or []) }
        # Expand count to include highest table number present in DB
        try:
            max_db_no = 0
            for r in (rows or []):
                try:
                    n = safe_table_number(getattr(r, 'table_number', None))
                    if n > max_db_no:
                        max_db_no = n
                except Exception:
                    continue
            if max_db_no > count:
                count = max_db_no
        except Exception:
            pass
    except Exception:
        db_status = {}
    for i in range(1, count+1):
        draft = kv_get(f'draft:{branch_code}:{i}', {}) or {}
        has_draft = bool(draft.get('items') or [])
        tbl_st = (db_status.get(str(i)) or 'available').lower()
        status = 'occupied' if (has_draft or tbl_st == 'occupied') else 'available'
        items.append({'table_number': i, 'status': status})
    return jsonify(items)


@main.route('/api/menu/<cat_id>/items', methods=['GET'], endpoint='api_menu_items')
@login_required
def api_menu_items(cat_id):
    # Prefer DB; fallback to KV/demo — warm up once
    warmup_db_once()
    cat = None
    try:
        cid = int(cat_id)
        cat = MenuCategory.query.get(cid)
    except Exception:
        pass
    if not cat:
        try:
            cat = MenuCategory.query.filter(db.func.lower(MenuCategory.name) == (cat_id or '').lower()).first()
        except Exception:
            cat = None
    def _slug(s):
        try:
            import re
            s = (s or '').strip().lower()
            s = re.sub(r"[^a-z0-9]+", "-", s)
            s = re.sub(r"-+", "-", s).strip('-')
            return s or 'item'
        except Exception:
            return 'item'
    def _static_image_url(*candidates):
        try:
            import os
            for rel in candidates:
                if not rel:
                    continue
                fp = os.path.join(current_app.static_folder, rel.replace('/', os.sep))
                if os.path.exists(fp):
                    return '/static/' + rel
        except Exception:
            pass
        return None
    def _uploads_image_url(kind, slug):
        try:
            if not slug:
                return None
            rels = [
                f'uploads/{kind}/{slug}.webp',
                f'uploads/{kind}/{slug}.jpg',
                f'uploads/{kind}/{slug}.png',
            ]
            for r in rels:
                u = _static_image_url(r)
                if u:
                    return u
        except Exception:
            pass
        return None
    def _default_image_url():
        try:
            u = kv_get('menu:default_image', None) or kv_get('menu:default_item_image', None) or kv_get('menu:default_category_image', None)
            if isinstance(u, str) and u:
                return u
        except Exception:
            pass
        return 'https://also3odyah.com/wp-content/uploads/2024/08/Best-Asian-Restaurants-in-Riyadh-Yauatcha-1024x768-1.jpg'
    def _remote_image_for_keywords(name, kind):
        try:
            n = (name or '').lower()
            keys = {
                'category': {
                    'appetizers': 'https://images.unsplash.com/photo-1547592180-85f173990554?q=80&w=1600&auto=format&fit=crop',
                    'biryani': 'https://images.unsplash.com/photo-1604908176997-2846b8dce8e7?q=80&w=1600&auto=format&fit=crop',
                    'rice': 'https://images.unsplash.com/photo-1547592180-85f173990554?q=80&w=1600&auto=format&fit=crop',
                    'noodles': 'https://images.unsplash.com/photo-1526318472351-c75fd2f0703d?q=80&w=1600&auto=format&fit=crop',
                    'soup': 'https://images.unsplash.com/photo-1505577058444-a3dab90d4253?q=80&w=1600&auto=format&fit=crop',
                    'salad': 'https://images.unsplash.com/photo-1555939594-58d7cb561ad1?q=80&w=1600&auto=format&fit=crop',
                    'seafood': 'https://images.unsplash.com/photo-1551183053-bf91a1d81141?q=80&w=1600&auto=format&fit=crop',
                    'prawn': 'https://images.unsplash.com/photo-1611253135999-daddb98d97a2?q=80&w=1600&auto=format&fit=crop',
                    'chicken': 'https://images.unsplash.com/photo-1544025162-d76694265947?q=80&w=1600&auto=format&fit=crop',
                    'beef': 'https://images.unsplash.com/photo-1553163147-622ab57be1c7?q=80&w=1600&auto=format&fit=crop',
                    'lamb': 'https://images.unsplash.com/photo-1625944521360-4f098c2e8a04?q=80&w=1600&auto=format&fit=crop',
                    'kebab': 'https://images.unsplash.com/photo-1604908554043-3df28557f9a5?q=80&w=1600&auto=format&fit=crop',
                    'grill': 'https://images.unsplash.com/photo-1558031239-aa4ae7329230?q=80&w=1600&auto=format&fit=crop',
                    'drinks': 'https://images.unsplash.com/photo-1509228627152-72ae9ae6848d?q=80&w=1600&auto=format&fit=crop',
                    'chinese': 'https://images.unsplash.com/photo-1544510808-8b0b0f7de841?q=80&w=1600&auto=format&fit=crop',
                    'indian': 'https://images.unsplash.com/photo-1543352634-8730a9b5e570?q=80&w=1600&auto=format&fit=crop',
                },
                'item': {
                    'biryani': 'https://images.unsplash.com/photo-1604908176997-2846b8dce8e7?q=80&w=1600&auto=format&fit=crop',
                    'noodles': 'https://images.unsplash.com/photo-1526318472351-c75fd2f0703d?q=80&w=1600&auto=format&fit=crop',
                    'samosa': 'https://images.unsplash.com/photo-1625944519087-2a9c7b0f1ff4?q=80&w=1600&auto=format&fit=crop',
                    'hummus': 'https://images.unsplash.com/photo-1552332386-f8dd00dc2f85?q=80&w=1600&auto=format&fit=crop',
                    'potato': 'https://images.unsplash.com/photo-1518806118471-f66f1e488eae?q=80&w=1600&auto=format&fit=crop',
                    'naan': 'https://images.unsplash.com/photo-1601050693721-03bba7dd6e61?q=80&w=1600&auto=format&fit=crop',
                    'tempura': 'https://images.unsplash.com/photo-1562967915-6a88f1f68047?q=80&w=1600&auto=format&fit=crop',
                    'kebab': 'https://images.unsplash.com/photo-1604908554043-3df28557f9a5?q=80&w=1600&auto=format&fit=crop',
                    'prawn': 'https://images.unsplash.com/photo-1611253135999-daddb98d97a2?q=80&w=1600&auto=format&fit=crop',
                    'shrimp': 'https://images.unsplash.com/photo-1611253135999-daddb98d97a2?q=80&w=1600&auto=format&fit=crop',
                    'chicken': 'https://images.unsplash.com/photo-1544025162-d76694265947?q=80&w=1600&auto=format&fit=crop',
                    'fish': 'https://images.unsplash.com/photo-1551183053-bf91a1d81141?q=80&w=1600&auto=format&fit=crop'
                }
            }[kind]
            syn = {
                'biryani': ['برياني','بريان','biryani'],
                'noodles': ['noodles','نودلز','معكرونة','chow','mein','chopsuey'],
                'samosa': ['سمبوسة','سمبوسا','سموسا','samosa','samusa'],
                'hummus': ['حمص','homous','hummus','humus','homouse'],
                'potato': ['بطاطس','بطاطا','potato','chips','chope','chop','choap','frenchfry','french fry','fries','finger chips'],
                'naan': ['خبز نان','نان','naan','nan','garlic nan','garlic naan','plain nan','plain naan'],
                'tempura': ['تمبورا','تيمبورا','tempura'],
                'kebab': ['كباب','kebab'],
                'prawn': ['روبيان','ربيان','ribyan','prawn','prwn','prawns','fried prawns','gold.fri.prawns','gold fried prawns','prwnballs','prawn balls','prwn balls'],
                'shrimp': ['جمبري','shrimp','fried shrimp'],
                'chicken': ['دجاج','chicken','kaichai'],
                'appetizers': ['appetizers','starter','starters','مقبلات','مقبا','mixed app','mixeed app','mixed appetizer','mix app','mixeed app(l)','mixeed app(m)','platter','مقبلات مشكل'],
                'seafood': ['seafood','مأكولات بحرية'],
                'fish': ['fish','fish finger','finger fish','fishfinger','fish-finger','سمك','فيش'],
                'salad': ['سلطة','salad'],
                'soup': ['شوربة','soup']
            }
            for k, lst in syn.items():
                for t in lst:
                    if t and t in n and k in keys:
                        url = keys.get(k)
                        if url:
                            return url
            for k, url in keys.items():
                if k in n:
                    return url
        except Exception:
            pass
        return None
    def _item_image_url(m):
        return _default_image_url()
    if cat:
        items = MenuItem.query.filter_by(category_id=cat.id).order_by(MenuItem.name).all()
        return jsonify([{'id': m.id, 'name': m.name, 'price': float(m.price), 'image_url': _item_image_url(m)} for m in items])
    # KV fallback
    data = kv_get(f'menu:items:{cat_id}', None)
    if isinstance(data, list):
        return jsonify(data)
    # Demo fallback
    demo_items = [{'id': None, 'name': nm, 'price': float(pr)} for (nm, pr) in _DEF_MENU.get(cat_id, [])]


    out = []
    # Force default image for demo items as well
    for d in demo_items:
        dd = dict(d)
        dd['image_url'] = _default_image_url()
        out.append(dd)
    return jsonify(out)

# ---------- Branch Settings API used by POS ----------
@main.route('/api/branch-settings/<branch_code>', methods=['GET'], endpoint='api_branch_settings')
@login_required
def api_branch_settings(branch_code):
    try:
        from models import Settings
        try:
            s = Settings.query.first()
        except Exception:
            s = None

        # Defaults
        void_pwd = '1991'
        vat_rate = 15.0
        discount_rate = 0.0

        if s:
            if branch_code == 'china_town':
                try:
                    void_pwd = (getattr(s, 'china_town_void_password', void_pwd) or void_pwd)
                except Exception:
                    pass
                try:
                    vat_rate = float(getattr(s, 'china_town_vat_rate', vat_rate) or vat_rate)
                except Exception:
                    pass
                try:
                    discount_rate = float(getattr(s, 'china_town_discount_rate', discount_rate) or discount_rate)
                except Exception:
                    pass
            elif branch_code == 'place_india':
                try:
                    void_pwd = (getattr(s, 'place_india_void_password', void_pwd) or void_pwd)
                except Exception:
                    pass
                try:
                    vat_rate = float(getattr(s, 'place_india_vat_rate', vat_rate) or vat_rate)
                except Exception:
                    pass
                try:
                    discount_rate = float(getattr(s, 'place_india_discount_rate', discount_rate) or discount_rate)
                except Exception:
                    pass

        return jsonify({
            'branch': branch_code,
            'void_password': str(void_pwd),
            'vat_rate': vat_rate,
            'discount_rate': discount_rate
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/draft-order/<branch_code>/<int:table_number>', methods=['POST'], endpoint='api_draft_create_or_update')
@login_required
def api_draft_create_or_update(branch_code, table_number):
    # Return 401 for unauthenticated XHR so the front-end can redirect cleanly
    try:
        if not getattr(current_user, 'is_authenticated', False):
            return jsonify({'success': False, 'error': 'unauthenticated'}), 401
    except Exception:
        return jsonify({'success': False, 'error': 'unauthenticated'}), 401
    if not user_can('sales','view', branch_code):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    try:
        payload = request.get_json(silent=True) or {}
        items = payload.get('items') or []
        draft_id = f"{branch_code}:{table_number}"
        kv_set(f'draft:{branch_code}:{table_number}', {
            'draft_id': draft_id,
            'items': items,
            'customer': payload.get('customer') or {},
            'discount_pct': float((payload.get('discount_pct') or 0) or 0),
            'tax_pct': float((payload.get('tax_pct') or 15) or 15),
            'payment_method': (payload.get('payment_method') or '')
        })
        # Persist table status in DB for multi-user consistency (transactional helper)
        _set_table_status_concurrent(branch_code, str(table_number), 'available' if not items else 'occupied')
        return jsonify({'success': True, 'draft_id': draft_id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


def _parse_draft_id(draft_id):


    # supports both branch:table and branch-table
    if ':' in draft_id:
        branch, num = draft_id.split(':', 1)
    elif '-' in draft_id:
        branch, num = draft_id.rsplit('-', 1)
    else:
        return None, None
    try:
        return branch, int(num)
    except Exception:
        return None, None


@main.route('/api/draft_orders/<draft_id>/update', methods=['POST'], endpoint='api_draft_update')
@login_required
def api_draft_update(draft_id):
    try:
        branch, table = _parse_draft_id(draft_id)
        if not branch:
            return jsonify({'success': False, 'error': 'invalid_draft_id'}), 400
        # Return 401 for unauthenticated XHR so the front-end can redirect cleanly
        try:
            if not getattr(current_user, 'is_authenticated', False):
                return jsonify({'success': False, 'error': 'unauthenticated'}), 401
        except Exception:
            return jsonify({'success': False, 'error': 'unauthenticated'}), 401
        payload = request.get_json(silent=True) or {}
        if not user_can('sales','view', branch):
            return jsonify({'success': False, 'error': 'forbidden'}), 403

        rec = kv_get(f'draft:{branch}:{table}', {}) or {}
        # map items to unified structure while preserving name/price
        items = payload.get('items') or []
        existing = rec.get('items') or []
        by_id = {}
        for eit in existing:
            key = eit.get('meal_id') or eit.get('id')
            if key is not None:
                by_id[int(key)] = {
                    'name': eit.get('name') or '',
                    'price': float(eit.get('price') or eit.get('unit') or 0.0)
                }
        norm = []
        for it in items:
            mid = it.get('meal_id') or it.get('id')
            qty = it.get('qty') or it.get('quantity') or 1
            nm = it.get('name')
            pr = it.get('price') or it.get('unit')
            if (not nm or pr in [None, '', 0, 0.0]) and mid:
                try:
                    cached = by_id.get(int(mid))
                except Exception:
                    cached = None
                if cached:
                    nm = nm or cached.get('name')
                    pr = pr or cached.get('price')
            if (not nm or pr in [None, '', 0, 0.0]) and mid:
                try:
                    m = db.session.get(MenuItem, int(mid))
                    if m:
                        nm = nm or m.name
                        pr = pr or float(m.price)
                except Exception:
                    pass
            norm.append({
                'meal_id': mid,
                'name': nm or '',
                'price': float(pr or 0.0),
                'qty': qty
            })
        rec['items'] = norm
        # update optional fields
        if 'customer_name' in payload or 'customer_phone' in payload:
            rec['customer'] = {
                'name': (payload.get('customer_name') or '').strip(),
                'phone': (payload.get('customer_phone') or '').strip(),
            }
        if 'payment_method' in payload:
            rec['payment_method'] = payload.get('payment_method') or ''
        if 'discount_pct' in payload:
            try: rec['discount_pct'] = float(payload.get('discount_pct') or 0)
            except Exception: rec['discount_pct'] = 0.0
        if 'tax_pct' in payload:
            try: rec['tax_pct'] = float(payload.get('tax_pct') or 15)
            except Exception: rec['tax_pct'] = 15.0
        kv_set(f'draft:{branch}:{table}', rec)
        # Also ensure DB table status reflects occupied/available based on items
        _set_table_status_concurrent(branch, str(table), 'occupied' if rec.get('items') else 'available')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@main.route('/api/draft_orders/<draft_id>/cancel', methods=['POST'], endpoint='api_draft_cancel')
@login_required
def api_draft_cancel(draft_id):
    try:
        branch, table = _parse_draft_id(draft_id)


        if not branch:
            return jsonify({'success': False, 'error': 'invalid_draft_id'}), 400
        if not user_can('sales','view', branch):
            return jsonify({'success': False, 'error': 'forbidden'}), 403

        # clear draft and mark table available in DB
        kv_set(f'draft:{branch}:{table}', {'draft_id': draft_id, 'items': []})
        _set_table_status_concurrent(branch, str(table), 'available')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


@main.route('/api/draft/checkout', methods=['POST'], endpoint='api_draft_checkout')


@login_required
def api_draft_checkout():
    payload = request.get_json(force=True) or {}
    draft_id = payload.get('draft_id') or ''
    branch, table = _parse_draft_id(draft_id)
    if not branch:
        return jsonify({'error': 'invalid_draft_id'}), 400
    if not user_can('sales','view', branch):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    warmup_db_once()
    draft = kv_get(f'draft:{branch}:{table}', {}) or {}
    items = draft.get('items') or []
    subtotal = 0.0
    for it in items:
        qty = float(it.get('qty') or it.get('quantity') or 1)
        price = float(it.get('price') or it.get('unit') or 0.0)
        if price <= 0 and it.get('meal_id'):
            m = MenuItem.query.get(int(it.get('meal_id')))
            if m:
                price = float(m.price)
        subtotal += qty * (price or 0.0)
    discount_pct = float(payload.get('discount_pct') or 0)
    tax_pct = float(payload.get('tax_pct') or 15)
    discount_amount = subtotal * (discount_pct/100.0)
    taxable_amount = max(subtotal - discount_amount, 0.0)
    vat_amount = taxable_amount * (tax_pct/100.0)
    total_after = taxable_amount + vat_amount
    payment_method = (payload.get('payment_method') or '').strip().upper()
    if payment_method not in ['CASH','CARD']:
        return jsonify({'success': False, 'error': 'اختر طريقة الدفع (CASH أو CARD)'}), 400
    # Reuse preview invoice number if present to keep display number consistent
    draft_data = kv_get(f'draft:{branch}:{table}', {}) or {}
    preview_no = (draft_data.get('preview_invoice_number') or '').strip()
    invoice_number = preview_no if preview_no else f"INV-{int(datetime.utcnow().timestamp())}-{branch[:2]}{table}"
    try:
        inv = SalesInvoice(
            invoice_number=invoice_number,
            branch=branch,
            table_number=int(table),
            customer_name=(payload.get('customer_name') or '').strip(),
            customer_phone=(payload.get('customer_phone') or '').strip(),
            payment_method=payment_method,
            total_before_tax=round(subtotal, 2),
            tax_amount=round(vat_amount, 2),
            discount_amount=round(discount_amount, 2),
            total_after_tax_discount=round(total_after, 2),
            user_id=int(getattr(current_user, 'id', 1) or 1),
            status='issued',
        )
        db.session.add(inv)
        db.session.flush()
        for it in items:
            qty = float(it.get('qty') or it.get('quantity') or 1)
            price = float(it.get('price') or it.get('unit') or 0.0)
            if price <= 0 and it.get('meal_id'):
                m = MenuItem.query.get(int(it.get('meal_id')))
                if m:
                    price = float(m.price)
            db.session.add(SalesInvoiceItem(
                invoice_id=inv.id,
                product_name=(it.get('name') or ''),
                quantity=qty,
                price_before_tax=price or 0.0,
                tax=0,
                discount=0,
                total_price=round((price or 0.0) * qty, 2),
            ))
        db.session.commit()
        try:
            base_amt = float(inv.total_before_tax or 0.0) - float(inv.discount_amount or 0.0)
            tax_amt = float(inv.tax_amount or 0.0)
            try:
                cust = (customer_name or '').strip().lower()
                grp = _platform_group(cust)
                if grp == 'keeta':
                    ar_code = _acc_override('AR_KEETA', SHORT_TO_NUMERIC['AR_KEETA'][0])
                elif grp == 'hunger':
                    ar_code = _acc_override('AR_HUNGER', SHORT_TO_NUMERIC['AR_HUNGER'][0])
                else:
                    ar_code = _acc_override('AR', SHORT_TO_NUMERIC['AR_KEETA'][0])
            except Exception:
                ar_code = SHORT_TO_NUMERIC['AR_KEETA'][0]
            _post_ledger(inv.date, ar_code, 'Accounts Receivable', 'asset', float(inv.total_after_tax_discount or 0.0), 0.0, f'SALE {inv.invoice_number}')
            try:
                pm = (payment_method or '').strip().upper()
                # Prefer platform-specific revenue when customer matches
                cust = (customer_name or '').strip().lower()
                grp = _platform_group(cust)
                if grp == 'keeta':
                    rev_code = _acc_override('REV_KEETA', SHORT_TO_NUMERIC['REV_KEETA'][0])
                elif grp == 'hunger':
                    rev_code = _acc_override('REV_HUNGER', SHORT_TO_NUMERIC['REV_HUNGER'][0])
                else:
                    if branch == 'place_india':
                        rev_code = _acc_override('REV_PI', SHORT_TO_NUMERIC['REV_PI'][0])
                    elif branch == 'china_town':
                        rev_code = _acc_override('REV_CT', SHORT_TO_NUMERIC['REV_CT'][0])
                    else:
                        rev_code = _acc_override('REV_CT', SHORT_TO_NUMERIC['REV_CT'][0])
                _post_ledger(inv.date, rev_code, CHART_OF_ACCOUNTS[rev_code]['name'], 'revenue', 0.0, base_amt, f'SALE {inv.invoice_number}')
            except Exception:
                _post_ledger(inv.date, rev_code, 'Revenue', 'revenue', 0.0, base_amt, f'SALE {inv.invoice_number}')
            if tax_amt > 0:
                _post_ledger(inv.date, 'VAT_OUT', 'VAT Output', 'tax', 0.0, tax_amt, f'SALE {inv.invoice_number}')
        except Exception:
            pass
        try:
            base_amt = float(inv.total_before_tax or 0.0) - float(inv.discount_amount or 0.0)
            tax_amt = float(inv.tax_amount or 0.0)
            _post_ledger(inv.date, 'AR', 'Accounts Receivable', 'asset', float(inv.total_after_tax_discount or 0.0), 0.0, f'SALE {inv.invoice_number}')
            try:
                pm = (payment_method or '').strip().upper()
                # Prefer platform-specific revenue when customer matches
                cust = (customer_name or '').strip().lower()
                grp = _platform_group(cust)
                if grp == 'keeta':
                    rev_code = _acc_override('REV_KEETA', SHORT_TO_NUMERIC['REV_KEETA'][0])
                elif grp == 'hunger':
                    rev_code = _acc_override('REV_HUNGER', SHORT_TO_NUMERIC['REV_HUNGER'][0])
                else:
                    if branch == 'place_india':
                        rev_code = _acc_override('REV_PI', SHORT_TO_NUMERIC['REV_PI'][0])
                    elif branch == 'china_town':
                        rev_code = _acc_override('REV_CT', SHORT_TO_NUMERIC['REV_CT'][0])
                    else:
                        rev_code = _acc_override('REV_CT', SHORT_TO_NUMERIC['REV_CT'][0])
                _post_ledger(inv.date, rev_code, CHART_OF_ACCOUNTS[rev_code]['name'], 'revenue', 0.0, base_amt, f'SALE {inv.invoice_number}')
            except Exception:
                _post_ledger(inv.date, rev_code, 'Revenue', 'revenue', 0.0, base_amt, f'SALE {inv.invoice_number}')
            if tax_amt > 0:
                _post_ledger(inv.date, 'VAT_OUT', 'VAT Output', 'tax', 0.0, tax_amt, f'SALE {inv.invoice_number}')
        except Exception:
            pass
        # Auto-mark paid for non Keeta/Hunger and create Payment record
        try:
            from models import Payment
            cust = (payload.get('customer_name') or '').strip().lower()
            grp = _platform_group(cust)
            amt = float(inv.total_after_tax_discount or 0.0)
            if grp in ('keeta','hunger'):
                inv.status = 'unpaid'
                db.session.commit()
            else:
                db.session.add(Payment(
                    invoice_id=inv.id,
                    invoice_type='sales',
                    amount_paid=amt,
                    payment_method=(payment_method or 'CASH').upper(),
                    payment_date=get_saudi_now()
                ))
                inv.status = 'paid'
                db.session.commit()
                # Post payment ledger: DR Cash/Bank, CR Accounts Receivable (platform-specific when applicable)
                cash_acc = _pm_account(payment_method)
                try:
                    cust = (getattr(inv, 'customer_name', '') or '').lower()
                    grp = _platform_group(cust)
                    if grp == 'keeta':
                        ar_code = _acc_override('AR_KEETA', SHORT_TO_NUMERIC['AR_KEETA'][0])
                    elif grp == 'hunger':
                        ar_code = _acc_override('AR_HUNGER', SHORT_TO_NUMERIC['AR_HUNGER'][0])
                    else:
                        ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
                except Exception:
                    ar_code = CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020')
                _post_ledger(inv.date, ar_code, 'Accounts Receivable', 'asset', 0.0, amt, f'PAY SALE {inv.invoice_number}')
                if cash_acc:
                    _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', amt, 0.0, f'PAY SALE {inv.invoice_number}')
        except Exception:
            pass
        # Mark table available only after we confirm print+pay (handled in api_invoice_confirm_print)
    except Exception as e:


        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    return jsonify({'ok': True, 'invoice_id': invoice_number, 'payment_method': payment_method, 'total_amount': round(total_after, 2), 'print_url': url_for('main.print_receipt', invoice_number=invoice_number), 'branch_code': branch, 'table_number': int(table)})


@main.route('/api/sales/checkout', methods=['POST'], endpoint='api_sales_checkout')
@login_required
def api_sales_checkout():
    payload = request.get_json(force=True) or {}
    warmup_db_once()
    branch = (payload.get('branch_code') or '').strip() or 'unknown'
    table = int(payload.get('table_number') or 0)
    if not user_can('sales','view', branch):
        return jsonify({'success': False, 'error': 'forbidden'}), 403

    items = payload.get('items') or []
    # get prices from DB when missing
    subtotal = 0.0
    resolved = []
    for it in items:
        qty = float(it.get('qty') or 1)
        price = float(it.get('price') or 0.0)
        if price <= 0 and it.get('meal_id'):
            m = MenuItem.query.get(int(it.get('meal_id')))
            if m:
                price = float(m.price)
                name = m.name
            else:
                name = it.get('name') or ''
        else:
            name = it.get('name') or ''
        subtotal += qty * (price or 0.0)
        resolved.append({'meal_id': it.get('meal_id'), 'name': name, 'price': price or 0.0, 'qty': qty})
    discount_pct = float(payload.get('discount_pct') or 0)
    tax_pct = float(payload.get('tax_pct') or 15)
    discount_amount = subtotal * (discount_pct/100.0)
    taxable_amount = max(subtotal - discount_amount, 0.0)
    vat_amount = taxable_amount * (tax_pct/100.0)
    total_after = taxable_amount + vat_amount
    payment_method = (payload.get('payment_method') or '').strip().upper()
    if payment_method not in ['CASH','CARD']:
        return jsonify({'success': False, 'error': '\u0627\u062e\u062a\u0631 \u0637\u0631\u064a\u0642\u0629 \u0627\u0644\u062f\u0641\u0639 (CASH \u0623\u0648 CARD)'}), 400
    invoice_number = f"INV-{int(datetime.utcnow().timestamp())}-{branch[:2]}{table}"
    try:
        inv = SalesInvoice(
            invoice_number=invoice_number,
            branch=branch,
            table_number=table,
            customer_name=(payload.get('customer_name') or '').strip(),
            customer_phone=(payload.get('customer_phone') or '').strip(),
            payment_method=payment_method,
            total_before_tax=round(subtotal, 2),
            tax_amount=round(vat_amount, 2),
            discount_amount=round(discount_amount, 2),
            total_after_tax_discount=round(total_after, 2),
            user_id=int(getattr(current_user, 'id', 1) or 1),
            status='issued',
        )
        db.session.add(inv)
        db.session.flush()
        for it in resolved:
            db.session.add(SalesInvoiceItem(
                invoice_id=inv.id,
                product_name=it.get('name'),
                quantity=float(it.get('qty') or 1),
                price_before_tax=float(it.get('price') or 0.0),
                tax=0,
                discount=0,
                total_price=round(float(it.get('price') or 0.0) * float(it.get('qty') or 1), 2),
            ))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    try:
        from models import Payment
        cust = (payload.get('customer_name') or '').strip().lower()
        grp = _platform_group(cust)
        amt = float(inv.total_after_tax_discount or 0.0)
        if grp:
            inv.status = 'unpaid'
            db.session.commit()
        else:
            db.session.add(Payment(
                invoice_id=inv.id,
                invoice_type='sales',
                amount_paid=amt,
                payment_method=(payment_method or 'CASH').upper(),
                payment_date=get_saudi_now()
            ))
            inv.status = 'paid'
            db.session.commit()
            cash_acc = _pm_account(payment_method)
            try:
                cust2 = (getattr(inv, 'customer_name', '') or '').lower()
                grp2 = _platform_group(cust2)
                if grp2 == 'keeta':
                    ar_code = _acc_override('AR_KEETA', SHORT_TO_NUMERIC['AR_KEETA'][0])
                elif grp2 == 'hunger':
                    ar_code = _acc_override('AR_HUNGER', SHORT_TO_NUMERIC['AR_HUNGER'][0])
                else:
                    ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
            except Exception:
                ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
            _post_ledger(inv.date, ar_code, 'Accounts Receivable', 'asset', 0.0, amt, f'PAY SALE {inv.invoice_number}')
            if cash_acc:
                _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', amt, 0.0, f'PAY SALE {inv.invoice_number}')
    except Exception:
        pass
    try:
        _create_sale_journal(inv)
    except Exception:
        pass
    return jsonify({'ok': True, 'invoice_id': invoice_number, 'payment_method': payment_method, 'total_amount': round(total_after, 2), 'print_url': url_for('main.invoice_print', invoice_id=invoice_number), 'branch_code': branch, 'table_number': table})


@main.route('/api/invoice/confirm-print', methods=['POST'], endpoint='api_invoice_confirm_print')
@login_required
def api_invoice_confirm_print():
    # Mark invoice paid and free the table. Create payment record.
    try:
        from models import SalesInvoice, Payment
        payload = request.get_json(force=True) or {}
        branch = (payload.get('branch_code') or '').strip()
        table = int(payload.get('table_number') or 0)
        raw_invoice_id = payload.get('invoice_id')
        payment_method = (payload.get('payment_method') or '').upper() or None
        total_amount = float(payload.get('total_amount') or 0)

        # Resolve invoice by number or numeric id
        inv = None
        if isinstance(raw_invoice_id, str) and not raw_invoice_id.isdigit():
            inv = SalesInvoice.query.filter_by(invoice_number=raw_invoice_id).first()
        else:
            try:
                inv = SalesInvoice.query.get(int(raw_invoice_id))
            except Exception:
                inv = None
        if not inv:
            return jsonify({'success': False, 'error': 'Invoice not found'}), 404

        if (inv.status or '').lower() != 'paid':
            grp = _platform_group(inv.customer_name or '')
            if grp in ('keeta','hunger'):
                pass
            else:
                if payment_method:
                    db.session.add(Payment(
                        invoice_id=inv.id,
                        invoice_type='sales',
                        amount_paid=total_amount or float(inv.total_after_tax_discount or 0),
                        payment_method=payment_method,
                        payment_date=get_saudi_now()
                    ))
                inv.status = 'paid'
                db.session.commit()
            try:
                key = f"pdf_path:sales:{inv.invoice_number}"
                cur = kv_get(key, {}) or {}
                if not (cur.get('saved_at')):
                    kv_set(key, {'path': cur.get('path'), 'saved_at': get_saudi_now().isoformat()})
            except Exception:
                pass
            try:
                import re
                s = re.sub(r'[^a-z]', '', (inv.customer_name or '').lower())
                is_special = s.startswith('hunger') or s.startswith('keeta') or s.startswith('keet')
                if not is_special:
                    amt = float(total_amount or float(inv.total_after_tax_discount or 0))
                    cash_acc = _pm_account(payment_method)
                    try:
                        cust = (getattr(inv, 'customer_name', '') or '').lower()
                        grp = _platform_group(cust)
                        if grp == 'keeta':
                            ar_code = _acc_override('AR_KEETA', SHORT_TO_NUMERIC['AR_KEETA'][0])
                        elif grp == 'hunger':
                            ar_code = _acc_override('AR_HUNGER', SHORT_TO_NUMERIC['AR_HUNGER'][0])
                        else:
                            ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
                    except Exception:
                        ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
                    _post_ledger(inv.date, ar_code, 'Accounts Receivable', 'asset', 0.0, amt, f'PAY SALE {inv.invoice_number}')
                    if cash_acc:
                        _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', amt, 0.0, f'PAY SALE {inv.invoice_number}')
            except Exception:
                pass
            try:
                _archive_invoice_pdf(inv)
            except Exception as e:
                current_app.logger.error('Archive PDF failed: %s', e)

        # Clear draft to free the table and update DB table status to available
        if branch and table:
            kv_set(f'draft:{branch}:{table}', {'draft_id': f'{branch}:{table}', 'items': []})
            try:
                from models import Table
                t = Table.query.filter_by(branch_code=branch, table_number=str(table)).first()
                if t:
                    t.status = 'available'
                    t.updated_at = get_saudi_now()
                    db.session.commit()
            except Exception:
                db.session.rollback()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'error': str(e)}), 400
    return jsonify({'success': True})

@main.route('/api/sales/mark-unpaid-paid', methods=['POST'], endpoint='api_sales_mark_unpaid_paid')
@login_required
def api_sales_mark_unpaid_paid():
    try:
        from models import SalesInvoice, Payment
        from sqlalchemy import func
        q = SalesInvoice.query
        # Only consider invoices not marked as paid
        q = q.filter(func.lower(SalesInvoice.status) != 'paid') if hasattr(SalesInvoice, 'status') else q
        rows = q.order_by(SalesInvoice.date.desc()).limit(5000).all()
        updated = 0
        for inv in (rows or []):
            try:
                total = float(getattr(inv, 'total_after_tax_discount', 0) or 0)
                paid_sum = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                                 .filter(Payment.invoice_type == 'sales', Payment.invoice_id == inv.id)
                                 .scalar() or 0.0)
                remaining = max(total - paid_sum, 0.0)
                if remaining <= 0:
                    inv.status = 'paid'
                    db.session.commit()
                    updated += 1
                    continue
                pm = (getattr(inv, 'payment_method', '') or 'CASH').strip().upper()
                db.session.add(Payment(
                    invoice_id=inv.id,
                    invoice_type='sales',
                    amount_paid=remaining,
                    payment_method=pm,
                    payment_date=get_saudi_now()
                ))
                inv.status = 'paid'
                db.session.commit()
                # Post payment ledger: DR Cash/Bank, CR AR
                try:
                    cash_acc = _pm_account(pm)
                    _post_ledger(inv.date, 'AR', 'Accounts Receivable', 'asset', 0.0, remaining, f'PAY SALE {inv.invoice_number}')
                    if cash_acc:
                        _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', remaining, 0.0, f'PAY SALE {inv.invoice_number}')
                except Exception:
                    pass
                updated += 1
            except Exception:
                try:
                    db.session.rollback()
                except Exception:
                    pass
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400


@main.route('/api/sales/mark-platform-unpaid', methods=['GET','POST'], endpoint='api_sales_mark_platform_unpaid')
@main.route('/api/sales/mark-platform-unpaid/', methods=['GET','POST'])
@login_required
def api_sales_mark_platform_unpaid():
    try:
        from models import SalesInvoice, Payment
        q = SalesInvoice.query
        rows = q.order_by(SalesInvoice.date.desc()).limit(3000).all()
        updated = 0
        for inv in (rows or []):
            grp = _norm_group(getattr(inv, 'customer_name', '') or '')
            if grp in ('keeta','hunger'):
                try:
                    Payment.query.filter(Payment.invoice_type=='sales', Payment.invoice_id==inv.id).delete(synchronize_session=False)
                except Exception:
                    pass
                try:
                    inv.status = 'unpaid'
                    db.session.commit()
                except Exception:
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                updated += 1
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/sales/batch-pay', methods=['GET','POST'], endpoint='api_sales_batch_pay')
@main.route('/api/sales/batch-pay/', methods=['GET','POST'])
@login_required
def api_sales_batch_pay():
    try:
        from models import SalesInvoice, Payment
        payload = request.get_json(silent=True) or {}
        customer = (request.form.get('customer') or payload.get('customer') or 'all').strip().lower()
        method = (request.form.get('payment_method') or payload.get('payment_method') or 'CASH').strip().upper()
        limit = int((request.form.get('limit') or payload.get('limit') or 3000))
        def _norm_group(n: str):
            s = (n or '').lower()
            if ('hunger' in s) or ('هنقر' in s) or ('هونقر' in s):
                return 'hunger'
            if ('keeta' in s) or ('كيتا' in s) or ('كيت' in s):
                return 'keeta'
            return ''
        q = SalesInvoice.query
        # Consider only not-paid invoices
        try:
            q = q.filter(func.lower(SalesInvoice.status) != 'paid')
        except Exception:
            pass
        rows = q.order_by(SalesInvoice.date.desc()).limit(limit).all()
        updated = 0
        for inv in (rows or []):
            grp = _platform_group(getattr(inv, 'customer_name', '') or '')
            if customer not in ('all','') and grp != customer:
                continue
            # Compute remaining
            total = float(getattr(inv, 'total_after_tax_discount', 0.0) or 0.0)
            paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                         .filter(Payment.invoice_type == 'sales', Payment.invoice_id == inv.id)
                         .scalar() or 0.0)
            remaining = max(total - paid, 0.0)
            if remaining <= 0:
                try:
                    inv.status = 'paid'
                    db.session.commit()
                except Exception:
                    try: db.session.rollback()
                    except Exception: pass
                continue
            try:
                db.session.add(Payment(
                    invoice_id=inv.id,
                    invoice_type='sales',
                    amount_paid=remaining,
                    payment_method=method,
                    payment_date=get_saudi_now()
                ))
                inv.status = 'paid'
                db.session.commit()
                # Post payment ledger: DR Cash/Bank, CR AR (platform-specific when applicable)
                cash_acc = _pm_account(method)
                try:
                    cust = (getattr(inv, 'customer_name', '') or '').lower()
                    grp = _platform_group(cust)
                    if grp == 'keeta':
                        ar_code = _acc_override('AR_KEETA', SHORT_TO_NUMERIC['AR_KEETA'][0])
                    elif grp == 'hunger':
                        ar_code = _acc_override('AR_HUNGER', SHORT_TO_NUMERIC['AR_HUNGER'][0])
                    else:
                        ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
                except Exception:
                    ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
                _post_ledger(inv.date, ar_code, 'Accounts Receivable', 'asset', 0.0, remaining, f'PAY SALE {inv.invoice_number}')
                if cash_acc:
                    _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', remaining, 0.0, f'PAY SALE {inv.invoice_number}')
                updated += 1
            except Exception:
                try:
                    db.session.rollback()
                except Exception:
                    pass
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/invoice/print-log', methods=['POST'], endpoint='api_invoice_print_log')
@login_required
def api_invoice_print_log():
    try:
        payload = request.get_json(force=True) or {}
        invoice_id = (payload.get('invoice_id') or '').strip()
        if not invoice_id:
            return jsonify({'ok': False, 'error': 'missing_invoice_id'}), 400
        key = f"print_log:{invoice_id}"
        logs = kv_get(key, []) or []
        logs.append({'ts': get_saudi_now().isoformat(), 'user_id': int(getattr(current_user, 'id', 0) or 0)})
        kv_set(key, logs)
        return jsonify({'ok': True, 'count': len(logs)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@main.route('/api/sales/void-check', methods=['POST'], endpoint='api_sales_void_check')
@login_required
def api_sales_void_check():
    payload = request.get_json(force=True) or {}
    ok = (str(payload.get('password') or '').strip() == '1991')
    return jsonify({'ok': ok})


@main.route('/api/payments/register', methods=['POST'], endpoint='register_payment_ajax')
@login_required
def register_payment_ajax():
    # Accept both form and JSON
    payload = request.get_json(silent=True) or {}
    invoice_id = request.form.get('invoice_id') or payload.get('invoice_id')
    invoice_type = (request.form.get('invoice_type') or payload.get('invoice_type') or '').strip().lower()
    amount = request.form.get('amount') or payload.get('amount')
    payment_method = (request.form.get('payment_method') or payload.get('payment_method') or 'CASH').strip().upper()
    try:
        inv_id = int(invoice_id)
        amt = float(amount or 0)
    except Exception:
        return jsonify({'status': 'error', 'message': 'Invalid invoice id or amount'}), 400
    if amt <= 0:
        return jsonify({'status': 'error', 'message': 'Amount must be > 0'}), 400
    if invoice_type not in ('purchase','expense'):
        return jsonify({'status': 'error', 'message': 'Unsupported invoice type'}), 400

    try:
        # Create payment record
        p = Payment(invoice_id=inv_id, invoice_type=invoice_type, amount_paid=amt, payment_method=payment_method)
        db.session.add(p)
        db.session.flush()

        # Fetch invoice and totals
        if invoice_type == 'purchase':
            inv = PurchaseInvoice.query.get(inv_id)
            total = float(inv.total_after_tax_discount or 0.0) if inv else 0.0
        else:
            inv = ExpenseInvoice.query.get(inv_id)
            total = float(inv.total_after_tax_discount or 0.0) if inv else 0.0

        # Sum paid so far (including this payment)
        paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                     .filter(Payment.invoice_id == inv_id, Payment.invoice_type == invoice_type).scalar() or 0.0)

        # Robust status calculation with rounding tolerance (0.01)
        from decimal import Decimal, ROUND_HALF_UP
        def to_cents(value):
            try:
                return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            except Exception:
                return Decimal('0.00')
        total_c = to_cents(total)
        paid_c = to_cents(paid)
        if inv:
            if (total_c - paid_c) <= Decimal('0.01') and total_c > Decimal('0.00'):
                inv.status = 'paid'
            elif paid_c > Decimal('0.00'):
                inv.status = 'partial'
            else:
                inv.status = inv.status or ('unpaid' if invoice_type=='purchase' else 'paid')
        db.session.commit()
        try:
            if invoice_type == 'purchase':
                _post_ledger(get_saudi_now().date(), 'AP', 'Accounts Payable', 'liability', amt, 0.0, f'PAY PUR {inv_id}')
            else:
                _post_ledger(get_saudi_now().date(), 'AP', 'Accounts Payable', 'liability', amt, 0.0, f'PAY EXP {inv_id}')
            ca = _pm_account(payment_method)
            if ca:
                _post_ledger(get_saudi_now().date(), ca.code, ca.name, 'asset', 0.0, amt, f'PAY {invoice_type.upper()} {inv_id}')
        except Exception:
            pass
        return jsonify({'status': 'success', 'invoice_id': inv_id, 'amount': amt, 'paid': paid, 'total': total, 'new_status': getattr(inv, 'status', None)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 400



@main.route('/api/payments/supplier/register', methods=['POST'], endpoint='register_payment_supplier')
@login_required
def register_payment_supplier():
    payload = request.get_json(silent=True) or {}
    supplier_name = (request.form.get('supplier') or payload.get('supplier') or '').strip()
    amount = request.form.get('amount') or payload.get('amount')
    method = (request.form.get('payment_method') or payload.get('payment_method') or 'CASH').strip().upper()
    sd_str = (request.form.get('start_date') or payload.get('start_date') or '2025-10-01').strip()
    ed_str = (request.form.get('end_date') or payload.get('end_date') or get_saudi_now().date().isoformat()).strip()
    try:
        amt = float(amount or 0)
    except Exception:
        return jsonify({'status': 'error', 'message': 'Invalid amount'}), 400
    if not supplier_name:
        return jsonify({'status': 'error', 'message': 'Supplier required'}), 400
    if amt <= 0:
        return jsonify({'status': 'error', 'message': 'Amount must be > 0'}), 400
    try:
        sd_dt = datetime.fromisoformat(sd_str)
    except Exception:
        sd_dt = datetime(get_saudi_now().year, 10, 1)
    try:
        ed_dt = datetime.fromisoformat(ed_str)
    except Exception:
        ed_dt = get_saudi_now()

    from decimal import Decimal, ROUND_HALF_UP
    def to_cents(value):
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        except Exception:
            return Decimal('0.00')

    allocations = []
    remaining_pay = amt
    try:
        q = PurchaseInvoice.query
        try:
            q = q.filter(or_(PurchaseInvoice.created_at.between(sd_dt, ed_dt), PurchaseInvoice.date.between(sd_dt.date(), ed_dt.date())))
        except Exception:
            pass
        # Case-insensitive supplier name match
        q = q.filter(func.lower(PurchaseInvoice.supplier_name) == supplier_name.lower())
        # Oldest-first FIFO allocation by date/created_at
        q = q.order_by(PurchaseInvoice.date.asc(), PurchaseInvoice.created_at.asc())
        rows = q.all()
        for inv in rows:
            if remaining_pay <= 0:
                break
            total = float(inv.total_after_tax_discount or 0.0)
            paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                         .filter(Payment.invoice_id == inv.id, Payment.invoice_type == 'purchase').scalar() or 0.0)
            total_c = to_cents(total); paid_c = to_cents(paid)
            remaining = float(max(total_c - paid_c, Decimal('0.00')))
            if remaining <= 0.0:
                continue
            alloc = min(remaining_pay, remaining)
            db.session.add(Payment(invoice_id=inv.id, invoice_type='purchase', amount_paid=float(alloc), payment_method=method))
            new_paid = float(paid + alloc)
            # Update status robustly
            np_c = to_cents(new_paid)
            if (total_c - np_c) <= Decimal('0.01') and total_c > Decimal('0.00'):
                inv.status = 'paid'
            elif np_c > Decimal('0.00'):
                inv.status = 'partial'
            else:
                inv.status = 'unpaid'
            allocations.append({'invoice_id': inv.id, 'invoice_number': getattr(inv, 'invoice_number', None) or inv.id, 'allocated': float(alloc)})
            remaining_pay -= float(alloc)
        db.session.commit()
        # Ledger postings: Debit AP (reduce liability), Credit cash/bank
        try:
            _post_ledger(get_saudi_now().date(), 'AP', 'Accounts Payable', 'liability', amt, 0.0, f'PAY SUP {supplier_name}')
            ca = _pm_account(method)
            if ca:
                _post_ledger(get_saudi_now().date(), ca.code, ca.name, 'asset', 0.0, amt, f'PAY SUP {supplier_name}')
        except Exception:
            pass
        return jsonify({'status': 'success', 'supplier': supplier_name, 'amount': amt, 'allocated': allocations, 'unallocated': float(max(remaining_pay, 0.0))})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 400

@main.route('/api/payments/pay_all', methods=['POST'], endpoint='api_payments_pay_all')
@login_required
def api_payments_pay_all():
    # Bulk pay remaining for filtered invoices
    payload = request.get_json(silent=True) or {}
    type_f = (payload.get('type') or request.form.get('type') or '').strip().lower()
    method = (payload.get('payment_method') or request.form.get('payment_method') or 'CASH').strip().upper()
    # Supported: purchase, expense
    kinds = []
    if type_f in ('purchase','expense'):
        kinds = [type_f]
    else:
        kinds = ['purchase','expense']

    from decimal import Decimal, ROUND_HALF_UP
    def to_cents(value):
        try:
            return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        except Exception:
            return Decimal('0.00')

    affected = []
    try:
        for k in kinds:
            if k == 'purchase':
                rows = PurchaseInvoice.query.order_by(PurchaseInvoice.created_at.desc()).all()
            else:
                rows = ExpenseInvoice.query.order_by(ExpenseInvoice.created_at.desc()).all()
            for inv in rows:
                total = to_cents(float(getattr(inv, 'total_after_tax_discount', 0) or 0))
                paid = to_cents(float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                          .filter(Payment.invoice_id == inv.id, Payment.invoice_type == k).scalar() or 0.0))
                remaining = total - paid
                if remaining > Decimal('0.01'):
                    amt = float(remaining)
                    db.session.add(Payment(invoice_id=inv.id, invoice_type=k, amount_paid=amt, payment_method=method))
                    inv.status = 'paid'
                    affected.append({'id': inv.id, 'type': k, 'amount': amt})
        db.session.commit()
        return jsonify({'status':'success', 'count': len(affected), 'items': affected})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status':'error', 'message': str(e)}), 400


# ---- Customers lightweight search API ----
@main.route('/api/customers/search', methods=['GET'], endpoint='api_customers_search')
@login_required
def api_customers_search():
    q = (request.args.get('q') or '').strip()
    if not q:
        rows = Customer.query.filter_by(active=True).order_by(Customer.name.asc()).limit(10).all()
    else:
        like = f"%{q}%"
        rows = Customer.query.filter(
            Customer.active == True,
            (Customer.name.ilike(like)) | (Customer.phone.ilike(like))
        ).order_by(Customer.name.asc()).limit(20).all()
    data = [{
        'id': c.id,
        'name': c.name,
        'phone': c.phone,
        'discount_percent': float(c.discount_percent or 0)
    } for c in rows]
    return jsonify({'results': data})


# ---- Print routes for sales receipts ----
@main.route('/print/receipt/<invoice_number>', methods=['GET'], endpoint='print_receipt')
@login_required
def print_receipt(invoice_number):
    inv = SalesInvoice.query.filter_by(invoice_number=invoice_number).first()
    if not inv:
        return 'Invoice not found', 404
    items = SalesInvoiceItem.query.filter_by(invoice_id=inv.id).all()
    # compute items context and use stored totals
    items_ctx = []
    for it in items:
        line = float(it.price_before_tax or 0) * float(it.quantity or 0)
        items_ctx.append({
            'product_name': it.product_name or '',
            'quantity': float(it.quantity or 0),
            'total_price': line,
        })

    inv_ctx = {
        'invoice_number': inv.invoice_number,
        'table_number': inv.table_number,
        'customer_name': inv.customer_name,
        'customer_phone': inv.customer_phone,
        'payment_method': inv.payment_method,
        'status': 'PAID',
        'total_before_tax': float(inv.total_before_tax or 0.0),
        'tax_amount': float(inv.tax_amount or 0.0),
        'discount_amount': float(inv.discount_amount or 0.0),
        'total_after_tax_discount': float(inv.total_after_tax_discount or 0.0),
        'branch': getattr(inv, 'branch', None),
        'branch_code': getattr(inv, 'branch', None),
    }
    try:
        s = Settings.query.first()
    except Exception:
        s = None
    branch_name = BRANCH_LABELS.get(getattr(inv, 'branch', None) or '', getattr(inv, 'branch', ''))
    # Use first print timestamp if available; fallback to archive saved_at or invoice.created_at
    dt_str = None
    try:
        logs = kv_get(f"print_log:{inv.invoice_number}", []) or []
    except Exception:
        logs = []
    try:
        if logs:
            from datetime import datetime as _dti
            def _p(x):
                try:
                    return _dti.fromisoformat((x or {}).get('ts') or '')
                except Exception:
                    return None
            cand = [_p(x) for x in logs]
            cand = [c for c in cand if c]
            if cand:
                dtp = min(cand)
                dt_str = dtp.strftime('%H:%M:%S %Y-%m-%d')
    except Exception:
        dt_str = None
    if not dt_str:
        try:
            meta = kv_get(f"pdf_path:sales:{inv.invoice_number}", {}) or {}
            sa = meta.get('saved_at')
            if sa:
                from datetime import datetime as _dti
                try:
                    dtp = _dti.fromisoformat(sa)
                    dt_str = dtp.strftime('%H:%M:%S %Y-%m-%d')
                except Exception:
                    dt_str = None
        except Exception:
            dt_str = None
    if not dt_str:
        try:
            from models import Payment
            dtp_row = db.session.query(Payment.payment_date).\
                filter(Payment.invoice_type == 'sales', Payment.invoice_id == inv.id).\
                order_by(Payment.payment_date.asc()).first()
            if dtp_row and dtp_row[0]:
                dt_str = dtp_row[0].strftime('%H:%M:%S %Y-%m-%d')
        except Exception:
            pass
    if not dt_str:
        try:
            if getattr(inv, 'created_at', None):
                dt_str = inv.created_at.strftime('%H:%M:%S %Y-%m-%d')
            else:
                dt_str = get_saudi_now().strftime('%H:%M:%S %Y-%m-%d')
        except Exception:
            dt_str = get_saudi_now().strftime('%H:%M:%S %Y-%m-%d')
    # Prepare embedded logo as data URL for more reliable thermal printing
    logo_data_url = None
    try:
        if s and getattr(s, 'receipt_show_logo', False) and (s.logo_url or '').strip():
            url = s.logo_url.strip()
            import base64, mimetypes, os
            fpath = None
            # Static folder
            if url.startswith('/static/'):
                rel = url.split('/static/', 1)[1]
                fpath = os.path.join(current_app.static_folder, rel)
            # Persistent uploads folder
            elif url.startswith('/uploads/'):
                upload_dir = os.getenv('UPLOAD_DIR') or os.path.join(current_app.static_folder, 'uploads')
                rel = url.split('/uploads/', 1)[1]
                fpath = os.path.join(upload_dir, rel)
            if fpath and os.path.exists(fpath):
                with open(fpath, 'rb') as f:
                    b = f.read()
                mime = mimetypes.guess_type(fpath)[0] or 'image/png'
                logo_data_url = f'data:{mime};base64,' + base64.b64encode(b).decode('ascii')
    except Exception:
        logo_data_url = None
    # Generate ZATCA-compliant QR (server-side) if possible
    qr_data_url = None
    try:
        from utils.qr import generate_zatca_qr_from_invoice
        b64 = generate_zatca_qr_from_invoice(inv, s, None)
        if b64:
            qr_data_url = 'data:image/png;base64,' + b64
    except Exception:
        qr_data_url = None
    return render_template('print/receipt.html', inv=inv_ctx, items=items_ctx,
                           settings=s, branch_name=branch_name, date_time=dt_str,
                           display_invoice_number=inv.invoice_number,
                           qr_data_url=qr_data_url,
                           logo_data_url=logo_data_url,
                           paid=True)


@main.route('/print/order-preview/<branch>/<int:table>', methods=['GET'], endpoint='print_order_preview')
@login_required
def print_order_preview(branch, table):
    # Read draft
    rec = kv_get(f'draft:{branch}:{table}', {}) or {}
    raw_items = rec.get('items') or []
    items_ctx = []
    subtotal = 0.0
    for it in raw_items:
        meal_id = it.get('meal_id') or it.get('id')
        qty = float(it.get('qty') or it.get('quantity') or 1)
        name = it.get('name') or ''
        price = it.get('price') or 0.0
        if (not name) or (not price) and meal_id:
            m = MenuItem.query.get(int(meal_id))
            if m:
                name = name or m.name
                price = price or float(m.price)
        line = float(price or 0) * qty
        subtotal += line
        items_ctx.append({'product_name': name, 'quantity': qty, 'total_price': line})

    tax_pct = float(rec.get('tax_pct') or 15)
    discount_pct = float(rec.get('discount_pct') or 0)
    vat_amount = subtotal * (tax_pct/100.0)
    discount_amount = (subtotal + vat_amount) * (discount_pct/100.0)
    total_after = subtotal + vat_amount - discount_amount

    # Resolve payment method from draft if chosen in POS
    try:
        pm_raw = (rec.get('payment_method') or '').strip()
    except Exception:
        pm_raw = ''
    inv_ctx = {
        'invoice_number': '',  # no real invoice number for preview
        'table_number': table,
        'customer_name': (rec.get('customer') or {}).get('name') or '',
        'customer_phone': (rec.get('customer') or {}).get('phone') or '',
        'payment_method': pm_raw.upper() if pm_raw else '',
        'status': 'DRAFT',
        'total_before_tax': round(subtotal, 2),
        'tax_amount': round(vat_amount, 2),
        'discount_amount': round(discount_amount, 2),
        'total_after_tax_discount': round(total_after, 2),
        'branch': branch,
        'branch_code': branch,
    }
    try:
        s = Settings.query.first()
    except Exception:
        s = None
    branch_name = BRANCH_LABELS.get(branch, branch)
    dt_str = get_saudi_now().strftime('%Y-%m-%d %H:%M:%S')
    # Keep preview number generation as before, but with INV prefix to match final
    order_no = f"INV-{int(datetime.utcnow().timestamp())}-{branch[:2]}{table}"
    # Persist this preview number so checkout reuses exactly the same value
    try:
        rec['preview_invoice_number'] = order_no
        kv_set(f'draft:{branch}:{table}', rec)
    except Exception:
        pass
    # Save an OrderInvoice record to track pre-payment prints
    try:
        from models import OrderInvoice
        # Create table if missing (first run)
        try:
            OrderInvoice.__table__.create(bind=db.engine, checkfirst=True)
        except Exception:
            pass

        # Avoid duplicates if user reopens/refreshes preview
        existing = OrderInvoice.query.filter_by(invoice_no=order_no).first()
        if not existing:
            # Build items JSON with name, qty, unit, line_total
            items_json = []
            try:
                for it in raw_items:
                    name = it.get('name') or ''
                    qty = float(it.get('qty') or it.get('quantity') or 1)
                    price = it.get('price')
                    if (not price):
                        meal_id = it.get('meal_id') or it.get('id')
                        if meal_id:
                            m = MenuItem.query.get(int(meal_id))
                            price = float(getattr(m, 'price', 0) or 0)
                    unit = float(price or 0)
                    line_total = unit * qty
                    items_json.append({'name': name, 'qty': qty, 'unit': unit, 'line_total': line_total})
            except Exception:
                # Fallback from items_ctx if needed
                for r in items_ctx:
                    q = float(r.get('quantity') or 1)
                    t = float(r.get('total_price') or 0)
                    u = (t / q) if q else 0
                    items_json.append({'name': r.get('product_name') or '-', 'qty': q, 'unit': u, 'line_total': t})

            order = OrderInvoice(
                branch=branch,
                invoice_no=order_no,
                customer=((rec.get('customer') or {}).get('name') or None),
                items=items_json,
                subtotal=float(subtotal),
                discount=float(discount_amount),
                vat=float(vat_amount),
                total=float(total_after),
                payment_method='PENDING'
            )
            db.session.add(order)
            db.session.commit()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        current_app.logger.error('Failed to save order invoice (preview): %s', e)
    # Generate ZATCA QR as base64 PNG and pass to template
    qr_data_url = None
    try:
        from utils.qr import generate_zatca_qr_base64
        b64 = generate_zatca_qr_base64(
            getattr(s, 'company_name', 'Restaurant'),
            getattr(s, 'tax_number', '123456789012345'),
            get_saudi_now(),
            float(total_after),
            float(vat_amount)
        )
        if b64:
            qr_data_url = 'data:image/png;base64,' + b64
    except Exception:
        qr_data_url = None

    return render_template('print/receipt.html', inv=inv_ctx, items=items_ctx,
                           settings=s, branch_name=branch_name, date_time=dt_str,
                           display_invoice_number=order_no,
                           qr_data_url=qr_data_url,
                           paid=False)


@main.route('/print/order-slip/<branch>/<int:table>', methods=['GET'], endpoint='print_order_slip')
@login_required
def print_order_slip(branch, table):
    rec = kv_get(f'draft:{branch}:{table}', {}) or {}
    raw_items = rec.get('items') or []
    items_ctx = []
    subtotal = 0.0
    for it in raw_items:
        meal_id = it.get('meal_id') or it.get('id')
        qty = float(it.get('qty') or it.get('quantity') or 1)
        name = it.get('name') or ''
        price = it.get('price') or 0.0
        if (not name) or ((not price) and meal_id):
            m = MenuItem.query.get(int(meal_id))
            if m:
                name = name or m.name
                price = price or float(m.price)
        line = float(price or 0) * qty
        subtotal += line
        items_ctx.append({'product_name': name, 'quantity': qty, 'unit_price': float(price or 0), 'discount': 0.0, 'tax': 0.0, 'line_total': line})
    tax_pct = float(rec.get('tax_pct') or 15)
    discount_pct = float(rec.get('discount_pct') or 0)
    vat_amount = subtotal * (tax_pct/100.0)
    discount_amount = (subtotal + vat_amount) * (discount_pct/100.0)
    total_after = subtotal + vat_amount - discount_amount
    order_seq = None
    today_str = get_saudi_now().date().isoformat()
    rec_seq_date = (rec.get('order_seq_date') or '') if isinstance(rec, dict) else ''
    try:
        order_seq = rec.get('order_seq')
    except Exception:
        order_seq = None
    if not order_seq or rec_seq_date != today_str:
        key = f"order_seq:{today_str}"
        cur = kv_get(key, 1) or 1
        order_seq = int(cur)
        kv_set(key, order_seq + 1)
        rec['order_seq'] = order_seq
        rec['order_seq_date'] = today_str
        kv_set(f'draft:{branch}:{table}', rec)
    branch_name = BRANCH_LABELS.get(branch, branch)
    dt_str = get_saudi_now().strftime('%Y-%m-%d %H:%M:%S')
    return render_template('print/order_slip.html',
                           order_number=order_seq,
                           branch_name=branch_name,
                           date_time=dt_str,
                           items=items_ctx,
                           subtotal=round(subtotal, 2),
                           discount=round(discount_amount, 2),
                           vat=round(vat_amount, 2),
                           total=round(total_after, 2))

@main.route('/invoice/print/<invoice_id>', methods=['GET'], endpoint='invoice_print')
@login_required
def invoice_print(invoice_id):
    """
    New standalone invoice print page for issued invoices
    This is a read-only print view that opens in a new window
    """
    # Find invoice by ID or invoice number
    inv = None
    try:
        inv = SalesInvoice.query.get(int(invoice_id))
    except Exception:
        inv = SalesInvoice.query.filter_by(invoice_number=str(invoice_id)).first()
    
    if not inv:
        return 'Invoice not found', 404
    
    # Check if invoice is issued (status should be 'issued' or 'finalized')
    if inv.status not in ['issued', 'finalized', 'paid']:
        return 'Invoice not issued yet', 403
    
    items = SalesInvoiceItem.query.filter_by(invoice_id=inv.id).all()
    
    # Prepare items context
    items_ctx = []
    for it in items:
        line = float(it.price_before_tax or 0) * float(it.quantity or 0)
        items_ctx.append({
            'product_name': it.product_name or '',
            'quantity': float(it.quantity or 0),
            'total_price': line,
        })

    inv_ctx = {
        'invoice_number': inv.invoice_number,
        'table_number': inv.table_number,
        'customer_name': inv.customer_name,
        'customer_phone': inv.customer_phone,
        'payment_method': inv.payment_method,
        'status': inv.status.upper(),
        'total_before_tax': float(inv.total_before_tax or 0.0),
        'tax_amount': float(inv.tax_amount or 0.0),
        'discount_amount': float(inv.discount_amount or 0.0),
        'total_after_tax_discount': float(inv.total_after_tax_discount or 0.0),
        'branch': getattr(inv, 'branch', None),
        'branch_code': getattr(inv, 'branch', None),
    }
    
    try:
        s = Settings.query.first()
    except Exception:
        s = None
        
    branch_name = BRANCH_LABELS.get(getattr(inv, 'branch', None) or '', getattr(inv, 'branch', ''))
    try:
        if inv.created_at:
            base_dt = inv.created_at
            if getattr(base_dt, 'tzinfo', None):
                dt_str = base_dt.astimezone(KSA_TZ).strftime('%Y-%m-%d %H:%M:%S')
            else:
                try:
                    dt_str = base_dt.replace(tzinfo=_pytz.UTC).astimezone(KSA_TZ).strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    dt_str = KSA_TZ.localize(base_dt).strftime('%Y-%m-%d %H:%M:%S')
        else:
            dt_str = get_saudi_now().strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        dt_str = get_saudi_now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Prepare embedded logo as data URL for thermal printing
    logo_data_url = None
    try:
        if s and getattr(s, 'receipt_show_logo', False) and (s.logo_url or '').strip():
            url = s.logo_url.strip()
            import base64, mimetypes, os
            fpath = None
            # Static folder
            if url.startswith('/static/'):
                rel = url.split('/static/', 1)[1]
                fpath = os.path.join(current_app.static_folder, rel)
            # Persistent uploads folder
            elif url.startswith('/uploads/'):
                upload_dir = os.getenv('UPLOAD_DIR') or os.path.join(current_app.static_folder, 'uploads')
                rel = url.split('/uploads/', 1)[1]
                fpath = os.path.join(upload_dir, rel)
            if fpath and os.path.exists(fpath):
                with open(fpath, 'rb') as f:
                    b = f.read()
                mime = mimetypes.guess_type(fpath)[0] or 'image/png'
                logo_data_url = f'data:{mime};base64,' + base64.b64encode(b).decode('ascii')
    except Exception:
        logo_data_url = None
    
    # Generate ZATCA-compliant QR
    qr_data_url = None
    try:
        from utils.qr import generate_zatca_qr_from_invoice
        b64 = generate_zatca_qr_from_invoice(inv, s, None)
        if b64:
            qr_data_url = 'data:image/png;base64,' + b64
    except Exception:
        qr_data_url = None
    
    # Render standalone print template
    return render_template('print/invoice_print.html', 
                         inv=inv_ctx, 
                         items=items_ctx,
                         settings=s, 
                         branch_name=branch_name, 
                         date_time=dt_str,
                         display_invoice_number=inv.invoice_number,
                         qr_data_url=qr_data_url,
                         logo_data_url=logo_data_url,
                         paid=(inv.status in ['paid', 'finalized']))



@main.route('/customers', methods=['GET', 'POST'], endpoint='customers')
@login_required
def customers():
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        phone = (request.form.get('phone') or '').strip() or None
        discount = request.form.get('discount_percent', type=float) or 0.0
        if not name:
            flash('Name is required', 'danger')
        else:
            try:
                c = Customer(name=name, phone=phone, discount_percent=float(discount or 0))
                db.session.add(c)
                db.session.commit()
                flash('Customer added', 'success')
            except Exception as e:
                db.session.rollback()
                flash('Error adding customer', 'danger')
        return redirect(url_for('main.customers'))
    # GET
    q = (request.args.get('q') or '').strip()
    qry = Customer.query.filter_by(active=True)
    if q:
        like = f"%{q}%"
        qry = qry.filter((Customer.name.ilike(like)) | (Customer.phone.ilike(like)))
    customers = qry.order_by(Customer.name.asc()).all()
    return render_template('customers.html', customers=customers, q=q)

@main.route('/customers/<int:cid>/toggle', methods=['POST'], endpoint='customer_toggle')
@login_required
def customer_toggle(cid):
    try:
        c = db.session.get(Customer, cid)
        if not c:
            flash('Customer not found', 'warning')
            return redirect(url_for('main.customers'))
        c.active = not bool(getattr(c, 'active', True))
        db.session.commit()
        flash('Customer status updated', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating customer: {e}', 'danger')
    return redirect(url_for('main.customers'))


@main.route('/customers/<int:cid>/delete', methods=['POST'], endpoint='customer_delete')
@login_required
def customer_delete(cid):
    # Try hard delete; if constrained, fall back to soft deactivate
    try:
        c = db.session.get(Customer, cid)
        if not c:
            flash('Customer not found', 'warning')
        else:
            try:
                db.session.delete(c)
                db.session.commit()
                flash('Customer deleted', 'success')
            except IntegrityError:
                db.session.rollback()
                c.active = False
                db.session.commit()
                flash('Customer has related records; deactivated instead', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting customer: {e}', 'danger')
    return redirect(url_for('main.customers'))


# ---- Customers: POS-friendly search alias and AJAX create ----
@main.route('/api/pos/<branch>/customers/search', methods=['GET'], endpoint='api_pos_customers_search')
@login_required
def api_pos_customers_search(branch):
    q = (request.args.get('q') or '').strip()
    if not q:
        rows = Customer.query.filter_by(active=True).order_by(Customer.name.asc()).limit(10).all()
    else:
        like = f"%{q}%"
        rows = Customer.query.filter(
            Customer.active == True,
            (Customer.name.ilike(like)) | (Customer.phone.ilike(like))
        ).order_by(Customer.name.asc()).limit(20).all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'phone': c.phone,
        'discount_percent': float(c.discount_percent or 0)
    } for c in rows])


@main.route('/api/customers', methods=['POST'], endpoint='api_customers_create')
@login_required
@csrf.exempt
def api_customers_create():
    try:
        if request.is_json:
            data = request.get_json() or {}
            name = (data.get('name') or '').strip()
            phone = (data.get('phone') or '').strip() or None
            discount = float((data.get('discount_percent') or 0) or 0)
        else:
            name = (request.form.get('name') or '').strip()
            phone = (request.form.get('phone') or '').strip() or None
            discount = request.form.get('discount_percent', type=float) or 0.0
        if not name:
            return jsonify({'ok': False, 'error': 'Name is required'}), 400
        c = Customer(name=name, phone=phone, discount_percent=float(discount or 0))
        db.session.add(c)
        db.session.commit()
        return jsonify({'ok': True, 'customer': {
            'id': c.id, 'name': c.name, 'phone': c.phone,
            'discount_percent': float(c.discount_percent or 0)
        }})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/salaries/accrual', methods=['POST'], endpoint='salaries_accrual')
@login_required
def salaries_accrual():
    if not user_can('salaries','add'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    try:
        emp_id_raw = (request.form.get('employee_id') or '').strip().lower()
        month_str = (request.form.get('month') or get_saudi_now().strftime('%Y-%m')).strip()
        y, m = month_str.split('-'); year = int(y); month = int(m)
        if emp_id_raw == 'all':
            employee_ids = [int(e.id) for e in Employee.query.all()]
        else:
            employee_ids = [int(emp_id_raw)] if emp_id_raw else []
        created = 0
        for _emp_id in employee_ids:
            s = Salary.query.filter_by(employee_id=_emp_id, year=year, month=month).first()
            if not s:
                try:
                    from models import EmployeeSalaryDefault
                    d = EmployeeSalaryDefault.query.filter_by(employee_id=_emp_id).first()
                    base = float(getattr(d, 'base_salary', 0) or 0)
                    allow = float(getattr(d, 'allowances', 0) or 0)
                    ded = float(getattr(d, 'deductions', 0) or 0)
                except Exception:
                    base = allow = ded = 0.0
                total = max(0.0, base + allow - ded)
                s = Salary(employee_id=_emp_id, year=year, month=month, basic_salary=base, allowances=allow, deductions=ded, previous_salary_due=0.0, total_salary=total, status='due')
                db.session.add(s)
                db.session.flush()
            amount = float(s.total_salary or 0.0)
            if amount <= 0:
                continue
            try:
                from models import JournalEntry, JournalLine
                exp_acc = _account(SHORT_TO_NUMERIC['SAL_EXP'][0], CHART_OF_ACCOUNTS['5310']['name'], CHART_OF_ACCOUNTS['5310']['type'])
                liab_acc = _account(SHORT_TO_NUMERIC['PAYROLL_LIAB'][0], CHART_OF_ACCOUNTS['2130']['name'], CHART_OF_ACCOUNTS['2130']['type'])
                je = JournalEntry(entry_number=f"JE-SALACC-{_emp_id}-{year}{month:02d}", date=get_saudi_now().date(), branch_code=None, description=f"Payroll accrual {_emp_id} {year}-{month}", status='posted', total_debit=amount, total_credit=amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None), salary_id=s.id)
                db.session.add(je); db.session.flush()
                if exp_acc:
                    db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=exp_acc.id, debit=amount, credit=0, description='Payroll expense', line_date=get_saudi_now().date(), employee_id=_emp_id))
                if liab_acc:
                    db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=liab_acc.id, debit=0, credit=amount, description='Payroll liability', line_date=get_saudi_now().date(), employee_id=_emp_id))
                db.session.commit()
                try:
                    _post_ledger(get_saudi_now().date(), 'SAL_EXP', 'مصروف رواتب', 'expense', amount, 0.0, f'ACCRUAL {_emp_id} {year}-{month}')
                    _post_ledger(get_saudi_now().date(), 'PAYROLL_LIAB', 'رواتب مستحقة', 'liability', 0.0, amount, f'ACCRUAL {_emp_id} {year}-{month}')
                except Exception:
                    pass
                created += 1
            except Exception:
                db.session.rollback()
        return jsonify({'ok': True, 'count': created})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400


@main.route('/suppliers', methods=['GET','POST'], endpoint='suppliers')
@login_required
def suppliers():
    if request.method == 'POST':
        try:
            # Accept optional CR/IBAN even if model lacks fields; store in notes
            cr = (request.form.get('cr_number') or '').strip() or None
            iban = (request.form.get('iban') or '').strip() or None
            notes_in = (request.form.get('notes') or '').strip() or None
            notes_extra = []
            if notes_in:
                notes_extra.append(notes_in)
            if cr:
                notes_extra.append(f"CR: {cr}")
            if iban:
                notes_extra.append(f"IBAN: {iban}")
            notes_text = ' | '.join(notes_extra) if notes_extra else None

            s = Supplier(
                name=(request.form.get('name') or '').strip(),
                contact_person=request.form.get('contact_person') or None,
                phone=request.form.get('phone') or None,
                email=request.form.get('email') or None,
                tax_number=request.form.get('tax_number') or None,
                address=request.form.get('address') or None,
                payment_method=(request.form.get('payment_method') or 'CASH').strip().upper(),
                notes=notes_text,
                active=True if str(request.form.get('active')).lower() in ['1','true','yes','on'] else False,
            )
            if not s.name:
                raise ValueError('Name is required')
            if ext_db is not None:
                ext_db.session.add(s)
                ext_db.session.commit()
            else:
                db.session.add(s)
                db.session.commit()
            flash('Supplier added', 'success')
        except Exception as e:
            if ext_db is not None:
                ext_db.session.rollback()
            else:
                db.session.rollback()
            flash(f'Could not add supplier: {e}', 'danger')
        return redirect(url_for('main.suppliers'))
    try:
        q = (request.args.get('q') or '').strip()
        query = Supplier.query
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Supplier.name.ilike(like)) |
                (Supplier.contact_person.ilike(like)) |
                (Supplier.phone.ilike(like)) |
                (Supplier.email.ilike(like)) |
                (Supplier.tax_number.ilike(like)) |
                (Supplier.address.ilike(like)) |
                (Supplier.notes.ilike(like))
            )
        all_suppliers = query.order_by(Supplier.name.asc()).all()
    except Exception:
        all_suppliers = []
    return render_template('suppliers.html', suppliers=all_suppliers)

@main.route('/suppliers/list', methods=['GET'], endpoint='suppliers_list')
@login_required
def suppliers_list():
    try:
        q = (request.args.get('q') or '').strip()
        query = Supplier.query
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Supplier.name.ilike(like)) |
                (Supplier.contact_person.ilike(like)) |
                (Supplier.phone.ilike(like)) |
                (Supplier.email.ilike(like)) |
                (Supplier.tax_number.ilike(like)) |
                (Supplier.address.ilike(like)) |
                (Supplier.notes.ilike(like))
            )
        all_suppliers = query.order_by(Supplier.name.asc()).all()
    except Exception:
        all_suppliers = []
    return render_template('suppliers_list.html', suppliers=all_suppliers)

@main.route('/suppliers/<int:sid>/toggle', methods=['POST'], endpoint='supplier_toggle')
@login_required
def supplier_toggle(sid):
    try:
        s = Supplier.query.get(sid)
        if not s:
            flash('Supplier not found', 'warning')
            return redirect(url_for('main.suppliers'))
        s.active = not bool(getattr(s, 'active', True))
        if ext_db is not None:
            ext_db.session.commit()
        else:
            db.session.commit()
        flash('Supplier status updated', 'success')
    except Exception as e:
        if ext_db is not None:
            ext_db.session.rollback()
        else:
            db.session.rollback()
        flash(f'Error updating supplier: {e}', 'danger')
    return redirect(url_for('main.suppliers'))

@main.route('/suppliers/export', methods=['GET'], endpoint='suppliers_export')
@login_required
def suppliers_export():
    try:
        q = (request.args.get('q') or '').strip()
        query = Supplier.query
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Supplier.name.ilike(like)) |
                (Supplier.contact_person.ilike(like)) |
                (Supplier.phone.ilike(like)) |
                (Supplier.email.ilike(like)) |
                (Supplier.tax_number.ilike(like)) |
                (Supplier.address.ilike(like)) |
                (Supplier.notes.ilike(like))
            )
        rows = query.order_by(Supplier.name.asc()).all()

        def escape_csv(val: str) -> str:
            if val is None:
                return ''
            s = str(val)
            if any(c in s for c in [',','"','\n','\r']):
                s = '"' + s.replace('"','""') + '"'
            return s

        parts = []
        header = [
            'Name','Contact Person','Phone','Email','Tax Number','CR Number','IBAN','Address','Notes','Active'
        ]
        parts.append(','.join(header))
        for s in rows:
            cr = getattr(s, 'cr_number', None)
            iban = getattr(s, 'iban', None)
            if (not cr or not iban) and (s.notes):
                try:
                    for part in (s.notes or '').split('|'):
                        p = (part or '').strip()
                        if (not cr) and p[:3] == 'CR:':
                            cr = p[3:].strip()
                        if (not iban) and p[:5] == 'IBAN:':
                            iban = p[5:].strip()
                except Exception:
                    pass
            row = [
                escape_csv(s.name),
                escape_csv(s.contact_person or ''),
                escape_csv(s.phone or ''),
                escape_csv(s.email or ''),
                escape_csv(s.tax_number or ''),
                escape_csv(cr or ''),
                escape_csv(iban or ''),
                escape_csv(s.address or ''),
                escape_csv(s.notes or ''),
                '1' if getattr(s, 'active', True) else '0'
            ]
            parts.append(','.join(row))
        csv_data = '\n'.join(parts)
        return current_app.response_class(csv_data, mimetype='text/csv; charset=utf-8', headers={
            'Content-Disposition': 'attachment; filename="suppliers.csv"'
        })
    except Exception as e:
        flash(f'Failed to export suppliers: {e}', 'danger')
        return redirect(url_for('main.suppliers'))

@main.route('/suppliers/<int:sid>/delete', methods=['POST'], endpoint='supplier_delete')
@login_required
def supplier_delete(sid):
    # Try hard delete; if constrained, fall back to soft deactivate
    try:
        s = Supplier.query.get(sid)
        if not s:
            flash('Supplier not found', 'warning')
        else:
            try:
                if ext_db is not None:
                    ext_db.session.delete(s)
                    ext_db.session.commit()
                else:
                    db.session.delete(s)
                    db.session.commit()
                flash('Supplier deleted', 'success')
            except IntegrityError:
                # Rollback then soft deactivate
                if ext_db is not None:
                    ext_db.session.rollback()
                    s.active = False
                    ext_db.session.commit()
                else:
                    db.session.rollback()
                    s.active = False
                    db.session.commit()
                flash('Supplier has related records; deactivated instead', 'warning')
    except Exception as e:
        if ext_db is not None:
            ext_db.session.rollback()
        else:
            db.session.rollback()
        flash(f'Error deleting supplier: {e}', 'danger')
    return redirect(url_for('main.suppliers'))

@main.route('/menu', methods=['GET'], endpoint='menu')
@login_required
def menu():
    warmup_db_once()
    try:
        cats = MenuCategory.query.order_by(MenuCategory.sort_order, MenuCategory.name).all()
    except Exception:
        try:
            cats = MenuCategory.query.order_by(MenuCategory.name).all()
        except Exception:
            cats = MenuCategory.query.all()
    current = None
    cat_id = request.args.get('cat_id', type=int)
    if cat_id:
        try:
            current = db.session.get(MenuCategory, int(cat_id))
        except Exception:
            current = None
    if not current and cats:
        current = cats[0]

    # Cleanup exact duplicates (same name and same price) within the current section
    if current:
        try:
            seen = {}
            dups = []
            for it in MenuItem.query.filter_by(category_id=current.id).order_by(MenuItem.id.asc()).all():
                key = ((it.name or '').strip().lower(), round(float(it.price or 0), 2))
                if key in seen:
                    dups.append(it)
                else:
                    seen[key] = it.id
            if dups:
                for it in dups:
                    db.session.delete(it)
                db.session.commit()
        except Exception:
            db.session.rollback()

    items = []
    if current:
        items = MenuItem.query.filter_by(category_id=current.id).order_by(MenuItem.name).all()
    # meals for dropdown (active first; fallback to all)
    try:
        meals = Meal.query.filter_by(active=True).order_by(Meal.name.asc()).all()
        if not meals:
            meals = Meal.query.order_by(Meal.name.asc()).all()
    except Exception:
        meals = []
    # counts per category
    item_counts = {c.id: MenuItem.query.filter_by(category_id=c.id).count() for c in cats}
    return render_template('menu.html', sections=cats, current_section=current, items=items, item_counts=item_counts, meals=meals)


@main.route('/menu/category/add', methods=['POST'], endpoint='menu_category_add')
@login_required
def menu_category_add():
    warmup_db_once()
    name = (request.form.get('name') or '').strip()
    sort = request.form.get('display_order', type=int) or 0
    if not name:
        flash('Name is required', 'danger')
        return redirect(url_for('main.menu'))
    try:
        c = MenuCategory(name=name, sort_order=sort)
        db.session.add(c)
        db.session.commit()
        return redirect(url_for('main.menu', cat_id=c.id))
    except Exception as e:
        db.session.rollback()
        flash('Error creating category', 'danger')
        return redirect(url_for('main.menu'))


@main.route('/menu/category/<int:cat_id>/delete', methods=['POST'], endpoint='menu_category_delete')
@login_required
def menu_category_delete(cat_id):
    warmup_db_once()
    try:
        c = db.session.get(MenuCategory, int(cat_id))
        if c:
            MenuItem.query.filter_by(category_id=c.id).delete()
            db.session.delete(c)
            db.session.commit()
            flash('Category deleted', 'info')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting category', 'danger')
    return redirect(url_for('main.menu'))


@main.route('/menu/item/add', methods=['POST'], endpoint='menu_item_add')
@login_required
def menu_item_add():
    warmup_db_once()
    cat_id = request.form.get('section_id', type=int)
    meal_id = request.form.get('meal_id', type=int)
    name = (request.form.get('name') or '').strip()
    price_raw = request.form.get('price')
    # parse price robustly (support comma decimal)
    try:
        price = float((price_raw or '').replace(',', '.')) if price_raw not in [None, ''] else None
    except Exception:
        price = None
    if not cat_id:
        flash('Missing category', 'danger')
        return redirect(url_for('main.menu'))
    try:
        # Resolve meal if provided
        meal = db.session.get(Meal, int(meal_id)) if meal_id else None
        if meal:
            disp_name = f"{meal.name} / {meal.name_ar}" if getattr(meal, 'name_ar', None) else (meal.name or name)
            final_name = (name or '').strip() or disp_name
            final_price = float(price) if price is not None else float(meal.selling_price or 0.0)
        else:
            final_name = name
            final_price = float(price or 0.0)
        final_name = (final_name or '').strip()[:150]
        if not final_name:
            flash('Missing item name', 'danger')
            return redirect(url_for('main.menu', cat_id=cat_id))
        it = MenuItem(name=final_name, price=final_price, category_id=int(cat_id), meal_id=(meal.id if meal else None))
        db.session.add(it)
        db.session.commit()
        return redirect(url_for('main.menu', cat_id=cat_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating item: {e}', 'danger')
        return redirect(url_for('main.menu', cat_id=cat_id))


@main.route('/menu/item/<int:item_id>/update', methods=['POST'], endpoint='menu_item_update')
@login_required
def menu_item_update(item_id):
    warmup_db_once()
    name = (request.form.get('name') or '').strip()
    price = request.form.get('price', type=float)
    try:
        it = db.session.get(MenuItem, int(item_id))
        if not it:
            flash('Item not found', 'warning')
            return redirect(url_for('main.menu'))
        if name:
            it.name = name
        if price is not None:
            it.price = float(price)
        db.session.commit()
        return redirect(url_for('main.menu', cat_id=it.category_id))
    except Exception:
        db.session.rollback()
        flash('Error updating item', 'danger')
        return redirect(url_for('main.menu'))


@main.route('/menu/item/<int:item_id>/delete', methods=['POST'], endpoint='menu_item_delete')
@login_required
def menu_item_delete(item_id):
    warmup_db_once()
    try:
        it = db.session.get(MenuItem, int(item_id))
        if it:
            cat_id = it.category_id
            db.session.delete(it)
            db.session.commit()
            flash('Item deleted', 'info')
            return redirect(url_for('main.menu', cat_id=cat_id))
    except Exception:
        db.session.rollback()
        flash('Error deleting item', 'danger')
    return redirect(url_for('main.menu'))


# API endpoints for JS delete flows (optional)
@main.route('/api/items/<int:item_id>/delete', methods=['POST'], endpoint='api_item_delete')
@login_required
def api_item_delete(item_id):
    payload = request.get_json(silent=True) or {}
    if (payload.get('password') or '') != '1991':
        return jsonify({'ok': False, 'error': 'invalid_password'}), 403
    try:
        it = db.session.get(MenuItem, int(item_id))
        if not it:
            return jsonify({'ok': False, 'error': 'not_found'}), 404
        db.session.delete(it)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400


@main.route('/api/items/delete-all', methods=['POST'], endpoint='api_item_delete_all')
@login_required
def api_item_delete_all():
    payload = request.get_json(silent=True) or {}
    if (payload.get('password') or '') != '1991':
        return jsonify({'ok': False, 'error': 'invalid_password'}), 403
    try:
        deleted = db.session.query(MenuItem).delete()
        db.session.commit()
        return jsonify({'ok': True, 'deleted': int(deleted or 0)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400



# --- Users Management API (minimal) ---
@main.route('/api/users', methods=['POST'], endpoint='api_users_create')
@login_required
def api_users_create():
    try:
        data = request.get_json(silent=True) or {}
        username = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()
        active = bool(data.get('active', True))
        if not username or not password:
            return jsonify({'ok': False, 'error': 'username_password_required'}), 400
        # uniqueness
        if User.query.filter_by(username=username).first():
            return jsonify({'ok': False, 'error': 'username_taken'}), 400
        u = User(username=username)
        u.set_password(password)
        db.session.add(u)
        db.session.flush()  # to get u.id before commit
        try:
            kv_set(f"user_active:{u.id}", {'active': bool(active)})
        except Exception:
            pass
        db.session.commit()
        return jsonify({'ok': True, 'id': u.id, 'active': bool(active)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/users/<int:uid>', methods=['PATCH'], endpoint='api_users_update')
@login_required
def api_users_update(uid):
    try:
        u = db.session.get(User, uid)
        if not u:
            return jsonify({'ok': False, 'error': 'not_found'}), 404
        data = request.get_json(silent=True) or {}
        # Optional: allow password change
        pw = (data.get('password') or '').strip()
        if pw:
            u.set_password(pw)
        # Active status persisted in AppKV to avoid schema change
        if 'active' in data:
            try:
                kv_set(f"user_active:{u.id}", {'active': bool(data.get('active'))})
            except Exception:
                pass
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/users', methods=['DELETE'], endpoint='api_users_delete')
@login_required
def api_users_delete():
    try:
        data = request.get_json(silent=True) or {}
        ids = data.get('ids') or []
        if not isinstance(ids, list):
            return jsonify({'ok': False, 'error': 'invalid_ids'}), 400
        deleted = 0
        for i in ids:
            try:
                i = int(i)
            except Exception:
                continue
            if hasattr(current_user, 'id') and i == current_user.id:
                continue
            u = db.session.get(User, i)
            if u:
                db.session.delete(u)
                deleted += 1
        db.session.commit()
        return jsonify({'ok': True, 'deleted': deleted})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400

# Permissions: persist to AppKV to avoid schema changes
PERM_SCREENS = ['dashboard','sales','purchases','inventory','expenses','employees','salaries','financials','vat','reports','invoices','payments','customers','menu','settings','suppliers','table_settings','users','sample_data','archive','journal']

def _perms_k(uid, scope):
    return f"user_perms:{scope or 'all'}:{int(uid)}"

@main.route('/api/users/<int:uid>/permissions', methods=['GET'], endpoint='api_users_permissions_get')
@login_required
def api_users_permissions_get(uid):
    try:
        scope = (request.args.get('branch_scope') or 'all').strip().lower()
        k = _perms_k(uid, scope)
        row = AppKV.query.filter_by(k=k).first()
        if row:
            try:
                saved = json.loads(row.v)
                items = saved.get('items') or []
            except Exception:
                items = []
        else:
            items = [{ 'screen_key': s, 'view': False, 'add': False, 'edit': False, 'delete': False, 'print': False } for s in PERM_SCREENS]
            try:
                u = User.query.get(int(uid))
            except Exception:
                u = None
            if u and ((getattr(u, 'username','') == 'admin') or (getattr(u,'id',None) == 1) or (getattr(u,'role','') == 'admin')):
                for it in items:
                    if it.get('screen_key') == 'archive':
                        it['view'] = True
        return jsonify({'ok': True, 'items': items})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/users/<int:uid>/permissions', methods=['POST'], endpoint='api_users_permissions_set')
@login_required
def api_users_permissions_set(uid):
    try:
        data = request.get_json(silent=True) or {}
        scope = (data.get('branch_scope') or 'all').strip().lower()
        items = data.get('items') or []
        # Basic validation
        norm_items = []
        for it in items:
            key = (it.get('screen_key') or '').strip()
            if not key:
                continue
            norm_items.append({
                'screen_key': key,
                'view': bool(it.get('view')),
                'add': bool(it.get('add')),
                'edit': bool(it.get('edit')),
                'delete': bool(it.get('delete')),
                'print': bool(it.get('print')),
            })
        payload = {'items': norm_items, 'saved_at': datetime.utcnow().isoformat()}
        k = _perms_k(uid, scope)
        row = AppKV.query.filter_by(k=k).first()
        if not row:
            row = AppKV(k=k, v=json.dumps(payload, ensure_ascii=False))
            db.session.add(row)
        else:
            row.v = json.dumps(payload, ensure_ascii=False)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/settings', methods=['GET', 'POST'], endpoint='settings')
@login_required
def settings():
    # Load first (and only) settings row
    s = None
    try:
        s = Settings.query.first()
    except Exception:
        s = None
    if request.method == 'POST':
        try:
            if not s:
                s = Settings()
                (ext_db.session.add(s) if ext_db is not None else db.session.add(s))

            form_data = request.form if request.form else (request.get_json(silent=True) or {})

            # Helpers for coercion
            def as_bool(val):
                return True if str(val).lower() in ['1','true','yes','on'] else False
            def as_int(val, default=None):
                try:
                    return int(str(val).strip())
                except Exception:
                    return default
            def as_float(val, default=None):
                try:
                    return float(str(val).strip())
                except Exception:
                    return default

            # Strings (do not overwrite with empty string)
            for fld in ['company_name','tax_number','phone','address','email','currency','place_india_label','china_town_label','logo_url','default_theme','printer_type','footer_message','receipt_footer_text','china_town_phone1','china_town_phone2','place_india_phone1','place_india_phone2','china_town_logo_url','place_india_logo_url']:
                if hasattr(s, fld) and fld in form_data:
                    val = (form_data.get(fld) or '').strip()
                    if val != '':
                        setattr(s, fld, val)

            # Booleans (explicitly set False when absent)
            if hasattr(s, 'receipt_show_logo'):
                s.receipt_show_logo = as_bool(form_data.get('receipt_show_logo'))
            if hasattr(s, 'receipt_show_tax_number'):
                s.receipt_show_tax_number = as_bool(form_data.get('receipt_show_tax_number'))

            # Integers
            if hasattr(s, 'receipt_font_size'):
                s.receipt_font_size = as_int(form_data.get('receipt_font_size'), s.receipt_font_size if s.receipt_font_size is not None else 12)
            if hasattr(s, 'receipt_logo_height'):
                s.receipt_logo_height = as_int(form_data.get('receipt_logo_height'), s.receipt_logo_height if s.receipt_logo_height is not None else 40)
            if hasattr(s, 'receipt_extra_bottom_mm'):
                s.receipt_extra_bottom_mm = as_int(form_data.get('receipt_extra_bottom_mm'), s.receipt_extra_bottom_mm if s.receipt_extra_bottom_mm is not None else 15)

            # Keep as string for width ('80' or '58')
            if hasattr(s, 'receipt_paper_width') and 'receipt_paper_width' in form_data:
                s.receipt_paper_width = (form_data.get('receipt_paper_width') or '80').strip()

            # Floats / numerics
            if hasattr(s, 'vat_rate') and 'vat_rate' in form_data:
                s.vat_rate = as_float(form_data.get('vat_rate'), s.vat_rate if s.vat_rate is not None else 15.0)
            # Branch-specific
            if hasattr(s, 'china_town_vat_rate') and 'china_town_vat_rate' in form_data:
                s.china_town_vat_rate = as_float(form_data.get('china_town_vat_rate'), s.china_town_vat_rate if s.china_town_vat_rate is not None else 15.0)
            if hasattr(s, 'china_town_discount_rate') and 'china_town_discount_rate' in form_data:
                s.china_town_discount_rate = as_float(form_data.get('china_town_discount_rate'), s.china_town_discount_rate if s.china_town_discount_rate is not None else 0.0)
            if hasattr(s, 'place_india_vat_rate') and 'place_india_vat_rate' in form_data:
                s.place_india_vat_rate = as_float(form_data.get('place_india_vat_rate'), s.place_india_vat_rate if s.place_india_vat_rate is not None else 15.0)
            if hasattr(s, 'place_india_discount_rate') and 'place_india_discount_rate' in form_data:
                s.place_india_discount_rate = as_float(form_data.get('place_india_discount_rate'), s.place_india_discount_rate if s.place_india_discount_rate is not None else 0.0)
            # Passwords (strings)
            # Passwords: do not overwrite with empty string (avoid accidental clearing)
            if hasattr(s, 'china_town_void_password') and 'china_town_void_password' in form_data:
                _pwd = (form_data.get('china_town_void_password') or '').strip()
                if _pwd != '':
                    s.china_town_void_password = _pwd
            if hasattr(s, 'place_india_void_password') and 'place_india_void_password' in form_data:
                _pwd2 = (form_data.get('place_india_void_password') or '').strip()
                if _pwd2 != '':
                    s.place_india_void_password = _pwd2

            # Map currency_image_url -> currency_image
            if hasattr(s, 'currency_image'):
                cur_url = (form_data.get('currency_image_url') or '').strip()
                if cur_url:
                    s.currency_image = cur_url

            # Handle file uploads (logo, currency PNG)
            try:
                # Support persistent uploads directory via env var
                upload_dir = os.getenv('UPLOAD_DIR') or os.path.join(current_app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)
                path_prefix = '/uploads/' if os.getenv('UPLOAD_DIR') else '/static/uploads/'
                # Logo file (global)
                logo_file = request.files.get('logo_file')
                if logo_file and getattr(logo_file, 'filename', ''):
                    ext = os.path.splitext(logo_file.filename)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.svg', '.gif']:
                        fname = f"logo_{int(datetime.utcnow().timestamp())}{ext}"
                        fpath = os.path.join(upload_dir, fname)
                        logo_file.save(fpath)
                        s.logo_url = f"{path_prefix}{fname}"
                # China Town branch logo
                china_logo = request.files.get('china_logo_file')
                if china_logo and getattr(china_logo, 'filename', ''):
                    ext = os.path.splitext(china_logo.filename)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.svg', '.gif']:
                        fname = f"china_logo_{int(datetime.utcnow().timestamp())}{ext}"
                        fpath = os.path.join(upload_dir, fname)
                        china_logo.save(fpath)
                        if hasattr(s, 'china_town_logo_url'):
                            s.china_town_logo_url = f"{path_prefix}{fname}"
                # Palace India branch logo
                india_logo = request.files.get('india_logo_file')
                if india_logo and getattr(india_logo, 'filename', ''):
                    ext = os.path.splitext(india_logo.filename)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.svg', '.gif']:
                        fname = f"india_logo_{int(datetime.utcnow().timestamp())}{ext}"
                        fpath = os.path.join(upload_dir, fname)
                        india_logo.save(fpath)
                        if hasattr(s, 'place_india_logo_url'):
                            s.place_india_logo_url = f"{path_prefix}{fname}"
                # Currency PNG
                cur_file = request.files.get('currency_file')
                if cur_file and getattr(cur_file, 'filename', ''):


                    ext = os.path.splitext(cur_file.filename)[1].lower()
                    if ext == '.png':
                        fname = f"currency_{int(datetime.utcnow().timestamp())}{ext}"
                        fpath = os.path.join(upload_dir, fname)
                        cur_file.save(fpath)
                        if hasattr(s, 'currency_image'):
                            s.currency_image = f"{path_prefix}{fname}"
            except Exception:
                # Ignore upload errors but continue saving other fields
                pass

            try:
                acc_map = kv_get('acc_map', {}) or {}
                def _set_acc(key_alias, form_key):
                    val = (form_data.get(form_key) or '').strip()
                    if val:
                        acc_map[key_alias] = val
                _set_acc('AR_KEETA', 'acc_ar_keeta')
                _set_acc('AR_HUNGER', 'acc_ar_hunger')
                _set_acc('AR', 'acc_ar_default')
                _set_acc('REV_KEETA', 'acc_rev_keeta')
                _set_acc('REV_HUNGER', 'acc_rev_hunger')
                _set_acc('REV_CT', 'acc_rev_ct')
                _set_acc('REV_PI', 'acc_rev_pi')
                kv_set('acc_map', acc_map)
            except Exception:
                pass

            try:
                platforms = kv_get('platforms_map', []) or []
                pk = (form_data.get('platform_key') or '').strip().lower()
                if pk:
                    kws_raw = (form_data.get('platform_keywords') or '').strip()
                    keywords = [x.strip().lower() for x in kws_raw.split(',') if x.strip()]
                    entry = {
                        'key': pk,
                        'keywords': keywords,
                        'ar_code': (form_data.get('platform_ar_code') or '').strip(),
                        'rev_code': (form_data.get('platform_rev_code') or '').strip(),
                        'auto_unpaid': True
                    }
                    idx = None
                    for i, e in enumerate(platforms):
                        if (e.get('key') or '').strip().lower() == pk:
                            idx = i; break
                    if idx is not None:
                        # merge update
                        prev = platforms[idx]
                        prev.update({k: v for k, v in entry.items() if v})
                        platforms[idx] = prev
                    else:
                        platforms.append(entry)
                    kv_set('platforms_map', platforms)
            except Exception:
                pass

            (ext_db.session.commit() if ext_db is not None else db.session.commit())
            flash('Settings saved', 'success')
        except Exception as e:
            (ext_db.session.rollback() if ext_db is not None else db.session.rollback())
            flash(f'Could not save settings: {e}', 'danger')
        return redirect(url_for('main.settings'))
    return render_template('settings.html', s=s)

@main.route('/table-settings', endpoint='table_settings')
@login_required
def table_settings():
    # Define branches for table management
    branches = {
        'china_town': 'CHINA TOWN',
        'place_india': 'PALACE INDIA'
    }
    return render_template('table_settings.html', branches=branches)

@main.route('/table-manager/<branch_code>', endpoint='table_manager')
@login_required
def table_manager(branch_code):
    # Define branch labels
    branch_labels = {
        'china_town': 'CHINA TOWN',
        'place_india': 'PALACE INDIA'
    }
    
    # Check if branch exists
    if branch_code not in branch_labels:
        flash('Branch not found', 'error')
        return redirect(url_for('main.table_settings'))
    
    branch_label = branch_labels[branch_code]
    return render_template('table_manager.html', branch_code=branch_code, branch_label=branch_label)

@csrf.exempt
@main.route('/api/table-layout/<branch_code>', methods=['GET', 'POST'], endpoint='api_table_layout_fixed')
@login_required
def api_table_layout_fixed(branch_code):
    """Table layout API that integrates with the actual POS table system"""
    # Define branch labels
    branch_labels = {
        'china_town': 'CHINA TOWN',
        'place_india': 'PALACE INDIA'
    }
    
    # Check if branch exists
    if branch_code not in branch_labels:
        return jsonify({'error': 'Branch not found'}), 404
    
    if request.method == 'GET':
        # Load layout from the actual table sections system
        try:
            from models import TableSection, TableSectionAssignment

            def decode_name(name: str):
                name = name or ''
                visible = name
                layout = ''
                if '[rows:' in name and name.endswith(']'):
                    try:
                        before, bracket = name.rsplit('[rows:', 1)
                        visible = before.rstrip()
                        layout = bracket[:-1].strip()  # drop trailing ]
                    except Exception:
                        pass
                return visible, layout

            sections = TableSection.query.filter_by(branch_code=branch_code).order_by(TableSection.sort_order, TableSection.id).all()
            assignments = TableSectionAssignment.query.filter_by(branch_code=branch_code).all()

            # Convert to table manager format (preserve actual table numbers; do not synthesize placeholders)
            layout_sections = []
            for s in sections:
                visible, layout = decode_name(s.name)
                section_data = {
                    'id': s.id,
                    'name': visible,
                    'rows': []
                }

                # Collect actual tables assigned to this section (preserve string numbers)
                assigned = [a for a in assignments if a.section_id == s.id]
                # Sort by natural order: numeric if possible, else lexicographic
                def _key(a):
                    try:
                        return (0, int(str(a.table_number).strip()))
                    except Exception:
                        return (1, str(a.table_number) or '')
                assigned.sort(key=_key)

                # Map to table objects
                tables = [{ 'id': f'table_{idx+1}', 'number': a.table_number, 'seats': 4 } for idx, a in enumerate(assigned)]

                if layout:
                    try:
                        counts = [int(x.strip()) for x in layout.split(',') if x.strip()]
                    except Exception:
                        counts = []
                    i = 0
                    for cnt in counts:
                        if cnt <= 0: continue
                        row_tables = tables[i:i+cnt]
                        section_data['rows'].append({ 'id': f'row_{len(section_data["rows"])+1}', 'tables': row_tables })
                        i += cnt
                    if i < len(tables):
                        section_data['rows'].append({ 'id': f'row_{len(section_data["rows"])+1}', 'tables': tables[i:] })
                else:
                    # Single row with all assigned tables
                    section_data['rows'].append({ 'id': 'row_1', 'tables': tables })

                layout_sections.append(section_data)

            return jsonify({'sections': layout_sections})
        except Exception as e:
            print(f"Error loading layout: {e}")
            return jsonify({'sections': []})
    
    elif request.method == 'POST':
        # Save layout to the actual table sections system
        try:
            from models import TableSection, TableSectionAssignment
            
            # Get the data from request
            data = None
            if request.is_json:
                data = request.get_json()
            else:
                # Try to get raw data and parse it
                raw_data = request.get_data(as_text=True)
                if raw_data:
                    import json
                    data = json.loads(raw_data)
            
            if not data:
                return jsonify({'error': 'No data provided'}), 400
            
            sections_data = data.get('sections', [])
            
            # Clear existing sections and assignments for this branch
            TableSectionAssignment.query.filter_by(branch_code=branch_code).delete()
            TableSection.query.filter_by(branch_code=branch_code).delete()
            
            # Create new sections and assignments
            for section_idx, section in enumerate(sections_data):
                section_name = section.get('name', '')
                rows = section.get('rows', [])
                
                # Calculate layout string from rows
                layout_parts = []
                for row in rows:
                    table_count = len(row.get('tables', []))
                    if table_count > 0:
                        layout_parts.append(str(table_count))
                
                layout_string = ','.join(layout_parts) if layout_parts else ''
                encoded_name = f"{section_name} [rows:{layout_string}]" if layout_string else section_name
                
                # Create section
                new_section = TableSection(
                    branch_code=branch_code,
                    name=encoded_name,
                    sort_order=section.get('sort_order', section_idx)
                )
                db.session.add(new_section)
                db.session.flush()  # Get the ID
                
                # Create assignments for tables
                for row in rows:
                    for table in row.get('tables', []):
                        table_number = table.get('number')
                        if table_number:
                            assignment = TableSectionAssignment(
                                branch_code=branch_code,
                                table_number=str(table_number),
                                section_id=new_section.id
                            )
                            db.session.add(assignment)
            
            db.session.commit()
            return jsonify({'success': True, 'message': 'Layout saved successfully'})
        except Exception as e:
            db.session.rollback()
            print(f"Error saving layout: {e}")
            return jsonify({'error': f'Failed to save layout: {str(e)}'}), 500


# Serve uploaded files from a persistent directory when UPLOAD_DIR is set
@main.route('/uploads/<path:filename>', methods=['GET'], endpoint='serve_uploads')
def serve_uploads(filename):
    try:
        upload_dir = os.getenv('UPLOAD_DIR') or os.path.join(current_app.static_folder, 'uploads')
        return send_from_directory(upload_dir, filename)
    except Exception:
        return 'File not found', 404

@main.route('/users', endpoint='users')
@login_required
def users():
    # Provide users list to template so toolbar/actions work
    try:
        users_list = User.query.order_by(User.id.asc()).all()
    except Exception:
        users_list = []
    # Populate 'active' per user from AppKV (default True)
    for u in users_list:
        try:
            info = kv_get(f"user_active:{u.id}", {'active': True}) or {}
            setattr(u, 'active', bool(info.get('active', True)))
        except Exception:
            setattr(u, 'active', True)
    return render_template('users.html', users=users_list)

 

# ---------- VAT blueprint ----------
vat = Blueprint('vat', __name__, url_prefix='/vat')

@vat.route('/', endpoint='vat_dashboard')
@login_required
def vat_dashboard():
    # Period handling (quarterly default; supports monthly as well)
    period = (request.args.get('period') or 'quarterly').strip().lower()
    branch = (request.args.get('branch') or 'all').strip()

    try:
        y = request.args.get('year', type=int) or get_saudi_now().year
    except Exception:
        y = get_saudi_now().year

    start_date: date
    end_date: date
    if period == 'monthly':
        ym = (request.args.get('month') or '').strip()  # 'YYYY-MM'
        try:
            if ym and '-' in ym:
                yy, mm = ym.split('-')
                yy = int(yy); mm = int(mm)
                start_date = date(yy, mm, 1)
                nm_y = yy + (1 if mm == 12 else 0)
                nm_m = 1 if mm == 12 else mm + 1
                end_date = date(nm_y, nm_m, 1) - timedelta(days=1)
            else:
                # Fallback to current month
                today = get_saudi_now().date()
                start_date = date(today.year, today.month, 1)
                nm_y = today.year + (1 if today.month == 12 else 0)
                nm_m = 1 if today.month == 12 else today.month + 1
                end_date = date(nm_y, nm_m, 1) - timedelta(days=1)
        except Exception:
            today = get_saudi_now().date()
            start_date = date(today.year, today.month, 1)
            nm_y = today.year + (1 if today.month == 12 else 0)
            nm_m = 1 if today.month == 12 else today.month + 1
            end_date = date(nm_y, nm_m, 1) - timedelta(days=1)
        q = None
    else:
        # Quarterly
        try:
            q = request.args.get('quarter', type=int) or 1
        except Exception:
            q = 1
        q = q if q in (1, 2, 3, 4) else 1
        start_month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        end_month = {1: 3, 2: 6, 3: 9, 4: 12}[q]
        start_date = date(y, start_month, 1)
        next_month_first = date(y + (1 if end_month == 12 else 0), 1 if end_month == 12 else end_month + 1, 1)
        end_date = next_month_first - timedelta(days=1)

    # VAT rate and header info from Settings
    try:
        s = Settings.query.first()
    except Exception:
        s = None
    vat_rate = float(getattr(s, 'vat_rate', 15) or 15) / 100.0

    # SALES: category breakdown (based on tax_amount presence)
    sales_standard_base = sales_standard_vat = sales_standard_total = 0.0
    sales_zero_base = sales_zero_vat = sales_zero_total = 0.0
    sales_exempt_base = sales_exempt_vat = sales_exempt_total = 0.0
    sales_exports_base = sales_exports_vat = sales_exports_total = 0.0

    if hasattr(SalesInvoice, 'total_before_tax') and hasattr(SalesInvoice, 'tax_amount') and hasattr(SalesInvoice, 'total_after_tax_discount') and hasattr(SalesInvoice, 'date'):
        q_sales = db.session.query(SalesInvoice).filter(SalesInvoice.date.between(start_date, end_date))
        if hasattr(SalesInvoice, 'branch') and branch in ('place_india', 'china_town'):
            q_sales = q_sales.filter(SalesInvoice.branch == branch)

        # Standard rated: tax_amount > 0
        sales_standard_base = float((db.session
            .query(func.coalesce(func.sum(SalesInvoice.total_before_tax), 0))
            .filter(SalesInvoice.date.between(start_date, end_date))
            .filter((SalesInvoice.tax_amount > 0))
            .filter((SalesInvoice.branch == branch) if hasattr(SalesInvoice, 'branch') and branch in ('place_india','china_town') else True)
            .scalar() or 0.0))
        sales_standard_vat = float((db.session
            .query(func.coalesce(func.sum(SalesInvoice.tax_amount), 0))
            .filter(SalesInvoice.date.between(start_date, end_date))
            .filter((SalesInvoice.tax_amount > 0))
            .filter((SalesInvoice.branch == branch) if hasattr(SalesInvoice, 'branch') and branch in ('place_india','china_town') else True)
            .scalar() or 0.0))
        sales_standard_total = float((db.session
            .query(func.coalesce(func.sum(SalesInvoice.total_after_tax_discount), 0))
            .filter(SalesInvoice.date.between(start_date, end_date))
            .filter((SalesInvoice.tax_amount > 0))
            .filter((SalesInvoice.branch == branch) if hasattr(SalesInvoice, 'branch') and branch in ('place_india','china_town') else True)
            .scalar() or 0.0))

        # Zero rated: tax_amount == 0 (we cannot distinguish exempt/exports without extra fields)
        sales_zero_base = float((db.session
            .query(func.coalesce(func.sum(SalesInvoice.total_before_tax), 0))
            .filter(SalesInvoice.date.between(start_date, end_date))
            .filter((SalesInvoice.tax_amount == 0))
            .filter((SalesInvoice.branch == branch) if hasattr(SalesInvoice, 'branch') and branch in ('place_india','china_town') else True)
            .scalar() or 0.0))
        sales_zero_vat = 0.0
        sales_zero_total = float((db.session
            .query(func.coalesce(func.sum(SalesInvoice.total_after_tax_discount), 0))
            .filter(SalesInvoice.date.between(start_date, end_date))
            .filter((SalesInvoice.tax_amount == 0))
            .filter((SalesInvoice.branch == branch) if hasattr(SalesInvoice, 'branch') and branch in ('place_india','china_town') else True)
            .scalar() or 0.0))

        # Optional split of zero-tax sales into exempt vs exports using keyword mapping (AppKV)
        try:
            m = kv_get('vat_category_map', {}) or {}
            exempt_kw = [str(x).strip().lower() for x in (m.get('exempt_keywords') or []) if str(x).strip()]
            export_kw = [str(x).strip().lower() for x in (m.get('exports_keywords') or []) if str(x).strip()]
            zero_kw = [str(x).strip().lower() for x in (m.get('zero_keywords') or []) if str(x).strip()]
            q_items = (
                db.session.query(SalesInvoiceItem, SalesInvoice)
                .join(SalesInvoice, SalesInvoiceItem.invoice_id == SalesInvoice.id)
                .filter(SalesInvoice.date.between(start_date, end_date))
            )
            if hasattr(SalesInvoice, 'branch') and branch in ('place_india','china_town'):
                q_items = q_items.filter(SalesInvoice.branch == branch)
            for it, inv_row in q_items.all():
                name = (getattr(it, 'product_name', '') or '').strip().lower()
                qty = float(getattr(it, 'quantity', 0) or 0)
                price = float(getattr(it, 'price_before_tax', 0) or 0)
                disc = float(getattr(it, 'discount', 0) or 0)
                tax = float(getattr(it, 'tax', 0) or 0)
                base = max(0.0, (price * qty) - disc)
                if tax > 0:
                    sales_standard_base += base
                    sales_standard_vat += tax
                    sales_standard_total += (base + tax)
                else:
                    cat = 'zero'
                    if name and export_kw and any(kw in name for kw in export_kw):
                        cat = 'exports'
                    elif name and exempt_kw and any(kw in name for kw in exempt_kw):
                        cat = 'exempt'
                    elif name and zero_kw and any(kw in name for kw in zero_kw):
                        cat = 'zero'
                    if cat == 'exports':
                        sales_exports_base += base
                        sales_exports_vat += 0.0
                        sales_exports_total += base
                    elif cat == 'exempt':
                        sales_exempt_base += base
                        sales_exempt_vat += 0.0
                        sales_exempt_total += base
                    else:
                        sales_zero_base += base
                        sales_zero_vat += 0.0
                        sales_zero_total += base
        except Exception:
            pass

    # PURCHASES & EXPENSES: combine into deductible vs non-deductible
    purchases_deductible_base = purchases_deductible_vat = purchases_deductible_total = 0.0
    purchases_non_deductible_base = purchases_non_deductible_vat = purchases_non_deductible_total = 0.0

    try:
        # Purchases (items)
        # Deductible: items with tax > 0
        p_ded_base = db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.total_price - PurchaseInvoiceItem.tax), 0)) \
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id) \
            .filter(PurchaseInvoice.date.between(start_date, end_date)) \
            .filter(PurchaseInvoiceItem.tax > 0).scalar() or 0.0
        p_ded_vat = db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.tax), 0)) \
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id) \
            .filter(PurchaseInvoice.date.between(start_date, end_date)) \
            .filter(PurchaseInvoiceItem.tax > 0).scalar() or 0.0
        # Non-deductible: items with tax == 0
        p_nd_base = db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.total_price), 0)) \
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id) \
            .filter(PurchaseInvoice.date.between(start_date, end_date)) \
            .filter(PurchaseInvoiceItem.tax == 0).scalar() or 0.0

        # Expenses
        e_ded_base = db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_before_tax), 0)) \
            .filter(ExpenseInvoice.date.between(start_date, end_date)) \
            .filter(ExpenseInvoice.tax_amount > 0).scalar() or 0.0
        e_ded_vat = db.session.query(func.coalesce(func.sum(ExpenseInvoice.tax_amount), 0)) \
            .filter(ExpenseInvoice.date.between(start_date, end_date)) \
            .filter(ExpenseInvoice.tax_amount > 0).scalar() or 0.0
        e_nd_base = db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_before_tax), 0)) \
            .filter(ExpenseInvoice.date.between(start_date, end_date)) \
            .filter(ExpenseInvoice.tax_amount == 0).scalar() or 0.0

        purchases_deductible_base = float(p_ded_base or 0.0) + float(e_ded_base or 0.0)
        purchases_deductible_vat = float(p_ded_vat or 0.0) + float(e_ded_vat or 0.0)
        purchases_deductible_total = purchases_deductible_base + purchases_deductible_vat

        purchases_non_deductible_base = float(p_nd_base or 0.0) + float(e_nd_base or 0.0)
        purchases_non_deductible_vat = 0.0
        purchases_non_deductible_total = purchases_non_deductible_base
    except Exception:
        purchases_deductible_base = purchases_deductible_vat = purchases_deductible_total = 0.0
        purchases_non_deductible_base = purchases_non_deductible_vat = purchases_non_deductible_total = 0.0

    # Summary
    output_vat = float(sales_standard_vat or 0.0)
    # Include deductible expenses VAT in input VAT
    input_vat = float(purchases_deductible_vat or 0.0)
    net_vat = output_vat - input_vat

    data = {
        'period': period,
        'year': y,
        'quarter': q,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'branch': branch,
        # Sales breakdown
        'sales_standard_base': float(sales_standard_base or 0.0),
        'sales_standard_vat': float(sales_standard_vat or 0.0),
        'sales_standard_total': float(sales_standard_total or 0.0),
        'sales_zero_base': float(sales_zero_base or 0.0),
        'sales_zero_vat': float(sales_zero_vat or 0.0),
        'sales_zero_total': float(sales_zero_total or 0.0),
        'sales_exempt_base': float(sales_exempt_base or 0.0),
        'sales_exempt_vat': float(sales_exempt_vat or 0.0),
        'sales_exempt_total': float(sales_exempt_total or 0.0),
        'sales_exports_base': float(sales_exports_base or 0.0),
        'sales_exports_vat': float(sales_exports_vat or 0.0),
        'sales_exports_total': float(sales_exports_total or 0.0),
        # Purchases/Expenses breakdown (combined)
        'purchases_deductible_base': float(purchases_deductible_base or 0.0),
        'purchases_deductible_vat': float(purchases_deductible_vat or 0.0),
        'purchases_deductible_total': float(purchases_deductible_total or 0.0),
        'purchases_non_deductible_base': float(purchases_non_deductible_base or 0.0),
        'purchases_non_deductible_total': float(purchases_non_deductible_total or 0.0),
        # Expenses breakdown (explicit rows)
        'expenses_deductible_base': float(e_ded_base or 0.0),
        'expenses_deductible_vat': float(e_ded_vat or 0.0),
        'expenses_non_deductible_base': float(e_nd_base or 0.0),
        # Summary
        'output_vat': float(output_vat or 0.0),
        'input_vat': float(input_vat or 0.0),
        'net_vat': float(net_vat or 0.0),
        # Header info
        'company_name': (getattr(s, 'company_name', '') or ''),
        'tax_number': (getattr(s, 'tax_number', '') or ''),
        'vat_rate': vat_rate,
    }

    return render_template('vat/vat_dashboard.html', data=data)

@vat.route('/print', methods=['GET'])
@login_required
def vat_print():
    # Compute quarter boundaries
    from datetime import date as _date
    try:
        year = int(request.args.get('year') or 0) or _date.today().year
    except Exception:
        year = _date.today().year
    try:
        quarter = int(request.args.get('quarter') or 0) or ((_date.today().month - 1)//3 + 1)
    except Exception:
        quarter = ((_date.today().month - 1)//3 + 1)

    if quarter not in (1,2,3,4):
        quarter = 1
    start_month = {1:1, 2:4, 3:7, 4:10}[quarter]
    end_month = {1:3, 2:6, 3:9, 4:12}[quarter]
    start_date = _date(year, start_month, 1)
    next_month_first = _date(year + (1 if end_month == 12 else 0), 1 if end_month == 12 else end_month + 1, 1)
    end_date = next_month_first
    from datetime import timedelta as _td
    end_date = end_date - _td(days=1)

    # Aggregate totals safely across possible model variants
    sales_place_india = 0
    sales_china_town = 0

    try:
        if hasattr(SalesInvoice, 'total_amount') and hasattr(SalesInvoice, 'created_at') and hasattr(SalesInvoice, 'branch_code'):
            # Simplified POS model (app.models)
            sales_place_india = db.session.query(func.coalesce(func.sum(SalesInvoice.total_amount), 0)) \
                .filter(SalesInvoice.branch_code == 'place_india') \
                .filter(SalesInvoice.created_at.between(start_date, end_date)).scalar()
            sales_china_town = db.session.query(func.coalesce(func.sum(SalesInvoice.total_amount), 0)) \
                .filter(SalesInvoice.branch_code == 'china_town') \
                .filter(SalesInvoice.created_at.between(start_date, end_date)).scalar()
        elif hasattr(SalesInvoice, 'total_before_tax') and hasattr(SalesInvoice, 'date') and hasattr(SalesInvoice, 'branch'):
            # Rich model (models.py)
            sales_place_india = db.session.query(func.coalesce(func.sum(SalesInvoice.total_before_tax), 0)) \
                .filter(SalesInvoice.branch == 'place_india') \
                .filter(SalesInvoice.date.between(start_date, end_date)).scalar()
            sales_china_town = db.session.query(func.coalesce(func.sum(SalesInvoice.total_before_tax), 0)) \
                .filter(SalesInvoice.branch == 'china_town') \
                .filter(SalesInvoice.date.between(start_date, end_date)).scalar()
    except Exception:
        sales_place_india = 0
        sales_china_town = 0

    # Branch filter for sales total
    branch = (request.args.get('branch') or 'all').strip()
    if branch == 'place_india':
        sales_total = float(sales_place_india or 0)
    elif branch == 'china_town':
        sales_total = float(sales_china_town or 0)
    else:
        sales_total = float(sales_place_india or 0) + float(sales_china_town or 0)

    # Aggregate purchases and expenses within the period
    try:
        purchases_total = float(db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.total_price), 0))
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id)
            .filter(PurchaseInvoice.date.between(start_date, end_date)).scalar() or 0.0)
    except Exception:
        purchases_total = 0.0
    try:
        expenses_total = float(db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_after_tax_discount), 0))
            .filter(ExpenseInvoice.date.between(start_date, end_date)).scalar() or 0.0)
    except Exception:
        expenses_total = 0.0

    s = None
    try:
        s = Settings.query.first()
    except Exception:
        s = None
    vat_rate = float(getattr(s, 'vat_rate', 15) or 15)/100.0

    try:
        output_vat = float(db.session.query(func.coalesce(func.sum(SalesInvoice.tax_amount), 0))
            .filter(SalesInvoice.date.between(start_date, end_date))
            .scalar() or 0.0)
    except Exception:
        output_vat = float(sales_total or 0) * vat_rate
    try:
        input_vat = float(db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.tax), 0))
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id)
            .filter(PurchaseInvoice.date.between(start_date, end_date))
            .scalar() or 0.0) + float(db.session.query(func.coalesce(func.sum(ExpenseInvoice.tax_amount), 0))
            .filter(ExpenseInvoice.date.between(start_date, end_date))
            .scalar() or 0.0)
    except Exception:
        input_vat = float((purchases_total or 0) + (expenses_total or 0)) * vat_rate
    net_vat = output_vat - input_vat

    # Render unified print report (support CSV export)
    report_title = f"VAT Report — Q{quarter} {year}"
    if branch and branch != 'all':
        report_title += f" — {branch}"
    columns = ['Metric', 'Amount']
    data_rows = [
        {'Metric': 'Sales (Net Base)', 'Amount': float(sales_total or 0.0)},
        {'Metric': 'Purchases', 'Amount': float(purchases_total or 0.0)},
        {'Metric': 'Expenses', 'Amount': float(expenses_total or 0.0)},
        {'Metric': 'Output VAT', 'Amount': float(output_vat or 0.0)},
        {'Metric': 'Input VAT', 'Amount': float(input_vat or 0.0)},
        {'Metric': 'Net VAT', 'Amount': float(net_vat or 0.0)},
    ]
    totals = {'Amount': float(net_vat or 0.0)}
    fmt = (request.args.get('format') or '').strip().lower()
    if fmt == 'csv':
        try:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Metric','Amount'])
            for r in data_rows:
                writer.writerow([r.get('Metric',''), r.get('Amount',0.0)])
            from flask import Response
            return Response(output.getvalue(), mimetype='text/csv',
                            headers={'Content-Disposition': f'attachment; filename="vat_q{quarter}_{year}.csv"'})
        except Exception:
            pass
    return render_template('print_report.html', report_title=report_title, settings=s,
                           generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
                           start_date=start_date.isoformat(), end_date=end_date.isoformat(),
                           payment_method=None, branch=branch or 'all',
                           columns=columns, data=data_rows, totals=totals, totals_columns=['Amount'],
                           totals_colspan=1, payment_totals=None)

# ---------- Financials blueprint ----------
financials = Blueprint('financials', __name__, url_prefix='/financials')

@financials.route('/income-statement', endpoint='income_statement')
@login_required
def income_statement():
    period = (request.args.get('period') or 'today').strip()
    today = get_saudi_now().date()
    if period == 'today':
        start_date = end_date = today
    elif period == 'this_week':
        start_date = today - timedelta(days=today.isoweekday() - 1)
        end_date = start_date + timedelta(days=6)
    elif period == 'this_month':
        start_date = date(today.year, today.month, 1)
        nm = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)
        end_date = nm - timedelta(days=1)
    elif period == 'this_year':
        start_date = date(today.year, 1, 1)
        end_date = date(today.year, 12, 31)
    else:
        sd = request.args.get('start_date')
        ed = request.args.get('end_date')
        try:
            start_date = datetime.strptime(sd, '%Y-%m-%d').date() if sd else today
            end_date = datetime.strptime(ed, '%Y-%m-%d').date() if ed else today
        except Exception:
            start_date = end_date = today
        period = 'custom'

    # Branch filter (optional)
    branch = (request.args.get('branch') or 'all').strip()

    # Revenue from Sales (net of invoice discount; excludes VAT)
    revenue = 0.0
    try:
        q = db.session.query(SalesInvoice, SalesInvoiceItem).join(
            SalesInvoiceItem, SalesInvoiceItem.invoice_id == SalesInvoice.id
        )
        if hasattr(SalesInvoice, 'created_at'):
            q = q.filter(SalesInvoice.created_at.between(start_date, end_date))
        elif hasattr(SalesInvoice, 'date'):
            q = q.filter(SalesInvoice.date.between(start_date, end_date))
        if branch and branch != 'all':
            if hasattr(SalesInvoice, 'branch_code'):
                q = q.filter(SalesInvoice.branch_code == branch)
            elif hasattr(SalesInvoice, 'branch'):
                q = q.filter(SalesInvoice.branch == branch)
        for inv, it in q.all():
            if hasattr(it, 'unit_price') and hasattr(it, 'qty'):
                line = float((it.unit_price or 0.0) * (it.qty or 0.0))
            elif hasattr(it, 'price_before_tax') and hasattr(it, 'quantity'):
                line = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            else:
                line = 0.0
            disc_pct = float(getattr(inv, 'discount_pct', 0.0) or 0.0)
            revenue += line * (1.0 - disc_pct/100.0)
    except Exception:
        revenue = 0.0

    # COGS approximated by purchase costs in period (before tax)
    cogs = 0.0
    try:
        q_p = db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.price_before_tax * PurchaseInvoiceItem.quantity), 0)) \
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id) \
            .filter(PurchaseInvoice.date.between(start_date, end_date))
        cogs = float(q_p.scalar() or 0.0)
    except Exception:
        cogs = 0.0

    # Operating expenses from Expense invoices (after tax and discount)
    operating_expenses = 0.0
    try:
        q_e = db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_after_tax_discount), 0)) \
            .filter(ExpenseInvoice.date.between(start_date, end_date))
        operating_expenses = float(q_e.scalar() or 0.0)
    except Exception:
        operating_expenses = 0.0

    gross_profit = max(revenue - cogs, 0.0)
    operating_profit = gross_profit - operating_expenses
    other_income = 0.0
    other_expenses = 0.0
    net_profit_before_tax = operating_profit + other_income - other_expenses
    tax = 0.0
    net_profit_after_tax = net_profit_before_tax - tax

    try:
        vat_out = float(db.session.query(func.coalesce(func.sum(LedgerEntry.credit - LedgerEntry.debit), 0))
            .join(Account, LedgerEntry.account_id == Account.id)
            .filter(Account.code == '6020')
            .filter(LedgerEntry.date.between(start_date, end_date)).scalar() or 0.0)
    except Exception:
        vat_out = 0.0
    try:
        vat_in = float(db.session.query(func.coalesce(func.sum(LedgerEntry.debit - LedgerEntry.credit), 0))
            .join(Account, LedgerEntry.account_id == Account.id)
            .filter(Account.code.in_(['6010','1300']))
            .filter(LedgerEntry.date.between(start_date, end_date)).scalar() or 0.0)
    except Exception:
        vat_in = 0.0
    try:
        rev_pi = float(db.session.query(func.coalesce(func.sum(LedgerEntry.credit - LedgerEntry.debit), 0))
            .join(Account, LedgerEntry.account_id == Account.id)
            .filter(Account.code == '4010')
            .filter(LedgerEntry.date.between(start_date, end_date)).scalar() or 0.0)
    except Exception:
        rev_pi = 0.0
    try:
        rev_ct = float(db.session.query(func.coalesce(func.sum(LedgerEntry.credit - LedgerEntry.debit), 0))
            .join(Account, LedgerEntry.account_id == Account.id)
            .filter(Account.code == '4020')
            .filter(LedgerEntry.date.between(start_date, end_date)).scalar() or 0.0)
    except Exception:
        rev_ct = 0.0

    pl_types = ['REVENUE','OTHER_INCOME','EXPENSE','OTHER_EXPENSE','COGS','TAX']
    type_totals = {}
    for t in pl_types:
        try:
            if t in ['REVENUE','OTHER_INCOME']:
                amt = float(db.session.query(func.coalesce(func.sum(LedgerEntry.credit - LedgerEntry.debit), 0))
                    .join(Account, LedgerEntry.account_id == Account.id)
                    .filter(Account.type == t)
                    .filter(LedgerEntry.date.between(start_date, end_date)).scalar() or 0.0)
            else:
                amt = float(db.session.query(func.coalesce(func.sum(LedgerEntry.debit - LedgerEntry.credit), 0))
                    .join(Account, LedgerEntry.account_id == Account.id)
                    .filter(Account.type == t)
                    .filter(LedgerEntry.date.between(start_date, end_date)).scalar() or 0.0)
        except Exception:
            amt = 0.0
        type_totals[t] = amt


    data = {
        'period': period,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'branch': branch,
        'revenue': float(revenue or 0.0),
        'cogs': float(cogs or 0.0),
        'gross_profit': float(gross_profit or 0.0),
        'operating_expenses': float(operating_expenses or 0.0),
        'operating_profit': float(operating_profit or 0.0),
        'other_income': float(other_income or 0.0),
        'other_expenses': float(other_expenses or 0.0),
        'net_profit_before_tax': float(net_profit_before_tax or 0.0),
        'tax': float(tax or 0.0),
        'zakat': 0.0,
        'income_tax': 0.0,
        'net_profit_after_tax': float(net_profit_after_tax or 0.0),
        'vat_out': float(vat_out or 0.0),
        'vat_in': float(vat_in or 0.0),
        'revenue_by_branch': {'Place India': float(rev_pi or 0.0), 'China Town': float(rev_ct or 0.0)},
        'type_totals': type_totals,
    }
    return render_template('financials/income_statement.html', data=data)


@financials.route('/print/income_statement', methods=['GET'], endpoint='print_income_statement')
@login_required
def print_income_statement():
    # Period params
    period = (request.args.get('period') or 'today').strip()
    today = get_saudi_now().date()
    if period == 'today':
        start_date = end_date = today
    elif period == 'this_week':
        start_date = today - timedelta(days=today.isoweekday() - 1)
        end_date = start_date + timedelta(days=6)
    elif period == 'this_month':
        start_date = date(today.year, today.month, 1)
        nm = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)
        end_date = nm - timedelta(days=1)
    elif period == 'this_year':
        start_date = date(today.year, 1, 1)
        end_date = date(today.year, 12, 31)
    else:
        sd = request.args.get('start_date')
        ed = request.args.get('end_date')
        try:
            start_date = datetime.strptime(sd, '%Y-%m-%d').date() if sd else today
            end_date = datetime.strptime(ed, '%Y-%m-%d').date() if ed else today
        except Exception:
            start_date = end_date = today
        period = 'custom'

    branch = (request.args.get('branch') or 'all').strip()

    # Compute revenue
    revenue = 0.0
    try:
        q = db.session.query(SalesInvoice, SalesInvoiceItem).join(
            SalesInvoiceItem, SalesInvoiceItem.invoice_id == SalesInvoice.id
        )
        if hasattr(SalesInvoice, 'created_at'):
            q = q.filter(SalesInvoice.created_at.between(start_date, end_date))
        elif hasattr(SalesInvoice, 'date'):
            q = q.filter(SalesInvoice.date.between(start_date, end_date))
        if branch and branch != 'all':
            if hasattr(SalesInvoice, 'branch_code'):
                q = q.filter(SalesInvoice.branch_code == branch)
            elif hasattr(SalesInvoice, 'branch'):
                q = q.filter(SalesInvoice.branch == branch)
        for inv, it in q.all():
            if hasattr(it, 'unit_price') and hasattr(it, 'qty'):
                line = float((it.unit_price or 0.0) * (it.qty or 0.0))
            elif hasattr(it, 'price_before_tax') and hasattr(it, 'quantity'):
                line = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            else:
                line = 0.0
            disc_pct = float(getattr(inv, 'discount_pct', 0.0) or 0.0)
            revenue += line * (1.0 - disc_pct/100.0)
    except Exception:
        revenue = 0.0

    # COGS
    cogs = 0.0
    try:
        q_p = db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.price_before_tax * PurchaseInvoiceItem.quantity), 0)) \
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id) \
            .filter(PurchaseInvoice.date.between(start_date, end_date))
        cogs = float(q_p.scalar() or 0.0)
    except Exception:
        cogs = 0.0

    # Operating expenses
    operating_expenses = 0.0
    try:
        q_e = db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_after_tax_discount), 0)) \
            .filter(ExpenseInvoice.date.between(start_date, end_date))
        operating_expenses = float(q_e.scalar() or 0.0)
    except Exception:
        operating_expenses = 0.0

    gross_profit = max(revenue - cogs, 0.0)
    operating_profit = gross_profit - operating_expenses
    other_income = 0.0
    other_expenses = 0.0
    net_profit_before_tax = operating_profit + other_income - other_expenses
    tax = 0.0
    net_profit_after_tax = net_profit_before_tax - tax

    # Settings for header
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None

    # Prepare print data
    report_title = 'Income Statement'
    if branch and branch != 'all':
        report_title += f' — {branch}'
    columns = ['Item', 'Amount']
    rows = [
        {'Item': 'Revenue', 'Amount': float(revenue or 0.0)},
        {'Item': 'COGS', 'Amount': float(cogs or 0.0)},
        {'Item': 'Gross Profit', 'Amount': float(gross_profit or 0.0)},
        {'Item': 'Operating Expenses', 'Amount': float(operating_expenses or 0.0)},
        {'Item': 'Operating Profit', 'Amount': float(operating_profit or 0.0)},
        {'Item': 'Other Income', 'Amount': float(other_income or 0.0)},
        {'Item': 'Other Expenses', 'Amount': float(other_expenses or 0.0)},
        {'Item': 'Net Profit Before Tax', 'Amount': float(net_profit_before_tax or 0.0)},
        {'Item': 'Tax', 'Amount': float(tax or 0.0)},
        {'Item': 'Net Profit After Tax', 'Amount': float(net_profit_after_tax or 0.0)},
    ]
    totals = {'Amount': float(net_profit_after_tax or 0.0)}

    return render_template('print_report.html', report_title=report_title, settings=settings,
                           generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
                           start_date=start_date.isoformat(), end_date=end_date.isoformat(),
                           payment_method=None, branch=branch or 'all',
                           columns=columns, data=rows, totals=totals, totals_columns=['Amount'],
                           totals_colspan=1, payment_totals=None)



@financials.route('/balance-sheet', endpoint='balance_sheet')
@login_required
def balance_sheet():
    d_str = (request.args.get('date') or get_saudi_now().date().isoformat())
    try:
        asof = datetime.strptime(d_str, '%Y-%m-%d').date()
    except Exception:
        asof = get_saudi_now().date()

    rows = db.session.query(
        Account.code.label('code'),
        Account.name.label('name'),
        Account.type.label('type'),
        func.coalesce(func.sum(LedgerEntry.debit), 0).label('debit'),
        func.coalesce(func.sum(LedgerEntry.credit), 0).label('credit'),
    ).outerjoin(LedgerEntry, LedgerEntry.account_id == Account.id) \
     .filter((LedgerEntry.date <= asof) | (LedgerEntry.id == None)) \
     .group_by(Account.id) \
     .order_by(Account.type.asc(), Account.code.asc()).all()

    ca_codes = {'1010','1020','1110','1210','1220','1310','1320'}
    cl_codes = {'2010','2020','2030','2040','2050'}
    current_assets = 0.0
    noncurrent_assets = 0.0
    current_liabilities = 0.0
    noncurrent_liabilities = 0.0
    asset_rows_detail = []
    liability_rows_detail = []
    equity_rows_detail = []
    for r in rows:
        if r.type == 'ASSET':
            bal = float(r.debit or 0) - float(r.credit or 0)
            if (r.code or '') in ca_codes:
                current_assets += bal
            else:
                noncurrent_assets += bal
            asset_rows_detail.append({'code': r.code, 'name': r.name, 'balance': bal, 'class': 'Current' if (r.code or '') in ca_codes else 'Non-current'})
        elif r.type == 'LIABILITY':
            bal = float(r.credit or 0) - float(r.debit or 0)
            if (r.code or '') in cl_codes:
                current_liabilities += bal
            else:
                noncurrent_liabilities += bal
            liability_rows_detail.append({'code': r.code, 'name': r.name, 'balance': bal, 'class': 'Current' if (r.code or '') in cl_codes else 'Non-current'})
        elif r.type == 'EQUITY':
            bal = float(r.credit or 0) - float(r.debit or 0)
            equity_rows_detail.append({'code': r.code, 'name': r.name, 'balance': bal})
    assets = current_assets + noncurrent_assets
    liabilities = current_liabilities + noncurrent_liabilities
    equity = assets - liabilities
    type_totals = {}
    for r in rows:
        t = (getattr(r, 'type', None) or '').upper()
        d = float(r.debit or 0)
        c = float(r.credit or 0)
        if t not in type_totals:
            type_totals[t] = {'debit': 0.0, 'credit': 0.0}
        type_totals[t]['debit'] += d
        type_totals[t]['credit'] += c
    data = {
        'date': asof,
        'assets': assets, 'liabilities': liabilities, 'equity': equity,
        'current_assets': current_assets, 'noncurrent_assets': noncurrent_assets,
        'current_liabilities': current_liabilities, 'noncurrent_liabilities': noncurrent_liabilities,
        'asset_rows_detail': asset_rows_detail,
        'liability_rows_detail': liability_rows_detail,
        'equity_rows_detail': equity_rows_detail,
        'type_totals': type_totals
    }
    return render_template('financials/balance_sheet.html', data=data)

@financials.route('/trial-balance', endpoint='trial_balance')
@login_required
def trial_balance():
    d_str = (request.args.get('date') or get_saudi_now().date().isoformat())
    try:
        asof = datetime.strptime(d_str, '%Y-%m-%d').date()
    except Exception:
        asof = get_saudi_now().date()
    rows = db.session.query(
        Account.code.label('code'),
        Account.name.label('name'),
        Account.type.label('type'),
        func.coalesce(func.sum(LedgerEntry.debit), 0).label('debit'),
        func.coalesce(func.sum(LedgerEntry.credit), 0).label('credit'),
    ).outerjoin(LedgerEntry, LedgerEntry.account_id == Account.id) \
     .filter((LedgerEntry.date <= asof) | (LedgerEntry.id == None)) \
     .group_by(Account.id) \
     .order_by(Account.type.asc(), Account.code.asc()).all()
    if not rows:
        rev = float(db.session.query(func.coalesce(func.sum(SalesInvoice.total_before_tax), 0)).filter(SalesInvoice.date <= asof).scalar() or 0.0)
        cgs = float(db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.price_before_tax * PurchaseInvoiceItem.quantity), 0))
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id)
            .filter(PurchaseInvoice.date <= asof).scalar() or 0.0)
        opex = float(db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_after_tax_discount), 0)).filter(ExpenseInvoice.date <= asof).scalar() or 0.0)
        rows = [
            type('R', (), {'code': 'REV', 'name': 'Revenue', 'debit': 0.0, 'credit': rev}),
            type('R', (), {'code': 'COGS', 'name': 'COGS', 'debit': cgs, 'credit': 0.0}),
            type('R', (), {'code': 'EXP', 'name': 'Operating Expenses', 'debit': opex, 'credit': 0.0}),
        ]
    total_debit = float(sum([float(getattr(r, 'debit', 0) or 0) for r in rows]))
    total_credit = float(sum([float(getattr(r, 'credit', 0) or 0) for r in rows]))
    type_totals = {}
    for r in rows:
        t = (getattr(r, 'type', None) or '').upper()
        d = float(getattr(r, 'debit', 0) or 0)
        c = float(getattr(r, 'credit', 0) or 0)
        if t not in type_totals:
            type_totals[t] = {'debit': 0.0, 'credit': 0.0}
        type_totals[t]['debit'] += d
        type_totals[t]['credit'] += c
    data = {
        'date': asof,
        'rows': rows,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'type_totals': type_totals,
    }
    return render_template('financials/trial_balance.html', data=data)


@financials.route('/ledger/backfill', methods=['GET', 'POST'], endpoint='ledger_backfill')
@login_required
def ledger_backfill():
    sd = request.args.get('start_date')
    ed = request.args.get('end_date')
    scope = (request.args.get('scope') or 'all').strip().lower()
    today = get_saudi_now().date()
    try:
        start_date = datetime.strptime(sd, '%Y-%m-%d').date() if sd else date(today.year, 1, 1)
        end_date = datetime.strptime(ed, '%Y-%m-%d').date() if ed else date(today.year, 12, 31)
    except Exception:
        start_date = date(today.year, 1, 1)
        end_date = date(today.year, 12, 31)

    def get_or_create(code, name, type_):
        acc = Account.query.filter_by(code=code).first()
        if not acc:
            acc = Account(code=code, name=name, type=type_)
            db.session.add(acc)
            db.session.flush()
        return acc

    def pm_code(pm):
        p = (pm or 'CASH').strip().upper()
        if p in ('CASH',):
            return '1040'
        if p in ('BANK','CARD','VISA','MASTERCARD','MADA','ONLINE','TRANSFER'):
            return '1050'
        return '1020'

    created = {'sales': 0, 'purchase': 0, 'expense': 0}

    if scope in ('all','sales'):
        try:
            for inv in SalesInvoice.query.filter(SalesInvoice.date.between(start_date, end_date)).all():
                desc_key = f"Sales {inv.invoice_number}"
                exists = db.session.query(LedgerEntry.id).filter(LedgerEntry.description == desc_key).first()
                if exists:
                    continue
                gross = float(inv.total_after_tax_discount or 0.0)
                base = max(0.0, float(inv.total_before_tax or 0.0) - float(inv.discount_amount or 0.0))
                vat = float(inv.tax_amount or 0.0)
                cash_code = pm_code(inv.payment_method)
                cash_acc = get_or_create(cash_code, 'Cash' if cash_code=='1040' else ('Bank' if cash_code=='1050' else 'Accounts Receivable'), 'ASSET')
                rev_code = '4010' if (inv.branch or '').strip() == 'place_india' else ('4000' if (inv.branch or '').strip() == 'china_town' else '4010')
                rev_name = 'مبيعات Place India' if rev_code=='4010' else 'مبيعات China Town'
                rev_acc = get_or_create(rev_code, rev_name, 'REVENUE')
                vat_out_acc = get_or_create('2060', 'Output VAT – ضريبة المخرجات', 'LIABILITY')
                db.session.add(LedgerEntry(date=inv.date, account_id=cash_acc.id, debit=gross, credit=0, description=desc_key))
                if base > 0:
                    db.session.add(LedgerEntry(date=inv.date, account_id=rev_acc.id, debit=0, credit=base, description=desc_key))
                if vat > 0:
                    db.session.add(LedgerEntry(date=inv.date, account_id=vat_out_acc.id, debit=0, credit=vat, description=f"VAT Output {inv.invoice_number}"))
                created['sales'] += 1
            db.session.commit()
        except Exception:
            db.session.rollback()

    if scope in ('all','purchase'):
        try:
            for inv in PurchaseInvoice.query.filter(PurchaseInvoice.date.between(start_date, end_date)).all():
                desc_key = f"Purchase {inv.invoice_number}"
                exists = db.session.query(LedgerEntry.id).filter(LedgerEntry.description == desc_key).first()
                if exists:
                    continue
                base = float(inv.total_before_tax or 0.0)
                vat = float(inv.tax_amount or 0.0)
                gross = float(inv.total_after_tax_discount or 0.0)
                inv_acc = get_or_create('1200', 'Inventory', 'ASSET')
                vat_in_acc = get_or_create('1300', 'VAT Input', 'ASSET')
                ap_acc = get_or_create('2000', 'Accounts Payable', 'LIABILITY')
                db.session.add(LedgerEntry(date=inv.date, account_id=inv_acc.id, debit=base, credit=0, description=desc_key))
                if vat > 0:
                    db.session.add(LedgerEntry(date=inv.date, account_id=vat_in_acc.id, debit=vat, credit=0, description=f"VAT Input {inv.invoice_number}"))
                db.session.add(LedgerEntry(date=inv.date, account_id=ap_acc.id, credit=gross, debit=0, description=f"AP for {inv.invoice_number}"))
                created['purchase'] += 1
            db.session.commit()
        except Exception:
            db.session.rollback()

    if scope in ('all','expense'):
        try:
            for inv in ExpenseInvoice.query.filter(ExpenseInvoice.date.between(start_date, end_date)).all():
                desc_key = f"Expense {inv.invoice_number}"
                exists = db.session.query(LedgerEntry.id).filter(LedgerEntry.description == desc_key).first()
                if exists:
                    continue
                base = max(0.0, float(inv.total_before_tax or 0.0) - float(inv.discount_amount or 0.0))
                vat = float(inv.tax_amount or 0.0)
                gross = float(inv.total_after_tax_discount or 0.0)
                exp_acc = get_or_create('5070', 'مصروفات أخرى', 'EXPENSE')
                vat_in_acc = get_or_create('1300', 'VAT Input', 'ASSET')
                ap_acc = get_or_create('2000', 'Accounts Payable', 'LIABILITY')
                if base > 0:
                    db.session.add(LedgerEntry(date=inv.date, account_id=exp_acc.id, debit=base, credit=0, description=desc_key))
                if vat > 0:
                    db.session.add(LedgerEntry(date=inv.date, account_id=vat_in_acc.id, debit=vat, credit=0, description=f"VAT Input {inv.invoice_number}"))
                db.session.add(LedgerEntry(date=inv.date, account_id=ap_acc.id, credit=gross, debit=0, description=f"AP for {inv.invoice_number}"))
                created['expense'] += 1
            db.session.commit()
        except Exception:
            db.session.rollback()

    return jsonify({'status': 'ok', 'created': created, 'start_date': start_date.isoformat(), 'end_date': end_date.isoformat(), 'scope': scope})


@financials.route('/balance-sheet/print', endpoint='print_balance_sheet')
@login_required
def print_balance_sheet():
    d = (request.args.get('date') or date.today().isoformat())
    # Simple placeholders (until full ledger is implemented)
    assets = 0.0
    liabilities = 0.0
    equity = float(assets - liabilities)
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    columns = ['Metric', 'Amount']
    data_rows = [
        {'Metric': 'Assets', 'Amount': float(assets or 0.0)},
        {'Metric': 'Liabilities', 'Amount': float(liabilities or 0.0)},
        {'Metric': 'Equity', 'Amount': float(equity or 0.0)},
    ]
    totals = {'Amount': float(equity or 0.0)}
    return render_template('print_report.html', report_title=f"Balance Sheet — {d}", settings=settings,
                           generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
                           start_date=d, end_date=d,
                           payment_method=None, branch='all',
                           columns=columns, data=data_rows, totals=totals, totals_columns=['Amount'],
                           totals_colspan=1, payment_totals=None)

@financials.route('/trial-balance/print', endpoint='print_trial_balance')
@login_required
def print_trial_balance():
    d_str = (request.args.get('date') or get_saudi_now().date().isoformat())
    try:
        asof = datetime.strptime(d_str, '%Y-%m-%d').date()
    except Exception:
        asof = get_saudi_now().date()
    rows_q = db.session.query(
        Account.code.label('code'),
        Account.name.label('name'),
        func.coalesce(func.sum(LedgerEntry.debit), 0).label('debit'),
        func.coalesce(func.sum(LedgerEntry.credit), 0).label('credit'),
    ).outerjoin(LedgerEntry, LedgerEntry.account_id == Account.id) \
     .filter((LedgerEntry.date <= asof) | (LedgerEntry.id == None)) \
     .group_by(Account.id) \
     .order_by(Account.code.asc()).all()
    rows = []
    if rows_q:
        for r in rows_q:
            rows.append({'Code': r.code, 'Account': r.name, 'Debit': float(r.debit or 0.0), 'Credit': float(r.credit or 0.0)})
    else:
        rev = float(db.session.query(func.coalesce(func.sum(SalesInvoice.total_before_tax), 0)).filter(SalesInvoice.date <= asof).scalar() or 0.0)
        cgs = float(db.session.query(func.coalesce(func.sum(PurchaseInvoiceItem.price_before_tax * PurchaseInvoiceItem.quantity), 0))
            .join(PurchaseInvoice, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id)
            .filter(PurchaseInvoice.date <= asof).scalar() or 0.0)
        opex = float(db.session.query(func.coalesce(func.sum(ExpenseInvoice.total_after_tax_discount), 0)).filter(ExpenseInvoice.date <= asof).scalar() or 0.0)
        rows = [
            {'Code': 'REV', 'Account': 'Revenue', 'Debit': 0.0, 'Credit': rev},
            {'Code': 'COGS', 'Account': 'COGS', 'Debit': cgs, 'Credit': 0.0},
            {'Code': 'EXP', 'Account': 'Operating Expenses', 'Debit': opex, 'Credit': 0.0},
        ]
    total_debit = float(sum([float(r.get('Debit') or 0.0) for r in rows]))
    total_credit = float(sum([float(r.get('Credit') or 0.0) for r in rows]))
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    columns = ['Code', 'Account', 'Debit', 'Credit']
    totals = {'Debit': total_debit, 'Credit': total_credit}
    return render_template('print_report.html', report_title=f"Trial Balance — {asof}", settings=settings,
                           generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
                           start_date=asof, end_date=asof,
                           payment_method=None, branch='all',
                           columns=columns, data=rows, totals=totals, totals_columns=['Debit','Credit'],
                           totals_colspan=2, payment_totals=None)


# --------- Minimal report APIs to avoid 404s and support UI tables/prints ---------
@main.route('/api/all-invoices', methods=['GET'], endpoint='api_all_invoices')
@login_required
def api_all_invoices():
    try:
        payment_method = (request.args.get('payment_method') or '').strip().lower()
        branch_f = (request.args.get('branch') or '').strip().lower()
        rows = []
        branch_totals = {}
        overall = {'subtotal': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0}

        q = db.session.query(SalesInvoice)
        if payment_method and payment_method != 'all':
            q = q.filter(func.lower(SalesInvoice.payment_method) == payment_method)
        if branch_f and branch_f != 'all':
            q = q.filter(func.lower(SalesInvoice.branch) == branch_f)
        if hasattr(SalesInvoice, 'created_at'):
            q = q.order_by(SalesInvoice.created_at.desc())
        elif hasattr(SalesInvoice, 'date'):
            q = q.order_by(SalesInvoice.date.desc())

        results = q.all()

        for inv in results:
            branch = getattr(inv, 'branch', None) or getattr(inv, 'branch_code', None) or 'unknown'
            if getattr(inv, 'created_at', None):
                date_s = inv.created_at.date().isoformat()
            elif getattr(inv, 'date', None):
                date_s = inv.date.isoformat()
            else:
                date_s = ''
            subtotal = float(getattr(inv, 'total_before_tax', 0.0) or 0.0)
            discount = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            vat = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            total = float(getattr(inv, 'total_after_tax_discount', 0.0) or max(subtotal - discount, 0.0) + vat)
            pm = (inv.payment_method or '').upper()
            rows.append({
                'branch': branch,
                'date': date_s,
                'invoice_number': inv.invoice_number,
                'subtotal': round(subtotal, 2),
                'discount': round(discount, 2),
                'vat': round(vat, 2),
                'total': round(total, 2),
                'payment_method': pm,
            })
            bt = branch_totals.setdefault(branch, {'subtotal': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0})
            bt['subtotal'] += subtotal; bt['discount'] += discount; bt['vat'] += vat; bt['total'] += total
            overall['subtotal'] += subtotal; overall['discount'] += discount; overall['vat'] += vat; overall['total'] += total

        return jsonify({'invoices': rows, 'branch_totals': branch_totals, 'overall_totals': overall})
    except Exception as e:
        try:
            current_app.logger.exception(f"/api/all-invoices failed: {e}")
        except Exception:
            pass
        return jsonify({'invoices': [], 'branch_totals': {}, 'overall_totals': {'subtotal':0,'discount':0,'vat':0,'total':0}, 'note': 'stub', 'error': str(e)}), 200


@main.route('/api/reports/all-purchases', methods=['GET'], endpoint='api_reports_all_purchases')
@login_required
def api_reports_all_purchases():
    try:
        payment_method = (request.args.get('payment_method') or '').strip().lower()
        rows = []
        overall = {'amount': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0}
        q = db.session.query(PurchaseInvoice, PurchaseInvoiceItem).join(
            PurchaseInvoiceItem, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id
        )
        if payment_method and payment_method != 'all':
            q = q.filter(func.lower(PurchaseInvoice.payment_method) == payment_method)
        # ترتيب آمن
        if hasattr(PurchaseInvoice, 'created_at'):
            q = q.order_by(PurchaseInvoice.created_at.desc())
        elif hasattr(PurchaseInvoice, 'date'):
            q = q.order_by(PurchaseInvoice.date.desc())
        q = q.limit(1000)

        results = q.all()
        # اجمع أساس البنود لكل فاتورة لتوزيع الخصم/الضريبة من رأس الفاتورة
        inv_base = {}
        for inv, it in results:
            line_base = float(it.price_before_tax or 0.0) * float(it.quantity or 0.0)
            inv_base[inv.id] = inv_base.get(inv.id, 0.0) + line_base

        for inv, it in results:
            date_s = (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')
            amount = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            base_total = float(inv_base.get(inv.id, 0.0) or 0.0)
            # استخدم خصم وضريبة رأس الفاتورة إن توفّرا، وإلا ارجع لقيم البند
            inv_disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            inv_vat = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            if base_total > 0 and (inv_disc > 0 or inv_vat > 0):
                disc = inv_disc * (amount / base_total)
                vat = inv_vat * (amount / base_total)
            else:
                disc = float(it.discount or 0.0)
                base = max(amount - disc, 0.0)
                vat = float(it.tax or 0.0)
            base_after_disc = max(amount - disc, 0.0)
            total = float(getattr(inv, 'total_after_tax_discount', 0.0) or 0.0)
            if total <= 0:
                total = base_after_disc + vat
            rows.append({
                'date': date_s,
                'invoice_number': inv.invoice_number,
                'item_name': it.raw_material_name,
                'quantity': float(it.quantity or 0.0),
                'price': float(it.price_before_tax or 0.0),
                'amount': amount,
                'discount': round(disc, 2),
                'vat': round(vat, 2),
                'total': round(base_after_disc + vat, 2),
                'payment_method': (inv.payment_method or '').upper(),
            })
            overall['amount'] += amount; overall['discount'] += disc; overall['vat'] += vat; overall['total'] += (base_after_disc + vat)
        return jsonify({'purchases': rows, 'overall_totals': overall})
    except Exception:
        return jsonify({'purchases': [], 'overall_totals': {'amount':0,'discount':0,'vat':0,'total':0}, 'note': 'stub'}), 200


@main.route('/api/all-expenses', methods=['GET'], endpoint='api_all_expenses')
@login_required
def api_all_expenses():
    try:
        payment_method = (request.args.get('payment_method') or '').strip().lower()
        rows = []
        overall = {'amount': 0.0}
        q = ExpenseInvoice.query
        if payment_method and payment_method != 'all':
            q = q.filter(func.lower(ExpenseInvoice.payment_method) == payment_method)
        q = q.order_by(ExpenseInvoice.created_at.desc()).limit(1000)
        for exp in q.all():
            # Build full items list for this expense invoice
            items_list = []
            try:
                for it in ExpenseInvoiceItem.query.filter_by(invoice_id=exp.id).all():
                    qty = float(getattr(it, 'quantity', 0.0) or 0.0)
                    price = float(getattr(it, 'price_before_tax', 0.0) or 0.0)
                    tax = float(getattr(it, 'tax', 0.0) or 0.0)
                    disc = float(getattr(it, 'discount', 0.0) or 0.0)
                    line_total = float(getattr(it, 'total_price', None) or (max(price*qty - disc, 0.0) + tax) or 0.0)
                    items_list.append({
                        'description': getattr(it, 'description', '') or '',
                        'quantity': qty,
                        'price': price,
                        'tax': tax,
                        'discount': disc,
                        'total': line_total,
                    })
            except Exception:
                items_list = []
            rows.append({
                'date': (exp.date.strftime('%Y-%m-%d') if getattr(exp, 'date', None) else ''),
                'expense_number': exp.invoice_number,
                'items': items_list,
                'amount': float(exp.total_after_tax_discount or 0.0),
                'payment_method': (exp.payment_method or '').upper(),
            })
            overall['amount'] += float(exp.total_after_tax_discount or 0.0)
        return jsonify({'expenses': rows, 'overall_totals': overall})
    except Exception:
        return jsonify({'expenses': [], 'overall_totals': {'amount':0}, 'note': 'stub'}), 200


@main.route('/reports/print/all-invoices/sales', methods=['GET'], endpoint='reports_print_all_invoices_sales')
@login_required
def reports_print_all_invoices_sales():
    # Reuse same data generation as api_all_invoices but render print template
    payment_method = (request.args.get('payment_method') or 'all').strip().lower()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    fmt = (request.args.get('format') or '').strip().lower()
    rows = []
    branch_totals = {}
    overall = {'amount': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0}
    try:
        q = db.session.query(SalesInvoice, SalesInvoiceItem).join(
            SalesInvoiceItem, SalesInvoiceItem.invoice_id == SalesInvoice.id
        )
        if payment_method and payment_method != 'all':
            q = q.filter(func.lower(SalesInvoice.payment_method) == payment_method)
        # Optional branch filter (sales only)
        branch = (request.args.get('branch') or 'all').strip().lower()
        if branch and branch != 'all':
            q = q.filter(func.lower(SalesInvoice.branch) == branch)
        # Optional date range filter
        if start_date:
            try:
                if hasattr(SalesInvoice, 'created_at'):
                    q = q.filter(func.date(SalesInvoice.created_at) >= start_date)
                elif hasattr(SalesInvoice, 'date'):
                    q = q.filter(SalesInvoice.date >= start_date)
            except Exception:
                pass
        if end_date:
            try:
                if hasattr(SalesInvoice, 'created_at'):
                    q = q.filter(func.date(SalesInvoice.created_at) <= end_date)
                elif hasattr(SalesInvoice, 'date'):
                    q = q.filter(SalesInvoice.date <= end_date)
            except Exception:
                pass
        if hasattr(SalesInvoice, 'created_at'):
            q = q.order_by(SalesInvoice.created_at.desc())
        elif hasattr(SalesInvoice, 'date'):
            q = q.order_by(SalesInvoice.date.desc())

        results = q.all()
        # اجمع أساس البنود لكل فاتورة لتوزيع الخصم/الضريبة من رأس الفاتورة
        inv_base = {}
        for inv, it in results:
            line_base = float(it.price_before_tax or 0.0) * float(it.quantity or 0.0)
            inv_base[inv.id] = inv_base.get(inv.id, 0.0) + line_base

        for inv, it in results:
            branch = getattr(inv, 'branch', None) or getattr(inv, 'branch_code', None) or 'unknown'
            if getattr(inv, 'created_at', None):
                date_s = inv.created_at.date().isoformat()
            elif getattr(inv, 'date', None):
                date_s = inv.date.isoformat()
            else:
                date_s = ''
            amount = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            base_total = float(inv_base.get(inv.id, 0.0) or 0.0)
            inv_disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            inv_vat  = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            if base_total > 0 and (inv_disc > 0 or inv_vat > 0):
                disc = inv_disc * (amount / base_total)
                vat  = inv_vat  * (amount / base_total)
            else:
                disc = float(getattr(it, 'discount', 0.0) or 0.0)
                base = max(amount - disc, 0.0)
                vat  = float(getattr(it, 'tax', 0.0) or 0.0)
            base_after_disc = max(amount - disc, 0.0)
            total = base_after_disc + vat
            pm = (inv.payment_method or '').upper()
            rows.append({
                'branch': branch,
                'date': date_s,
                'invoice_number': inv.invoice_number,
                'item_name': it.product_name,
                'quantity': float(it.quantity or 0.0),
                'amount': amount,
                'discount': round(disc, 2),
                'vat': round(vat, 2),
                'total': round(total, 2),
                'payment_method': pm,
            })
            bt = branch_totals.setdefault(branch, {'amount': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0})
            bt['amount'] += amount; bt['discount'] += disc; bt['vat'] += vat; bt['total'] += total
            overall['amount'] += amount; overall['discount'] += disc; overall['vat'] += vat; overall['total'] += total
    except Exception:
        pass

    # Settings for header
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None

    # Payment method totals
    payment_totals = {}
    for r in rows:
        pm = (r.get('payment_method') or '').upper()
        payment_totals[pm] = payment_totals.get(pm, 0.0) + float(r.get('total') or 0.0)

    # Aggregate item totals (by quantity) for display
    item_totals = {}
    try:
        for r in rows:
            name = (r.get('item_name') or '').strip()
            qty = float(r.get('quantity') or 0.0)
            if name:
                item_totals[name] = item_totals.get(name, 0.0) + qty
    except Exception:
        item_totals = {}

    grouped_invoices = []
    try:
        inv_map = {}
        for inv, it in results:
            key = getattr(inv, 'invoice_number', None) or str(getattr(inv, 'id', ''))
            if key not in inv_map:
                if getattr(inv, 'created_at', None):
                    date_s = inv.created_at.date().isoformat()
                elif getattr(inv, 'date', None):
                    date_s = inv.date.isoformat()
                else:
                    date_s = ''
                branch = getattr(inv, 'branch', None) or getattr(inv, 'branch_code', None) or 'unknown'
                subtotal = float(inv_base.get(inv.id, 0.0) or 0.0)
                disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
                vat = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
                total = float(getattr(inv, 'total_after_tax_discount', 0.0) or (subtotal - disc + vat))
                inv_map[key] = {
                    'branch': branch,
                    'date': date_s,
                    'invoice_number': getattr(inv, 'invoice_number', key),
                    'payment_method': (inv.payment_method or '').upper(),
                    'subtotal': round(subtotal, 2),
                    'discount': round(disc, 2),
                    'vat': round(vat, 2),
                    'total': round(total, 2),
                    'items': []
                }
            amount = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            base_total = float(inv_base.get(inv.id, 0.0) or 0.0)
            inv_disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            inv_vat  = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            if base_total > 0 and (inv_disc > 0 or inv_vat > 0):
                disc_part = inv_disc * (amount / base_total)
                vat_part  = inv_vat  * (amount / base_total)
            else:
                disc_part = float(getattr(it, 'discount', 0.0) or 0.0)
                vat_part  = float(getattr(it, 'tax', 0.0) or 0.0)
            line_total = max(amount - disc_part, 0.0) + vat_part
            inv_map[key]['items'].append({
                'item_name': it.product_name,
                'quantity': float(it.quantity or 0.0),
                'amount': round(amount, 2),
                'discount': round(disc_part, 2),
                'vat': round(vat_part, 2),
                'total': round(line_total, 2),
            })
        grouped_invoices = list(inv_map.values())
    except Exception:
        grouped_invoices = []

    meta = {
        'title': 'Sales Invoices (All) — Print',
        'payment_method': payment_method or 'all',
        'branch': (branch or 'all'),
        'start_date': request.args.get('start_date') or '',
        'end_date': request.args.get('end_date') or '',
        'generated_at': get_saudi_now().strftime('%Y-%m-%d %H:%M')
    }
    # Use unified print_report.html like purchases/expenses
    columns = ['Branch','Date','Invoice No.','Item','Qty','Amount','Discount','VAT','Total','Payment']
    data = []
    for r in rows:
        data.append({
            'Branch': r.get('branch') or '',
            'Date': r.get('date') or '',
            'Invoice No.': r.get('invoice_number') or '',
            'Item': r.get('item_name') or '',
            'Qty': float(r.get('quantity') or 0.0),
            'Amount': float(r.get('amount') or 0.0),
            'Discount': float(r.get('discount') or 0.0),
            'VAT': float(r.get('vat') or 0.0),
            'Total': float(r.get('total') or 0.0),
            'Payment': r.get('payment_method') or '',
        })
    totals = {
        'Amount': float(overall.get('amount') or 0.0),
        'Discount': float(overall.get('discount') or 0.0),
        'VAT': float(overall.get('vat') or 0.0),
        'Total': float(overall.get('total') or 0.0),
    }
    totals_columns = ['Amount','Discount','VAT','Total']
    totals_colspan = len(columns) - len(totals_columns)
    # CSV export
    if fmt == 'csv':
        try:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for r in data:
                writer.writerow([r.get('Branch',''), r.get('Date',''), r.get('Invoice No.',''), r.get('Item',''),
                                 r.get('Qty',0.0), r.get('Amount',0.0), r.get('Discount',0.0), r.get('VAT',0.0),
                                 r.get('Total',0.0), r.get('Payment','')])
            from flask import Response
            return Response(output.getvalue(), mimetype='text/csv',
                            headers={'Content-Disposition': 'attachment; filename="sales_invoices.csv"'})
        except Exception:
            pass

    return render_template('print_report.html', report_title=meta['title'], settings=settings,
                           generated_at=meta['generated_at'], start_date=meta['start_date'], end_date=meta['end_date'],
                           payment_method=meta['payment_method'], branch=meta['branch'],
                           columns=columns, data=data, totals=totals, totals_columns=totals_columns,
                           totals_colspan=totals_colspan, payment_totals=payment_totals,
                           item_totals=item_totals, grouped_invoices=grouped_invoices)


@main.route('/reports/print/daily-sales', methods=['GET'], endpoint='reports_print_daily_sales')
@login_required
def reports_print_daily_sales():
    pm = (request.args.get('payment_method') or 'all').strip().lower()
    branch = (request.args.get('branch') or 'all').strip().lower()
    fmt = (request.args.get('format') or '').strip().lower()
    now = get_saudi_now()
    from datetime import datetime as _dt, timedelta as _td
    try:
        from models import KSA_TZ
        tz = KSA_TZ
    except Exception:
        tz = getattr(now, 'tzinfo', None)
    anchor = now.date() if now.hour >= 11 else (now - _td(days=1)).date()
    start_dt = tz.localize(_dt(anchor.year, anchor.month, anchor.day, 11, 0, 0)) if hasattr(tz, 'localize') else _dt(anchor.year, anchor.month, anchor.day, 11, 0, 0)
    end_dt = start_dt + _td(hours=14)
    rows = []
    results = []
    inv_base = {}
    branch_totals = {}
    overall = {'amount': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0}
    try:
        q = db.session.query(SalesInvoice, SalesInvoiceItem).join(
            SalesInvoiceItem, SalesInvoiceItem.invoice_id == SalesInvoice.id
        )
        if pm and pm != 'all':
            q = q.filter(func.lower(SalesInvoice.payment_method) == pm)
        if branch and branch != 'all':
            q = q.filter(func.lower(SalesInvoice.branch) == branch)
        if hasattr(SalesInvoice, 'created_at'):
            q = q.filter(SalesInvoice.created_at >= start_dt, SalesInvoice.created_at < end_dt)
            q = q.order_by(SalesInvoice.created_at.asc())
        else:
            q = q.filter(SalesInvoice.date.in_([start_dt.date(), end_dt.date()]))
            q = q.order_by(SalesInvoice.date.asc())
        results = q.all()
        for inv, it in results:
            line_base = float(it.price_before_tax or 0.0) * float(it.quantity or 0.0)
            inv_base[inv.id] = inv_base.get(inv.id, 0.0) + line_base
        for inv, it in results:
            b = getattr(inv, 'branch', None) or getattr(inv, 'branch_code', None) or 'unknown'
            if getattr(inv, 'created_at', None):
                try:
                    base_dt = inv.created_at
                    if getattr(base_dt, 'tzinfo', None):
                        date_s = base_dt.astimezone(tz).strftime('%Y-%m-%d %H:%M')
                    else:
                        date_s = tz.localize(base_dt).strftime('%Y-%m-%d %H:%M')
                except Exception:
                    date_s = get_saudi_now().strftime('%Y-%m-%d %H:%M')
            elif getattr(inv, 'date', None):
                date_s = inv.date.isoformat()
            else:
                date_s = ''
            amount = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            base_total = float(inv_base.get(inv.id, 0.0) or 0.0)
            inv_disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            inv_vat  = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            if base_total > 0 and (inv_disc > 0 or inv_vat > 0):
                disc = inv_disc * (amount / base_total)
                vat  = inv_vat  * (amount / base_total)
            else:
                disc = float(getattr(it, 'discount', 0.0) or 0.0)
                vat  = float(getattr(it, 'tax', 0.0) or 0.0)
            base_after_disc = max(amount - disc, 0.0)
            total = base_after_disc + vat
            pmu = (inv.payment_method or '').upper()
            rows.append({
                'branch': b,
                'date': date_s,
                'invoice_number': inv.invoice_number,
                'item_name': it.product_name,
                'quantity': float(it.quantity or 0.0),
                'amount': amount,
                'discount': round(disc, 2),
                'vat': round(vat, 2),
                'total': round(total, 2),
                'payment_method': pmu,
            })
            bt = branch_totals.setdefault(b, {'amount': 0.0, 'discount': 0.0, 'vat': 0.0, 'total': 0.0})
            bt['amount'] += amount; bt['discount'] += disc; bt['vat'] += vat; bt['total'] += total
            overall['amount'] += amount; overall['discount'] += disc; overall['vat'] += vat; overall['total'] += total
    except Exception:
        pass
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    payment_totals = {}
    for r in rows:
        pmu = (r.get('payment_method') or '').upper()
        payment_totals[pmu] = payment_totals.get(pmu, 0.0) + float(r.get('total') or 0.0)
    item_totals = {}
    try:
        for r in rows:
            name = (r.get('item_name') or '').strip()
            qty = float(r.get('quantity') or 0.0)
            if name:
                item_totals[name] = item_totals.get(name, 0.0) + qty
    except Exception:
        item_totals = {}
    meta = {
        'title': 'Daily Sales — Print',
        'payment_method': pm or 'all',
        'branch': branch or 'all',
        'start_date': start_dt.strftime('%Y-%m-%d %H:%M'),
        'end_date': end_dt.strftime('%Y-%m-%d %H:%M'),
        'generated_at': get_saudi_now().strftime('%Y-%m-%d %H:%M')
    }
    inv_map = {}
    try:
        for inv, it in results:
            key = getattr(inv, 'invoice_number', None) or str(getattr(inv, 'id', ''))
            if key not in inv_map:
                b = getattr(inv, 'branch', None) or getattr(inv, 'branch_code', None) or 'unknown'
                if getattr(inv, 'created_at', None):
                    try:
                        base_dt = inv.created_at
                        if getattr(base_dt, 'tzinfo', None):
                            date_s = base_dt.astimezone(tz).strftime('%Y-%m-%d %H:%M')
                        else:
                            date_s = tz.localize(base_dt).strftime('%Y-%m-%d %H:%M')
                    except Exception:
                        date_s = get_saudi_now().strftime('%Y-%m-%d %H:%M')
                elif getattr(inv, 'date', None):
                    date_s = inv.date.isoformat()
                else:
                    date_s = ''
                subtotal = float(inv_base.get(inv.id, 0.0) or 0.0)
                disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
                vat = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
                total = float(getattr(inv, 'total_after_tax_discount', 0.0) or (subtotal - disc + vat))
                inv_map[key] = {
                    'branch': b,
                    'date': date_s,
                    'invoice_number': getattr(inv, 'invoice_number', key),
                    'payment_method': (inv.payment_method or '').upper(),
                    'subtotal': round(subtotal, 2),
                    'discount': round(disc, 2),
                    'vat': round(vat, 2),
                    'total': round(total, 2),
                    'items': [],
                    'item_count': 0
                }
            amount = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            base_total = float(inv_base.get(inv.id, 0.0) or 0.0)
            inv_disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            inv_vat  = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            if base_total > 0 and (inv_disc > 0 or inv_vat > 0):
                disc_part = inv_disc * (amount / base_total)
                vat_part  = inv_vat  * (amount / base_total)
            else:
                disc_part = float(getattr(it, 'discount', 0.0) or 0.0)
                vat_part  = float(getattr(it, 'tax', 0.0) or 0.0)
            line_total = max(amount - disc_part, 0.0) + vat_part
            inv_map[key]['items'].append({
                'item_name': it.product_name,
                'quantity': float(it.quantity or 0.0),
                'amount': round(amount, 2),
                'discount': round(disc_part, 2),
                'vat': round(vat_part, 2),
                'total': round(line_total, 2),
            })
            inv_map[key]['item_count'] = int(inv_map[key].get('item_count', 0)) + 1
        grouped_invoices = list(inv_map.values())
    except Exception:
        grouped_invoices = []
    if fmt == 'csv':
        try:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Branch','Date','Invoice No.','Items','Payment','Subtotal','Discount','VAT','Total'])
            for inv in grouped_invoices:
                writer.writerow([
                    inv.get('branch',''),
                    inv.get('date',''),
                    inv.get('invoice_number',''),
                    int(inv.get('item_count') or (len(inv.get('items') or []))),
                    inv.get('payment_method',''),
                    inv.get('subtotal',0.0),
                    inv.get('discount',0.0),
                    inv.get('vat',0.0),
                    inv.get('total',0.0)
                ])
            from flask import Response
            fname = f"daily_sales_{anchor.isoformat()}.csv"
            return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename="{fname}"'})
        except Exception:
            pass
    if fmt == 'pdf':
        try:
            import io
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            w, h = A4
            y = h - 20*mm
            p.setFont("Helvetica-Bold", 14)
            p.drawString(20*mm, y, "Daily Sales Report")
            y -= 8*mm
            p.setFont("Helvetica", 10)
            p.drawString(20*mm, y, f"Period: {meta['start_date']} → {meta['end_date']}")
            y -= 6*mm
            p.drawString(20*mm, y, f"Generated: {meta['generated_at']}")
            y -= 10*mm
            p.setFont("Helvetica-Bold", 9)
            headers = ["Branch","Date","Invoice","Items","PM","Subtotal","Discount","VAT","Total"]
            xs = [10*mm, 30*mm, 70*mm, 100*mm, 115*mm, 130*mm, 150*mm, 170*mm, 190*mm]
            for i, hdr in enumerate(headers):
                p.drawString(xs[i], y, hdr)
            y -= 5*mm
            p.setFont("Helvetica", 9)
            for inv in grouped_invoices:
                row = [
                    str(inv.get('branch','')),
                    str(inv.get('date','')),
                    str(inv.get('invoice_number','')),
                    str(int(inv.get('item_count') or (len(inv.get('items') or [])))) ,
                    str(inv.get('payment_method','')),
                    f"{inv.get('subtotal',0.0):.2f}",
                    f"{inv.get('discount',0.0):.2f}",
                    f"{inv.get('vat',0.0):.2f}",
                    f"{inv.get('total',0.0):.2f}",
                ]
                if y < 20*mm:
                    p.showPage(); y = h - 20*mm; p.setFont("Helvetica", 9)
                for i, val in enumerate(row):
                    p.drawString(xs[i], y, val)
                y -= 5*mm
            y -= 6*mm
            p.setFont("Helvetica-Bold", 10)
            p.drawString(130*mm, y, f"Totals: {overall['total']:.2f}")
            p.showPage(); p.save()
            pdf_bytes = buffer.getvalue()
            from flask import Response
            fname = f"daily_sales_{anchor.isoformat()}.pdf"
            return Response(pdf_bytes, mimetype='application/pdf', headers={'Content-Disposition': f'attachment; filename="{fname}"'})
        except Exception:
            pass
    return render_template('print_report.html', report_title=meta['title'], settings=settings,
                           generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'), start_date=meta['start_date'], end_date=meta['end_date'],
                           payment_method=meta['payment_method'], branch=meta['branch'],
                           columns=['Branch','Date','Invoice No.','Item','Qty','Amount','Discount','VAT','Total','Payment'], data=rows,
                           totals={'Amount': overall['amount'], 'Discount': overall['discount'], 'VAT': overall['vat'], 'Total': overall['total']},
                           totals_columns=['Amount','Discount','VAT','Total'], totals_colspan=6,
                           payment_totals=payment_totals, item_totals=item_totals,
                           branch_totals=branch_totals, overall=overall,
                           grouped_invoices=grouped_invoices, summary_mode=True)


@main.route('/reports/print/all-invoices/purchases', methods=['GET'], endpoint='reports_print_all_invoices_purchases')
@login_required
def reports_print_all_invoices_purchases():
    payment_method = (request.args.get('payment_method') or 'all').strip().lower()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    branch = (request.args.get('branch') or 'all').strip().lower()
    fmt = (request.args.get('format') or '').strip().lower()
    rows = []
    totals = {'Amount': 0.0, 'Discount': 0.0, 'VAT': 0.0, 'Total': 0.0}
    payment_totals = {}
    try:
        q = db.session.query(PurchaseInvoice, PurchaseInvoiceItem).join(
            PurchaseInvoiceItem, PurchaseInvoiceItem.invoice_id == PurchaseInvoice.id
        )
        if payment_method and payment_method != 'all':
            q = q.filter(func.lower(PurchaseInvoice.payment_method) == payment_method)
        # Date filtering if fields exist
        if start_date:
            try:
                q = q.filter(PurchaseInvoice.date >= start_date)
            except Exception:
                pass
        if end_date:
            try:
                q = q.filter(PurchaseInvoice.date <= end_date)
            except Exception:
                pass
        if branch and branch != 'all' and hasattr(PurchaseInvoice, 'branch'):
            q = q.filter(func.lower(PurchaseInvoice.branch) == branch)
        if hasattr(PurchaseInvoice, 'created_at'):
            q = q.order_by(PurchaseInvoice.created_at.desc())
        elif hasattr(PurchaseInvoice, 'date'):
            q = q.order_by(PurchaseInvoice.date.desc())
        q = q.limit(2000)

        results = q.all()
        # Aggregate base per invoice for proportional discount/VAT allocation
        inv_base = {}
        for inv, it in results:
            line_base = float(it.price_before_tax or 0.0) * float(it.quantity or 0.0)
            inv_base[inv.id] = inv_base.get(inv.id, 0.0) + line_base

        for inv, it in results:
            d = (inv.date.strftime('%Y-%m-%d') if getattr(inv, 'date', None) else '')
            amount = float((it.price_before_tax or 0.0) * (it.quantity or 0.0))
            base_total = float(inv_base.get(inv.id, 0.0) or 0.0)
            inv_disc = float(getattr(inv, 'discount_amount', 0.0) or 0.0)
            inv_vat = float(getattr(inv, 'tax_amount', 0.0) or 0.0)
            if base_total > 0 and (inv_disc > 0 or inv_vat > 0):
                disc = inv_disc * (amount / base_total)
                vat = inv_vat * (amount / base_total)
            else:
                disc = float(it.discount or 0.0)
                base = max(amount - disc, 0.0)
                vat = float(it.tax or 0.0)
            base_after_disc = max(amount - disc, 0.0)
            line_total = base_after_disc + vat
            pm = (inv.payment_method or '').upper()
            rows.append({
                'Date': d,
                'Invoice No.': inv.invoice_number,
                'Item': it.raw_material_name,
                'Qty': float(it.quantity or 0.0),
                'Amount': amount,
                'Discount': round(disc, 2),
                'VAT': round(vat, 2),
                'Total': round(line_total, 2),
                'Payment': pm,
            })
            totals['Amount'] += amount; totals['Discount'] += disc; totals['VAT'] += vat; totals['Total'] += line_total
            payment_totals[pm] = payment_totals.get(pm, 0.0) + line_total
    except Exception:
        pass

    # Settings & meta
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    meta = {
        'title': 'Purchase Invoices — Print',
        'payment_method': payment_method or 'all',
        'branch': branch or 'all',
        'start_date': start_date,
        'end_date': end_date,
        'generated_at': get_saudi_now().strftime('%Y-%m-%d %H:%M')
    }
    columns = ['Date','Invoice No.','Item','Qty','Amount','Discount','VAT','Total','Payment']
    # CSV export
    if fmt == 'csv':
        try:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for r in rows:
                writer.writerow([r.get('Date',''), r.get('Invoice No.',''), r.get('Item',''), r.get('Qty',0.0),
                                 r.get('Amount',0.0), r.get('Discount',0.0), r.get('VAT',0.0), r.get('Total',0.0),
                                 r.get('Payment','')])
            from flask import Response
            return Response(output.getvalue(), mimetype='text/csv',
                            headers={'Content-Disposition': 'attachment; filename="purchase_invoices.csv"'})
        except Exception:
            pass
    return render_template('print_report.html', report_title=meta['title'], settings=settings,
                           generated_at=meta['generated_at'], start_date=meta['start_date'], end_date=meta['end_date'],
                           payment_method=meta['payment_method'], branch=meta['branch'],
                           columns=columns, data=rows, totals=totals, totals_columns=['Amount','Discount','VAT','Total'],
                           totals_colspan=4, payment_totals=payment_totals)


@main.route('/reports/print/daily-items-summary', methods=['GET'], endpoint='reports_print_daily_items_summary')
@login_required
def reports_print_daily_items_summary():
    """Print a daily summary of items (by quantity) for a given day.
    By default, summarizes from Order Invoices (pre-payment) so printed orders show up.
    Optional source param: source=orders (default) or source=sales.
    """
    from sqlalchemy import func
    source = (request.args.get('source') or 'orders').strip().lower()
    branch_f = (request.args.get('branch') or '').strip()

    # Determine target date (YYYY-MM-DD) - defaults to today
    date_str = (request.args.get('date') or '').strip()
    from datetime import date as _date, datetime as _dt, time as _time
    try:
        target_date = _date.fromisoformat(date_str) if date_str else _date.today()
    except Exception:
        target_date = _date.today()

    # Compute start/end of day (00:00:00 -> 23:59:59)
    start_dt = _dt.combine(target_date, _time.min)
    end_dt = _dt.combine(target_date, _time.max)

    data = []
    try:
        if source == 'orders':
            from models import OrderInvoice
            # Fetch orders within day window and aggregate from JSON items
            q = OrderInvoice.query.filter(
                OrderInvoice.invoice_date >= start_dt,
                OrderInvoice.invoice_date <= end_dt
            )
            if branch_f:
                q = q.filter(OrderInvoice.branch == branch_f)
            orders = q.limit(5000).all()

            totals_map = {}
            for o in orders:
                try:
                    for it in (o.items or []):
                        name = (it.get('name') if isinstance(it, dict) else None) or '-'
                        qty_val = it.get('qty') if isinstance(it, dict) else 0
                        try:
                            qty = float(qty_val or 0)
                        except Exception:
                            qty = 0.0
                        totals_map[name] = totals_map.get(name, 0.0) + qty
                except Exception:
                    continue
            data = [
                {'Item': k, 'Total Qty': float(v)}
                for k, v in sorted(totals_map.items(), key=lambda x: (-x[1], x[0]))
            ]
        else:
            from models import SalesInvoice, SalesInvoiceItem
            # Sales invoices use date (Date) column; filter by equality if Date only
            q = db.session.query(
                SalesInvoiceItem.product_name.label('item_name'),
                func.coalesce(func.sum(SalesInvoiceItem.quantity), 0).label('total_qty')
            ).join(SalesInvoice, SalesInvoiceItem.invoice_id == SalesInvoice.id)
            q = q.filter(SalesInvoice.date == target_date)
            if branch_f and hasattr(SalesInvoice, 'branch'):
                q = q.filter(SalesInvoice.branch == branch_f)
            if hasattr(SalesInvoice, 'status'):
                q = q.filter(func.lower(SalesInvoice.status) == 'paid')
            q = q.group_by(SalesInvoiceItem.product_name).order_by(func.sum(SalesInvoiceItem.quantity).desc())
            rows = q.all()
            data = [
                {'Item': (r.item_name or '-'), 'Total Qty': float(r.total_qty or 0.0)} for r in rows
            ]
    except Exception:
        data = []

    # Settings & meta
    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    meta_title = f"Daily Items Summary — {target_date.isoformat()} ({'orders' if source=='orders' else 'sales'})"
    columns = ['Item', 'Total Qty']

    return render_template(
        'print_report.html',
        report_title=meta_title,
        settings=settings,
        generated_at=get_saudi_now().strftime('%Y-%m-%d %H:%M'),
        start_date=target_date.isoformat(),
        end_date=target_date.isoformat(),
        payment_method=source,
        branch=(branch_f or 'all'),
        columns=columns,
        data=data,
        totals={},
        totals_columns=[],
        totals_colspan=len(columns)
    )

@main.route('/reports/print/all-invoices/expenses', methods=['GET'], endpoint='reports_print_all_invoices_expenses')
@login_required
def reports_print_all_invoices_expenses():
    payment_method = (request.args.get('payment_method') or 'all').strip().lower()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    branch = (request.args.get('branch') or 'all').strip().lower()
    fmt = (request.args.get('format') or '').strip().lower()
    rows = []
    totals = {'Amount': 0.0}
    payment_totals = {}
    try:
        q = ExpenseInvoice.query
        if payment_method and payment_method != 'all':
            q = q.filter(func.lower(ExpenseInvoice.payment_method) == payment_method)
        if start_date:
            try:
                q = q.filter(ExpenseInvoice.date >= start_date)
            except Exception:
                pass
        if end_date:
            try:
                q = q.filter(ExpenseInvoice.date <= end_date)
            except Exception:
                pass
        if branch and branch != 'all' and hasattr(ExpenseInvoice, 'branch'):
            q = q.filter(func.lower(ExpenseInvoice.branch) == branch)
        q = q.order_by(ExpenseInvoice.created_at.desc()).limit(2000)
        for exp in q.all():
            d = (exp.date.strftime('%Y-%m-%d') if getattr(exp, 'date', None) else '')
            pm = (exp.payment_method or '').upper()
            # Print per item
            try:
                items = ExpenseInvoiceItem.query.filter_by(invoice_id=exp.id).all()
            except Exception:
                items = []
            for it in items:
                qty = float(getattr(it, 'quantity', 0.0) or 0.0)
                price = float(getattr(it, 'price_before_tax', 0.0) or 0.0)
                tax = float(getattr(it, 'tax', 0.0) or 0.0)
                disc = float(getattr(it, 'discount', 0.0) or 0.0)
                line_total = float(getattr(it, 'total_price', None) or (max(price*qty - disc, 0.0) + tax) or 0.0)
                rows.append({
                    'Date': d,
                    'Expense No.': exp.invoice_number,
                    'Description': getattr(it, 'description', '') or '',
                    'Qty': qty,
                    'Amount': price * qty,
                    'Tax': tax,
                    'Discount': disc,
                    'Line Total': line_total,
                    'Payment': pm,
                })
                totals['Amount'] += (price * qty)
                totals['Tax'] = float(totals.get('Tax', 0.0) or 0.0) + tax
                totals['Discount'] = float(totals.get('Discount', 0.0) or 0.0) + disc
                totals['Line Total'] = float(totals.get('Line Total', 0.0) or 0.0) + line_total
                payment_totals[pm] = payment_totals.get(pm, 0.0) + line_total
    except Exception:
        pass

    try:
        settings = Settings.query.first()
    except Exception:
        settings = None
    meta = {
        'title': 'Expenses — Print',
        'payment_method': payment_method or 'all',
        'branch': branch or 'all',
        'start_date': start_date,
        'end_date': end_date,
        'generated_at': get_saudi_now().strftime('%Y-%m-%d %H:%M')
    }
    columns = ['Date','Expense No.','Description','Qty','Amount','Tax','Discount','Line Total','Payment']
    totals_columns = ['Amount','Tax','Discount','Line Total']
    totals_colspan = len(columns) - len(totals_columns)
    # CSV export
    if fmt == 'csv':
        try:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for r in rows:
                writer.writerow([r.get('Date',''), r.get('Expense No.',''), r.get('Description',''), r.get('Qty',0.0),
                                 r.get('Amount',0.0), r.get('Tax',0.0), r.get('Discount',0.0), r.get('Line Total',0.0), r.get('Payment','')])
            from flask import Response
            return Response(output.getvalue(), mimetype='text/csv',
                            headers={'Content-Disposition': 'attachment; filename="expense_invoices.csv"'})
        except Exception:
            pass

    return render_template('print_report.html', report_title=meta['title'], settings=settings,
                           generated_at=meta['generated_at'], start_date=meta['start_date'], end_date=meta['end_date'],
                           payment_method=meta['payment_method'], branch=meta['branch'],
                           columns=columns, data=rows, totals=totals, totals_columns=totals_columns,
                           totals_colspan=totals_colspan, payment_totals=payment_totals)


# ================= Salaries/Payroll Print Reports =================
@main.route('/reports/print/payroll', methods=['GET'], endpoint='print_payroll')
@login_required
def print_payroll():
    # Params
    from datetime import date as _date
    try:
        y_from = int(request.args.get('year_from') or 0) or _date.today().year
    except Exception:
        y_from = _date.today().year
    try:
        m_from = int(request.args.get('month_from') or 0) or 1
    except Exception:
        m_from = 1
    try:
        y_to = int(request.args.get('year_to') or 0) or _date.today().year
    except Exception:
        y_to = _date.today().year
    try:
        m_to = int(request.args.get('month_to') or 0) or _date.today().month
    except Exception:
        m_to = _date.today().month
    raw_codes = (request.args.get('employee_codes') or '').strip()
    selected_codes = [c.strip() for c in raw_codes.split(',') if c.strip()] if raw_codes else []

    # Models
    try:
        from models import Employee, EmployeeSalaryDefault, Salary, Payment
    except Exception:
        try:
            from app.models import Employee, EmployeeSalaryDefault, Salary, Payment
        except Exception:
            Employee = None; Salary = None; Payment = None; EmployeeSalaryDefault = None

    if not Employee or not Salary or not Payment:
        flash('Payroll models not available', 'danger')
        return redirect(url_for('main.dashboard'))

    def iter_months(y1, m1, y2, m2):
        y = int(y1); m = int(m1)
        end_key = (int(y2) * 100 + int(m2))
        while (y * 100 + m) <= end_key:
            yield y, m
            if m == 12:
                y += 1; m = 1
            else:
                m += 1

    # Resolve employees
    emp_q = Employee.query
    if selected_codes:
        try:
            emp_q = emp_q.filter(Employee.employee_code.in_(selected_codes))
        except Exception:
            pass
    employees = emp_q.order_by(Employee.full_name.asc()).all()

    # Helper: defaults
    def get_defaults(emp_id):
        try:
            d = EmployeeSalaryDefault.query.filter_by(employee_id=emp_id).first()
            if d:
                return float(d.base_salary or 0.0), float(d.allowances or 0.0), float(d.deductions or 0.0)
        except Exception:
            pass
        return 0.0, 0.0, 0.0

    employees_ctx = []
    grand = {'basic': 0.0, 'allowances': 0.0, 'deductions': 0.0, 'prev_due': 0.0, 'total': 0.0, 'paid': 0.0, 'remaining': 0.0}
    for emp in employees:
        basic_sum = allow_sum = ded_sum = prev_sum = total_sum = paid_sum = 0.0
        months_rows = []
        unpaid_count = 0
        for (yy, mm) in iter_months(y_from, m_from, y_to, m_to):
            s = Salary.query.filter_by(employee_id=emp.id, year=int(yy), month=int(mm)).first()
            if not s:
                b, a, d = get_defaults(emp.id)
                prev = 0.0
                tot = float(b + a - d + prev)
                sal_id = None
            else:
                b = float(s.basic_salary or 0.0)
                a = float(s.allowances or 0.0)
                d = float(s.deductions or 0.0)
                prev = float(s.previous_salary_due or 0.0)
                tot = float(s.total_salary or (b + a - d + prev))
                sal_id = int(s.id)
            # Paid for this month
            if sal_id:
                paid = float(db.session.query(db.func.coalesce(db.func.sum(Payment.amount_paid), 0))
                              .filter(Payment.invoice_type == 'salary', Payment.invoice_id == sal_id).scalar() or 0.0)
            else:
                paid = 0.0
            remaining = max(tot - paid, 0.0)
            status = 'paid' if remaining <= 0.01 and tot > 0 else ('partial' if paid > 0 else 'due')
            if status != 'paid':
                unpaid_count += 1
            months_rows.append({
                'year': yy, 'month': mm,
                'basic': b, 'allowances': a, 'deductions': d,
                'prev_due': prev, 'total': tot, 'paid': paid, 'remaining': remaining, 'status': status
            })
            basic_sum += b; allow_sum += a; ded_sum += d; prev_sum += prev; total_sum += tot; paid_sum += paid

@main.route('/api/vat/categories-map', methods=['GET','POST'], endpoint='api_vat_categories_map')
@login_required
def api_vat_categories_map():
    try:
        default_map = {
            'zero_keywords': [],
            'exempt_keywords': [],
            'exports_keywords': []
        }
        if request.method == 'GET':
            m = kv_get('vat_category_map', default_map) or default_map
            return jsonify({'ok': True, 'map': m})
        data = request.get_json(silent=True) or {}
        out = {
            'zero_keywords': [str(x).strip() for x in (data.get('zero_keywords') or []) if str(x).strip()],
            'exempt_keywords': [str(x).strip() for x in (data.get('exempt_keywords') or []) if str(x).strip()],
            'exports_keywords': [str(x).strip() for x in (data.get('exports_keywords') or []) if str(x).strip()],
        }
        kv_set('vat_category_map', out)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
        emp_row = {
            'name': getattr(emp, 'full_name', '') or getattr(emp, 'employee_code', ''),
            'unpaid_months': unpaid_count,
            'basic_total': basic_sum,
            'allowances_total': allow_sum,
            'deductions_total': ded_sum,
            'prev_due': prev_sum,
            'total': total_sum,
            'paid': paid_sum,
            'remaining': max(total_sum - paid_sum, 0.0),
            'months': months_rows,
        }
        employees_ctx.append(emp_row)
        grand['basic'] += basic_sum; grand['allowances'] += allow_sum; grand['deductions'] += ded_sum
        grand['prev_due'] += prev_sum; grand['total'] += total_sum; grand['paid'] += paid_sum
    grand['remaining'] = max(grand['total'] - grand['paid'], 0.0)

    return render_template('payroll_report.html',
                           employees=employees_ctx,
                           grand=grand,
                           mode=('all' if len(employees_ctx) != 1 else 'single'),
                           start_year=y_from, start_month=m_from,
                           end_year=y_to, end_month=m_to,
                           selected_codes=(','.join(selected_codes) if selected_codes else ''),
                           auto_print=True,
                           close_after_print=False)


@main.route('/reports/print/payroll/selected', methods=['POST'], endpoint='print_selected')
@login_required
def print_selected():
    # Accept codes (employee_code) from text field, split by comma
    raw = (request.form.get('employee_ids') or '').strip()
    y_from = request.form.get('year_from') or request.args.get('year_from') or ''
    m_from = request.form.get('month_from') or request.args.get('month_from') or ''
    y_to = request.form.get('year_to') or request.args.get('year_to') or ''
    m_to = request.form.get('month_to') or request.args.get('month_to') or ''
    qs = {
        'year_from': y_from or '',
        'month_from': m_from or '',
        'year_to': y_to or '',
        'month_to': m_to or '',
    }
    if raw:
        qs['employee_codes'] = ','.join([c.strip() for c in raw.split(',') if c.strip()])
    # Redirect to GET printer with params to avoid resubmission issues
    return redirect(url_for('main.print_payroll', **qs))
def _archive_invoice_pdf(inv):
    import os, io
    try:
        dt = getattr(inv, 'created_at', None) or get_saudi_now()
    except Exception:
        dt = get_saudi_now()
    year = dt.year; month = dt.month; day = dt.day
    quarter_num = ((month - 1) // 3) + 1
    quarter = f"Q{quarter_num}"
    base_dir = os.getenv('ARCHIVE_DIR') or os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Archive')
    rel_dir = os.path.join(str(year), quarter, f"{month:02d}", f"{day:02d}")
    out_dir = os.path.join(base_dir, rel_dir)
    os.makedirs(out_dir, exist_ok=True)
    fname = f"INV-{inv.invoice_number}.pdf"
    fpath = os.path.join(out_dir, fname)
    # Render receipt HTML
    try:
        items = SalesInvoiceItem.query.filter_by(invoice_id=inv.id).all()
        items_ctx = []
        for it in items:
            line = float(getattr(it, 'price_before_tax', 0) or 0) * float(getattr(it, 'quantity', 0) or 0)
            items_ctx.append({'product_name': getattr(it, 'product_name', '') or '', 'quantity': float(getattr(it, 'quantity', 0) or 0), 'total_price': line})
        s = None
        try:
            s = Settings.query.first()
        except Exception:
            s = None
        dt_str = get_saudi_now().strftime('%Y-%m-%d %H:%M:%S')
        html = render_template('print/receipt.html', inv={
            'invoice_number': inv.invoice_number,
            'table_number': inv.table_number,
            'customer_name': inv.customer_name,
            'customer_phone': inv.customer_phone,
            'payment_method': inv.payment_method,
            'status': 'PAID',
            'total_before_tax': float(inv.total_before_tax or 0.0),
            'tax_amount': float(inv.tax_amount or 0.0),
            'discount_amount': float(inv.discount_amount or 0.0),
            'total_after_tax_discount': float(inv.total_after_tax_discount or 0.0),
            'branch': getattr(inv, 'branch', None),
            'branch_code': getattr(inv, 'branch', None),
        }, items=items_ctx, settings=s, branch_name=BRANCH_LABELS.get(getattr(inv,'branch',None) or '', getattr(inv,'branch','')), date_time=dt_str, display_invoice_number=inv.invoice_number)
    except Exception:
        html = '<html><body><h3>Receipt</h3><p>Render failed.</p></body></html>'
    # Convert to PDF
    pdf_bytes = None
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=current_app.root_path).write_pdf()
    except Exception:
        try:
            import pdfkit
            pdf_bytes = pdfkit.from_string(html, False)
        except Exception:
            pdf_bytes = None
    if pdf_bytes:
        with open(fpath, 'wb') as f:
            f.write(pdf_bytes)
    else:
        # Fallback: save HTML for now
        fpath = fpath[:-4] + '.html'
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(html)
    try:
        kv_set(f"pdf_path:sales:{inv.invoice_number}", {'path': fpath, 'saved_at': get_saudi_now().isoformat()})
    except Exception:
        pass

@main.route('/api/archive/list', methods=['GET'], endpoint='api_archive_list')
@login_required
def api_archive_list():
    try:
        if not user_can('archive','view'):
            return jsonify({'ok': False, 'error': 'forbidden'}), 403
        year = request.args.get('year', type=int)
        quarter = (request.args.get('quarter') or '').strip().upper() or None
        month = request.args.get('month', type=int)
        day = request.args.get('day', type=int)
        branch = (request.args.get('branch') or '').strip()
        page = request.args.get('page', type=int) or 1
        page_size = request.args.get('page_size', type=int) or 100
        if page_size > 500:
            page_size = 500
        if page_size < 1:
            page_size = 100
        if (not year) and (month or day or quarter):
            year = get_saudi_now().year
        start_date = None
        end_date = None
        if year and (not quarter and not month and not day):
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31, 23, 59, 59)
        if year and quarter and (not month and not day):
            qmap = {'Q1': (1, 3), 'Q2': (4, 6), 'Q3': (7, 9), 'Q4': (10, 12)}
            rng = qmap.get(quarter)
            if rng:
                start_date = datetime(year, rng[0], 1)
                if rng[1] == 12:
                    end_date = datetime(year, 12, 31, 23, 59, 59)
                else:
                    end_date = datetime(year, rng[1] + 1, 1) - timedelta(seconds=1)
        if year and month and (not day):
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year, 12, 31, 23, 59, 59)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(seconds=1)
        if year and month and day:
            start_date = datetime(year, month, day)
            end_date = start_date + timedelta(days=1) - timedelta(seconds=1)

        q = db.session.query(
            SalesInvoice.invoice_number,
            SalesInvoice.created_at,
            SalesInvoice.date,
            SalesInvoice.payment_method,
            SalesInvoice.total_after_tax_discount,
            SalesInvoice.branch
        ).filter((SalesInvoice.status == 'paid'))
        if branch:
            q = q.filter(SalesInvoice.branch == branch)
        if start_date and end_date:
            q = q.filter(SalesInvoice.date.between(start_date.date(), end_date.date()))
        q = q.order_by(SalesInvoice.date.desc(), SalesInvoice.created_at.desc())
        rows = q.offset((page - 1) * page_size).limit(page_size).all()
        keys = [f"pdf_path:sales:{r[0]}" for r in rows]
        meta_rows = AppKV.query.filter(AppKV.k.in_(keys)).all() if keys else []
        log_keys = [f"print_log:{r[0]}" for r in rows]
        log_rows = AppKV.query.filter(AppKV.k.in_(log_keys)).all() if log_keys else []
        meta_map = {}
        for kv in meta_rows:
            try:
                meta_map[kv.k] = json.loads(kv.v) if kv.v else {}
            except Exception:
                meta_map[kv.k] = {}
        logs_map = {}
        for kv in log_rows:
            try:
                logs_map[kv.k] = json.loads(kv.v) if kv.v else []
            except Exception:
                logs_map[kv.k] = []
        out = []
        for r in rows:
            inv_no, created_at, date_f, pm, total_amt, branch_code = r
            # Use saved print timestamp if available
            meta = meta_map.get(f"pdf_path:sales:{inv_no}", {}) or {}
            logs = logs_map.get(f"print_log:{inv_no}", []) or []
            dt_print = None
            try:
                if logs:
                    try:
                        from datetime import datetime as _dti
                        def _p(x):
                            try:
                                return _dti.fromisoformat((x or {}).get('ts') or '')
                            except Exception:
                                return None
                        cand = [_p(x) for x in logs]
                        cand = [c for c in cand if c]
                        if cand:
                            dt_print = min(cand)
                    except Exception:
                        dt_print = None
                if (dt_print is None):
                    sa = meta.get('saved_at')
                    if sa:
                        try:
                            from datetime import datetime as _dti
                            dt_print = _dti.fromisoformat(sa)
                        except Exception:
                            dt_print = None
            except Exception:
                dt_print = None
            if dt_print is not None:
                dt_date = dt_print.date()
                tstr = dt_print.strftime('%H:%M:%S')
            else:
                try:
                    from models import SalesInvoice as SI, Payment
                    sid = db.session.query(SI.id).filter(SI.invoice_number == inv_no).scalar()
                    dtp_row = None
                    if sid:
                        dtp_row = db.session.query(Payment.payment_date).\
                            filter(Payment.invoice_type == 'sales', Payment.invoice_id == sid).\
                            order_by(Payment.payment_date.asc()).first()
                    if dtp_row and dtp_row[0]:
                        dt_date = dtp_row[0].date()
                        tstr = dtp_row[0].strftime('%H:%M:%S')
                    else:
                        raise Exception('no_payment_date')
                except Exception:
                    dt_date = (date_f or (created_at and created_at.date()) or get_saudi_now().date())
                    try:
                        tstr = created_at.strftime('%H:%M:%S') if created_at else '00:00:00'
                    except Exception:
                        tstr = '00:00:00'
            m = int(dt_date.month)
            qn = ((m - 1) // 3) + 1
            out.append({
                'invoice_number': inv_no,
                'date': dt_date.strftime('%Y-%m-%d'),
                'time': tstr,
                'payment_method': pm,
                'total_amount': float(total_amt or 0.0),
                'branch': branch_code,
                'quarter': f"Q{qn}",
                'pdf_path': meta.get('path')
            })
        if quarter in ('Q1','Q2','Q3','Q4'):
            out = [x for x in out if x['quarter'] == quarter]
        return jsonify({'ok': True, 'items': out, 'page': page, 'page_size': page_size})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/archive/open/<invoice_number>', methods=['GET'], endpoint='archive_open')
@login_required
def archive_open(invoice_number):
    try:
        if not user_can('archive','view'):
            return 'Forbidden', 403
        meta = kv_get(f"pdf_path:sales:{invoice_number}", {}) or {}
        p = meta.get('path')
        if not p or (not os.path.exists(p)):
            return 'File not found', 404
        return send_file(p, as_attachment=False)
    except Exception as e:
        return f'Error: {e}', 500

@main.route('/archive', methods=['GET'], endpoint='archive')
@login_required
def archive_page():
    try:
        if not user_can('archive','view'):
            return 'Forbidden', 403
        return render_template('archive.html', now=get_saudi_now(), branches=BRANCH_LABELS)
    except Exception as e:
        return f'Error loading archive: {e}', 500

@main.route('/archive/download', methods=['GET'], endpoint='archive_download')
@login_required
def archive_download():
    try:
        if not user_can('archive','view'):
            return jsonify({'ok': False, 'error': 'forbidden'}), 403
        import io, zipfile, csv
        year = request.args.get('year', type=int)
        quarter = (request.args.get('quarter') or '').strip().upper() or None
        month = request.args.get('month', type=int)
        day = request.args.get('day', type=int)
        branch = (request.args.get('branch') or '').strip()
        fmt = (request.args.get('fmt') or '').strip().lower() or 'zip'
        if (not year) and (month or day or quarter):
            year = get_saudi_now().year
        start_date = None
        end_date = None
        if year and (not quarter and not month and not day):
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31, 23, 59, 59)
        if year and quarter and (not month and not day):
            qmap = {'Q1': (1, 3), 'Q2': (4, 6), 'Q3': (7, 9), 'Q4': (10, 12)}
            rng = qmap.get(quarter)
            if rng:
                start_date = datetime(year, rng[0], 1)
                if rng[1] == 12:
                    end_date = datetime(year, 12, 31, 23, 59, 59)
                else:
                    end_date = datetime(year, rng[1] + 1, 1) - timedelta(seconds=1)
        if year and month and (not day):
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year, 12, 31, 23, 59, 59)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(seconds=1)
        if year and month and day:
            start_date = datetime(year, month, day)
            end_date = start_date + timedelta(days=1) - timedelta(seconds=1)
        q = db.session.query(
            SalesInvoice.invoice_number,
            SalesInvoice.created_at,
            SalesInvoice.date,
            SalesInvoice.payment_method,
            SalesInvoice.total_after_tax_discount,
            SalesInvoice.branch
        ).filter((SalesInvoice.status == 'paid'))
        if branch:
            q = q.filter(SalesInvoice.branch == branch)
        if start_date and end_date:
            q = q.filter(SalesInvoice.date.between(start_date.date(), end_date.date()))
        rows = q.order_by(SalesInvoice.date.desc(), SalesInvoice.created_at.desc()).limit(2000).all()
        items = []
        keys = [f"pdf_path:sales:{r[0]}" for r in rows]
        meta_rows = AppKV.query.filter(AppKV.k.in_(keys)).all() if keys else []
        log_keys = [f"print_log:{r[0]}" for r in rows]
        log_rows = AppKV.query.filter(AppKV.k.in_(log_keys)).all() if log_keys else []
        meta_map = {}
        for kv in meta_rows:
            try:
                meta_map[kv.k] = json.loads(kv.v) if kv.v else {}
            except Exception:
                meta_map[kv.k] = {}
        logs_map = {}
        for kv in log_rows:
            try:
                logs_map[kv.k] = json.loads(kv.v) if kv.v else []
            except Exception:
                logs_map[kv.k] = []
        for r in rows:
            inv_no, created_at, date_f, pm, total_amt, branch_code = r
            meta = meta_map.get(f"pdf_path:sales:{inv_no}", {}) or {}
            logs = logs_map.get(f"print_log:{inv_no}", []) or []
            dt_print = None
            try:
                if logs:
                    from datetime import datetime as _dti
                    def _p(x):
                        try:
                            return _dti.fromisoformat((x or {}).get('ts') or '')
                        except Exception:
                            return None
                    cand = [_p(x) for x in logs]
                    cand = [c for c in cand if c]
                    if cand:
                        dt_print = min(cand)
                if (dt_print is None):
                    sa = meta.get('saved_at')
                    if sa:
                        from datetime import datetime as _dti
                        try:
                            dt_print = _dti.fromisoformat(sa)
                        except Exception:
                            dt_print = None
            except Exception:
                dt_print = None
            if dt_print is None:
                try:
                    from models import SalesInvoice as SI, Payment
                    sid = db.session.query(SI.id).filter(SI.invoice_number == inv_no).scalar()
                    dtp_row = None
                    if sid:
                        dtp_row = db.session.query(Payment.payment_date).\
                            filter(Payment.invoice_type == 'sales', Payment.invoice_id == sid).\
                            order_by(Payment.payment_date.asc()).first()
                    if dtp_row and dtp_row[0]:
                        dt_date = dtp_row[0].date()
                    else:
                        dt_date = (date_f or (created_at and created_at.date()) or get_saudi_now().date())
                except Exception:
                    dt_date = (date_f or (created_at and created_at.date()) or get_saudi_now().date())
            else:
                dt_date = dt_print.date()
            m = int(dt_date.month)
            qn = ((m - 1) // 3) + 1
            if quarter in ('Q1','Q2','Q3','Q4') and f"Q{qn}" != quarter:
                continue
            p = meta.get('path')
            items.append({
                'invoice_number': inv_no,
                'date': dt_date.strftime('%Y-%m-%d'),
                'payment_method': pm,
                'total_amount': float(total_amt or 0.0),
                'branch': branch_code,
                'quarter': f"Q{qn}",
                'pdf_path': p if (p and os.path.exists(p)) else None
            })
        ts = get_saudi_now().strftime('%Y%m%d-%H%M%S')
        if fmt == 'csv':
            out = io.StringIO()
            w = csv.writer(out)
            w.writerow(['invoice_number','date','payment_method','total_amount','branch','quarter','pdf_path'])
            for it in items:
                w.writerow([it['invoice_number'], it['date'], it['payment_method'] or '', f"{it['total_amount']:.2f}", it['branch'] or '', it['quarter'], it['pdf_path'] or ''])
            data = out.getvalue().encode('utf-8')
            buf = io.BytesIO(data)
            return send_file(buf, mimetype='text/csv', as_attachment=True, download_name=f"invoices-{ts}.csv")
        else:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as z:
                manifest = io.StringIO()
                w = csv.writer(manifest)
                w.writerow(['invoice_number','date','payment_method','total_amount','branch','quarter','pdf_path'])
                for it in items:
                    w.writerow([it['invoice_number'], it['date'], it['payment_method'] or '', f"{it['total_amount']:.2f}", it['branch'] or '', it['quarter'], it['pdf_path'] or ''])
                z.writestr('manifest.csv', manifest.getvalue())
                for it in items:
                    if it['pdf_path']:
                        name = f"{it['invoice_number']}.pdf"
                        try:
                            z.write(it['pdf_path'], arcname=name)
                        except Exception:
                            pass
            buf.seek(0)
            return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=f"invoices-{ts}.zip")
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
def _account(code, name, kind):
    try:
        a = Account.query.filter(func.lower(Account.code) == code.lower()).first()
        if not a:
            a = Account(code=code.upper(), name=name, type=kind)
            db.session.add(a)
            db.session.flush()
        return a
    except Exception:
        return None

def _pm_account(pm):
    p = (pm or 'CASH').strip().upper()
    try:
        from app.routes import SHORT_TO_NUMERIC as _short
    except Exception:
        _short = {}
    if p == 'CASH':
        tgt = _short.get('CASH') or ('1110','Cash','ASSET')
    else:
        tgt = _short.get('BANK') or ('1120','Bank','ASSET')
    return _account(tgt[0], tgt[1], tgt[2])

def _acc_override(name: str, default_code: str) -> str:
    try:
        m = kv_get('acc_map', {}) or {}
        code = (m.get(name) or '').strip()
        if code:
            return code
    except Exception:
        pass
    return default_code

def _platform_group(name: str) -> str:
    try:
        s = (name or '').strip().lower()
        # Configured platforms from settings (AppKV)
        platforms = kv_get('platforms_map', []) or []
        for p in platforms:
            key = (p.get('key') or '').strip().lower()
            kws = p.get('keywords') or []
            for kw in kws:
                k = (kw or '').strip().lower()
                if k and (k in s):
                    return key
        # Built-in fallback for known platforms
        if ('hunger' in s) or ('هنقر' in s) or ('هونقر' in s):
            return 'hunger'
        if ('keeta' in s) or ('كيتا' in s) or ('كيت' in s):
            return 'keeta'
        return ''
    except Exception:
        return ''

def _create_sale_journal(inv):
    try:
        from models import JournalEntry, JournalLine
        base_amt = float(inv.total_before_tax or 0.0) - float(inv.discount_amount or 0.0)
        tax_amt = float(inv.tax_amount or 0.0)
        total_inc_tax = round(max(base_amt, 0.0) + max(tax_amt, 0.0), 2)
        cust = (getattr(inv, 'customer_name', '') or '').strip().lower()
        grp = _platform_group(cust)
        if grp == 'keeta':
            rev_code = _acc_override('REV_KEETA', SHORT_TO_NUMERIC['REV_KEETA'][0])
        elif grp == 'hunger':
            rev_code = _acc_override('REV_HUNGER', SHORT_TO_NUMERIC['REV_HUNGER'][0])
        else:
            if (getattr(inv, 'branch', '') or '') == 'place_india':
                rev_code = _acc_override('REV_PI', SHORT_TO_NUMERIC['REV_PI'][0])
            elif (getattr(inv, 'branch', '') or '') == 'china_town':
                rev_code = _acc_override('REV_CT', SHORT_TO_NUMERIC['REV_CT'][0])
            else:
                rev_code = _acc_override('REV_CT', SHORT_TO_NUMERIC['REV_CT'][0])
        vat_out_code = SHORT_TO_NUMERIC.get('VAT_OUT', ('2060',))[0] if isinstance(SHORT_TO_NUMERIC.get('VAT_OUT'), tuple) else '2060'
        paid = ((getattr(inv,'status','') or '').lower() == 'paid')
        if paid:
            ca = _pm_account(getattr(inv,'payment_method','CASH'))
            je = JournalEntry(entry_number=f"JE-SAL-{inv.invoice_number}", date=(getattr(inv,'date',None) or get_saudi_now().date()), branch_code=getattr(inv,'branch',None), description=f"Sales {inv.invoice_number}", status='posted', total_debit=total_inc_tax, total_credit=total_inc_tax, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
            db.session.add(je); db.session.flush()
            if ca:
                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=ca.id, debit=total_inc_tax, credit=0.0, description=f"Receipt {inv.invoice_number}", line_date=(getattr(inv,'date',None) or get_saudi_now().date())))
        else:
            ar_code = _acc_override('AR', CHART_OF_ACCOUNTS.get('1020', {'code':'1020'}).get('code','1020'))
            from models import Account
            ar_acc = _account(ar_code, CHART_OF_ACCOUNTS.get(ar_code, {'name':'Accounts Receivable','type':'ASSET'}).get('name','Accounts Receivable'), CHART_OF_ACCOUNTS.get(ar_code, {'name':'Accounts Receivable','type':'ASSET'}).get('type','ASSET'))
            je = JournalEntry(entry_number=f"JE-SAL-{inv.invoice_number}", date=(getattr(inv,'date',None) or get_saudi_now().date()), branch_code=getattr(inv,'branch',None), description=f"Sales {inv.invoice_number}", status='posted', total_debit=total_inc_tax, total_credit=total_inc_tax, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
            db.session.add(je); db.session.flush()
            if ar_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=ar_acc.id, debit=total_inc_tax, credit=0.0, description=f"AR {inv.invoice_number}", line_date=(getattr(inv,'date',None) or get_saudi_now().date())))
        from models import Account
        rev_acc = _account(rev_code, CHART_OF_ACCOUNTS.get(rev_code, {'name':'Revenue','type':'REVENUE'}).get('name','Revenue'), 'REVENUE')
        vat_acc = _account(vat_out_code, CHART_OF_ACCOUNTS.get(vat_out_code, {'name':'VAT Output','type':'LIABILITY'}).get('name','VAT Output'), 'LIABILITY')
        ln = 2
        if rev_acc and base_amt > 0:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=rev_acc.id, debit=0.0, credit=base_amt, description=f"Revenue {inv.invoice_number}", line_date=(getattr(inv,'date',None) or get_saudi_now().date())));
            ln += 1
        if vat_acc and tax_amt > 0:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=vat_acc.id, debit=0.0, credit=tax_amt, description=f"VAT Output {inv.invoice_number}", line_date=(getattr(inv,'date',None) or get_saudi_now().date())))
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

def _create_purchase_journal(inv):
    try:
        from models import JournalEntry, JournalLine
        total_before = float(inv.total_before_tax or 0.0)
        tax_amt = float(inv.tax_amount or 0.0)
        total_inc_tax = round(total_before + tax_amt, 2)
        exp_acc = _account('1210', CHART_OF_ACCOUNTS.get('1210', {'name':'Inventory','type':'ASSET'}).get('name','Inventory'), CHART_OF_ACCOUNTS.get('1210', {'name':'Inventory','type':'ASSET'}).get('type','ASSET'))
        vat_in_acc = _account('6200', CHART_OF_ACCOUNTS.get('6200', {'name':'VAT Input','type':'ASSET'}).get('name','VAT Input'), CHART_OF_ACCOUNTS.get('6200', {'name':'VAT Input','type':'ASSET'}).get('type','ASSET'))
        ap_acc = _account('2110', CHART_OF_ACCOUNTS.get('2110', {'name':'Accounts Payable','type':'LIABILITY'}).get('name','Accounts Payable'), CHART_OF_ACCOUNTS.get('2110', {'name':'Accounts Payable','type':'LIABILITY'}).get('type','LIABILITY'))
        je = JournalEntry(entry_number=f"JE-PUR-{inv.invoice_number}", date=(getattr(inv,'date',None) or get_saudi_now().date()), branch_code=None, description=f"Purchase {inv.invoice_number}", status='posted', total_debit=total_inc_tax, total_credit=total_inc_tax, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
        db.session.add(je); db.session.flush()
        ln = 1
        if exp_acc and total_before > 0:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=exp_acc.id, debit=total_before, credit=0.0, description="Purchase", line_date=(getattr(inv,'date',None) or get_saudi_now().date()))); ln += 1
        if vat_in_acc and tax_amt > 0:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=vat_in_acc.id, debit=tax_amt, credit=0.0, description="VAT Input", line_date=(getattr(inv,'date',None) or get_saudi_now().date()))); ln += 1
        if ap_acc:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=ap_acc.id, debit=0.0, credit=total_inc_tax, description="Accounts Payable", line_date=(getattr(inv,'date',None) or get_saudi_now().date())))
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

def _create_expense_journal(inv):
    try:
        from models import JournalEntry, JournalLine
        total_before = float(inv.total_before_tax or 0.0)
        tax_amt = float(inv.tax_amount or 0.0)
        total_inc_tax = round(total_before + tax_amt, 2)
        exp_acc = _account('5100', CHART_OF_ACCOUNTS.get('5100', {'name':'رسوم حكومية','type':'EXPENSE'}).get('name','Expense'), 'EXPENSE')
        vat_in_acc = _account('6200', CHART_OF_ACCOUNTS.get('6200', {'name':'VAT Input','type':'ASSET'}).get('name','VAT Input'), 'ASSET')
        ap_acc = _account('2110', CHART_OF_ACCOUNTS.get('2110', {'name':'Accounts Payable','type':'LIABILITY'}).get('name','Accounts Payable'), 'LIABILITY')
        je = JournalEntry(entry_number=f"JE-EXP-{inv.invoice_number}", date=(getattr(inv,'date',None) or get_saudi_now().date()), branch_code=None, description=f"Expense {inv.invoice_number}", status='posted', total_debit=total_inc_tax, total_credit=total_inc_tax, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
        db.session.add(je); db.session.flush()
        ln = 1
        if exp_acc and total_before > 0:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=exp_acc.id, debit=total_before, credit=0.0, description="Expense", line_date=(getattr(inv,'date',None) or get_saudi_now().date()))); ln += 1
        if vat_in_acc and tax_amt > 0:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=vat_in_acc.id, debit=tax_amt, credit=0.0, description="VAT Input", line_date=(getattr(inv,'date',None) or get_saudi_now().date()))); ln += 1
        if ap_acc:
            db.session.add(JournalLine(journal_id=je.id, line_no=ln, account_id=ap_acc.id, debit=0.0, credit=total_inc_tax, description="Accounts Payable", line_date=(getattr(inv,'date',None) or get_saudi_now().date())))
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

def _post_ledger(date_val, acc_code, acc_name, acc_type, debit_amt, credit_amt, ref_text):
    try:
        mc, mn, mt = _resolve_account(acc_code, acc_name, acc_type)
    except Exception:
        mc, mn, mt = acc_code, acc_name, acc_type
    a = _account(mc, mn, mt)
    if not a:
        return
    try:
        le = LedgerEntry(account_id=a.id, date=date_val, description=ref_text, debit=float(debit_amt or 0.0), credit=float(credit_amt or 0.0))
        db.session.add(le)
        db.session.flush()
    except Exception:
        db.session.rollback()

def _expense_account_by_code(code):
    c = (code or '').strip().upper()
    table = {
        'RENT': ('RENT','Rent','EXPENSE'),
        'MAINT': ('MAINT','Maintenance','EXPENSE'),
        'UTIL': ('UTIL','Utilities','EXPENSE'),
        'LOG': ('LOG','Logistics','EXPENSE'),
        'MKT': ('MKT','Marketing','EXPENSE'),
        'TEL': ('TEL','Telecom & Internet','EXPENSE'),
        'STAT': ('STAT','Stationery','EXPENSE'),
        'CLEAN': ('CLEAN','Cleaning','EXPENSE'),
        'GOV': ('GOV','Government Payments','EXPENSE'),
        'EXP': ('EXP','Operating Expenses','EXPENSE'),
    }
    return table.get(c)

def _expense_account_for(desc):
    s = (desc or '').strip().lower()
    mapping = [
        (('rent','ايجار','إيجار'), ('RENT','Rent','EXPENSE')),
        (('maintenance','صيانة'), ('MAINT','Maintenance','EXPENSE')),
        (('electric','كهرب','كهرباء','water','ماء','utilities','فاتورة'), ('UTIL','Utilities','EXPENSE')),
        (('delivery','نقل','شحن','logistics','لوجست'), ('LOG','Logistics','EXPENSE')),
        (('marketing','تسويق','اعلان','إعلان','دعاية'), ('MKT','Marketing','EXPENSE')),
        (('internet','انترنت','شبكة','اتصالات','هاتف','phone'), ('TEL','Telecom & Internet','EXPENSE')),
        (('stationery','قرطاس','قرطاسية','مطبوعات'), ('STAT','Stationery','EXPENSE')),
        (('clean','نظافة'), ('CLEAN','Cleaning','EXPENSE')),
    ]
    for kws, acc in mapping:
        for kw in kws:
            if kw in s:
                return acc
    return ('EXP','Operating Expenses','EXPENSE')

# ---- Chart of Accounts (Clean Restaurant Chart 2025) ----
CHART_OF_ACCOUNTS = {
    # 1. ASSETS – الأصول
    '1010': {'name': 'النقدية العامة', 'type': 'ASSET'},
    '1040': {'name': 'نقد وصندوق (الصندوق العام)', 'type': 'ASSET'},
    '1110': {'name': 'صندوق نقطة البيع (POS Cash Drawer)', 'type': 'ASSET'},
    '1050': {'name': 'بنك – حساب رئيسي', 'type': 'ASSET'},
    '1120': {'name': 'بنك – حساب إضافي', 'type': 'ASSET'},
    '1020': {'name': 'المدينون / Accounts Receivable', 'type': 'ASSET'},
    '1130': {'name': 'حسابات التحصيل – Keeta', 'type': 'ASSET'},
    '1140': {'name': 'حسابات التحصيل – HungerStation', 'type': 'ASSET'},
    '1070': {'name': 'مخزون مواد غذائية', 'type': 'ASSET'},
    '1210': {'name': 'مخزون – مواد وأدوات تشغيل', 'type': 'ASSET'},
    '1025': {'name': 'مخزون – مستلزمات أخرى', 'type': 'ASSET'},
    '1220': {'name': 'هدر مواد – مراقب', 'type': 'ASSET'},
    '1030': {'name': 'سلف الموظفين المستحقة', 'type': 'ASSET'},
    '1090': {'name': 'مصروفات مدفوعة مسبقاً', 'type': 'ASSET'},
    '1310': {'name': 'إيجار مدفوع مقدماً', 'type': 'ASSET'},
    '1320': {'name': 'تأمين مدفوع مقدماً', 'type': 'ASSET'},
    '1060': {'name': 'أصول ثابتة (معدات/أثاث/أجهزة)', 'type': 'ASSET'},
    '1510': {'name': 'معدات وتجهيزات المطعم', 'type': 'ASSET'},
    '1520': {'name': 'أثاث وديكور', 'type': 'ASSET'},
    '1530': {'name': 'أجهزة كمبيوتر وبرمجيات', 'type': 'ASSET'},
    '1540': {'name': 'قيمة شراء المطعم من المالك السابق', 'type': 'ASSET'},

    # 2. LIABILITIES – الالتزامات
    '2000': {'name': 'الالتزامات العامة', 'type': 'LIABILITY'},
    '2020': {'name': 'التزامات أخرى', 'type': 'LIABILITY'},
    '2010': {'name': 'الموردون', 'type': 'LIABILITY'},
    '2110': {'name': 'الموردون – عام', 'type': 'LIABILITY'},
    '2030': {'name': 'الرواتب المستحقة', 'type': 'LIABILITY'},
    '2130': {'name': 'رواتب مستحقة', 'type': 'LIABILITY'},
    '2040': {'name': 'سلف الموظفين', 'type': 'LIABILITY'},
    '2050': {'name': 'VAT المستحق', 'type': 'LIABILITY'},
    '2060': {'name': 'ضريبة المخرجات (Output VAT)', 'type': 'LIABILITY'},
    '2120': {'name': 'ضريبة القيمة المضافة مستحقة', 'type': 'LIABILITY'},
    '2200': {'name': 'الالتزامات طويلة الأجل', 'type': 'LIABILITY'},

    # 3. EQUITY – حقوق الملكية
    '3000': {'name': 'حقوق الملكية', 'type': 'EQUITY'},
    '3010': {'name': 'رأس المال عند التأسيس', 'type': 'EQUITY'},
    '3100': {'name': 'رأس المال', 'type': 'EQUITY'},
    '3020': {'name': 'المسحوبات الشخصية', 'type': 'EQUITY'},
    '3200': {'name': 'الأرباح/الخسائر المحتجزة', 'type': 'EQUITY'},
    '3030': {'name': 'صافي الربح/الخسارة', 'type': 'EQUITY'},

    # 4. REVENUE – الإيرادات
    '4000': {'name': 'مبيعات China Town – رئيسي', 'type': 'REVENUE'},
    '4100': {'name': 'مبيعات China Town – فرعي', 'type': 'REVENUE'},
    '4010': {'name': 'مبيعات Place India – رئيسي', 'type': 'REVENUE'},
    '4110': {'name': 'مبيعات Place India – فرعي', 'type': 'REVENUE'},
    '4020': {'name': 'مبيعات Keeta – عبر الإنترنت', 'type': 'REVENUE'},
    '4120': {'name': 'مبيعات Keeta – فرعي', 'type': 'REVENUE'},
    '4030': {'name': 'مبيعات HungerStation – عبر الإنترنت', 'type': 'REVENUE'},
    '4040': {'name': 'عمولات بنكية مستلمة', 'type': 'REVENUE'},
    '4140': {'name': 'خصومات على المبيعات', 'type': 'EXPENSE'},

    # 5. COGS – تكلفة البضاعة المباعة
    '5005': {'name': 'تكلفة المواد الغذائية المباشرة', 'type': 'COGS'},

    # 6. EXPENSES – المصروفات التشغيلية
    '5010': {'name': 'رواتب وأجور الموظفين', 'type': 'EXPENSE'},
    '5300': {'name': 'رواتب ومكافآت', 'type': 'EXPENSE'},
    '5310': {'name': 'مصروف رواتب', 'type': 'EXPENSE'},
    '5320': {'name': 'سلف موظفين (مصروف)', 'type': 'EXPENSE'},
    '5330': {'name': 'تسوية سلف رواتب', 'type': 'EXPENSE'},
    '5020': {'name': 'كهرباء وماء وصيانة', 'type': 'EXPENSE'},
    '5150': {'name': 'صيانة عامة', 'type': 'EXPENSE'},
    '5030': {'name': 'إيجار', 'type': 'EXPENSE'},
    '5040': {'name': 'ديزل', 'type': 'EXPENSE'},
    '5050': {'name': 'إنترنت', 'type': 'EXPENSE'},
    '5060': {'name': 'أدوات مكتبية', 'type': 'EXPENSE'},
    '5070': {'name': 'مواد تنظيف', 'type': 'EXPENSE'},
    '5080': {'name': 'غسيل ملابس', 'type': 'EXPENSE'},
    '5090': {'name': 'عمولات بنكية', 'type': 'EXPENSE'},
    '5100': {'name': 'رسوم حكومية', 'type': 'EXPENSE'},
    '5110': {'name': 'تسويق', 'type': 'EXPENSE'},
    '5120': {'name': 'مصاريف حكومية', 'type': 'EXPENSE'},
    '5130': {'name': 'مصروفات أخرى', 'type': 'EXPENSE'},
    '5140': {'name': 'مصروفات غير تشغيلية', 'type': 'EXPENSE'},
    '5160': {'name': 'أدوات مكتبية (مكرر)', 'type': 'EXPENSE'},
    '5170': {'name': 'مواد تنظيف (مكرر)', 'type': 'EXPENSE'},
    '5180': {'name': 'غسيل ملابس (مكرر)', 'type': 'EXPENSE'},
    '5190': {'name': 'عمولات بنكية (مكرر)', 'type': 'EXPENSE'},
    '5200': {'name': 'رسوم حكومية (مكرر)', 'type': 'EXPENSE'},

    # 7. TAX – الضرائب
    '6000': {'name': 'تسوية ضريبة المدخلات والمخرجات', 'type': 'TAX'},
    '6300': {'name': 'تسوية ضريبة القيمة المضافة', 'type': 'TAX'},
    '1100': {'name': 'Input VAT (ضريبة المدخلات)', 'type': 'ASSET'}
}

SHORT_TO_NUMERIC = {
    'CASH': ('1040', CHART_OF_ACCOUNTS.get('1040', {'name':'نقد وصندوق','type':'ASSET'})['name'], CHART_OF_ACCOUNTS.get('1040', {'type':'ASSET'})['type']),
    'BANK': ('1050', CHART_OF_ACCOUNTS.get('1050', {'name':'بنك – حساب رئيسي','type':'ASSET'})['name'], CHART_OF_ACCOUNTS.get('1050', {'type':'ASSET'})['type']),
    'AP': ('2110', CHART_OF_ACCOUNTS['2110']['name'], CHART_OF_ACCOUNTS['2110']['type']),
    'AR_KEETA': ('1130', CHART_OF_ACCOUNTS.get('1130', {'name':'حسابات التحصيل (عميل – Keeta)','type':'ASSET'})['name'], CHART_OF_ACCOUNTS.get('1130', {'type':'ASSET'})['type']),
    'AR_HUNGER': ('1140', CHART_OF_ACCOUNTS.get('1140', {'name':'حسابات التحصيل (عميل – Hunger)','type':'ASSET'})['name'], CHART_OF_ACCOUNTS.get('1140', {'type':'ASSET'})['type']),
    'VAT_IN': ('1100', CHART_OF_ACCOUNTS.get('1100', {'name':'Input VAT (ضريبة المدخلات)','type':'ASSET'})['name'], CHART_OF_ACCOUNTS.get('1100', {'type':'ASSET'})['type']),
    'VAT_OUT': ('2060', CHART_OF_ACCOUNTS.get('2060', {'name':'Output VAT (ضريبة المخرجات)','type':'LIABILITY'})['name'], CHART_OF_ACCOUNTS.get('2060', {'type':'LIABILITY'})['type']),
    'VAT_SETTLE': ('6000', CHART_OF_ACCOUNTS.get('6000', {'name':'تسوية ضريبة المدخلات والمخرجات','type':'TAX'})['name'], CHART_OF_ACCOUNTS.get('6000', {'type':'TAX'})['type']),
    'REV_CT': ('4000', CHART_OF_ACCOUNTS.get('4000', {'name':'مبيعات China Town','type':'REVENUE'})['name'], CHART_OF_ACCOUNTS.get('4000', {'type':'REVENUE'})['type']),
    'REV_PI': ('4010', CHART_OF_ACCOUNTS.get('4010', {'name':'مبيعات Place India','type':'REVENUE'})['name'], CHART_OF_ACCOUNTS.get('4010', {'type':'REVENUE'})['type']),
    'REV_KEETA': ('4020', CHART_OF_ACCOUNTS.get('4020', {'name':'مبيعات Keeta (عبر الإنترنت)','type':'REVENUE'})['name'], CHART_OF_ACCOUNTS.get('4020', {'type':'REVENUE'})['type']),
    'REV_HUNGER': ('4030', CHART_OF_ACCOUNTS.get('4030', {'name':'مبيعات Hunger (عبر الإنترنت)','type':'REVENUE'})['name'], CHART_OF_ACCOUNTS.get('4030', {'type':'REVENUE'})['type']),
    'DISC_SALES': ('4140', CHART_OF_ACCOUNTS['4140']['name'], CHART_OF_ACCOUNTS['4140']['type']),
    'EXP_OP': ('5100', CHART_OF_ACCOUNTS['5100']['name'], CHART_OF_ACCOUNTS['5100']['type']),
    'RENT': ('5110', CHART_OF_ACCOUNTS['5110']['name'], CHART_OF_ACCOUNTS['5110']['type']),
    'ELEC': ('5120', CHART_OF_ACCOUNTS['5120']['name'], CHART_OF_ACCOUNTS['5120']['type']),
    'DIESEL': ('5130', CHART_OF_ACCOUNTS['5130']['name'], CHART_OF_ACCOUNTS['5130']['type']),
    'NET': ('5140', CHART_OF_ACCOUNTS['5140']['name'], CHART_OF_ACCOUNTS['5140']['type']),
    'MAINT': ('5150', CHART_OF_ACCOUNTS['5150']['name'], CHART_OF_ACCOUNTS['5150']['type']),
    'STATIONERY': ('5160', CHART_OF_ACCOUNTS['5160']['name'], CHART_OF_ACCOUNTS['5160']['type']),
    'CLEAN': ('5170', CHART_OF_ACCOUNTS['5170']['name'], CHART_OF_ACCOUNTS['5170']['type']),
    'LAUNDRY': ('5180', CHART_OF_ACCOUNTS['5180']['name'], CHART_OF_ACCOUNTS['5180']['type']),
    'BANK_COMM': ('5190', CHART_OF_ACCOUNTS['5190']['name'], CHART_OF_ACCOUNTS['5190']['type']),
    'GOV': ('5200', CHART_OF_ACCOUNTS['5200']['name'], CHART_OF_ACCOUNTS['5200']['type']),
    'SAL_ALL': ('5300', CHART_OF_ACCOUNTS['5300']['name'], CHART_OF_ACCOUNTS['5300']['type']),
    'SAL_EXP': ('5310', CHART_OF_ACCOUNTS['5310']['name'], CHART_OF_ACCOUNTS['5310']['type']),
    'EMP_ADV_EXP': ('5330', CHART_OF_ACCOUNTS['5330']['name'], CHART_OF_ACCOUNTS['5330']['type']),
    'SAL_DED': ('5330', CHART_OF_ACCOUNTS['5330']['name'], CHART_OF_ACCOUNTS['5330']['type']),
    'EXP_NONOP': ('5400', CHART_OF_ACCOUNTS.get('5400', {'name':'مصروفات غير تشغيلية','type':'EXPENSE'})['name'], CHART_OF_ACCOUNTS.get('5400', {'type':'EXPENSE'})['type']),
    # payroll liabilities and asset advances
    'PAYROLL': ('2030', CHART_OF_ACCOUNTS['2030']['name'], CHART_OF_ACCOUNTS['2030']['type']),
    'PAYROLL_LIAB': ('2130', CHART_OF_ACCOUNTS['2130']['name'], CHART_OF_ACCOUNTS['2130']['type']),
    'EMP_ADV': ('1030', CHART_OF_ACCOUNTS['1030']['name'], CHART_OF_ACCOUNTS['1030']['type']),
    'COGS': ('5005', CHART_OF_ACCOUNTS['5005']['name'], CHART_OF_ACCOUNTS['5005']['type']),
}

def _resolve_account(code: str, name: str, kind: str):
    c = (code or '').strip().upper()
    if c in SHORT_TO_NUMERIC:
        cc, nn, tt = SHORT_TO_NUMERIC[c]
        return cc, nn, tt
    # If numeric code already
    if c in CHART_OF_ACCOUNTS:
        return c, CHART_OF_ACCOUNTS[c]['name'], CHART_OF_ACCOUNTS[c]['type']
    return c, name, (kind or '').strip().upper()

def seed_chart_of_accounts():
    try:
        for cc, info in CHART_OF_ACCOUNTS.items():
            _account(cc, info['name'], info['type'])
        db.session.commit()
    except Exception:
        db.session.rollback()
@main.route('/financial', methods=['GET'], endpoint='financial_dashboard')
@login_required
def financial_dashboard():
    warmup_db_once()
    sd = (request.args.get('start_date') or '').strip()
    ed = (request.args.get('end_date') or '').strip()
    br = (request.args.get('branch') or 'all').strip()
    try:
        if sd:
            start_dt = datetime.strptime(sd, '%Y-%m-%d').date()
        else:
            today = get_saudi_now().date()
            start_dt = date(today.year, today.month, 1)
        end_dt = datetime.strptime(ed, '%Y-%m-%d').date() if ed else get_saudi_now().date()
    except Exception:
        today = get_saudi_now().date()
        start_dt = date(today.year, today.month, 1)
        end_dt = today

    def branch_filter(q, model):
        try:
            if br in ('place_india','china_town') and hasattr(model, 'branch'):
                return q.filter(model.branch == br)
            return q
        except Exception:
            return q

    sales_q = db.session.query(
        func.coalesce(func.sum(SalesInvoice.total_before_tax), 0).label('amount'),
        func.coalesce(func.sum(SalesInvoice.discount_amount), 0).label('discount'),
        func.coalesce(func.sum(SalesInvoice.tax_amount), 0).label('vat'),
        func.coalesce(func.sum(SalesInvoice.total_after_tax_discount), 0).label('total')
    ).filter(SalesInvoice.date.between(start_dt, end_dt))
    sales_q = branch_filter(sales_q, SalesInvoice)
    sales_row = sales_q.first()
    sales_summary = {
        'amount': float(sales_row.amount or 0),
        'discount': float(sales_row.discount or 0),
        'vat_out': float(sales_row.vat or 0),
        'net_sales': float(sales_row.total or 0),
    }

    pm_rows = db.session.query(SalesInvoice.payment_method, func.count('*'), func.coalesce(func.sum(SalesInvoice.total_after_tax_discount), 0))
    pm_rows = pm_rows.filter(SalesInvoice.date.between(start_dt, end_dt))
    pm_rows = branch_filter(pm_rows, SalesInvoice)
    pm_rows = pm_rows.group_by(SalesInvoice.payment_method).all()
    by_payment_sales = [{'method': (r[0] or 'unknown'), 'count': int(r[1] or 0), 'total': float(r[2] or 0)} for r in pm_rows]

    purch_q = db.session.query(
        func.coalesce(func.sum(PurchaseInvoice.total_before_tax), 0),
        func.coalesce(func.sum(PurchaseInvoice.tax_amount), 0),
        func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0)
    ).filter(PurchaseInvoice.date.between(start_dt, end_dt))
    purch_row = purch_q.first()
    purchases_summary = {
        'amount': float(purch_row[0] or 0),
        'vat_in': float(purch_row[1] or 0),
        'total': float(purch_row[2] or 0),
        'no_vat_amount': float(db.session.query(func.coalesce(func.sum(PurchaseInvoice.total_before_tax - PurchaseInvoice.discount_amount), 0))
            .filter(PurchaseInvoice.date.between(start_dt, end_dt), func.coalesce(PurchaseInvoice.tax_amount, 0) == 0).scalar() or 0)
    }

    exp_rows = db.session.query(
        ExpenseInvoiceItem.description,
        func.coalesce(func.sum((ExpenseInvoiceItem.price_before_tax * ExpenseInvoiceItem.quantity) - ExpenseInvoiceItem.discount), 0).label('amount')
    ).join(ExpenseInvoice, ExpenseInvoiceItem.invoice_id == ExpenseInvoice.id)
    exp_rows = exp_rows.filter(ExpenseInvoice.date.between(start_dt, end_dt)).group_by(ExpenseInvoiceItem.description).all()
    expenses_total = float(sum([float(r[1] or 0) for r in exp_rows]) or 0)
    expenses_summary = {
        'total': expenses_total,
        'items': [{'label': r[0], 'amount': float(r[1] or 0)} for r in exp_rows]
    }

    vat_summary = {
        'vat_out': float(sales_summary['vat_out']),
        'vat_in': float(purchases_summary['vat_in']),
        'net_vat_due': float(sales_summary['vat_out']) - float(purchases_summary['vat_in'])
    }

    try:
        ap_total = float(db.session.query(func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0))
            .filter(PurchaseInvoice.date.between(start_dt, end_dt)).scalar() or 0)
        ap_paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
            .filter(Payment.invoice_type == 'purchase').filter(Payment.created_at.between(start_dt, end_dt)).scalar() or 0)
        liabilities_summary = {
            'suppliers_due': max(0.0, ap_total - ap_paid),
            'prev_owner': 0.0,
            'other_dues': 0.0,
        }
    except Exception:
        liabilities_summary = {'suppliers_due': 0.0, 'prev_owner': 0.0, 'other_dues': 0.0}

    try:
        gross_profit = float(sales_summary['net_sales']) - float(purchases_summary['no_vat_amount'])
        net_profit = gross_profit - float(expenses_total)
    except Exception:
        gross_profit = 0.0
        net_profit = 0.0

    return render_template('financial_dashboard.html',
        start_date=start_dt, end_date=end_dt, branch=br,
        sales_summary=sales_summary, by_payment_sales=by_payment_sales,
        purchases_summary=purchases_summary,
        expenses_summary=expenses_summary,
        vat_summary=vat_summary,
        liabilities_summary=liabilities_summary,
        gross_profit=gross_profit, net_profit=net_profit)
@main.route('/financial/print', methods=['GET'], endpoint='financial_print')
@login_required
def financial_print():
    warmup_db_once()
    args = request.args.to_dict()
    with current_app.test_request_context(query_string=args):
        return financial_dashboard()

@main.route('/advances', methods=['GET'], endpoint='advances_page')
@login_required
def advances_page():
    warmup_db_once()
    try:
        employees = Employee.query.order_by(Employee.full_name.asc()).all()
    except Exception:
        employees = []
    return render_template('advances.html', employees=employees)

@main.route('/api/advances/list', methods=['GET'], endpoint='api_advances_list')
@login_required
def api_advances_list():
    try:
        from datetime import date as _date
        from datetime import datetime as _dt
        from models import JournalLine, JournalEntry, Employee
        emp_id = request.args.get('employee_id', type=int)
        dept_f = (request.args.get('dept') or '').strip()
        if dept_f:
            low = dept_f.lower()
            if low in ('kitchen','chef'):
                dept_f = 'شيف'
            elif low in ('hall','floor','wait'):
                dept_f = 'الصالة'
            elif low in ('admin','management'):
                dept_f = 'إداري'
        month_param = (request.args.get('month') or '').strip()
        start_date = (request.args.get('start') or '').strip()
        end_date = (request.args.get('end') or '').strip()

        q = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id)
        adv_code = SHORT_TO_NUMERIC['EMP_ADV'][0]
        acc = _account(adv_code, CHART_OF_ACCOUNTS[adv_code]['name'], CHART_OF_ACCOUNTS[adv_code]['type'])
        if acc:
            q = q.filter(JournalLine.account_id == acc.id)

        # AND logic: apply all provided filters
        if emp_id:
            q = q.filter(JournalLine.employee_id == emp_id)
        if dept_f:
            try:
                emp_ids = [int(e.id) for e in Employee.query.filter(func.lower(Employee.department) == dept_f.lower()).all()]
                if emp_ids:
                    q = q.filter(JournalLine.employee_id.in_(emp_ids))
                else:
                    q = q.filter(JournalLine.employee_id == -1)
            except Exception:
                pass
        # Date range: month or explicit start/end
        if month_param and '-' in month_param:
            try:
                y, m = month_param.split('-'); y = int(y); m = int(m)
                start_dt = _date(y, m, 1)
                end_dt = _date(y + (1 if m == 12 else 0), 1 if m == 12 else (m + 1), 1)
                q = q.filter(JournalEntry.date >= start_dt, JournalEntry.date < end_dt)
            except Exception:
                pass
        else:
            sd_dt = None; ed_dt = None
            try:
                sd_dt = _dt.fromisoformat(start_date).date() if start_date else None
            except Exception:
                sd_dt = None
            try:
                ed_dt = _dt.fromisoformat(end_date).date() if end_date else None
            except Exception:
                ed_dt = None
            if sd_dt and ed_dt:
                q = q.filter(JournalEntry.date.between(sd_dt, ed_dt))
            elif sd_dt and not ed_dt:
                q = q.filter(JournalEntry.date >= sd_dt)
            elif ed_dt and not sd_dt:
                q = q.filter(JournalEntry.date <= ed_dt)

        rows = q.order_by(JournalEntry.date.desc(), JournalLine.line_no.asc()).all()
        data_map = {}
        for jl, je in rows:
            k = int(getattr(jl, 'employee_id', 0) or 0)
            if k not in data_map:
                data_map[k] = {
                    'employee_id': k,
                    'employee_name': '',
                    'branch': getattr(je, 'branch_code', '') or '',
                    'department': '',
                    'granted': 0.0,
                    'paid': 0.0,
                    'remaining': 0.0,
                    'last_date': str(getattr(je, 'date', get_saudi_now().date()))
                }
            data_map[k]['granted'] += float(getattr(jl, 'debit', 0) or 0)
            data_map[k]['paid'] += float(getattr(jl, 'credit', 0) or 0)
            data_map[k]['remaining'] = max(0.0, data_map[k]['granted'] - data_map[k]['paid'])
            data_map[k]['last_date'] = str(getattr(je, 'date', get_saudi_now().date()))
        # Attach names and departments
        try:
            emps = {int(e.id): e for e in Employee.query.all()}
            for k, v in data_map.items():
                e = emps.get(k)
                v['employee_name'] = (getattr(e, 'full_name', '') or '') if e else ''
                v['department'] = (getattr(e, 'department', '') or '') if e else ''
        except Exception:
            pass
        # Filter to only current active employees and exclude orphan keys (e.g., employee_id=0)
        active_ids = set()
        try:
            active_ids = {int(e.id) for e in Employee.query.filter_by(active=True).all()}
        except Exception:
            active_ids = set(int(k) for k in (emps.keys() if 'emps' in locals() else []))
        data = [v for (k, v) in data_map.items() if int(k or 0) in active_ids]
        totals = {
            'granted': float(sum([d['granted'] for d in data]) or 0),
            'paid': float(sum([d['paid'] for d in data]) or 0),
            'remaining': float(sum([d['remaining'] for d in data]) or 0),
        }
        return jsonify({'ok': True, 'rows': data, 'totals': totals})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/advances/metrics', methods=['GET'], endpoint='api_advances_metrics')
@login_required
def api_advances_metrics():
    try:
        from models import JournalLine, JournalEntry, Employee
        adv_code = SHORT_TO_NUMERIC['EMP_ADV'][0]
        acc = _account(adv_code, CHART_OF_ACCOUNTS[adv_code]['name'], CHART_OF_ACCOUNTS[adv_code]['type'])
        if not acc:
            return jsonify({'ok': True, 'total': 0.0, 'unpaid': 0.0, 'last_date': None, 'series': []})
        rows = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id).filter(JournalLine.account_id == acc.id).order_by(JournalEntry.date.asc(), JournalLine.line_no.asc()).all()
        total_debit = float(sum([float(getattr(jl, 'debit', 0) or 0) for jl, _ in rows]) or 0)
        total_credit = float(sum([float(getattr(jl, 'credit', 0) or 0) for jl, _ in rows]) or 0)
        last_date = None
        for _, je in rows:
            last_date = str(getattr(je, 'date', get_saudi_now().date()))
        series_map = {}
        emp_depts = {}
        try:
            for e in Employee.query.all():
                emp_depts[int(e.id)] = (e.department or '').strip()
        except Exception:
            pass
        for jl, je in rows:
            d = getattr(je, 'date', get_saudi_now().date())
            y = int(getattr(d, 'year', get_saudi_now().year))
            m = int(getattr(d, 'month', get_saudi_now().month))
            dept = (emp_depts.get(int(getattr(jl, 'employee_id', 0) or 0)) or '')
            key = (y, m, dept)
            if key not in series_map:
                series_map[key] = {'year': y, 'month': m, 'department': dept, 'granted': 0.0, 'repaid': 0.0}
            series_map[key]['granted'] += float(getattr(jl, 'debit', 0) or 0)
            series_map[key]['repaid'] += float(getattr(jl, 'credit', 0) or 0)
        series = list(series_map.values())
        series.sort(key=lambda r: (r['year'], r['month']))
        return jsonify({'ok': True, 'total': total_debit, 'unpaid': max(0.0, total_debit - total_credit), 'last_date': last_date, 'series': series})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/advances/pay', methods=['POST'], endpoint='api_advances_pay')
@login_required
def api_advances_pay():
    try:
        emp_id = request.form.get('employee_id', type=int)
        amount = request.form.get('amount', type=float) or 0.0
        method = (request.form.get('method') or 'cash').strip().lower()
        date_s = (request.form.get('date') or get_saudi_now().strftime('%Y-%m-%d'))
        from datetime import datetime as _dt
        dval = _dt.strptime(date_s, '%Y-%m-%d').date()
        if not emp_id or amount <= 0:
            return jsonify({'ok': False, 'error': 'invalid'}), 400
        try:
            from models import JournalEntry, JournalLine
            cash_acc = _pm_account(method)
            emp_adv_acc = _account(SHORT_TO_NUMERIC['EMP_ADV'][0], CHART_OF_ACCOUNTS['1030']['name'], CHART_OF_ACCOUNTS['1030']['type'])
            je = JournalEntry(entry_number=f"JE-ADVPAY-{emp_id}-{int(amount)}", date=dval, branch_code=None, description=f"Advance repayment {emp_id}", status='posted', total_debit=amount, total_credit=amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
            db.session.add(je); db.session.flush()
            if cash_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=cash_acc.id, debit=amount, credit=0.0, description='Advance repayment cash/bank', line_date=dval, employee_id=emp_id))
            if emp_adv_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=emp_adv_acc.id, debit=0.0, credit=amount, description='Advance repayment', line_date=dval, employee_id=emp_id))
            db.session.commit()
        except Exception:
            db.session.rollback(); return jsonify({'ok': False, 'error': 'post_failed'}), 400
        try:
            if method in ('bank','card','visa','mastercard'):
                _post_ledger(dval, 'BANK', 'Bank', 'asset', amount, 0.0, f'ADV REPAY {emp_id}')
            else:
                _post_ledger(dval, 'CASH', 'Cash', 'asset', amount, 0.0, f'ADV REPAY {emp_id}')
            _post_ledger(dval, 'EMP_ADV', 'سلف للموظفين', 'asset', 0.0, amount, f'ADV REPAY {emp_id}')
        except Exception:
            pass
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/employees', methods=['GET'], endpoint='api_employees')
@login_required
def api_employees():
    try:
        rows = Employee.query.order_by(Employee.full_name.asc()).all()
        data = []
        for e in rows:
            data.append({
                'id': int(getattr(e,'id',0) or 0),
                'name': getattr(e,'full_name','') or '',
                'department': getattr(e,'department','') or '',
                'position': getattr(e,'position','') or '',
                'branch_code': getattr(e,'branch_code','') or '',
                'status': getattr(e,'status','') or '',
                'basic': float(getattr(getattr(e,'salary_default',None),'base_salary',0) or 0),
                'last_salary': 0.0,
                'advance': 0.0,
            })
        if not data:
            data = [
                {"id":1, "name":"RIDOY", "department":"شيف", "basic":800.00, "last_salary":1727.00, "advance":0},
                {"id":2, "name":"SAIFUL ISLAM", "department":"شيف", "basic":800.00, "last_salary":1832.00, "advance":0}
            ]
        return jsonify({'ok': True, 'employees': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/employees', methods=['POST'], endpoint='api_employees_create')
@login_required
def api_employees_create():
    try:
        from models import Employee, EmployeeSalaryDefault
        full_name = (request.form.get('full_name') or '').strip()
        national_id = (request.form.get('national_id') or '').strip()
        department = (request.form.get('department') or '').strip()
        position = (request.form.get('position') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        email = (request.form.get('email') or '').strip()
        hire_date = (request.form.get('hire_date') or '').strip()
        branch_code = (request.form.get('branch_code') or '').strip()
        base_salary = request.form.get('base_salary', type=float) or 0.0
        salary_type = (request.form.get('salary_type') or 'fixed').strip().lower()
        hourly_rate = request.form.get('hourly_rate', type=float)
        if not full_name or not national_id:
            return jsonify({'ok': False, 'error': 'missing_fields'}), 400
        from datetime import datetime as _dt
        hd = None
        try:
            hd = _dt.strptime(hire_date, '%Y-%m-%d').date() if hire_date else None
        except Exception:
            hd = None
        emp = Employee(full_name=full_name, national_id=national_id, department=department, position=position, phone=phone, email=email, hire_date=hd, status='active', active=True)
        try:
            db.session.add(emp); db.session.flush()
            db.session.add(EmployeeSalaryDefault(employee_id=int(emp.id), base_salary=base_salary, allowances=0.0, deductions=0.0))
            db.session.commit()
            try:
                from app.models import AppKV
                settings = {'salary_type': ('hourly' if salary_type=='hourly' else 'fixed')}
                if hourly_rate is not None:
                    settings['hourly_rate'] = float(hourly_rate or 0.0)
                AppKV.set(f"emp_settings:{int(emp.id)}", settings)
            except Exception:
                pass
            return jsonify({'ok': True, 'id': int(emp.id)})
        except Exception as e:
            db.session.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/employees/<int:eid>', methods=['PUT'], endpoint='api_employees_update')
@login_required
def api_employees_update(eid: int):
    try:
        from models import Employee, EmployeeSalaryDefault
        emp = Employee.query.get_or_404(int(eid))
        full_name = (request.form.get('full_name') or '').strip() or emp.full_name
        department = (request.form.get('department') or '').strip() or (emp.department or '')
        position = (request.form.get('position') or '').strip() or (emp.position or '')
        branch_code = (request.form.get('branch_code') or '').strip() or ''
        base_salary = request.form.get('base_salary', type=float)
        salary_type = (request.form.get('salary_type') or '').strip().lower()
        hourly_rate = request.form.get('hourly_rate', type=float)
        emp.full_name = full_name
        emp.department = department
        emp.position = position
        try:
            emp.branch_code = branch_code
        except Exception:
            pass
        db.session.flush()
        try:
            d = EmployeeSalaryDefault.query.filter_by(employee_id=int(emp.id)).first()
            if not d:
                d = EmployeeSalaryDefault(employee_id=int(emp.id), base_salary=0.0, allowances=0.0, deductions=0.0)
                db.session.add(d)
            if base_salary is not None:
                d.base_salary = float(base_salary or 0.0)
        except Exception:
            pass
        try:
            from app.models import AppKV
            cur = AppKV.get(f"emp_settings:{int(emp.id)}") or {}
            if salary_type in ('fixed','hourly'):
                cur['salary_type'] = salary_type
            if hourly_rate is not None:
                cur['hourly_rate'] = float(hourly_rate or 0.0)
            if cur:
                AppKV.set(f"emp_settings:{int(emp.id)}", cur)
        except Exception:
            pass
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/employees/<int:eid>', methods=['DELETE'], endpoint='api_employees_delete')
@login_required
def api_employees_delete(eid: int):
    try:
        from models import Employee, Salary, Payment, EmployeeSalaryDefault, EmployeeHours, JournalLine, JournalEntry, LedgerEntry
        emp = Employee.query.get_or_404(int(eid))
        sal_rows = Salary.query.filter_by(employee_id=int(eid)).all()
        sal_ids = [s.id for s in sal_rows]
        if sal_ids:
            try:
                Payment.query.filter(Payment.invoice_type=='salary', Payment.invoice_id.in_(sal_ids)).delete(synchronize_session=False)
            except Exception:
                pass
            Salary.query.filter(Salary.id.in_(sal_ids)).delete(synchronize_session=False)
        # Remove employee hours records
        try:
            EmployeeHours.query.filter_by(employee_id=int(eid)).delete(synchronize_session=False)
        except Exception:
            pass
        try:
            EmployeeSalaryDefault.query.filter_by(employee_id=int(eid)).delete(synchronize_session=False)
        except Exception:
            pass
        # Remove journal lines linked to this employee
        jline_ids = []
        jentry_ids = set()
        try:
            rows = JournalLine.query.filter(JournalLine.employee_id == int(eid)).all()
            for jl in rows:
                try:
                    jline_ids.append(int(jl.id))
                except Exception:
                    pass
                try:
                    jentry_ids.add(int(getattr(jl, 'journal_id', 0) or 0))
                except Exception:
                    pass
            if jline_ids:
                JournalLine.query.filter(JournalLine.id.in_(jline_ids)).delete(synchronize_session=False)
        except Exception:
            pass
        # Remove journal entries tied to the employee's salaries (accruals, payments, advances)
        try:
            if sal_ids:
                JournalEntry.query.filter(JournalEntry.salary_id.in_(sal_ids)).delete(synchronize_session=False)
        except Exception:
            pass
        # Clean up orphan journal entries (no remaining lines)
        try:
            if jentry_ids:
                for jid in list(jentry_ids):
                    try:
                        cnt = JournalLine.query.filter(JournalLine.journal_id == int(jid)).count()
                        if int(cnt or 0) == 0:
                            je = JournalEntry.query.get(int(jid))
                            if je:
                                db.session.delete(je)
                    except Exception:
                        continue
        except Exception:
            pass
        # Remove ledger entries created for salary payments or advances for this employee
        try:
            # Match descriptions used when posting:
            # - 'PAY SAL ... EMP <eid>' (salary payment)
            # - 'ADV EMP <eid>' (advance grant)
            # - 'ADV REPAY <eid>' (advance repayment)
            like_emp = f"% EMP {int(eid)}%"
            like_adv = f"%ADV EMP {int(eid)}%"
            like_adv_repay = f"%ADV REPAY {int(eid)}%"
            LedgerEntry.query.filter(
                or_(
                    LedgerEntry.description.like(like_emp),
                    LedgerEntry.description.like(like_adv),
                    LedgerEntry.description.like(like_adv_repay)
                )
            ).delete(synchronize_session=False)
        except Exception:
            pass
        db.session.delete(emp)
        # Remove KV settings/notes for this employee
        try:
            from app.models import AppKV
            AppKV.query.filter(AppKV.k.in_([f"emp_settings:{int(eid)}", f"emp_note:{int(eid)}"]))\
                .delete(synchronize_session=False)
        except Exception:
            pass
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/maintenance/cleanup-dummy', methods=['POST'], endpoint='api_cleanup_dummy')
@login_required
def api_cleanup_dummy():
    try:
        # Remove journal lines whose employee_id no longer exists
        from models import Employee, JournalLine, JournalEntry, LedgerEntry
        existing_ids = {int(e.id) for e in Employee.query.all()}
        rem_lines = 0
        try:
            rows = JournalLine.query.all()
            for jl in rows:
                try:
                    eid = int(getattr(jl, 'employee_id', 0) or 0)
                except Exception:
                    eid = 0
                if eid and eid not in existing_ids:
                    db.session.delete(jl); rem_lines += 1
            db.session.flush()
        except Exception:
            pass
        # Remove orphan journal entries (no remaining lines)
        rem_entries = 0
        try:
            jrows = JournalEntry.query.all()
            for je in jrows:
                try:
                    cnt = JournalLine.query.filter(JournalLine.journal_id == int(je.id)).count()
                    if int(cnt or 0) == 0:
                        db.session.delete(je); rem_entries += 1
                except Exception:
                    continue
        except Exception:
            pass
        # Remove ledger entries for advances/payments referencing non-active employees via description patterns
        rem_ledgers = 0
        try:
            active_ids = {int(e.id) for e in Employee.query.filter_by(active=True).all()}
            lrows = LedgerEntry.query.all()
            for le in lrows:
                desc = str(getattr(le, 'description', '') or '')
                # Try to extract employee id from patterns
                cand = None
                for tag in (' EMP ', 'ADV EMP ', 'ADV REPAY '):
                    if tag in desc:
                        try:
                            tail = desc.split(tag)[-1].strip()
                            cand = int(''.join([ch for ch in tail if ch.isdigit()]))
                        except Exception:
                            cand = None
                        break
                if cand and cand not in active_ids:
                    db.session.delete(le); rem_ledgers += 1
        except Exception:
            pass
        # Optionally remove known dummy employees by name if present
        removed_emps = 0
        try:
            dummies = Employee.query.filter(Employee.full_name.in_(['RIDOY','SAIFUL ISLAM'])).all()
            for emp in dummies:
                db.session.delete(emp); removed_emps += 1
        except Exception:
            pass
        db.session.commit()
        return jsonify({'ok': True, 'removed': {'journal_lines': rem_lines, 'journal_entries': rem_entries, 'ledger_entries': rem_ledgers, 'employees': removed_emps}})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/payroll-templates', methods=['GET'], endpoint='api_payroll_templates')
@login_required
def api_payroll_templates():
    try:
        tpl = {
            'template_id': 'default',
            'components': ["basic","extra","day_off","bonus","ot","others","vac_eid","deduct"]
        }
        return jsonify({'ok': True, 'templates': [tpl]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/payroll-run', methods=['POST'], endpoint='api_payroll_run')
@login_required
def api_payroll_run():
    try:
        import json as _json
        month = request.form.get('month') or request.form.get('pay_month') or get_saudi_now().strftime('%Y-%m')
        y, m = month.split('-'); year = int(y); mon = int(m)
        employees_json = request.form.get('rows')
        if employees_json:
            rows = _json.loads(employees_json)
        else:
            rows = []
            for e in Employee.query.order_by(Employee.full_name.asc()).all():
                rows.append({'id': int(e.id), 'name': e.full_name, 'department': e.department or '', 'basic': float(getattr(getattr(e,'salary_default',None),'base_salary',0) or 0), 'extra': 0.0, 'day_off': 0.0, 'bonus': 0.0, 'ot': 0.0, 'others': 0.0, 'vac_eid': 0.0, 'deduct': 0.0})
        exp_code = (request.form.get('exp_code') or SHORT_TO_NUMERIC.get('SAL_EXP', ('5310',))[0])
        liab_code = (request.form.get('liab_code') or SHORT_TO_NUMERIC.get('PAYROLL_LIAB', ('2130',))[0])
        tpl_id = (request.form.get('template_id') or 'default')
        preview_only = (request.form.get('preview') or '').strip().lower() in ('1','true','yes')
        journal = []
        total_sum = 0.0
        details = []
        for r in rows:
            b = float(r.get('basic') or 0)
            ex = float(r.get('extra') or 0)
            doff = float(r.get('day_off') or 0)
            bon = float(r.get('bonus') or 0)
            otv = float(r.get('ot') or 0)
            oth = float(r.get('others') or 0)
            ve = float(r.get('vac_eid') or 0)
            ded = float(r.get('deduct') or 0)
            total = round(b + ex + bon + otv + oth + ve - doff - ded, 2)
            total_sum += total
            details.append({'employee_id': r.get('id'), 'employee_name': r.get('name'), 'department': r.get('department'), 'net': total})
            if not preview_only:
                s = Salary.query.filter_by(employee_id=int(r.get('id')), year=year, month=mon).first()
                if not s:
                    s = Salary(employee_id=int(r.get('id')), year=year, month=mon, basic_salary=b, allowances=ex+bon+otv+oth+ve, deductions=doff+ded, previous_salary_due=0.0, total_salary=total, status='due')
                    db.session.add(s)
                else:
                    s.basic_salary = b
                    s.allowances = ex+bon+otv+oth+ve
                    s.deductions = doff+ded
                    s.previous_salary_due = 0.0
                    s.total_salary = total
                db.session.flush()
        rounding = round(total_sum, 2) - total_sum
        journal.append({'account': exp_code, 'name': CHART_OF_ACCOUNTS.get(exp_code, {}).get('name', 'مصروف رواتب'), 'debit': round(total_sum, 2), 'credit': 0.0})
        journal.append({'account': liab_code, 'name': CHART_OF_ACCOUNTS.get(liab_code, {}).get('name', 'رواتب مستحقة'), 'debit': 0.0, 'credit': round(total_sum, 2)})
        if not preview_only:
            try:
                from models import JournalEntry, JournalLine
                je = JournalEntry(entry_number=f"JE-PR-{year}{mon:02d}", date=get_saudi_now().date(), branch_code=None, description=f"Payroll run {year}-{mon}", status='posted', total_debit=round(total_sum,2), total_credit=round(total_sum,2), created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
                db.session.add(je); db.session.flush()
                exp_acc = _account(exp_code, CHART_OF_ACCOUNTS.get(exp_code,{}).get('name','مصروف رواتب'), CHART_OF_ACCOUNTS.get(exp_code,{}).get('type','EXPENSE'))
                liab_acc = _account(liab_code, CHART_OF_ACCOUNTS.get(liab_code,{}).get('name','رواتب مستحقة'), CHART_OF_ACCOUNTS.get(liab_code,{}).get('type','LIABILITY'))
                if exp_acc:
                    db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=exp_acc.id, debit=round(total_sum,2), credit=0.0, description='Payroll expense', line_date=get_saudi_now().date()))
                if liab_acc:
                    db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=liab_acc.id, debit=0.0, credit=round(total_sum,2), description='Payroll liability', line_date=get_saudi_now().date()))
                db.session.commit()
            except Exception:
                db.session.rollback()
        return jsonify({'ok': True, 'template_id': tpl_id, 'month': {'year': year, 'month': mon}, 'details': details, 'journal_preview': journal, 'rounding_diff': round(rounding, 2), 'posted': (not preview_only)})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/advances', methods=['POST'], endpoint='api_advances')
@login_required
def api_advances():
    try:
        emp_id = request.form.get('employee_id', type=int)
        amount = request.form.get('amount', type=float) or 0.0
        method = (request.form.get('method') or 'cash').strip().lower()
        date_s = (request.form.get('date') or get_saudi_now().strftime('%Y-%m-%d'))
        from datetime import datetime as _dt
        dval = _dt.strptime(date_s, '%Y-%m-%d').date()
        if not emp_id or amount <= 0:
            return jsonify({'ok': False, 'error': 'invalid'}), 400
        try:
            from models import JournalEntry, JournalLine
            cash_acc = _pm_account(method)
            emp_adv_acc = _account(SHORT_TO_NUMERIC['EMP_ADV'][0], CHART_OF_ACCOUNTS['1030']['name'], CHART_OF_ACCOUNTS['1030']['type'])
            je = JournalEntry(entry_number=f"JE-ADV-{emp_id}-{int(amount)}", date=dval, branch_code=None, description=f"Employee advance {emp_id}", status='posted', total_debit=amount, total_credit=amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
            db.session.add(je); db.session.flush()
            if emp_adv_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=emp_adv_acc.id, debit=amount, credit=0.0, description='Employee advance', line_date=dval, employee_id=emp_id))
            if cash_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=cash_acc.id, debit=0.0, credit=amount, description='Cash/Bank', line_date=dval, employee_id=emp_id))
            db.session.commit()
        except Exception:
            db.session.rollback(); return jsonify({'ok': False, 'error': 'post_failed'}), 400
        try:
            _post_ledger(dval, 'EMP_ADV', 'سلف للموظفين', 'asset', amount, 0.0, f'ADV EMP {emp_id}')
            if method in ('bank','card','visa','mastercard'):
                _post_ledger(dval, 'BANK', 'Bank', 'asset', 0.0, amount, f'ADV EMP {emp_id}')
            else:
                _post_ledger(dval, 'CASH', 'Cash', 'asset', 0.0, amount, f'ADV EMP {emp_id}')
        except Exception:
            pass
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/employee-ledger', methods=['GET'], endpoint='api_employee_ledger')
@login_required
def api_employee_ledger():
    try:
        emp_id = request.args.get('emp_id', type=int)
        if not emp_id:
            return jsonify({'ok': False, 'error': 'emp_id_required'}), 400
        from models import JournalLine, JournalEntry
        rows = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id).filter(JournalLine.employee_id == emp_id).order_by(JournalEntry.date.asc(), JournalLine.line_no.asc()).all()
        data = []
        for jl, je in rows:
            data.append({
                'date': str(getattr(je,'date',get_saudi_now().date())),
                'entry': getattr(je,'entry_number',''),
                'desc': getattr(jl,'description',''),
                'debit': float(getattr(jl,'debit',0) or 0),
                'credit': float(getattr(jl,'credit',0) or 0)
            })
        return jsonify({'ok': True, 'rows': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/payroll-import', methods=['POST'], endpoint='api_payroll_import')
@login_required
def api_payroll_import():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'ok': False, 'error': 'no_file'}), 400
        import csv, io
        content = file.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(io.StringIO(content))
        count = 0
        for row in reader:
            emp_id = int(row.get('employee_id') or 0)
            month = int(row.get('month') or get_saudi_now().month)
            year = int(row.get('year') or get_saudi_now().year)
            def to_f(x):
                try:
                    return float(x or 0)
                except Exception:
                    return 0.0
            b = to_f(row.get('basic'))
            ex = to_f(row.get('extra'))
            doff = to_f(row.get('day_off'))
            bon = to_f(row.get('bonus'))
            otv = to_f(row.get('ot'))
            oth = to_f(row.get('others'))
            ve = to_f(row.get('vac_eid'))
            ded = to_f(row.get('deduct'))
            total = to_f(row.get('total')) or round(b + ex + bon + otv + oth + ve - doff - ded, 2)
            s = Salary.query.filter_by(employee_id=emp_id, year=year, month=month).first()
            if not s:
                s = Salary(employee_id=emp_id, year=year, month=month, basic_salary=b, allowances=ex+bon+otv+oth+ve, deductions=doff+ded, previous_salary_due=0.0, total_salary=total, status='due')
                db.session.add(s)
            else:
                s.basic_salary = b
                s.allowances = ex+bon+otv+oth+ve
                s.deductions = doff+ded
                s.previous_salary_due = 0.0
                s.total_salary = total
            count += 1
        db.session.commit()
        return jsonify({'ok': True, 'imported': count})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/payroll/previous', methods=['GET'], endpoint='payroll_previous_page')
@login_required
def payroll_previous_page():
    warmup_db_once()
    try:
        employees = Employee.query.order_by(Employee.full_name.asc()).all()
    except Exception:
        employees = []
    return render_template('payroll_previous.html', employees=employees)

@main.route('/api/payroll/history', methods=['GET'], endpoint='api_payroll_history')
@login_required
def api_payroll_history():
    try:
        rows = db.session.query(Salary.year, Salary.month, func.sum(func.coalesce(Salary.total_salary, 0))).group_by(Salary.year, Salary.month).order_by(Salary.year.desc(), Salary.month.desc()).all()
        total = float(sum([float(r[2] or 0) for r in rows]) or 0)
        last = rows[0] if rows else None
        months_count = len(rows)
        last_label = None
        if last:
            last_label = f"{last[1]:02d}/{last[0]}"
        series = [{'year': int(r[0]), 'month': int(r[1]), 'total': float(r[2] or 0)} for r in reversed(rows)]
        return jsonify({'ok': True, 'total': total, 'last_month': last_label, 'months': months_count, 'series': series})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/payroll/history/list', methods=['GET'], endpoint='api_payroll_history_list')
@login_required
def api_payroll_history_list():
    try:
        month = (request.args.get('month') or '').strip()
        dept = (request.args.get('department') or '').strip().lower()
        status = (request.args.get('status') or '').strip().lower()
        if month:
            y, m = month.split('-')
            y = int(y); m = int(m)
            q = Salary.query.filter_by(year=y, month=m)
        else:
            q = Salary.query
        rows = q.all()
        data = []
        for s in rows:
            try:
                e = Employee.query.get(int(s.employee_id))
            except Exception:
                e = None
            dep = (getattr(e, 'department', '') or '')
            if dept and dept not in dep.lower():
                continue
            st = (getattr(s, 'status', '') or '').strip().lower()
            if status and status not in st:
                continue
            data.append({
                'employee_id': int(getattr(s, 'employee_id', 0) or 0),
                'employee_name': getattr(e, 'full_name', '') if e else '',
                'month_label': f"{int(getattr(s,'month',0)):02d}/{int(getattr(s,'year',0))}",
                'basic': float(getattr(s, 'basic_salary', 0) or 0),
                'ot': 0.0,
                'bonus': 0.0,
                'total': float(getattr(s, 'total_salary', 0) or 0),
                'status': getattr(s, 'status', '') or 'due'
            })
        return jsonify({'ok': True, 'rows': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/ledger', methods=['GET'], endpoint='ledger_page')
@login_required
def ledger_page():
    warmup_db_once()
    try:
        employees = Employee.query.order_by(Employee.full_name.asc()).all()
    except Exception:
        employees = []
    return render_template('ledger.html', employees=employees)

@main.route('/chart-of-accounts', methods=['GET'], endpoint='chart_of_accounts')
@login_required
def chart_of_accounts():
    return render_template('chart.html')

def _gen_code_for_type(acc_type: str) -> str:
    try:
        import re
        t = (acc_type or '').strip().upper()
        ranges = {
            'ASSET': (1000, 1999),
            'LIABILITY': (2000, 2999),
            'EQUITY': (3000, 3999),
            'REVENUE': (4000, 4999),
            'EXPENSE': (5000, 5999),
            'TAX': (6000, 6999),
            'COGS': (5000, 5999)
        }
        base, end = ranges.get(t, (7000, 7999))
        seq_key = f"chart_code_seq:{t}"
        last = kv_get(seq_key, None)
        if isinstance(last, int):
            cand = min(last + 10, end)
        else:
            codes = []
            for k in CHART_OF_ACCOUNTS.keys():
                try:
                    x = int(re.sub(r'[^0-9]', '', str(k)))
                except Exception:
                    x = 0
                if x >= base and x <= end:
                    codes.append(x)
            extra = kv_get('chart_accounts', []) or []
            for e in extra:
                try:
                    x = int(re.sub(r'[^0-9]', '', str(e.get('code'))))
                except Exception:
                    x = 0
                if x >= base and x <= end:
                    codes.append(x)
            cand = max(codes) + 10 if codes else base
        if cand > end:
            cand = end
        kv_set(seq_key, cand)
        return str(cand)
    except Exception:
        return str(int(datetime.utcnow().timestamp()) % 10000)

@main.route('/api/chart/list', methods=['GET'], endpoint='api_chart_list')
@login_required
def api_chart_list():
    try:
        overrides = kv_get('chart_overrides', {}) or {}
        extra = kv_get('chart_accounts', []) or []
        out = []
        for code, info in CHART_OF_ACCOUNTS.items():
            parent_guess = ''
            if code in ('1000','2000','3000','4000','5000','6000'): parent_guess=''
            else:
                t = (info.get('type') or '').upper()
                parent_guess = {'ASSET':'1000','LIABILITY':'2000','EQUITY':'3000','REVENUE':'4000','EXPENSE':'5000','TAX':'6000','COGS':'5000'}.get(t,'')
            out.append({
                'code': code,
                'name': info.get('name'),
                'type': info.get('type'),
                'enabled': bool((overrides.get(code) or {}).get('enabled', True)),
                'parent': parent_guess,
                'balance': 0.0,
                'notes': ''
            })
        for e in extra:
            out.append({
                'code': e.get('code'),
                'name': e.get('name'),
                'type': e.get('type'),
                'enabled': bool(e.get('enabled', True)),
                'parent': e.get('parent') or '',
                'balance': float(e.get('balance') or 0.0),
                'notes': e.get('notes') or ''
            })
        try:
            from models import Account
            rows = Account.query.all()
            known = {x['code'] for x in out}
            for a in rows:
                c = (a.code or '').strip()
                if not c or c in known:
                    continue
                out.append({
                    'code': c,
                    'name': a.name,
                    'type': (a.type or 'EXPENSE'),
                    'enabled': True,
                    'parent': ({'ASSET':'1000','LIABILITY':'2000','EQUITY':'3000','REVENUE':'4000','EXPENSE':'5000','TAX':'6000','COGS':'5000'}).get((a.type or '').upper(), ''),
                    'balance': 0.0,
                    'notes': ''
                })
        except Exception:
            pass
        return jsonify({'ok': True, 'items': out})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/chart/balances', methods=['GET'], endpoint='api_chart_balances')
@login_required
def api_chart_balances():
    try:
        from models import JournalLine, Account, JournalEntry
        start_arg = (request.args.get('start_date') or '').strip()
        end_arg = (request.args.get('end_date') or '').strip()
        branch = (request.args.get('branch') or '').strip()
        q = db.session.query(Account.code.label('code'), func.coalesce(func.sum(JournalLine.debit), 0.0).label('debit_total'), func.coalesce(func.sum(JournalLine.credit), 0.0).label('credit_total')).join(Account, JournalLine.account_id == Account.id).join(JournalEntry, JournalLine.journal_id == JournalEntry.id)
        if start_arg and end_arg:
            try:
                from datetime import datetime as _dt
                sd = _dt.strptime(start_arg, '%Y-%m-%d').date()
                ed = _dt.strptime(end_arg, '%Y-%m-%d').date()
                q = q.filter(JournalLine.line_date.between(sd, ed))
            except Exception:
                pass
        if branch in ('china_town','place_india'):
            q = q.filter(JournalEntry.branch_code == branch)
        rows = q.group_by(Account.code).all()
        items = []
        for code, d, c in rows:
            dd = float(d or 0.0); cc = float(c or 0.0)
            items.append({'code': code, 'debit_total': dd, 'credit_total': cc, 'net': round(dd - cc, 2)})
        return jsonify({'ok': True, 'items': items})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/chart/opening_balance', methods=['POST'], endpoint='api_chart_opening_balance')
@login_required
@csrf.exempt
def api_chart_opening_balance():
    try:
        data = request.get_json(force=True) or {}
        code = (data.get('code') or '').strip()
        side = (data.get('side') or 'debit').strip().lower()
        amount = float(data.get('amount') or 0.0)
        date_str = (data.get('date') or '').strip()
        note = (data.get('note') or '').strip()
        if not code or amount <= 0:
            return jsonify({'ok': False, 'error': 'invalid_input'}), 400
        # Resolve date
        try:
            if date_str:
                from datetime import datetime as _dt
                je_date = _dt.strptime(date_str, '%Y-%m-%d').date()
            else:
                today = get_saudi_now().date()
                je_date = today.replace(month=1, day=1)
        except Exception:
            je_date = get_saudi_now().date()
        from models import Account, JournalEntry, JournalLine
        # Ensure main account exists
        acc = Account.query.filter(Account.code == code).first()
        if not acc:
            info = CHART_OF_ACCOUNTS.get(code, {'name': 'Account', 'type': 'EXPENSE'})
            acc = Account(code=code, name=info.get('name'), type=info.get('type'))
            db.session.add(acc); db.session.flush()
        # Ensure offset account (Owners’ Equity) exists
        off_code = '3000'
        off = Account.query.filter(Account.code == off_code).first()
        if not off:
            info = CHART_OF_ACCOUNTS.get(off_code, {'name':'حقوق الملكية','type':'EQUITY'})
            off = Account(code=off_code, name=info.get('name'), type=info.get('type'))
            db.session.add(off); db.session.flush()
        # Create journal entry
        entry_no = f"JE-OPEN-{code}-{int(get_saudi_now().timestamp())}"
        je = JournalEntry(entry_number=entry_no, date=je_date, branch_code=None, description=f"Opening Balance {code}", status='posted', total_debit=amount, total_credit=amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
        db.session.add(je); db.session.flush()
        # Lines
        if side == 'debit':
            db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=acc.id, debit=amount, credit=0.0, description=note or f"Opening {code}", line_date=je_date))
            db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=off.id, debit=0.0, credit=amount, description=note or f"Opening offset", line_date=je_date))
        else:
            db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=acc.id, debit=0.0, credit=amount, description=note or f"Opening {code}", line_date=je_date))
            db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=off.id, debit=amount, credit=0.0, description=note or f"Opening offset", line_date=je_date))
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/chart/toggle/<code>', methods=['POST'], endpoint='api_chart_toggle')
@login_required
@csrf.exempt
def api_chart_toggle(code):
    try:
        overrides = kv_get('chart_overrides', {}) or {}
        cur = overrides.get(code) or {}
        cur['enabled'] = not bool(cur.get('enabled', True))
        overrides[code] = cur
        kv_set('chart_overrides', overrides)
        return jsonify({'ok': True, 'enabled': cur['enabled']})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/chart/add', methods=['POST'], endpoint='api_chart_add')
@login_required
@csrf.exempt
def api_chart_add():
    try:
        data = request.get_json(force=True) or {}
        name = (data.get('name') or '').strip()
        acc_type = (data.get('type') or '').strip().upper()
        role = (data.get('role') or '').strip().lower()
        enabled = bool(data.get('enabled', True))
        parent = (data.get('parent_code') or '').strip()
        opening_balance = float(data.get('opening_balance') or 0.0)
        notes = (data.get('notes') or '').strip()
        if not name or not acc_type:
            return jsonify({'ok': False, 'error': 'missing_fields'}), 400
        code = _gen_code_for_type(acc_type)
        cur = kv_get('chart_accounts', []) or []
        cur.append({'code': code, 'name': name, 'type': acc_type, 'enabled': enabled, 'parent': parent, 'balance': opening_balance, 'notes': notes})
        kv_set('chart_accounts', cur)
        if role in ('platform_ar','platform_revenue'):
            pk = (data.get('platform_key') or '').strip().lower()
            if pk:
                platforms = kv_get('platforms_map', []) or []
                kws_raw = (data.get('platform_keywords') or '').strip()
                keywords = [x.strip().lower() for x in kws_raw.split(',') if x.strip()]
                entry = {'key': pk, 'keywords': keywords, 'auto_unpaid': True}
                if role == 'platform_ar':
                    entry['ar_code'] = code
                else:
                    entry['rev_code'] = code
                idx = None
                for i, e in enumerate(platforms):
                    if (e.get('key') or '').strip().lower() == pk:
                        idx = i; break
                if idx is not None:
                    prev = platforms[idx]
                    prev.update(entry)
                    if keywords:
                        prev['keywords'] = keywords
                    platforms[idx] = prev
                else:
                    platforms.append(entry)
                kv_set('platforms_map', platforms)
        elif role in ('branch_rev_ct','branch_rev_pi','default_ar'):
            acc_map = kv_get('acc_map', {}) or {}
            if role == 'branch_rev_ct':
                acc_map['REV_CT'] = code
            elif role == 'branch_rev_pi':
                acc_map['REV_PI'] = code
            elif role == 'default_ar':
                acc_map['AR'] = code
            kv_set('acc_map', acc_map)
        return jsonify({'ok': True, 'code': code})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/chart/refresh', methods=['POST'], endpoint='api_chart_refresh')
@login_required
@csrf.exempt
def api_chart_refresh():
    try:
        overrides = kv_get('chart_overrides', {}) or {}
        cur = kv_get('chart_accounts', []) or []
        try:
            from models import Account
            rows = Account.query.all()
        except Exception:
            rows = []
        existing_codes = set([str(x.get('code')) for x in cur if x.get('code')])
        default_codes = set(list(CHART_OF_ACCOUNTS.keys()))
        added = 0
        for a in rows:
            c = (a.code or '').strip()
            if not c or c in existing_codes or c in default_codes:
                continue
            cur.append({'code': c, 'name': a.name, 'type': (a.type or 'EXPENSE'), 'enabled': bool((overrides.get(c) or {}).get('enabled', True))})
            existing_codes.add(c)
            added += 1
        kv_set('chart_accounts', cur)
        return jsonify({'ok': True, 'added': added, 'total': len(cur)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/chart/merge_rules', methods=['GET'], endpoint='api_chart_merge_rules')
@login_required
def api_chart_merge_rules():
    try:
        rules = kv_get('chart_name_merge', []) or []
        return jsonify({'ok': True, 'items': rules})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/chart/add_merge_rule', methods=['POST'], endpoint='api_chart_add_merge_rule')
@login_required
@csrf.exempt
def api_chart_add_merge_rule():
    try:
        data = request.get_json(force=True) or {}
        target_code = (data.get('target_code') or '').strip()
        patterns_raw = (data.get('patterns') or '').strip()
        acc_type = (data.get('type') or '').strip().upper() if data.get('type') else None
        if not target_code or not patterns_raw:
            return jsonify({'ok': False, 'error': 'missing_fields'}), 400
        patterns = [x.strip() for x in patterns_raw.split(',') if x.strip()]
        rules = kv_get('chart_name_merge', []) or []
        rules.append({'target_code': target_code, 'patterns': patterns, 'type': acc_type})
        kv_set('chart_name_merge', rules)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/ledger/metrics', methods=['GET'], endpoint='api_ledger_metrics')
@login_required
def api_ledger_metrics():
    try:
        month = (request.args.get('month') or '').strip()
        start_dt = None; end_dt = None
        if month:
            y, m = month.split('-'); y = int(y); m = int(m)
            start_dt = date(y, m, 1)
            if m == 12:
                end_dt = date(y+1, 1, 1)
            else:
                end_dt = date(y, m+1, 1)
        from models import JournalLine, JournalEntry
        q = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id)
        if start_dt and end_dt:
            q = q.filter(JournalEntry.date >= start_dt, JournalEntry.date < end_dt)
        rows = q.all()
        total_adv = float(sum([float(getattr(jl, 'debit', 0) or 0) for jl, _ in rows]) or 0)
        total_ded = float(sum([float(getattr(jl, 'credit', 0) or 0) for jl, _ in rows]) or 0)
        count_entries = len(rows)
        ser = {}
        for jl, je in rows:
            d = getattr(je, 'date', get_saudi_now().date())
            y = int(getattr(d, 'year', get_saudi_now().year))
            m = int(getattr(d, 'month', get_saudi_now().month))
            k = (y, m)
            if k not in ser:
                ser[k] = {'year': y, 'month': m, 'debit_total': 0.0, 'credit_total': 0.0, 'entries_count': 0}
            ser[k]['debit_total'] += float(getattr(jl, 'debit', 0) or 0)
            ser[k]['credit_total'] += float(getattr(jl, 'credit', 0) or 0)
            ser[k]['entries_count'] += 1
        series = list(ser.values())
        series.sort(key=lambda r: (r['year'], r['month']))
        return jsonify({'ok': True, 'deductions_total': total_ded, 'advances_outstanding': total_adv, 'entries_count': count_entries, 'series': series})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/ledger/list', methods=['GET'], endpoint='api_ledger_list')
@login_required
def api_ledger_list():
    try:
        emp_id = request.args.get('emp_id', type=int)
        month = (request.args.get('month') or '').strip()
        acc_type = (request.args.get('type') or '').strip().lower()
        start_dt = None; end_dt = None
        if month:
            y, m = month.split('-'); y = int(y); m = int(m)
            start_dt = date(y, m, 1)
            if m == 12:
                end_dt = date(y+1, 1, 1)
            else:
                end_dt = date(y, m+1, 1)
        from models import JournalLine, JournalEntry
        q = db.session.query(JournalLine, JournalEntry).join(JournalEntry, JournalLine.journal_id == JournalEntry.id)
        if emp_id:
            q = q.filter(JournalLine.employee_id == emp_id)
        if start_dt and end_dt:
            q = q.filter(JournalEntry.date >= start_dt, JournalEntry.date < end_dt)
        rows = q.order_by(JournalEntry.date.asc(), JournalLine.line_no.asc()).all()
        data = []
        for jl, je in rows:
            item = {
                'employee_id': int(getattr(jl, 'employee_id', 0) or 0),
                'employee_name': '',
                'date': str(getattr(je, 'date', get_saudi_now().date())),
                'desc': getattr(jl, 'description', '') or '',
                'debit': float(getattr(jl, 'debit', 0) or 0),
                'credit': float(getattr(jl, 'credit', 0) or 0)
            }
            dsc = (item['desc'] or '').lower()
            typ = 'advance' if ('advance' in dsc or 'سلفة' in item['desc']) else ('deduction' if ('deduct' in dsc or 'خصم' in item['desc']) else 'other')
            bal = float(item['debit']) - float(item['credit'])
            st = 'paid' if item['debit']>0 and item['credit']==0 else ('partial' if item['debit']>0 and item['credit']>0 else ('due' if item['credit']>0 and item['debit']==0 else 'posted'))
            item['type'] = typ
            item['balance'] = bal
            item['status'] = st
            data.append(item)
        try:
            emps = {int(e.id): e.full_name for e in Employee.query.all()}
            for d in data:
                d['employee_name'] = emps.get(d['employee_id'], '')
        except Exception:
            pass
        if acc_type:
            if acc_type == 'advance':
                data = [d for d in data if d.get('type') == 'advance']
            elif acc_type == 'deduction':
                data = [d for d in data if d.get('type') == 'deduction']
        return jsonify({'ok': True, 'rows': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@main.route('/api/ledger/create', methods=['POST'], endpoint='api_ledger_create')
@login_required
def api_ledger_create():
    try:
        from datetime import datetime as _dt
        from models import JournalEntry, JournalLine
        emp_id = request.form.get('employee_id', type=int)
        amount = request.form.get('amount', type=float) or 0.0
        date_s = (request.form.get('date') or get_saudi_now().strftime('%Y-%m-%d'))
        typ = (request.form.get('type') or '').strip().lower()
        method = (request.form.get('method') or 'cash').strip().lower()
        desc = (request.form.get('description') or '').strip() or f'{typ} entry'
        if not emp_id or amount <= 0 or typ not in ('advance','deduction'):
            return jsonify({'ok': False, 'error': 'invalid'}), 400
        dval = _dt.strptime(date_s, '%Y-%m-%d').date()
        cash_acc = _pm_account(method)
        if typ == 'advance':
            emp_adv_acc = _account(SHORT_TO_NUMERIC['EMP_ADV'][0], CHART_OF_ACCOUNTS['1030']['name'], CHART_OF_ACCOUNTS['1030']['type'])
            je = JournalEntry(entry_number=f"JE-LED-ADV-{emp_id}-{int(amount)}", date=dval, branch_code=None, description=desc, status='posted', total_debit=amount, total_credit=amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
            db.session.add(je); db.session.flush()
            if emp_adv_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=emp_adv_acc.id, debit=amount, credit=0.0, description='Employee advance', line_date=dval, employee_id=emp_id))
            if cash_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=cash_acc.id, debit=0.0, credit=amount, description='Cash/Bank', line_date=dval, employee_id=emp_id))
        elif typ == 'deduction':
            ded_acc = _account(SHORT_TO_NUMERIC['SAL_DED'][0], CHART_OF_ACCOUNTS['5330']['name'], CHART_OF_ACCOUNTS['5330']['type'])
            je = JournalEntry(entry_number=f"JE-LED-DED-{emp_id}-{int(amount)}", date=dval, branch_code=None, description=desc, status='posted', total_debit=amount, total_credit=amount, created_by=getattr(current_user,'id',None), posted_by=getattr(current_user,'id',None))
            db.session.add(je); db.session.flush()
            if cash_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=1, account_id=cash_acc.id, debit=amount, credit=0.0, description='Cash/Bank', line_date=dval, employee_id=emp_id))
            if ded_acc:
                db.session.add(JournalLine(journal_id=je.id, line_no=2, account_id=ded_acc.id, debit=0.0, credit=amount, description='Salary deduction', line_date=dval, employee_id=emp_id))
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400

@main.route('/api/ledger/delete/<int:jeid>', methods=['DELETE'], endpoint='api_ledger_delete')
@login_required
def api_ledger_delete(jeid: int):
    try:
        from models import JournalEntry
        je = JournalEntry.query.get_or_404(int(jeid))
        db.session.delete(je)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
@main.route('/employees/create', methods=['GET'], endpoint='employees_create_page')
@login_required
def employees_create_page():
    return redirect(url_for('main.employee_uvd', mode='create'))

@main.route('/api/sales/pay', methods=['POST'], endpoint='api_sales_pay')
@login_required
def api_sales_pay():
    try:
        from models import SalesInvoice, Payment
        from sqlalchemy import func
        payload = request.get_json(silent=True) or {}
        invoice_id = request.form.get('invoice_id') or payload.get('invoice_id')
        amount = request.form.get('amount') or payload.get('amount')
        method = (request.form.get('payment_method') or payload.get('payment_method') or 'CASH').strip().upper()
        try:
            inv_id = int(invoice_id)
            amt = float(amount or 0)
        except Exception:
            return jsonify({'ok': False, 'error': 'invalid_parameters'}), 400
        if amt <= 0:
            return jsonify({'ok': False, 'error': 'amount_must_be_positive'}), 400
        inv = SalesInvoice.query.get(inv_id)
        if not inv:
            return jsonify({'ok': False, 'error': 'invoice_not_found'}), 404
        total = float(getattr(inv, 'total_after_tax_discount', 0.0) or 0.0)
        paid_before = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0))
                           .filter(Payment.invoice_type == 'sales', Payment.invoice_id == inv.id)
                           .scalar() or 0.0)
        remaining = max(total - paid_before, 0.0)
        pay_amount = amt if amt < remaining else remaining
        if pay_amount <= 0:
            try:
                inv.status = 'paid'
                db.session.commit()
            except Exception:
                try: db.session.rollback()
                except Exception: pass
            return jsonify({'ok': True, 'paid': paid_before, 'remaining': 0.0, 'status': getattr(inv,'status','paid')})
        try:
            db.session.add(Payment(invoice_id=inv.id, invoice_type='sales', amount_paid=pay_amount, payment_method=method, payment_date=get_saudi_now()))
            db.session.commit()
        except Exception:
            db.session.rollback()
            return jsonify({'ok': False, 'error': 'db_error'}), 500
        paid_after = paid_before + pay_amount
        try:
            if paid_after >= total and total > 0:
                inv.status = 'paid'
            elif paid_after > 0:
                inv.status = 'partial'
            else:
                inv.status = 'unpaid'
            db.session.commit()
        except Exception:
            try: db.session.rollback()
            except Exception: pass
        try:
            cash_acc = _pm_account(method)
            _post_ledger(inv.date, 'AR', 'Accounts Receivable', 'asset', 0.0, pay_amount, f'PAY SALE {inv.invoice_number}')
            if cash_acc:
                _post_ledger(inv.date, cash_acc.code, cash_acc.name, 'asset', pay_amount, 0.0, f'PAY SALE {inv.invoice_number}')
        except Exception:
            pass
        return jsonify({'ok': True, 'invoice_id': inv.id, 'paid': paid_after, 'remaining': max(total - paid_after, 0.0), 'status': getattr(inv,'status','paid')})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500
