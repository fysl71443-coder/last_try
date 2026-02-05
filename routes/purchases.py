# Phase 2 – Purchases blueprint (purchases, raw materials, meals). Same URLs.
from __future__ import annotations

import json
import os
from datetime import datetime
from decimal import Decimal

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_babel import gettext as _
from flask_login import login_required, current_user

from app import db
from models import (
    PurchaseInvoice,
    PurchaseInvoiceItem,
    RawMaterial,
    Supplier,
    Meal,
    MealIngredient,
    MenuCategory,
    MenuItem,
    Payment,
    get_saudi_now,
)
from forms import PurchaseInvoiceForm, MealForm, RawMaterialForm
from app.routes import (
    warmup_db_once,
    _pm_account,
    _create_purchase_journal,
    _create_supplier_payment_journal,
    CHART_OF_ACCOUNTS,
    ext_db,
)

bp = Blueprint("purchases", __name__)


def _project_root():
    return os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

@bp.route('/meals/import', methods=['POST'], endpoint='meals_import')
@login_required
def meals_import():
    import os, csv
    from io import TextIOWrapper
    file = request.files.get('file')
    if not file or not file.filename:
        flash(_('لم يتم اختيار ملف'), 'warning')
        return redirect(url_for('main.menu'))

    # إن وُجد قسم حالي مرسل من النموذج نعيد التوجيه إليه بعد الاستيراد
    cat_id = request.form.get('cat_id', type=int)

    ext = os.path.splitext(file.filename)[1].lower()

    def upsert_meal(name_en, name_ar, price_val):
        name_en = (name_en or '').strip()
        name_ar = (name_ar or '').strip() or None
        try:
            price = float(str(price_val).replace(',', '').strip()) if price_val is not None else 0.0
        except Exception:
            price = 0.0
        # محاولة إيجاد وجبة موجودة بنفس الاسم (و/أو الاسم العربي)
        existing = None
        if name_en and name_ar:
            existing = Meal.query.filter_by(name=name_en, name_ar=name_ar).first()
        if not existing and name_en:
            existing = Meal.query.filter_by(name=name_en).first()
        if existing:
            # تأكد من وجود user_id، ثم حدّث سعر البيع فقط
            try:
                if getattr(existing, 'user_id', None) is None:
                    existing.user_id = current_user.id
            except Exception:
                pass
            try:
                existing.selling_price = price
            except Exception:
                pass
            return existing
        # إنشاء وجبة جديدة مع تعيين user_id
        m = Meal(name=name_en or 'Unnamed', name_ar=name_ar, description=None, category=None, user_id=current_user.id)
        try:
            m.selling_price = price
        except Exception:
            pass
        db.session.add(m)
        return m

    def add_meal_to_section(meal, category_id):
        """إضافة الوجبة كعنصر في القسم المختار (إن لم تكن موجودة)."""
        if not meal or not category_id:
            return
        try:
            existing = MenuItem.query.filter_by(category_id=category_id, meal_id=meal.id).first()
            if existing:
                try:
                    existing.name = (meal.name_ar and f"{meal.name} / {meal.name_ar}") or meal.name
                    existing.price = float(meal.selling_price or 0)
                except Exception:
                    pass
                return
            display_name = (meal.name_ar and f"{meal.name} / {meal.name_ar}") or meal.name
            price_val = float(meal.selling_price or 0)
            item = MenuItem(
                category_id=category_id,
                meal_id=meal.id,
                name=display_name,
                price=price_val,
            )
            db.session.add(item)
        except Exception:
            pass

    imported, errors = 0, 0

    if ext == '.csv':
        try:
            stream = TextIOWrapper(file.stream, encoding='utf-8')
            reader = csv.DictReader(stream)
            def norm(s):
                return (s or '').strip().lower()
            for row in reader:
                cols = {norm(k): v for k, v in row.items()}
                name = cols.get('name') or cols.get('اسم') or cols.get('product')
                name_ar = cols.get('name (arabic)') or cols.get('name_ar') or cols.get('arabic') or cols.get('الاسم العربي')
                price = cols.get('selling price') or cols.get('price') or cols.get('السعر')
                if not name and not name_ar:
                    continue
                try:
                    meal = upsert_meal(name, name_ar, price)
                    db.session.flush()
                    if cat_id and meal:
                        add_meal_to_section(meal, cat_id)
                    imported += 1
                except Exception:
                    errors += 1
            db.session.commit()
            flash(f'تم استيراد {imported} وجبة بنجاح' + (f'، أخطاء: {errors}' if errors else ''), 'success')
        except Exception as e:
            db.session.rollback()
            flash(_('فشل استيراد CSV: %(error)s', error=e), 'danger')
        return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))

    elif ext in ('.xlsx', '.xls'):
        try:
            try:
                import openpyxl
            except Exception:
                flash('لا يمكن قراءة ملفات Excel بدون تثبيت openpyxl. يمكنك رفع CSV بدلاً من ذلك أو اسمح لي بتثبيت openpyxl.', 'warning')
                return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
                h = (cell or '').strip().lower().replace('  ', ' ')
                headers.append(h)
            def norm_key(k):
                return (k or '').strip().lower().replace('  ', ' ')
            index = {norm_key(h): i for i, h in enumerate(headers)}
            def get_val(row, keys):
                row_list = list(row) if row else []
                for k in keys:
                    nk = norm_key(k)
                    if nk in index and index[nk] < len(row_list):
                        return row_list[index[nk]]
                return None
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = get_val(row, ['name', 'اسم', 'product'])
                name_ar = get_val(row, ['name (arabic)', 'name (arabic)', 'name_ar', 'arabic', 'الاسم العربي'])
                price = get_val(row, ['selling price', 'selling price', 'price', 'السعر'])
                if not name and not name_ar:
                    continue
                try:
                    meal = upsert_meal(name, name_ar, price)
                    db.session.flush()
                    if cat_id and meal:
                        add_meal_to_section(meal, cat_id)
                    imported += 1
                except Exception:
                    errors += 1
            db.session.commit()
            flash(_('تم استيراد %(n)s وجبة من Excel', n=imported) + (f'، أخطاء: {errors}' if errors else ''), 'success')
        except Exception as e:
            db.session.rollback()
            flash(_('فشل استيراد Excel: %(error)s', error=e), 'danger')
        return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))

    else:
        flash(_('صيغة الملف غير مدعومة. الرجاء رفع ملف CSV أو Excel (.xlsx/.xls).'), 'warning')
        return redirect(url_for('main.menu', cat_id=cat_id) if cat_id else url_for('main.menu'))


@bp.route('/meals', methods=['GET', 'POST'], endpoint='meals')
@login_required
def meals():
    # Build form and context similar to monolith, but within blueprint
    raw_materials = RawMaterial.query.filter_by(active=True).all()
    material_choices = [(m.id, m.display_name) for m in raw_materials]

    form = MealForm()
    # Ensure ingredient subforms have choices
    for ingredient_form in form.ingredients:
        ingredient_form.raw_material_id.choices = material_choices

    materials_json = json.dumps([
        {
            'id': m.id,
            'name': m.display_name,
            'cost_per_unit': float(m.cost_per_unit),
            'unit': m.unit,
        }
        for m in raw_materials
    ])

    if form.validate_on_submit():
        try:
            from decimal import Decimal
            meal = Meal(
                name=form.name.data,
                name_ar=form.name_ar.data,
                description=form.description.data,
                category=form.category.data,
                profit_margin_percent=form.profit_margin_percent.data,
                user_id=current_user.id,
            )
            db.session.add(meal)
            db.session.flush()  # get meal.id

            total_cost = 0
            # Parse dynamic ingredient rows from POST keys
            idxs = set()
            for k in request.form.keys():
                if k.startswith('ingredients-') and k.endswith('-raw_material_id'):
                    try:
                        idxs.add(int(k.split('-')[1]))
                    except Exception:
                        pass
            for i in sorted(idxs):
                try:
                    rm_id = int(request.form.get(f'ingredients-{i}-raw_material_id') or 0)
                    qty_raw = request.form.get(f'ingredients-{i}-quantity')
                    qty = Decimal(qty_raw) if qty_raw not in (None, '') else Decimal('0')
                except Exception:
                    rm_id, qty = 0, Decimal('0')
                if rm_id and qty > 0:
                    raw_material = RawMaterial.query.get(rm_id)
                    if raw_material:
                        ing = MealIngredient(meal_id=meal.id, raw_material_id=raw_material.id, quantity=qty)
                        try:
                            ing_cost = qty * raw_material.cost_per_unit
                            ing.total_cost = ing_cost
                        except Exception:
                            ing.total_cost = 0
                        db.session.add(ing)
                        total_cost += float(ing.total_cost)

            meal.total_cost = total_cost
            meal.calculate_selling_price()
            db.session.commit()
            flash('تم إنشاء الوجبة بنجاح', 'success')
            return redirect(url_for('purchases.meals'))
        except Exception as e:
            db.session.rollback()
            flash(_('فشل حفظ الوجبة'), 'danger')

    all_meals = Meal.query.filter_by(active=True).all()
    return render_template('meals.html', form=form, meals=all_meals, materials_json=materials_json)

# -------- Meals import (Excel/CSV): Name, Name (Arabic), Selling Price --------

@bp.route('/api/raw_materials/categories', methods=['GET'], endpoint='api_raw_materials_categories')
@login_required
def api_raw_materials_categories():
    try:
        cats = db.session.query(RawMaterial.category).filter(RawMaterial.active == True).distinct().all()
        items = [
            {
                'name': (c[0] or ''),
                'label': (c[0] or '')
            }
            for c in cats if c and c[0]
        ]
        return jsonify({'ok': True, 'categories': items})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/raw_materials', methods=['GET'], endpoint='api_raw_materials')
@login_required
def api_raw_materials():
    try:
        cat = (request.args.get('category') or '').strip()
        mats_q = RawMaterial.query.filter_by(active=True)
        if cat:
            mats_q = mats_q.filter(RawMaterial.category == cat)
        mats = mats_q.order_by(RawMaterial.name.asc()).all()
        data = [
            {
                'id': int(m.id),
                'name': (m.display_name if hasattr(m, 'display_name') else m.name),
                'name_ar': (getattr(m, 'name_ar', '') or ''),
                'unit': (m.unit or ''),
                'stock_quantity': float(getattr(m, 'stock_quantity', 0.0) or 0.0),
                'cost_per_unit': float(getattr(m, 'cost_per_unit', 0.0) or 0.0),
                'category': (m.category or '')
            }
            for m in mats
        ]
        return jsonify({'ok': True, 'items': data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/raw-materials', methods=['GET', 'POST'], endpoint='raw_materials')
@login_required
def raw_materials():
    form = RawMaterialForm()
    mode = (request.args.get('mode') or '').strip().lower()
    if request.method == 'POST' and (request.form.get('action') or '') == 'bulk_update_quantities':
        try:
            mats = RawMaterial.query.filter_by(active=True).all()
            for m in mats:
                qv = request.form.get(f'qty_{int(m.id)}')
                cv = request.form.get(f'cost_{int(m.id)}')
                try:
                    q = float(qv or 0)
                except Exception:
                    q = 0.0
                try:
                    c = float(cv or 0)
                except Exception:
                    c = 0.0
                if q > 0:
                    m.stock_quantity = float(m.stock_quantity or 0.0) + q
                if c > 0:
                    m.cost_per_unit = c
            db.session.commit()
            flash(_('تم تحديث الكميات بنجاح'), 'success')
        except Exception:
            db.session.rollback()
            flash('فشل تحديث الكميات', 'danger')
        return redirect(url_for('purchases.raw_materials', mode='quantities'))
    if request.method == 'POST' and form.validate_on_submit():
        try:
            rm = RawMaterial(
                name=form.name.data,
                name_ar=form.name_ar.data,
                unit=form.unit.data,
                cost_per_unit=form.cost_per_unit.data,
                category=form.category.data,
            )
            db.session.add(rm)
            db.session.commit()
            flash(_('تم حفظ المادة الخام بنجاح'), 'success')
            return redirect(url_for('purchases.raw_materials'))
        except Exception as e:
            db.session.rollback()
            flash(_('فشل حفظ المادة الخام'), 'danger')
    materials = RawMaterial.query.filter_by(active=True).all()
    return render_template('raw_materials.html', form=form, materials=materials, mode=mode)


@bp.route('/purchases', methods=['GET','POST'], endpoint='purchases')
@login_required
def purchases():
    warmup_db_once()
    # Prepare form and lists
    form = PurchaseInvoiceForm()
    try:
        if not form.date.data:
            form.date.data = get_saudi_now().date()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        pass
    suppliers = []
    try:
                    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    except Exception:
        suppliers = []
    supplier_balances = {}
    try:
        from sqlalchemy import func
        for s in suppliers:
            total_inv = db.session.query(func.coalesce(func.sum(PurchaseInvoice.total_after_tax_discount), 0)).filter(
                PurchaseInvoice.supplier_id == s.id
            ).scalar() or 0
            inv_ids = [r[0] for r in db.session.query(PurchaseInvoice.id).filter(PurchaseInvoice.supplier_id == s.id).all()]
            total_paid = 0
            if inv_ids:
                total_paid = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0)).filter(
                    Payment.invoice_type == 'purchase', Payment.invoice_id.in_(inv_ids)
                ).scalar() or 0)
            supplier_balances[s.id] = float(total_inv) - total_paid
    except Exception:
        pass
    suppliers_json = []
    for s in suppliers:
        cr = getattr(s, 'cr_number', None)
        iban = getattr(s, 'iban', None)
        if (not cr or not iban):
            try:
                notes = (getattr(s, 'notes', '') or '')
                if notes:
                    for part in notes.split('|'):
                        p = (part or '').strip()
                        if (not cr) and p[:3].upper() == 'CR:':
                            cr = p[3:].strip()
                        if (not iban) and p[:5].upper() == 'IBAN:':
                            iban = p[5:].strip()
            except Exception:
                pass
        suppliers_json.append({
            'id': s.id,
            'name': s.name,
            'phone': getattr(s, 'phone', None),
            'tax_number': getattr(s, 'tax_number', None),
            'address': getattr(s, 'address', None),
            'cr_number': cr,
            'iban': iban,
            'active': getattr(s, 'active', True),
            'balance': supplier_balances.get(s.id, 0),
        })
    # Ensure RawMaterial table contains bilingual items from purchase_categories
    try:
        base_dir = _project_root()
        data_path = os.path.join(base_dir, 'data', 'purchase_categories.json')
        if os.path.exists(data_path):
            with open(data_path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            pcs = payload.get('purchase_categories') or []
            existing = RawMaterial.query.all()
            # Build lookup by name and by name_ar (lowercased)
            name_lut = { (getattr(m,'name','') or '').strip().lower(): m for m in existing }
            ar_lut = { (getattr(m,'name_ar','') or '').strip().lower(): m for m in existing }
            def bi(sen, sar):
                sen = (sen or '').strip()
                sar = (sar or '').strip()
                if sen and sar:
                    return f"{sen} / {sar}"
                return sen or sar
            def _map_unit(nm):
                s = (nm or '').strip().lower()
                if any(k in s for k in ['meat','beef','lamb','chicken','fish','seafood','لحم','بقر','غنم','دجاج','سمك','بحرية']):
                    return 'kg'
                if any(k in s for k in ['vegetable','veg','greens','herb','خضار','أعشاب']):
                    return 'kg'
                if any(k in s for k in ['spice','spices','powder','tea','coffee','بهار','بهارات','بودرة','شاي','قهوة']):
                    return 'gram'
                if any(k in s for k in ['oil','sauce','liquid','milk','drink','juice','زيت','صلصة','سائل','حليب','مشروب','عصير']):
                    return 'liter'
                if any(k in s for k in ['rice','grain','pulses','lentil','flour','sugar','salt','أرز','حبوب','بقول','عدس','طحين','سكر','ملح']):
                    return 'kg'
                return 'piece'
            for cat in pcs:
                subs = cat.get('subcategories') or []
                for sub in subs:
                    nm = sub.get('name') or {}
                    items = sub.get('items') or []
                    for it in items:
                        if isinstance(it, dict):
                            en = (it.get('en') or '').strip()
                            ar = (it.get('ar') or '').strip()
                        else:
                            # if string, use as both
                            en = str(it).strip()
                            ar = str(it).strip()
                        disp = bi(en, ar)
                        key_en = disp.strip().lower()
                        key_ar = ar.strip().lower()
                        rm = name_lut.get(key_en) or ar_lut.get(key_ar)
                        if not rm:
                            try:
                                cat_name = bi((nm.get('en') if isinstance(nm, dict) else str(nm)), (nm.get('ar') if isinstance(nm, dict) else str(nm)))
                                unit = _map_unit(cat_name)
                                rm = RawMaterial(
                                    name=disp,
                                    name_ar=ar or None,
                                    unit=unit,
                                    cost_per_unit=0,
                                    category=cat_name,
                                )
                                db.session.add(rm)
                                db.session.flush()
                                name_lut[key_en] = rm
                                if key_ar:
                                    ar_lut[key_ar] = rm
                            except Exception:
                                try:
                                    db.session.rollback()
                                except Exception:
                                    pass
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
    except Exception:
        pass

    try:
        materials = RawMaterial.query.filter_by(active=True).order_by(RawMaterial.name.asc()).all()
    except Exception:
        materials = []
    materials_json = []
    for m in materials:
        name = m.name
        materials_json.append({
            'id': m.id,
            'name': name,
            'unit': m.unit,
            'cost_per_unit': float(m.cost_per_unit or 0),
            'stock_quantity': float(m.stock_quantity or 0),
            'category': getattr(m, 'category', None) or '',
        })

    if request.method == 'POST':
        pm = (request.form.get('payment_method') or 'cash').strip().lower()
        date_str = request.form.get('date') or get_saudi_now().date().isoformat()
        inv_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        from services.gl_truth import can_create_invoice_on_date
        ok, period_err = can_create_invoice_on_date(inv_date)
        if not ok:
            flash(period_err or _('الفترة المالية مغلقة لهذا التاريخ.'), 'danger')
            return redirect(url_for('purchases.purchases'))
        inv_type = (request.form.get('invoice_type') or 'VAT').strip().upper()
        supplier_name = (request.form.get('supplier_name') or '').strip() or None
        supplier_free = (request.form.get('supplier_name_free') or '').strip()
        if inv_type == 'NO_VAT' and supplier_free:
            supplier_name = supplier_free
        supplier_id = request.form.get('supplier_id', type=int)
        supplier_invoice_number = (request.form.get('supplier_invoice_number') or '').strip()
        notes_val = (request.form.get('notes') or '').strip()
        try:
            inv = PurchaseInvoice(
                invoice_number=f"INV-PUR-{get_saudi_now().year}-{(PurchaseInvoice.query.count()+1):04d}",
                date=inv_date,
                supplier_name=supplier_name,
                supplier_id=supplier_id,
                payment_method=pm,
                user_id=getattr(current_user, 'id', 1)
            )
            try:
                inv.supplier_invoice_number = supplier_invoice_number
                inv.notes = notes_val
            except Exception:
                pass
            inv.status = (request.form.get('status') or (getattr(PurchaseInvoice, 'status', None) and 'unpaid') or 'unpaid').strip().lower()
            if inv.status not in ('paid','partial','unpaid'):
                inv.status = 'unpaid'
            # Initialize totals; will be computed from items below
            inv.total_before_tax = 0.0
            inv.tax_amount = 0.0
            inv.discount_amount = 0.0
            inv.total_after_tax_discount = 0.0
            if ext_db is not None:
                ext_db.session.add(inv); ext_db.session.commit()
            else:
                db.session.add(inv); db.session.commit()
        except Exception as e:
            if ext_db is not None:
                ext_db.session.rollback()
            else:
                db.session.rollback()
            flash(_('Could not save purchase invoice: %(error)s', error=e), 'danger')
            return redirect(url_for('purchases.purchases'))

        # Save posted purchase items and update inventory (by item_name + category; find or create RawMaterial)
        try:
            from decimal import Decimal
            def _to_ascii_digits(s):
                try:
                    if s is None:
                        return ''
                    s = str(s)
                    trans = str.maketrans('٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹', '01234567890123456789')
                    s = s.translate(trans)
                    s = s.replace('٫', '.').replace('،', '.')
                    return s
                except Exception:
                    return str(s or '')
            idxs = set()
            for k in request.form.keys():
                if k.startswith('items-') and k.endswith('-item_name'):
                    try:
                        idxs.add(int(k.split('-')[1]))
                    except Exception:
                        pass
            total_before_tax_dec = Decimal('0')
            total_tax_dec = Decimal('0')
            total_discount_dec = Decimal('0')
            items_saved = 0
            VAT_RATE = Decimal('0.15') if inv_type == 'VAT' else Decimal('0')
            for i in sorted(idxs):
                item_name = (request.form.get(f'items-{i}-item_name') or '').strip()
                category = (request.form.get(f'items-{i}-category') or '').strip()
                if not item_name:
                    continue
                qty_raw = _to_ascii_digits(request.form.get(f'items-{i}-quantity'))
                price_raw = _to_ascii_digits(request.form.get(f'items-{i}-price_before_tax'))
                disc_raw = _to_ascii_digits(request.form.get(f'items-{i}-discount'))
                try:
                    qty = Decimal(str(qty_raw)) if (qty_raw not in (None, '')) else Decimal('0')
                except Exception:
                    qty = Decimal('0')
                try:
                    unit_price = Decimal(str(price_raw)) if (price_raw not in (None, '')) else Decimal('0')
                except Exception:
                    unit_price = Decimal('0')
                try:
                    disc = Decimal(str(disc_raw)) if (disc_raw not in (None, '')) else Decimal('0')
                except Exception:
                    disc = Decimal('0')
                if qty <= 0:
                    continue
                q = RawMaterial.query.filter(RawMaterial.name == item_name)
                if category:
                    q = q.filter(RawMaterial.category == category)
                else:
                    q = q.filter((RawMaterial.category == None) | (RawMaterial.category == ''))
                raw_material = q.first()
                if not raw_material:
                    raw_material = RawMaterial(
                        name=item_name,
                        name_ar=item_name,
                        unit='piece',
                        cost_per_unit=0,
                        stock_quantity=0,
                        category=category or None,
                        active=True
                    )
                    db.session.add(raw_material)
                    db.session.flush()
                line_base = (qty * unit_price) - disc
                if line_base < 0:
                    line_base = Decimal('0')
                line_tax = (line_base * VAT_RATE) if VAT_RATE > 0 else Decimal('0')
                line_total = line_base + line_tax
                prev_qty = Decimal(str(raw_material.stock_quantity or 0))
                prev_cost = Decimal(str(raw_material.cost_per_unit or 0))
                new_total_qty = prev_qty + qty
                if new_total_qty > 0:
                    new_total_cost = (prev_cost * prev_qty) + (unit_price * qty)
                    raw_material.cost_per_unit = (new_total_cost / new_total_qty).quantize(Decimal('0.0001'))
                raw_material.stock_quantity = new_total_qty
                display_name = getattr(raw_material, 'display_name', raw_material.name)
                inv_item = PurchaseInvoiceItem(
                    invoice_id=inv.id,
                    raw_material_id=raw_material.id,
                    raw_material_name=display_name,
                    quantity=float(qty),
                    price_before_tax=float(unit_price),
                    tax=float(line_tax),
                    discount=float(disc),
                    total_price=float(line_total)
                )
                db.session.add(inv_item)
                total_before_tax_dec += line_base
                total_tax_dec += line_tax
                total_discount_dec += disc
                items_saved += 1
            inv.total_before_tax = float(total_before_tax_dec)
            inv.tax_amount = float(total_tax_dec)
            inv.discount_amount = float(total_discount_dec)
            inv.total_after_tax_discount = float(total_before_tax_dec + total_tax_dec - total_discount_dec)
            if items_saved > 0:
                try:
                    inv.notes = ((inv.notes or '') + f" | ITEMS:{items_saved}").strip(' |')
                except Exception:
                    pass
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(_('Failed to save purchase items: %(error)s', error=e), 'warning')
            return redirect(url_for('purchases.purchases'))

        # لا حفظ بدون قيد. إذا مدفوعة فوراً: قيد واحد من حـ مخزون إلى حـ الصندوق (لا ذمة مورد). وإلا ذمة مورد ثم قيد الدفع للجزء المدفوع.
        st = (getattr(inv, 'status', '') or '').lower()
        total_amt = float(inv.total_after_tax_discount or 0.0)
        if st == 'paid' and total_amt > 0:
            db.session.add(Payment(invoice_id=inv.id, invoice_type='purchase', amount_paid=total_amt, payment_method=(pm or 'CASH').upper()))
        elif st == 'partial':
            amt_raw = request.form.get('partial_paid_amount')
            amt = float(amt_raw or 0.0)
            if total_amt > 0 and amt > 0:
                if amt > total_amt:
                    amt = total_amt
                db.session.add(Payment(invoice_id=inv.id, invoice_type='purchase', amount_paid=amt, payment_method=(pm or 'CASH').upper()))
                try:
                    fee = round(amt * 0.02, 2) if (pm or '').strip().lower() == 'bank' else 0
                    _create_supplier_payment_journal(inv.date, amt, inv.invoice_number, pm or 'CASH', bank_fee=fee if fee > 0 else 0)
                except Exception:
                    pass
        # تحديث الحالة من إجمالي المدفوع فعلياً (مدفوع كامل = paid، جزء = partial، غير مدفوع = unpaid)
        if total_amt > 0:
            from sqlalchemy import func
            paid_sum = float(db.session.query(func.coalesce(func.sum(Payment.amount_paid), 0)).filter(
                Payment.invoice_type == 'purchase', Payment.invoice_id == inv.id
            ).scalar() or 0)
            if paid_sum >= total_amt - 0.01:
                inv.status = 'paid'
            elif paid_sum > 0:
                inv.status = 'partial'
            else:
                inv.status = 'unpaid'
        try:
            _create_purchase_journal(inv)
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(_('فشل حفظ فاتورة المشتريات: لم يُنشأ القيد. %(error)s', error=str(e)), 'danger')
            return redirect(url_for('purchases.purchases'))
        flash(_('Purchase invoice saved'), 'success')
        return redirect(url_for('purchases.purchases'))

    return render_template('purchases.html', form=form, suppliers_list=suppliers, suppliers_json=suppliers_json, materials_json=materials_json)


@bp.route('/api/purchase-categories', methods=['GET'])
def api_purchase_categories():
    try:
        base_dir = _project_root()
        data_path = os.path.join(base_dir, 'data', 'purchase_categories.json')
        with open(data_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        return jsonify(payload)
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'error': str(e), 'purchase_categories': []}), 500


